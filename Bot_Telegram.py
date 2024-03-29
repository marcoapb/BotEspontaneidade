"""
Created on Thu Jul 16 11:21:16 2020

@author: 53363833172
"""

from __future__ import unicode_literals
from datetime import datetime, timedelta
import re
from mysql.connector.cursor import RE_SQL_ON_DUPLICATE
import schedule
import time
from random import randint
import threading
import sys
import os
import logging   
import mysql.connector
from mysql.connector import errorcode
from telegram.ext import Updater
from telegram.ext import CommandHandler, Filters, MessageHandler #,CallbackQueryHandler
from telegram import ReplyKeyboardMarkup #,InlineKeyboardButton, InlineKeyboardMarkup, KeyboardButton 
from calendar import weekday

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.utils import formatdate
from email import encoders
import smtplib

from openpyxl import Workbook
#from openpyxl.styles import colors
from openpyxl.styles import Font #, Color
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
############################### Bot ############################################


#formata  o número do TDPF
def formataTDPF(tdpf): #passar uma string
    tdpf = tdpf.strip()
    if len(tdpf)<11: #tdpf deve ter 16 dígitos; testo se o tamanho é menor do que 11, pois referencio esta posição
        return tdpf
    return tdpf[:7]+"-"+tdpf[7:11]+"-"+tdpf[11:]
    
#verifica se um CPF é válido
def validaCPF(cpfPar):
#The MIT License (MIT) Copyright (c) 2015 Derek Willian Stavis
    cpf = getAlgarismos(cpfPar)
    if len(cpf)!=11:
        return False

    if cpf in [s * 11 for s in [str(n) for n in range(10)]]:
        return False

    calc = lambda t: int(t[1]) * (t[0] + 2)
    d1 = (sum(map(calc, enumerate(reversed(cpf[:-2])))) * 10) % 11
    d2 = (sum(map(calc, enumerate(reversed(cpf[:-1])))) * 10) % 11
    if d1==10:
        d1 = 0
    if d2==10:
        d2 = 0      
    return str(d1) == cpf[-2] and str(d2) == cpf[-1]

def limpaMarkdown(texto): #deixa o texto adequado para negritar com *
    return texto.replace(".", "\.").replace("_", "\_").replace("[", "\[").replace("]", "\]").replace(")", "\)").replace("(", "\(").replace("-","\-").replace("|","\|")#.replace("*", "\\*")        

def getAlgarismos(texto): #retorna apenas os algarismos de uma string
    algarismos = ""
    for car in texto:
        if car.isdigit():
            algarismos = algarismos + car
    return algarismos

def isDate(data): #verifica se a string é uma data válida
    if data==None:
        return False
    if len(data)!=10: #exige no formato dd/mm/aaaa
        return False
    try:    
        # faz o split e transforma em números
        dia, mes, ano = map(int, data.split('/'))
    except:
        return False 
    # mês ou ano inválido (só considera do ano 1 em diante), retorna False   
    if mes < 1 or mes > 12 or ano <= 0:
        return False
    # verifica qual o último dia do mês
    if mes in (1, 3, 5, 7, 8, 10, 12):
        ultimo_dia = 31
    elif mes == 2:
        # verifica se é ano bissexto
        if (ano % 4 == 0) and (ano % 100 != 0 or ano % 400 == 0):
            ultimo_dia = 29
        else:
            ultimo_dia = 28
    else:
        ultimo_dia = 30
    # verifica se o dia é válido
    if dia < 1 or dia > ultimo_dia:
        return False
    return True

def getParametros(msg): #monta uma lista com os parâmetros a partir da msg do usuário
    msg = msg.strip()
    parametros = msg.split()
    i = 0
    tamLista = len(parametros)
    while i<tamLista:
        parametros[i] = parametros[i].strip()
        if parametros[i]=="":
            for j in range(i+1, tamLista):
                parametros[j-1] = parametros[j]
            parametros.pop()
            tamLista-=1
        i+=1    
    return parametros

def eliminaPendencia(userId):
    global pendencias, atividadeTxt
    #apaga todas as pendencias (usuário retornou ao menu, acionou alguma opção dele ou prestou todas as informações requeridas por uma funcionalidade)
    #logging.info(pendencias)
    if pendencias==None:
        return
    while pendencias.get(userId, "1")!="1": #há pendências (retona "1" em caso de não haver)
        try:
            del pendencias[userId]
        except:
            logging.info("Erro ao apagar pendencia do usuário "+str(userId))
            break
               
    while atividadeTxt.get(userId, "1")!="1": #há pendência (retona "1" em caso de não haver) de informação do texto da atividade
        try:
            del atividadeTxt[userId]
        except:
            logging.info("Erro (3) ao apagar pendencia do usuário "+str(userId))
            break         
    return      

#transforma uma data string de dd/mm/yyyy para yyyy/mm/dd para fins de consulta, inclusão ou atualização no BD SQL
#se o BD esperar a data em outro formato, basta alterarmos aqui - caiu em desuso depois do mysql.connector
#def converteAMD(data):
#    return data[6:]+"/"+data[3:5]+"/"+data[:2] 

def enviaEmail(email, texto, assunto, arquivo=None): #envia email, conforme parâmetros - se passar o arquivo (caminho e nome), ele vai como anexo 
    try:
        #server = smtplib.SMTP('INETRFOC.RFOC.SRF: 25') #servidor de email Notes
        server = smtplib.SMTP('exchangerfoc.rfoc.srf: 25')
        #pass
    except:
        return 1	
    # create message object instance
    msg = MIMEMultipart()
    # setup the parameters of the message
    msg['From'] = "botespontaneidade@rfb.gov.br"
    msg['To'] = email
    msg['Subject'] = assunto
    # add in the message body
    msg.attach(MIMEText(texto, 'plain'))  
    if arquivo!=None and arquivo!="":  
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(arquivo, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename='+arquivo)
        msg.attach(part)                 
    # send the message via the server.
    try:
        server.sendmail(msg['From'], msg['To'], msg.as_string())
        pass
    except:
        return 2
    server.quit()  
    return 3 #sucesso

def verificaUsuarioTelegram(conn, userId): #retorna True, CPF e "" se está cadastrado e ativo; False CPF/vazio e mensagem se não está no serviço ou está inativo (saiu) 
    if conn==None:
        return False, "", ""
    if userId==None or userId==0:
        return False, "", "userId inválido"
    comando = "Select CPF, Adesao, Saida, Ativo from Usuarios Where idTelegram=%s"
    cursor = conn.cursor(buffered=True)
    cursor.execute(comando, (userId,))
    row = cursor.fetchone()
    if row:
        cpf = row[0]
        if row[1]==None: #adesão
            return False, cpf, "Usuário não se registrou no serviço"
        if row[2]!=None or row[3]!='S': #saída
            return False, cpf, "Usuário está INATIVO ou DESABILITADO no serviço - saída em "+row[2].strftime("%d/%m/%Y") 
        return True, cpf, ""
    else:
        return False, "", "Usuário não está na base de dados"    

def verificaMonitoramento(conn, cpf, tdpf): #verifica se o usuário está monitorando o TDPF; retorna False, vazio, vazio e mensagem se não, 
                                            #e True, chaveTdpf, chaveFiscal e vazio caso afirmativo
    cursor = conn.cursor(buffered=True)
    comando = """Select CadastroTDPFs.Codigo, CadastroTDPFs.Fim, CadastroTDPFs.TDPF, Fiscais.Codigo from CadastroTDPFs, Fiscais, TDPFS 
                 Where Fiscais.CPF=%s and Fiscais.Codigo=CadastroTDPFs.Fiscal and CadastroTDPFs.TDPF=TDPFS.Codigo and TDPFS.Numero=%s"""
    cursor.execute(comando, (cpf, tdpf))
    row = cursor.fetchone()
    if not row:
        return False, None, None, "TDPF não está sendo monitorado pelo usuário."
    if row[1]!=None:
        return False, None, None, "O monitoramento do TDPF foi finalizado em "+row[1].strftime('%d/%m/%Y')+"."
    chaveTdpf = row[2]  
    chaveFiscal = row[3]  
    return True, chaveTdpf, chaveFiscal, ""    

def verificaAlocacao(conn, cpf, tdpf): #verifica se o usuário está alocado ao TDPF e se o TDPF está em andamento; 
                                       #retorna False, vazio, vazio e mensagem se não (ou TDPF inexistente), e True, chaveTdpf, chaveFiscal e vazio caso afirmativo
    cursor = conn.cursor(buffered=True)
    comando = """Select TDPFS.Numero, TDPFS.Codigo, Fiscais.Codigo from Alocacoes, TDPFS, Fiscais 
                 Where Fiscais.CPF=%s and Fiscais.Codigo=Alocacoes.Fiscal and TDPFS.Numero=%s and Alocacoes.TDPF=TDPFS.Codigo and (TDPFS.Tipo='F' or TDPFS.Tipo='D')
                 and TDPFS.Encerramento Is Null and Alocacoes.Desalocacao Is Null"""
    cursor.execute(comando, (cpf, tdpf))        
    row = cursor.fetchone()
    if not row:
        return False, None, None, "TDPF inexistente/encerrado, não é de fiscalização ou de diligência ou usuário não está alocado a ele ou foi desalocado."
    chaveTdpf = row[1]    
    chaveFiscal = row[2]
    return True, chaveTdpf, chaveFiscal, ""

def enviaMsgBot(bot, userId, text, parse_mode=None): #função intermediária para tratar erros
    try:
        bot.send_message(userId, text=text, parse_mode= parse_mode)
    except Exception as e:
        logging.info("Ocorreu um erro (1) ao enviar a mensagem para o userId "+str(userId)+" - "+str(e))
    return

def enviaMsgUpdater(userId, text, parse_mode=None): #função intermediária para tratar erros
    global updater
    try:
        updater.bot.send_message(userId, text, parse_mode=parse_mode)
    except Exception as e:
        logging.info("Ocorreu um erro (2) ao enviar a mensagem para o userId "+str(userId)+" - "+str(e))
    return        

def registra(update, context): #registra usuário no serviço
    global pendencias, textoRetorno, d1padrao, d2padrao, d3padrao #, cpfRegistro   
    response_message = ""
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=2:
        response_message = "Número de informações (parâmetros) inválido. Envie somente o CPF e o código de registro (chave)."
        response_message = response_message+textoRetorno
    else:
        cpf = getAlgarismos(parametros[0])
        chave = getAlgarismos(parametros[1])                     
        logging.info(cpf + " "+chave)
        if not validaCPF(cpf):
            response_message = "CPF inválido. Envie novamente CPF e o código de registro (chave)."
            response_message = response_message+textoRetorno
        elif not chave.isdigit() or chave==None or chave=="":
            response_message = "Código de registro inválido. Envie novamente CPF e o código de registro (chave)."
            response_message = response_message+textoRetorno 
        else:
            try:
                chave = int(chave)                  
            except:
                response_message = "Código de registro inválido. Envie novamente CPF e o código de registro (chave)."
                response_message = response_message+textoRetorno          
        if response_message!="":
            enviaMsgBot(bot, userId, text=response_message)
            return 
        conn = conecta()
        if not conn:   
            enviaMsgBot(bot, userId, text="Erro na conexão - registra")  
            eliminaPendencia(userId)
            mostraMenuPrincipal(update, context)                    
            return
        cursor = conn.cursor(buffered=True)
        cursor.execute("Select Codigo, CPF, Chave, Adesao, ValidadeChave, Tentativas from Usuarios where Ativo='S' and CPF=%s", (cpf,))  
        row = cursor.fetchone()
        if row:
            codigo = row[0]
            validade = row[4]
            tentativas = row[5]   
            if tentativas==None:
                tentativas = 0   
            if validade==None:
                validade = datetime.strptime("31/12/2050", "%d/%m/%Y")                    
            if validade.date()<datetime.today().date():
                eliminaPendencia(userId)                
                enviaMsgBot(bot, userId, text="Validade da chave está expirada.") 
                mostraMenuPrincipal(update, context)  
                conn.close()  
                return   
            if tentativas>3:
                eliminaPendencia(userId)                               
                enviaMsgBot(bot, userId, text="Número de tentativas excedida.") 
                mostraMenuPrincipal(update, context)  
                conn.close()  
                return                                   
            if row[2]==chave and chave>=100000: #chave tem que ser igual à registrada e ter 6 ou mais dígitos (atualmente, 6)
                eliminaPendencia(userId) #apaga a pendência de informação do usuário                
                try:
                    if row[3]==None: #usuário é novo no serviço (nunca havia se registrado)
                        comando =  "Update Usuarios set idTelegram=%s, Adesao=%s, Saida=Null, Tentativas=0, d1=%s, d2=%s, d3=%s where Codigo=%s"
                        cursor.execute(comando, (userId, datetime.now().date(), d1padrao, d2padrao, d3padrao, codigo))
                    else:   
                        comando =  "Update Usuarios set idTelegram=%s, Saida=Null, Tentativas=0 where Codigo=%s"
                        cursor.execute(comando, (userId, codigo))
                    conn.commit()
                    response_message = "Usuário registrado/reativado com sucesso."                     
                except:
                    conn.rollback()
                    response_message = "Erro ao registrar o usuário no serviço. Tente novamente mais tarde. Se o erro persistir, contacte o suporte."  
                enviaMsgBot(bot, userId, text=response_message) 
                mostraMenuPrincipal(update, context)  
                conn.close()  
                return
            else:
                try:
                    comando =  "Update Usuarios set Tentativas=%s Where Codigo=%s"
                    cursor.execute(comando, ((tentativas+1), codigo))   
                    conn.commit()  
                except:
                    conn.rollback()
                    logging.info('Erro na atualização de tentativas - registro '+cpf)
                conn.close()
                response_message = "Gere a chave primeiramente ou digite-a corretamente. Digite novamente o CPF e o código de registro." +textoRetorno                                   
        else:
            eliminaPendencia(userId) #apaga a pendência de informação do usuário            
            response_message = "Usuário (CPF) não foi cadastrado para registro no serviço ou está desabilitado."
            enviaMsgBot(bot, userId, text=response_message) 
            mostraMenuPrincipal(update, context)   
            conn.close() 
            return  
    print(response_message)          
    enviaMsgBot(bot, userId, text=response_message)
    return

def envioChave(update, context): #envia a chave de registro para o e-mail do usuário
    global pendencias, textoRetorno, ambiente 
    response_message = ""
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=1:
        response_message = "Número de informações (parâmetros) inválido. Envie somente o CPF do usuário."
        response_message = response_message+textoRetorno
        enviaMsgBot(bot, userId, text=response_message)        
    else:
        cpf = getAlgarismos(parametros[0])                   
        logging.info('Envio chave '+cpf)
        if not validaCPF(cpf):
            response_message = "CPF inválido. Envie novamente o CPF do usuário."
            response_message = response_message+textoRetorno
            enviaMsgBot(bot, userId, text=response_message)
            return 
        eliminaPendencia(userId)             
        conn = conecta()                       
        if not conn:   
            enviaMsgBot(bot, userId, text="Erro na conexão - envioChave")  
            mostraMenuPrincipal(update, context)                    
            return
        cursor = conn.cursor(buffered=True)  
        cursor.execute("Select Codigo, CPF, email, DataEnvio from Usuarios where Ativo='S'and CPF=%s", (cpf,))  
        row = cursor.fetchone()
        if row:
            codigo = row[0]
            email = row[2]
            erro = False
            if email==None or email=='':
                enviaMsgBot(bot, userId, text="CPF não tem e-mail associado na base de dados. Contacte o suporte ou cadastre um na opção 'Cadastros -> Cadastra/Exclui e-Mail' se você já estiver registrado no serviço.")  
                erro = True             
            elif not verificaEMail(email):
                enviaMsgBot(bot, userId, text="O e-mail do usuário é inválido. Contacte o suporte ou exclua o atual e cadastre um novo na opção 'Cadastros -> Cadastra/Exclui e-Mail' se você já estiver registrado no serviço.")
                erro = True
            elif not ("@rfb.gov.br" in email):
                enviaMsgBot(bot, userId, text="O e-mail do usuário não é institucional. Contacte o suporte ou exclua o atual e cadastre um novo na opção 'Cadastros -> Cadastra/Exclui e-Mail' se você já estiver registrado no serviço.")
                erro = True
            if erro:    
                conn.close()                 
                mostraMenuPrincipal(update, context)                    
                return                 
            dataEnvio = row[3]
            if dataEnvio==None:
                dataEnvio = datetime.strptime("01/01/1900", "%d/%m/%Y")
            dataEnvio = dataEnvio.date()    
            if dataEnvio<datetime.today().date():
                chave = randint(100000, 1000000) #a chave é um número inteiro de seis dígitos
                message = "Prezado(a),\n\nSua chave SIGILOSA de registro no Bot Espontaneidade (Telegram) é "+str(chave)+"\n\nEsta chave é utilizada também para acesso via ContÁgil no Script AlertasFiscalização e tem validade de 30 dias.\n\nAtenciosamente,\n\nDisav/Cofis\n\nAmbiente: "+ambiente 
                assunto = "Chave de Registro - Bot Espontaneidade"
                sucesso = enviaEmail(email, message, assunto)
                if sucesso==3:
                    comando = "Update Usuarios Set Chave=%s, ValidadeChave=%s, Tentativas=%s, DataEnvio=%s Where Codigo=%s"
                    validade = datetime.today().date()+timedelta(days=30) #a chave tem validade de 30 dias
                    try:
                        cursor.execute(comando, (chave, validade, 0, datetime.today().date(), codigo))
                        conn.commit()
                        enviaMsgBot(bot, userId, text="Chave foi enviada para o e-mail institucional do usuário.")                         
                    except:
                        conn.rollback()    
                        enviaMsgBot(bot, userId, text="Erro ao inserir a chave na tabela. A chave enviada não será reconhecida. Tente novamente mais tarde. Se o erro persistir, contacte o suporte.")  
                else:
                    enviaMsgBot(bot, userId, text="Houve erro no envio do e-mail. Tente novamente mais tarde. Se o erro persistir, contacte o suporte. "+str(sucesso))                                            
            else:
                enviaMsgBot(bot, userId, text="A chave já foi enviada hoje - é vedado o reenvio no mesmo dia.")       
        else:
            enviaMsgBot(bot, userId, text="CPF do usuário não foi encontrado na base de dados ou está desabilitado.")      
        mostraMenuPrincipal(update, context) 
        conn.close()    
    return
                

def acompanha(update, context): #inicia o monitoramente de um ou de TODOS os TDPFs que o usuário esteja alocado
    global pendencias, textoRetorno
    bot = update.effective_user.bot
    userId = update.message.from_user.id      
    conn = conecta()
    if not conn: 
        response_message = "Erro na conexão - acompanha."
        enviaMsgBot(bot, userId, text=response_message)
        eliminaPendencia(userId)
        mostraMenuPrincipal(update, context)
        return    
    cursor = conn.cursor(buffered=True)     
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=1:
        response_message = "Número de informações (parâmetros) inválido. Envie somente o nº do TDPF ou a palavra TODOS (incluirá apenas fiscalizações)."
        response_message = response_message+textoRetorno
        enviaMsgBot(bot, userId, text=response_message)
        conn.close()
        return        
    else:  
        tdpfs = None
        bAlocacao = False
        bUsuario, cpf, msg = verificaUsuarioTelegram(conn, userId)
        if not bUsuario:
            response_message = msg
            enviaMsgBot(bot, userId, text=response_message) 
            mostraMenuPrincipal(update, context)
            eliminaPendencia(userId) #apaga a pendência de informação do usuário
            conn.close()
            return                          
        info = parametros[0]
        if info.upper().strip() in ["TODOS", "TODAS"]: #se todos, apenas do tipo fiscalização
            comando = """Select TDPFS.Numero, TDPFS.Codigo, Fiscais.Codigo from Alocacoes, TDPFS, Fiscais 
                         Where Alocacoes.Desalocacao Is Null and Fiscais.CPF=%s and Alocacoes.Fiscal=Fiscais.Codigo and TDPFS.Tipo='F' and 
                         Alocacoes.TDPF=TDPFS.Codigo and TDPFS.Encerramento Is Null"""
            cursor.execute(comando, (cpf,))
            tdpfs = cursor.fetchall()
            if not tdpfs:
                response_message = "Não há TDPFs ativos em que o CPF {} esteja alocado.".format(cpf) #tb já foi testado                        
        else:
            tdpf = getAlgarismos(info.strip())
            if len(tdpf)!=16 or not tdpf.isdigit():
                response_message = "TDPF ou comando inválido. Envie novamente o nº do TDPF (16 dígitos) ou a palavra TODOS."
                response_message = response_message+textoRetorno  
                enviaMsgBot(bot, userId, text=response_message) 
                conn.close()
                return
            bAlocacao, chaveTdpf, chaveFiscal, msg = verificaAlocacao(conn, cpf, tdpf)                 
            if not bAlocacao:
                response_message = msg
            else:
                tdpfs = ([tdpf, chaveTdpf, chaveFiscal],)
        if tdpfs==None:
            eliminaPendencia(userId) #apaga a pendência de informação do usuário            
            enviaMsgBot(bot, userId, text=response_message) 
            mostraMenuPrincipal(update, context)
            conn.close()
            return
        atualizou = False 
        try:
            for tdpfObj in tdpfs:
                chaveTdpf = tdpfObj[1]
                chaveFiscal = tdpfObj[2]
                comando = """Select CadastroTDPFs.Codigo, CadastroTDPFs.Fim from CadastroTDPFs
                                Where CadastroTDPFs.Fiscal=%s and CadastroTDPFs.TDPF=%s"""
                cursor.execute(comando, (chaveFiscal, chaveTdpf))
                row = cursor.fetchone()
                if row:
                    if row[1]!=None:
                        comando = "Update CadastroTDPFs Set Fim=Null Where Codigo=%s"
                        cursor.execute(comando, (row[0],))
                        atualizou = True
                else:
                    comando = "Insert into CadastroTDPFs (Fiscal, TDPF, Inicio) Values (%s, %s, %s)"
                    cursor.execute(comando, (chaveFiscal, chaveTdpf, datetime.today().date()))
                    atualizou = True
            if atualizou:        
                conn.commit()
                response_message = "Operação efetivada com sucesso (houve atualização)."
            else:
                response_message = "Operação efetivada com sucesso (não houve atualização)."
        except:
            if atualizou:
                conn.rollback()
                response_message="Erro (1) ao tentar efetivar a operação. Tente novamente mais tarde."
            else:
                response_message="Erro (2) ao tentar efetivar a operação. Tente novamente mais tarde."                
        enviaMsgBot(bot, userId, text=response_message) 
        eliminaPendencia(userId) #apaga a pendência de informação do usuário        
        mostraMenuPrincipal(update, context)
        conn.close()
        return  

def efetivaAtividade(userId, tdpf, atividade, vencimento, inicio): #tenta efetivar uma atividade para um certo tdpf no BD
    conn = conecta()
    if not conn:         
        return False, "Erro na conexão - efetivaAtividade"
    cursor = conn.cursor(buffered=True)  
    try:
        bUsuario, cpf, msg = verificaUsuarioTelegram(conn, userId)
        if not bUsuario:
            conn.close()
            return False, msg #tb já foi testado  
        bAlocacao, chaveTdpf, chaveFiscal, msg = verificaAlocacao(conn, cpf, tdpf) 
        if not bAlocacao:
            conn.close()
            return False, msg         
        comando = "Select Codigo, Fim from CadastroTDPFs Where TDPF=%s and CadastroTDPFs.Fiscal=%s"
        cursor.execute(comando, (chaveTdpf, chaveFiscal))
        row = cursor.fetchone()        
        tdpfCadastrado = False
        fim = None
        if row:
            tdpfCadastrado = True  
            chaveCad = row[0]
            fim = row[1]
            #logging.info("TDPF cadastrado")          
    except:
        conn.close()
        return False, "Erro na consulta (efetivaAtividade)."
    try:
        if isDate(vencimento):
            dataVenc = datetime.strptime(vencimento, "%d/%m/%Y").date()
        else:
            dataVenc = datetime.today().date()+timedelta(days=int(vencimento)) 
        dataInicio = datetime.strptime(inicio, "%d/%m/%Y").date()          
        comando = "Insert into Atividades (TDPF, Atividade, Vencimento, Inicio) Values (%s, %s, %s, %s)"
        cursor.execute(comando, (chaveTdpf, atividade.upper(), dataVenc, dataInicio))
        msg = ""
        if fim!=None:
            msg = " Monitoramento deste TDPF foi reativado."
            comando = "Update CadastroTDPFs Set Fim=Null Where Codigo=%s"
            cursor.execute(comando, (chaveCad,))                         
        elif not tdpfCadastrado:
            msg = " Monitoramento deste TDPF foi iniciado."
            comando = "Insert into CadastroTDPFs (Fiscal, TDPF, Inicio) Values (%s, %s, %s)"
            cursor.execute(comando, (chaveFiscal, chaveTdpf, datetime.today().date()))
        conn.commit()
        conn.close()
        return True, msg
    except:
        conn.rollback()
        conn.close()
        return False, "Erro ao atualizar as tabelas (efetivaAtividade)."

def atividadeTexto(update, context): #obtém a descrição da atividade e chama a função que grava no BD
    global pendencias, textoRetorno, atividadeTxt
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    if len(msg)<3:
        response_message = "Descrição inválida (menos de 3 caracteres). Envie somente a descrição do texto (SEM informação protegida por sigilo) ou 'cancela'(sem aspas) para cancelar."
        response_message = response_message+textoRetorno
    if len(msg)>50:
        response_message = "Descrição inválida (mais de 50 caracteres). Envie somente a descrição do texto (SEM informação protegida por sigilo) ou 'cancela'(sem aspas) para cancelar."
        response_message = response_message+textoRetorno        
    else:
        atividade = msg.upper()
        if atividade=="CANCELA":
            eliminaPendencia(userId)
            mostraMenuPrincipal(update, context)
            return     
        efetivou, msgAtividade = efetivaAtividade(userId, atividadeTxt[userId][0], atividade, atividadeTxt[userId][1], atividadeTxt[userId][2])
        eliminaPendencia(userId) #apaga a pendência de informação do usuário        
        if efetivou:
            response_message = "Atividade registrada para o TDPF. No script do ContÁgil, você poder registrar informações p/ esta atividade sem restrição de sigilo no campo Observações."
            if msgAtividade!=None and msgAtividade!="":
                response_message = response_message+msgAtividade
        else:
            if msgAtividade==None or msgAtividade=='':
                msgAtividade = "Erro ao registrar a atividade. Tente novamente mais tarde."
            response_message = msgAtividade  #como são digitadas duas informações, desprezamos e retornamos ao menu pois o erro pode estar na anterior
        enviaMsgBot(bot, userId, text=response_message) 
        mostraMenuPrincipal(update, context)
        return            
    enviaMsgBot(bot, userId, text=response_message)  
    return                

def atividade(update, context): #critica e tenta efetivar a realização de uma atividade com prazo a vencer
    global pendencias, textoRetorno, atividadeTxt
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    msgAtividade = "Envie novamente o TDPF (16 dígitos), a data ou o prazo de vencimento em dias e a data de início da atividade."
    parametros = getParametros(msg)
    if len(parametros)!=3:
        response_message = "Número de informações (parâmetros) inválido. "+msgAtividade
        response_message = response_message+textoRetorno
    else:
        response_message = ""
        tdpf = getAlgarismos(parametros[0])
        data = parametros[1]
        inicio = parametros[2]
        if data.isdigit() and len(data)==8:
            data = data[:2]+"/"+data[2:4]+"/"+data[4:]   
        if inicio.isdigit() and len(inicio)==8:
            inicio = inicio[:2]+"/"+inicio[2:4]+"/"+inicio[4:]                      
        if len(tdpf)!=16 or not tdpf.isdigit():
            response_message = "TDPF inválido. "+msgAtividade
            response_message = response_message+textoRetorno                          
        elif not isDate(data) and not data.isdigit():
            response_message = "Data/prazo inválido. "+msgAtividade
            response_message = response_message+textoRetorno             
        if response_message=="" and not isDate(inicio):
            response_message = "Data de início da atividade inválida. "+msgAtividade
            response_message = response_message+textoRetorno 
        if len(response_message)>0:           
            enviaMsgBot(bot, userId, text=response_message)  
            return                                            
        prazo = 0
        dateTimeObj = None
        if isDate(data):
            try:
                dateTimeObj = datetime.strptime(data, '%d/%m/%Y').date()
                if dateTimeObj<=datetime.now().date():
                    response_message = "Data de vencimento deve ser futura. "+msgAtividade
                    response_message = response_message+textoRetorno                      
            except: #não deveria acontecer após o isDate, mas fazemos assim para não correr riscos
                response_message = "Erro na conversão da data de vencimento. "+msgAtividade
                response_message = response_message+textoRetorno   
        else:
            try:
                prazo = int(data)  
                if prazo>365 or prazo<=0:
                    response_message = "Prazo de vencimento deve ser superior a 0 e inferior a 366 dias. "+msgAtividade
                    response_message = response_message+textoRetorno                          
            except: 
                response_message = "Erro na conversão do prazo. "+msgAtividade
                response_message = response_message+textoRetorno 
        if len(response_message)==0:
            try:
                dataInicio =  datetime.strptime(inicio, '%d/%m/%Y').date()
                if dataInicio>datetime.now().date():
                    response_message = "Data de início deve ser atual ou passada. "+msgAtividade
                    response_message = response_message+textoRetorno                      
            except: #não deveria acontecer após o isDate, mas fazemos assim para não correr riscos
                response_message = "Erro na conversão da data de início. "+msgAtividade
                response_message = response_message+textoRetorno                                                  
        if len(response_message)==0:
            eliminaPendencia(userId)
            pendencias[userId] = 'atividadeTexto'
            atividadeTxt[userId] = [tdpf, data, inicio]
            response_message = "Informe a *descrição da atividade SEM QUALQUER INFORMAÇÃO PROTEGIDA POR SIGILO* (máximo de 50 caracteres):"                    
    enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    return     

def efetivaAnulacaoAtividade(userId, tdpf, codigo): #efetiva a anulação (apaga) de uma atividade de um tdpf (o código é a chave primária da atividade - opcional; se não houver, é o último registro daquele TDPF)
    conn = conecta()
    if not conn: 
        return False, "Erro na conexão - efetivaAnulacaoAtividade"
    cursor = conn.cursor(buffered=True)
    bUsuario, cpf, msg = verificaUsuarioTelegram(conn, userId)
    if not bUsuario:
        conn.close()
        return False, msg #tb já foi testado     
    try:
        bAlocacao, chaveTdpf, chaveFiscal, msg = verificaAlocacao(conn, cpf, tdpf)
        if not bAlocacao:
            conn.close()
            return False, msg        
        bMonitoramento, chaveTdpf, chaveFiscal, msg = verificaMonitoramento(conn, cpf, tdpf)
        if not bMonitoramento:
            conn.close()
            return False, msg
        if codigo==0:    
            comando = "Select Codigo, TDPF, Inicio from Atividades Where TDPF=%s Order by Codigo DESC"
            cursor.execute(comando, (chaveTdpf,))
            rows = cursor.fetchall()
        else:
            comando = "Select Codigo, TDPF, Inicio from Atividades Where Codigo=%s and TDPF=%s"
            cursor.execute(comando, (codigo, chaveTdpf))
            rows = cursor.fetchall()       
            if len(rows)==0:
                conn.close()
                return False, "Atividade (código) inexistente para o TDPF informado."         
    except:
        conn.close()
        return False, "Erro na consulta (efetivaAnulacaoAtividade)."
    if len(rows)==0:
        conn.close()
        return False, "Não há atividade informada para o TDPF." #Não havia data de ciência para o TDPF   
    if codigo == 0:         
        codigo = rows[0][0]    
    try:
        comando = "Delete from Atividades Where Codigo=%s and TDPF=%s"
        cursor.execute(comando, (codigo, chaveTdpf))
        conn.commit() 
        conn.close()  
        return True, ""
    except:
        conn.rollback()
        conn.close()
        return False, "Erro na atualização das tabelas. Tente novamente mais tarde."

def anulaAtividade(update, context): #anula (apaga) a última atividade cadastrada do TDPF
    global pendencias, textoRetorno
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=1 and len(parametros)!=2:
        response_message = "Envie o nº do TDPF (16 dígitos, sem espaços) e, opcionalmente, o código da atividade a ser excluída (se não for informado, será excluída a última cadastrada) - separe as informa;óes (TDPF e código) com espaço."
        response_message = response_message+textoRetorno      
    else:
        codigo = 0
        tdpf = getAlgarismos(parametros[0])
        response_message = ""
        if len(parametros)==2:
            try:
                codigo = int(parametros[1])
            except:
                response_message = "Código inválido. Envie o nº do TDPF (16 dígitos, sem espaços) e, opcionalmente, o código da atividade a ser excluída (se não for informado, será excluída a última cadastrada) - separe as informaçóes (TDPF e código) com espaço."
                response_message = response_message+textoRetorno      
        if len(tdpf)!=16 or not tdpf.isdigit():
            response_message = "TDPF inválido. Envie o nº do TDPF (16 dígitos, sem espaços) e, opcionalmente, o código da atividade a ser excluída (se não for informado, será excluída a última cadastrada) - separe as informaçóes (TDPF e código) com espaço."
            response_message = response_message+textoRetorno
        if response_message=="":
            efetivou, msgAnulacao = efetivaAnulacaoAtividade(userId, tdpf, codigo)
            eliminaPendencia(userId) #apaga a pendência de informação do usuário                
            if efetivou:
                response_message = "Última atividade ou atividade indicada foi excluída para o TDPF."         
            else:
                response_message = msgAnulacao
            enviaMsgBot(bot, userId, text=response_message) 
            mostraMenuPrincipal(update, context)
            return
    enviaMsgBot(bot, userId, text=response_message)  
    return 

def efetivaHorasAtividade(userId, codigo, horas):
    conn = conecta()
    if not conn: 
        return False, "Erro na conexão - efetivaHorasAtividade"
    cursor = conn.cursor(buffered=True)
    bUsuario, cpf, msg = verificaUsuarioTelegram(conn, userId)
    if not bUsuario:
        conn.close()
        return False, msg #tb já foi testado  
    try:
        comando = """Select Atividades.Atividade, Alocacoes.Desalocacao, Atividades.TDPF, Atividades.Inicio, Fiscais.Codigo
                     From Atividades, Alocacoes, Fiscais, CadastroTDPFs
                     Where Atividades.Codigo=%s and Atividades.TDPF=Alocacoes.TDPF and Fiscais.CPF=%s and Fiscais.Codigo=Alocacoes.Fiscal
                     and CadastroTDPFs.Fiscal=Fiscais.Codigo and CadastroTDPFs.TDPF=Atividades.TDPF and CadastroTDPFs.Fim Is Null"""
        cursor.execute(comando, (codigo, cpf))
        linhas = cursor.fetchall()
    except:
        conn.close()
        return False, "Erro nas consultas - efetivaHorasAtividade"    
    bAchou = True
    if linhas==None:
        bAchou = False
    if len(linhas)==0:
        bAchou = False
    if not bAchou:
        conn.close()
        return False, "Código da atividade não localizado ou usuário não está alocado ao TDPF ou o monitora"        
    linha = linhas[0]
    if linha[1]!=None:
        conn.close()
        return False, "Usuário não está mais alocado ao TDPF"
    chaveTdpf = linha[2]
    chaveFiscal = linha[4]    
    comando = "Select Codigo, Fim from CadastroTDPFs Where Fiscal=%s and TDPF=%s"
    try:
        cursor.execute(comando, (chaveFiscal, chaveTdpf))
        row = cursor.fetchone()
    except:
        conn.close()
        return False,  "Erro na consulta monitoramento - efetivaHorasAtividade"    
    if not row:
        conn.close()
        return False, "TDPF não está sendo monitorado para você."
    if row[1]!=None:
        conn.close()
        return False, "O acompanhamento do TDPF foi finalizado em "+row[1].strftime('%d/%m/%Y')+"."        
    try:
        comando = "Update Atividades Set Horas=%s Where Codigo=%s"
        cursor.execute(comando, (horas, codigo))   
        conn.commit()
        conn.close() 
        return True, linha[0]
    except:
        conn.rollback()
        conn.close()
        return False, "Erro na atualização da tabela - efetivaHorasAtividade"    

def informaHorasAtividade(update, context): #registra horas gastas em uma atividade
    global pendencias, textoRetorno
    msgHorasAtiv = "Envie o código da atividade e a quantidade de horas dispendidas (número inteiro) até o momento - separe as informações (código e data) com espaço."
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=2:
        response_message = msgHorasAtiv
        response_message = response_message+textoRetorno  
        enviaMsgBot(bot, userId, text=response_message)  
        return            
    codigo = parametros[0]
    if not codigo.isdigit():
        response_message = "Código inválido. "+msgHorasAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return
    try:
        codigo = int(codigo)
    except:
        response_message = "Código inválido (2). "+msgHorasAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return                  
    horas = parametros[1] 
    if not horas.isdigit():
        response_message = "Quantidade de horas inválida. "+msgHorasAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return
    if len(horas)>3:
        response_message = "Quantidade de horas inválida (2). "+msgHorasAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return         
    try:
        horas = int(horas)
    except:
        response_message = "Quantidade de horas inválida (3). "+msgHorasAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return    
    efetivou, msgAnulacao = efetivaHorasAtividade(userId, codigo, horas)
    eliminaPendencia(userId) #apaga a pendência de informação do usuário                
    if efetivou:
        response_message = "Horas dispendidas na atividade "+msgAnulacao+" foram registradas."         
    else:
        response_message = msgAnulacao
    enviaMsgBot(bot, userId, text=response_message) 
    mostraMenuPrincipal(update, context)
    return

def efetivaTerminoAtividade(userId, codigo, dataTermino, horas):
    conn = conecta()
    if not conn: 
        return False, "Erro na conexão - efetivaTerminoAtividade"
    cursor = conn.cursor(buffered=True)
    bUsuario, cpf, msg = verificaUsuarioTelegram(conn, userId)
    if not bUsuario:
        conn.close()
        return False, msg #tb já foi testado      
    try:
        comando = """Select Atividades.Atividade, Alocacoes.Desalocacao, Atividades.TDPF, Atividades.Inicio, Fiscais.Codigo, Atividades.TDPF
                     From Atividades, Alocacoes, Fiscais, CadastroTDPFs
                     Where Atividades.Codigo=%s and Atividades.TDPF=Alocacoes.TDPF and Fiscais.CPF=%s and Alocacoes.Fiscal=Fiscais.Codigo
                     and Atividades.TDPF=CadastroTDPFs.TDPF and Fiscais.Codigo=CadastroTDPFs.Fiscal and CadastroTDPFs.Fim Is Null"""
        cursor.execute(comando, (codigo, cpf))
        linhas = cursor.fetchall()
    except:
        conn.close()
        return False, "Erro nas consultas - efetivaTerminoAtividade"    
    bAchou = True
    if linhas==None:
        bAchou = False
    if len(linhas)==0:
        bAchou = False
    if not bAchou:
        conn.close()
        return False, "Código da atividade não localizado ou usuário não está alocado ao TDPF ou o monitora"        
    linha = linhas[0]
    if linha[1]!=None:
        conn.close()
        return False, "Usuário não está mais alocado ao TDPF"
    chaveFiscal = linha[4]
    chaveTdpf = linha[5]
    dataInicio = linha[3]    
    if dataInicio!=None:
        if dataTermino<dataInicio:
            conn.close()
            return False, "Data de término não pode ser anterior à data de início da atividade ("+dataInicio.strftime("%d/%m/%Y")+")"
    tdpf = linha[2]    
    comando = "Select Codigo, Fim from CadastroTDPFs Where Fiscal=%s and TDPF=%s"
    try:
        cursor.execute(comando, (chaveFiscal, chaveTdpf))
        row = cursor.fetchone()
    except:
        conn.close()
        return False,  "Erro na consulta monitoramento - efetivaTerminoAtividade"    
    if not row:
        conn.close()
        return False, "TDPF não está sendo monitorado para você."
    if row[1]!=None:
        conn.close()
        return False, "O acompanhamento do TDPF foi finalizado em "+row[1].strftime('%d/%m/%Y')+"."        
    try:
        comando = "Update Atividades Set Termino=%s, Horas=%s Where Codigo=%s"
        cursor.execute(comando, (dataTermino, horas, codigo))   
        conn.commit()
        conn.close() 
        return True, linha[0]
    except:
        conn.rollback()
        conn.close()
        return False, "Erro na atualização da tabela - efetivaTerminoAtividade"    

def terminoAtividade(update, context): #registra data de término de atividade cadastrada do TDPF
    global pendencias, textoRetorno
    msgTerminoAtiv = "Envie o código da atividade, a data de seu término (dd/mm/yyyy) e a quantidade de horas dispendidas (número inteiro) até o momento - separe as informações (código e data) com espaço."
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=3:
        response_message = msgTerminoAtiv
        response_message = response_message+textoRetorno  
        enviaMsgBot(bot, userId, text=response_message)  
        return            
    codigo = parametros[0]
    if not codigo.isdigit():
        response_message = "Código inválido. "+msgTerminoAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return
    try:
        codigo = int(codigo)
    except:
        response_message = "Código inválido (2). "+msgTerminoAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return                
    data = parametros[1] 
    if data.isdigit() and len(data)==8:
        data = data[:2]+"/"+data[2:4]+"/"+data[4:]                                     
    if not isDate(data):
        response_message = "Data inválida. "+msgTerminoAtiv
        response_message = response_message+textoRetorno  
        enviaMsgBot(bot, userId, text=response_message)  
        return                             
    dateTimeObj = None
    try:
        dateTimeObj = datetime.strptime(data, '%d/%m/%Y')
        if dateTimeObj.date()>datetime.now().date():
            response_message = "Data de término não pode ser futura. "+msgTerminoAtiv
            response_message = response_message+textoRetorno   
            enviaMsgBot(bot, userId, text=response_message) 
            return             
    except: #não deveria acontecer após o isDate, mas fazemos assim para não correr riscos
        response_message = "Erro na conversão da data. "+msgTerminoAtiv
        response_message = response_message+textoRetorno   
        enviaMsgBot(bot, userId, text=response_message) 
        return  
    horas = parametros[2] 
    if not horas.isdigit():
        response_message = "Quantidade de horas inválida. "+msgTerminoAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return
    if len(horas)>3:
        response_message = "Quantidade de horas inválida (2). "+msgTerminoAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return        
    try:
        horas = int(horas)
    except:
        response_message = "Quantidade de horas inválida (3). "+msgTerminoAtiv
        response_message = response_message+textoRetorno            
        enviaMsgBot(bot, userId, text=response_message)  
        return    
    efetivou, msgAnulacao = efetivaTerminoAtividade(userId, codigo, dateTimeObj, horas)
    eliminaPendencia(userId) #apaga a pendência de informação do usuário                
    if efetivou:
        response_message = "Término da atividade "+msgAnulacao+" foi registrado."         
    else:
        response_message = msgAnulacao
    enviaMsgBot(bot, userId, text=response_message) 
    mostraMenuPrincipal(update, context)
    return
            
def efetivaCiencia(userId, tdpf, data, vencimento, documento): #tenta efetivar a ciência de um tdpf no BD
    conn = conecta()
    if not conn: 
        return False, "Erro na conexão - efetivaCiencia"
    cursor = conn.cursor(buffered=True)  
    try:
        bUsuario, cpf, msg = verificaUsuarioTelegram(conn, userId)
        if not bUsuario:
            conn.close()
            return False, msg #tb já foi testado  
        bAlocacao, chaveTdpf, chaveFiscal, msg = verificaAlocacao(conn, cpf, tdpf)  
        if not bAlocacao:
            conn.close()
            return False, msg      
        comando = "Select Data from Ciencias Where TDPF=%s and Data>=%s Order by Data DESC"
        cursor.execute(comando, (chaveTdpf, data)) #datetime.strptime(data, "%d/%m/%Y")))
        row = cursor.fetchone()
        if row:
            conn.close()
            return False, "Data de ciência informada DEVE ser posterior à ultima informada para o TDPF ("+row[0].strftime('%d/%m/%Y')+")."
        comando = "Select Codigo, Fim from CadastroTDPFs Where TDPF=%s and Fiscal=%s"
        cursor.execute(comando, (chaveTdpf, chaveFiscal))
        row = cursor.fetchone()        
        tdpfCadastrado = False
        fim = None
        if row:
            tdpfCadastrado = True  
            chave = row[0]
            fim = row[1]
            #logging.info("TDPF cadastrado")           
    except:
        conn.close()
        return False, "Erro na consulta (3)."
    try:
        comando = "Insert into Ciencias (TDPF, Data, Documento, Vencimento) Values (%s, %s, %s, %s)"
        cursor.execute(comando, (chaveTdpf, data, documento, vencimento)) #datetime.strptime(data, "%d/%m/%Y"), documento))
        msg = ""
        if fim!=None:
            msg = " Monitoramento deste TDPF foi reativado."
            comando = "Update CadastroTDPFs Set Fim=Null Where Codigo=%s"
            cursor.execute(comando, (chave,))                         
        elif not tdpfCadastrado:
            msg = " Monitoramento deste TDPF foi iniciado."
            comando = "Insert into CadastroTDPFs (Fiscal, TDPF, Inicio) Values (%s, %s, %s)"
            cursor.execute(comando, (chaveFiscal, chaveTdpf, datetime.today().date()))
        conn.commit()
        conn.close()
        return True, msg
    except:
        conn.rollback()
        conn.close()
        return False, "Erro ao atualizar as tabelas (efetivaCiencia)."
            
def cienciaTexto(update, context): #obtém a descrição do documento e chama a função que grava no BD
    global pendencias, textoRetorno, cienciaTxt
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    if len(msg)<4:
        response_message = "Descrição inválida (menos de 4 caracteres). Envie somente a descrição do documento ou 'cancela'(sem aspas) para cancelar."
        response_message = response_message+textoRetorno
        enviaMsgBot(bot, userId, text=response_message)    
    elif len(msg)>50:
        response_message = "Descrição inválida (mais de 50 caracteres). Envie somente a descrição do documento ou 'cancela'(sem aspas) para cancelar."
        response_message = response_message+textoRetorno
        enviaMsgBot(bot, userId, text=response_message)                
    else:
        eliminaPendencia(userId) #apaga a pendência de informação do usuário        
        documento = msg.upper().strip()
        if documento=="CANCELA":
            mostraMenuPrincipal(update, context)
            return     
        efetivou, msgEfetivaCiencia= efetivaCiencia(userId, cienciaTxt[userId][0], cienciaTxt[userId][1], cienciaTxt[userId][2], documento)
        if efetivou:
            response_message = "Data de ciência registrada para o TDPF."
            if msgEfetivaCiencia!=None and msgEfetivaCiencia!="":
                response_message = response_message+msgEfetivaCiencia
        else:
            response_message = msgEfetivaCiencia         
        enviaMsgBot(bot, userId, text=response_message) 
        mostraMenuPrincipal(update, context)
    return   

def calculaVencimento(prazo, data): #calcula vencimento da intimação para o contribuinte; prazo: int (dias); data (ciência - início do prazo): datetime
    data = data + timedelta(days=1)
    while weekday(data.year, data.month, data.day)>=5: #contagem do prazo não inicia em sábado ou domingo
        data = data + timedelta(days=1)
    vencimento = data + timedelta(days=prazo-1)
    while weekday(vencimento.year, vencimento.month, vencimento.day)>=5: #contagem do prazo não termina em sábado ou domingo
        vencimento = vencimento + timedelta(days=1)
    return vencimento
    
def ciencia(update, context): #critica e tenta efetivar a ciência de um TDPF (registrar data)
    global pendencias, textoRetorno, cienciaTxt
    msgCiencia = "Envie novamente o TDPF, a data de ciência (dd/mm/aaaa) e, opcionalmente, o vencimento da intimação - data ou dias corridos (separados por espaço)."
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if not len(parametros) in [2, 3]:
        response_message = "Número de informações (parâmetros) inválido. Envie somente o nº do TDPF, a última data de ciência relativa a ele e, opcionalmente, o vencimento da intimação (data ou dias corridos)."
        response_message = response_message+textoRetorno
    else:
        response_message = ""
        tdpf = getAlgarismos(parametros[0])
        data = parametros[1]
        prazo = None        
        vencimentoObj = None        
        if len(parametros)==3:
            vencimento = parametros[2]
            if vencimento.isdigit() and len(vencimento)==8:
                vencimento = vencimento[:2]+"/"+vencimento[2:4]+"/"+vencimento[4:]
            if len(vencimento)!=10:
                if not vencimento.isdigit():
                    response_message = "Vencimento inválido. "+msgCiencia
                    response_message = response_message+textoRetorno 
                    enviaMsgBot(bot, userId, text=response_message)  
                    return                
                else:
                    try:
                        prazo = int(vencimento)
                        if prazo<1 or prazo>60:
                            response_message = "Vencimento (prazo) inválido - deve estar entre 1 e 60 dias "+msgCiencia
                            response_message = response_message+textoRetorno   
                            enviaMsgBot(bot, userId, text=response_message)  
                            return                                              
                    except:
                        response_message = "Vencimento (prazo) inválido. "+msgCiencia
                        response_message = response_message+textoRetorno   
                        enviaMsgBot(bot, userId, text=response_message)  
                        return                                               
            elif isDate(vencimento):
                try:
                    vencimentoObj = datetime.strptime(vencimento, '%d/%m/%Y')
                except: #não deveria acontecer após o isDate, mas fazemos assim para não correr riscos
                    logging.info("Erro na conversão do vencimento "+vencimento+" - UserId "+str(userId))
                    response_message = "Erro na conversão do vencimento. "+msgCiencia
                    response_message = response_message+textoRetorno   
                    enviaMsgBot(bot, userId, text=response_message)  
                    return
            else:
                response_message = "Vencimento inválido. "+msgCiencia
                response_message = response_message+textoRetorno   
                enviaMsgBot(bot, userId, text=response_message)  
                return                
        if data.isdigit() and len(data)==8:
            data = data[:2]+"/"+data[2:4]+"/"+data[4:]
        if len(tdpf)!=16 or not tdpf.isdigit():
            response_message = "TDPF inválido. "+msgCiencia
            response_message = response_message+textoRetorno            
        elif not isDate(data):
            response_message = "Data inválida. "+msgCiencia
            response_message = response_message+textoRetorno
        else:
            try:
                dateTimeObj = datetime.strptime(data, '%d/%m/%Y')
            except: #não deveria acontecer após o isDate, mas fazemos assim para não correr riscos
                logging.info("Erro na conversão da data "+data+" - UserId "+str(userId))
                response_message = "Erro na conversão da data. "+msgCiencia
                response_message = response_message+textoRetorno   
                enviaMsgBot(bot, userId, text=response_message)  
                return                
            if dateTimeObj.date()>datetime.now().date():
                response_message = "Data de ciência não pode ser futura. "+msgCiencia
                response_message = response_message+textoRetorno                    
            elif dateTimeObj.date()<datetime.now().date()-timedelta(days=60):
                response_message = "Data de ciência já está vencida para fins de recuperação da espontaneidade tributária. "+msgCiencia
                response_message = response_message+textoRetorno
            else:  
                if prazo!=None: #vemos qual vai ser o prazo em função do prazo em dias
                    vencimentoObj = calculaVencimento(prazo, dateTimeObj)
                    if vencimentoObj.date()<=dateTimeObj.date():
                        response_message = "Vencimento deve ser posterior à data de ciência. "+msgCiencia
                        response_message = response_message+textoRetorno   
                if response_message=="": 
                    eliminaPendencia(userId)
                    pendencias[userId] = 'cienciaTexto'
                    cienciaTxt[userId] = [tdpf, dateTimeObj, vencimentoObj]
                    response_message = "Informe a descrição do documento que efetivou a ciência (ex.: TIPF) (máximo de 50 caracteres):"                                        
    enviaMsgBot(bot, userId, text=response_message)  
    return 

def efetivaPrazos(userId, d): #altera no BD os prazos padrão do usuário (d é uma lista)
    conn = conecta()
    if not conn: 
        return False, "Erro na conexão - efetivaPrazos"
    cursor = conn.cursor(buffered=True)    
    #ordena os prazos (não tem utilidade, mas deixei assim pois pode ser que venha a ter)
    d.sort(reverse = True)
    d1 = d[0]
    d2 = d[1]
    d3 = d[2]            
    comando = "Update Usuarios set d1=%s, d2=%s, d3=%s where idTelegram=%s"
    try:
        logging.info(comando)
        cursor.execute(comando, (d1, d2, d3, userId))
        conn.commit()
        conn.close()
        return True, "" #retorna se operação foi um sucesso e mensagem de erro em caso de False no primeiro         
    except:
        conn.rollback()
        conn.close()
        return False, "Erro de atualização. Tente novamente mais tarde."
        
def prazos(update, context): #critica e tenta alterar os prazos padrão do usuário
    global pendencias, textoRetorno
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=3:
        response_message = "Número de informações (parâmetros) inválido. Envie 3 prazos em dias (1 a 50) antes da retomada da espontaneidade em que deseja ser alertado (separe cada um com um espaço)."
        response_message = response_message+textoRetorno
    else:
        d = parametros
        response_message = None
        if d[0].isdigit() and d[1].isdigit() and d[2].isdigit():
            diasInt = []
            for dn in d:
                dias = int(dn)
                diasInt.append(dias)
                if dias<1 or dias>50:
                    response_message = "Prazos devem estar compreendidos entre 1 e 50 dias. Envie 3 prazos em dias."
                    response_message = response_message+textoRetorno
                    break
            if not response_message:
                if d[0]==d[1] or d[1]==d[2] or d[0]==d[2]:
                    response_message = "Nenhum prazo pode ser igual a outro. Envie 3 prazos em dias diferentes um do outro."
                    response_message = response_message+textoRetorno                    
            if not response_message:
                efetivou, msgPrazos = efetivaPrazos(userId, diasInt)
                if efetivou:
                    eliminaPendencia(userId) #apaga a pendência de informação do usuário
                    response_message = "Prazos registrados."
                    enviaMsgBot(bot, userId, text=response_message) 
                    mostraMenuPrincipal(update, context)
                    return
                else:
                    response_message = msgPrazos+textoRetorno
        else:
            response_message = "Os dias devem ser números inteiros entre 1 e 50. Envie 3 prazos em dias (1 a 50) antes da retomada da espontaneidade em que deseja ser alertado (separe cada um com um espaço)."
            response_message = response_message+textoRetorno
    enviaMsgBot(bot, userId, text=response_message)  
    return

def efetivaAnulacao(userId, tdpf): #efetiva a anulação (apaga) de uma data de ciência de um tdpf - a data anterior, se houver, é que será considerada
    conn = conecta()
    if not conn: 
        return False, "Erro na conexão - efetivaAnulacao"
    cursor = conn.cursor(buffered=True)
    bUsuario, cpf, msg = verificaUsuarioTelegram(conn, userId)
    if not bUsuario:
        conn.close()
        return False, msg #tb já foi testado      
    try:
        bMonitoramento, chaveTdpf, chaveFiscal, msg = verificaMonitoramento(conn, cpf, tdpf)  
        if not bMonitoramento:
            conn.close()
            return False, msg      
        bAlocacao, chaveTdpf, chaveFiscal, msg = verificaAlocacao(conn, cpf, tdpf)
        if not bAlocacao:
            conn.close()
            return False, msg              
        comando = "Select Codigo, TDPF, Data from Ciencias Where TDPF=%s Order by Data"
        cursor.execute(comando, (chaveTdpf,))
        rows = cursor.fetchall()
    except:
        conn.close()
        return False, "Erro na consulta (5)."
    if len(rows)==0:
        conn.close()
        return False, "Não há data de ciência informada para o TDPF." #Não havia data de ciência para o TDPF
    if len(rows)==1:
        dataAnt = "Nenhuma Data Vigente" #não haverá data anterior
    else: #2 ou mais linhas
        dataAnt = rows[len(rows)-2][2].strftime('%d/%m/%Y')        
    chave = rows[len(rows)-1][0]    
    try:
        comando = "Delete from Ciencias Where Codigo=%s"
        cursor.execute(comando, (chave,))
        conn.commit()  
        conn.close() 
        return True, dataAnt #retorna se teve sucesso e a data anterior
    except:
        conn.rollback()
        conn.close()
        return False, "Erro na atualização das tabelas. Tente novamente mais tarde."

def anulaCiencia(update, context): #anula (apaga) a última data de ciência do TDPF
    global pendencias, textoRetorno
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=1:
        response_message = "Envie somente o nº do TDPF (16 dígitos), sem espaços."
        response_message = response_message+textoRetorno      
    else:
        tdpf = getAlgarismos(parametros[0])
        if len(tdpf)!=16 or not tdpf.isdigit():
            response_message = "TDPF inválido. Envie novamente o TDPF (16 dígitos)."
            response_message = response_message+textoRetorno
        else:
            efetivou, msgAnulacao = efetivaAnulacao(userId, tdpf)
            eliminaPendencia(userId) #apaga a pendência de informação do usuário                
            if efetivou:
                if msgAnulacao!=None: #se efetivou e retornou mensagem, esta é a data da ciência anterior
                    response_message = "Última data de ciência foi excluída para o TDPF. Retornou para a anterior: "+msgAnulacao
                    try:
                        dateTimeObj = datetime.strptime(msgAnulacao, '%d/%m/%Y')
                        if dateTimeObj.date()<datetime.now().date()-timedelta(days=59):
                            response_message = response_message+"\nA data de ciência anterior já está vencida e não irá gerar, na prática, monitoramento."
                    except :
                        response_message = response_message+"\nNa prática, não haverá agora monitoramento para este TDPF por falta de data de ciência."                   
                else:
                    response_message = "Este TDPF não estava sendo monitorado pelo serviço para você."
            else:
                response_message = msgAnulacao
            enviaMsgBot(bot, userId, text=response_message) 
            mostraMenuPrincipal(update, context)
            return
    enviaMsgBot(bot, userId, text=response_message)  
    return  

def formataEquipe(equipe):
    if equipe==None:
        return None
    if len(equipe)<14:
        return equipe
    return equipe[:7]+"."+equipe[7:11]+"."+equipe[11:].strip()

def relacionaPostagens (update, context): #relaciona as postagens do usuários feitas de uma quantidade de dias para trás até hoje
    global pendencias, textoRetorno
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    mensagem = "Envie a quantidade de dias (máximo de 90) para considerar na pesquisa e, separado por espaço, a indicação se considera (S) ou não (N) sua eventual qualidade de supervisor para pesquisar as postagens da equipe."
    if not len(parametros) in [1, 2]:
        response_message = "Quantidade de parâmetros inválida. "+mensagem
        response_message = response_message+textoRetorno      
    else:
        dias = parametros[0]
        if len(parametros)==2:
            qualidade = parametros[1].upper()
        else:
            qualidade = "N"
        if not len(dias) in [1, 2] or not dias.isdigit():
            response_message = "Quantidade de dias inválida. "+mensagem
            response_message = response_message+textoRetorno
        else: 
            dias = int(dias)
            if dias<0 or dias>90:
                response_message = "Quantidade de dias inválida (2). "+mensagem
                response_message = response_message+textoRetorno  
            else:
                if not qualidade in ["S", "SIM", "N", "NÃO", "NAO"]:
                    response_message = "Indicador de qualidade de supervisor inválido. "+mensagem
                    response_message = response_message+textoRetorno 
                else: #tudo ok - fazemos a consulta
                    eliminaPendencia(userId) #apaga a pendência de informação do usuário
                    conn = conecta()
                    if not conn: #não conectamos ao BD
                        response_message = "Erro na conexão com o banco de dados."
                        response_message = response_message
                        enviaMsgBot(bot, userId, text=response_message) 
                        mostraMenuPrincipal(update, context)
                        return                            
                    cursor = conn.cursor(buffered=True)                        
                    qualidade = qualidade[:1]       
                    consulta = """
                                Select TDPFS.Numero, 'U', TDPFS.Grupo as Equipe1, Documento, ControlePostal.DataEnvio as DtEnvio, SituacaoAtual, DataSituacao, DataRecebimento
                                from TDPFS, ControlePostal, Usuarios, Fiscais, Alocacoes
                                Where Usuarios.idTelegram=%s and Usuarios.CPF=Fiscais.CPF and Fiscais.Codigo=Alocacoes.Fiscal and Alocacoes.Desalocacao Is Null and
                                Alocacoes.TDPF=TDPFS.Codigo and TDPFS.Codigo=ControlePostal.TDPF and ControlePostal.DataEnvio>=cast((now() - interval %s day) as date)
                                and (TDPFS.Tipo='F' or TDPFS.Tipo='D')
                                """
                    if qualidade=="S": #é para pesquisar as postagens da equipe - considera o usuário um supervisor tb
                        consulta += """UNION
                                       Select TDPFS.Numero, 'S', TDPFS.Grupo as Equipe1, Documento, ControlePostal.DataEnvio as DtEnvio, SituacaoAtual, DataSituacao, DataRecebimento
                                       from TDPFS, ControlePostal, Usuarios, Fiscais, Supervisores
                                       Where Usuarios.idTelegram=%s and Usuarios.CPF=Fiscais.CPF and Fiscais.Codigo=Supervisores.Fiscal and Supervisores.Fim Is Null and
                                       Supervisores.Equipe=TDPFS.Grupo and TDPFS.Codigo=ControlePostal.TDPF and ControlePostal.DataEnvio>=cast((now() - interval %s day) as date)
                                       and (TDPFS.Tipo='F' or TDPFS.Tipo='D')
                                       Order by DtEnvio, Equipe1, Numero
                                    """
                        cursor.execute(consulta, (userId, dias, userId, dias))
                    else:
                        consulta += "Order by DtEnvio, Numero"
                        cursor.execute(consulta, (userId, dias)) 
                    linhas = cursor.fetchall()
                    resultado = ""
                    bSuperv = False
                    equipe = ""
                    for linha in linhas:
                        if linha[1]=="S" and not bSuperv:
                            bSuperv = True
                            resultado += "\n\nTDPFs sob sua supervisão:\n"                                    
                        else:
                            resultado += "\n"
                        if linha[1]=="S" and linha[2]!=equipe: #mudou a equipe e é supervisor
                            resultado += "\nEquipe "+ formataEquipe(linha[2])+":\n"
                            equipe = linha[2]
                        resultado += "*"+formataTDPF(linha[0])+"* | "
                        resultado += linha[3][:50].strip()+" | "
                        resultado += linha[4].strftime("%d/%m/%Y")+" | "
                        situacao = linha[5]
                        if situacao==None:
                            situacao = "ND"
                        dataSituacao = linha[6]
                        if dataSituacao==None:
                            dataSituacao = "ND"
                        else:
                            dataSituacao = dataSituacao.strftime("%d/%m/%Y")
                        resultado += situacao[:50].strip()+" | "
                        resultado += dataSituacao+" ! "
                        dataRecebimento = linha[7]
                        if dataRecebimento==None:
                            dataRecebimento = "ND"
                        else:
                            dataRecebimento = dataRecebimento.strftime("%d/%m/%Y")  
                        resultado += dataRecebimento+"\n"                      
                    if resultado!="":
                        resultado = "Relação de Postagens (TDPF | Documento | Envio | Situação Atual | Data | Rcbto AR/Correspond):\n"+resultado
                    else:
                        resultado = "Não foram encontradas postagens de documentos relativos a TDPFs sob sua responsabilidade"                                
                    enviaMsgBot(bot, userId, text=limpaMarkdown(resultado), parse_mode= 'MarkdownV2')
                    mostraMenuPrincipal(update, context)
                    conn.close()
                    return     
    enviaMsgBot(bot, userId, text=response_message) 
    return             
        

def efetivaFinalizacao(userId, tdpf=""): #finaliza monitoramento de um tdpf ou de todos do usuário (campo Fim da tabela CadastroTDPFs)
    conn = conecta()
    if not conn: 
        return False, "Erro na conexão - efetivaFinalização"
    cursor = conn.cursor(buffered=True)
    bUsuario, cpf, msg = verificaUsuarioTelegram(conn, userId)
    if not bUsuario:
        conn.close()
        return False, msg #tb já foi testado  
    try:
        if tdpf!="":
            comando = """Select CadastroTDPFs.Codigo, Fim, Fiscais.Codigo from CadastroTDPFs, Fiscais, TDPFS 
                         Where Fiscais.CPF=%s and CadastroTDPFs.Fiscal=Fiscais.Codigo and TDPFS.Numero=%s and TDPFS.Codigo=CadastroTDPFs.TDPF"""
            cursor.execute(comando, (cpf, tdpf))            
        else:
            comando = "Select CadastroTDPFs.Codigo, Fim, Fiscais.Codigo from CadastroTDPFs, Fiscais Where Fiscais.CPF=%s and CadastroTDPFs.Fiscal=Fiscais.Codigo and Fim Is Null"
            cursor.execute(comando, (cpf,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            if tdpf!="":
                return False, "TDPF não está sendo monitorado por você. Envie um novo nº de TDPF:"
            else:
                return True, "Nenhum TDPF está sendo monitorado por você atualmente."
        if row[1]!=None and tdpf!="":
            conn.close()
            return True, "O acompanhamento do TDPF já havia sido finalizado em "+row[1].strftime('%d/%m/%Y')+"."
        elif tdpf!="":
            #conn.close()    #<-------VERIFICAR - ACHO QUE ESTÁ ERRADO
            pass
    except:
        conn.close()
        return False, "Erro na consulta (7). Tente novamente mais tarde."
    try:
        chaveReg = row[0]        
        chaveFiscal = row[2]
        if tdpf!="":
            comando = "Update CadastroTDPFs Set Fim=%s Where Codigo=%s"
            cursor.execute(comando, (datetime.today().date(), chaveReg))
        else:
            comando = "Update CadastroTDPFs Set Fim=%s Where Fiscal=%s"
            cursor.execute(comando, (datetime.today().date(), chaveFiscal))
        conn.commit()  
        conn.close()          
        return True, None
    except:
        conn.rollback()
        conn.close()
        return False, "Erro na atualização das tabelas (efetivaFinalizacao)."

def fim(update, context): #exclui o TDPF do monitoramento (campo Fim da tabela CadastroTDPFs)
    global pendencias, textoRetorno
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=1:
        response_message = "Envie somente o nº do TDPF (16 dígitos), sem espaços, ou a palavra TODOS."
        response_message = response_message+textoRetorno
    else:
        efetivou = None
        parametro = parametros[0].strip().upper()
        tdpf = getAlgarismos(parametro)
        if parametro in ["TODOS", "TODAS"]:
            efetivou, msgAnulacao = efetivaFinalizacao(userId)
        elif len(tdpf)!=16 or not tdpf.isdigit():
            response_message = "TDPF ou comando inválido. Envie novamente o TDPF (16 dígitos), sem espaços, ou a palavra TODOS."
            response_message = response_message+textoRetorno
        else:
            efetivou, msgAnulacao = efetivaFinalizacao(userId, tdpf)
        if efetivou!=None:
            if efetivou:
                eliminaPendencia(userId) #apaga a pendência de informação do usuário
                if msgAnulacao==None or msgAnulacao=="":
                    response_message = "TDPF(s) excluído(s) do monitoramento deste serviço."
                else:
                    response_message = msgAnulacao
                enviaMsgBot(bot, userId, text=response_message) 
                mostraMenuPrincipal(update, context)
                return
            else:
                response_message = msgAnulacao+textoRetorno            
    enviaMsgBot(bot, userId, text=response_message)  
    return

def verificaEMail(email): #valida o e-mail se o usuário informou um completo
    regex1 = '^[a-zA-Z0-9]+[\._]?[a-zA-Z0-9\.\-]+[@]\w+[.]\w{2,3}$'
    regex2 = '^[a-zA-Z0-9]+[\._]?[a-zA-Z0-9\.\-]+[@]\w+[.]\w+[.]\w{2,3}$' 

    if(re.search(regex1,email)):  
        return True   
    elif(re.search(regex2,email)):  
        return True
    else:  
        return False 

def cadastraEMail(update, context): #cadastra o e-mail institucional
    global pendencias, textoRetorno
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    msg = update.message.text    
    parametros = getParametros(msg)
    response_message = ""
    if len(parametros)!=1:
        response_message = "Envie somente o nome de usuário do e-mail institucional @rfb.gov.br."
        response_message = response_message+textoRetorno
    else:        
        email = parametros[0].strip()
        if "@" in email:
            if not verificaEMail(email):
                response_message = "Email inválido. Envie somente o nome de usuário do e-mail institucional @rfb.gov.br."
                response_message = response_message+textoRetorno
            elif not "@rfb.gov.br" in email:
                response_message = "Email não é institucional. Envie somente o nome de usuário do e-mail institucional @rfb.gov.br."
                response_message = response_message+textoRetorno 
            else:
                email = email[:email.find("@")]
            logging.info(email) 
        elif len(email)<4:
            response_message = "Nome muito curto (min = 4 caracteres). Envie somente o nome de usuário do endereço @rfb.gov.br."
            response_message = response_message+textoRetorno 
        if response_message=="":    
            conn = conecta()
            if not conn: 
                return 
            comando = "Select Codigo, CPF, email from Usuarios Where Saida Is Null and idTelegram=%s and Ativo='S'"
            cursor = conn.cursor(buffered=True)
            try:
                cursor.execute(comando, (userId,))
                row = cursor.fetchone()
            except:
                response_message = "Erro na consulta (6)."+textoRetorno
                enviaMsgBot(bot, userId, text=response_message)  
                conn.close()
                return                
            if not row:
                response_message = "Usuário não registrado ou inativo/desabilitado no serviço."
                enviaMsgBot(bot, userId, text=response_message) 
                mostraMenuPrincipal(update, context)
                conn.close()
                return                 
            else:
                chave = row[0]
                cpf = row[1]
                cpf = cpf[:3]+"."+cpf[3:6]+"."+cpf[6:9]+"-"+cpf[9:]
                emailAnt = row[2]
                eliminaPendencia(userId) #apaga a pendência de informação do usuário                 
                if email.upper()=="NULO" and (emailAnt==None or emailAnt==""):
                    response_message = "Usuário não tinha e-mail cadastrado no serviço."
                    enviaMsgBot(bot, userId, text=response_message) 
                    mostraMenuPrincipal(update, context)
                    conn.close()
                    return
                try:
                    if email.upper()=="NULO":
                        comando = "Update Usuarios Set email=Null Where Codigo=%s"   
                        cursor.execute(comando, (chave,))
                    else:    
                        comando = "Update Usuarios Set email=%s Where Codigo=%s"  
                        cursor.execute(comando, (email+'@rfb.gov.br', chave))
                    conn.commit()                   
                    if emailAnt!=None and emailAnt!="":
                        if email.upper()=="NULO":
                            response_message = "Email anteriormente informado ("+emailAnt+") foi descadastrado."
                        else:    
                            response_message = "Email anteriormente cadastrado ("+emailAnt+") foi substituído.\n"                
                    if email.upper()!="NULO":
                        response_message = response_message+"Email cadastrado com sucesso - {}@rfb.gov.br.".format(email)    
                    enviaMsgBot(bot, userId, text=response_message) 
                    mostraMenuPrincipal(update, context)
                    conn.close()
                    return 
                except:
                    conn.close()
                    response_message = "Erro na atualização das tabelas. Tente novamente mais tarde."
                    enviaMsgBot(bot, userId, text=response_message) 
                    mostraMenuPrincipal(update, context)  
                    return                  
    enviaMsgBot(bot, userId, text=response_message)  
    return


def start(update, context): #comandos /start /menu /retorna acionam esta opção
    global pendencias
    
    userId = update.message.from_user.id  
    bot = update.effective_user.bot
    logging.info(update.message.from_user.first_name+" - "+str(userId))    
    if update.effective_user.is_bot:
        return #não atendemos bots     
    eliminaPendencia(userId)
    msg1 = 'Este serviço controla, dentre outros, os prazos p/ recuperação da espontaneidade, p/ vencimento do TDPF, vencimento de prazo de intimação e de atividades cadastradas:\n'
    msg2 = '- Alertas sobre a possível recuperação da espontaneidade de contribuintes, em prazos customizáveis (d1 [maior], d2 e d3 [menor] dias antes).\n'
    msg3 = '- Alertas sobre o vencimento do TDPF em duas datas distintas (separadas por não menos do que 8 dias e quando o vencimento se der em até 15 dias).\n'
    msg4 = '- Alertas sobre o vencimento do prazo para atendimento de intimação, se informado - no dia do vencimento.\n'
    msg5 = '- Alertas sobre o vencimento de atividades cadastradas - em d3 e no dia do vencimento.\n\n'
    msg6 = 'Atualmente, o vencimento do TDPF informado por este serviço poderá ocorrer com alguma antecedência devido ao cálculo baseado apenas na data de distribuição. '
    msg7 = 'Isto será corrigido futuramente.\n\n'
    msg8 = '*Digite a qualquer momento /menu para ver o menu principal e estas observações*, inclusive no caso de ocorrer alguma interrupção do serviço.'
    response_message = msg1+msg2+msg3+msg4+msg5+msg6+msg7+msg8
    enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    mostraMenuPrincipal(update, context)
    return
  
def unknown(update, context):
    global pendencias, dispatch
    
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    logging.info(update.message.from_user.first_name+" - "+str(userId)) 
    if update.effective_user.is_bot:
        return #não atendemos bots 
    mensagem = update.message.text
    logging.info(mensagem)    
    if pendencias.get(userId, "1")!="1": #há pendências - então usuário tinha acionado uma opção do menu e preencheu informações (válidas = parâmetros, ou não) 
        dispatch[pendencias[userId]](update, context) #encaminha para a função adequada à pendência
    else:            
        response_message = update.message.from_user.first_name+", esta mensagem/comando não significou nada para mim. Mostrarei o menu para lhe auxiliar."  
        enviaMsgBot(bot, userId, text=response_message)
        mostraMenuPrincipal(update, context) 
    return

def mostraMenuCadastro(update, context):
    global pendencias
    userId = update.effective_user.id   
    if update.effective_user.is_bot:
        return #não atendemos bots      
    opcao1 =  tipoOpcao1(userId) 
    bot = update.effective_user.bot     
    if opcao1[:4] == 'Erro':
        enviaMsgBot(bot, userId, text="Erro na consulta ao seu id")
        return
    menu = [[opcao1], ['Solicita Chave de Registro e do ContÁgil'], ['Prazos Para Receber Avisos'], ['Cadastra/Exclui e-Mail'], ['Menu Principal']] 
    #mensagem = enviaMsgBot(bot, userId, text="Teste apaga mensagem")
    #time.sleep(5)
    #bot.delete_message(userId, mensagem.message_id)
    #return              
    update.message.reply_text("Menu Cadastro:", reply_markup=ReplyKeyboardMarkup(menu, one_time_keyboard=True)) 
    return 

def mostraMenuTDPF(update, context):
    global pendencias
    #menu = [['Informa Data de Ciência Relativa a TDPF', 'Anula Ciência Relativa a TDPF'], 
    #        ['Mostra TDPFs Monitorados', 'Mostra TDPFs Supervisionados'],
    #        ['Monitora TDPF(s)', 'Finaliza Monitoramento de TDPF'], 
    #        ['Menu Principal']]
    menu = [['Espontaneidade e Atividades', 'Relaciona Postagens'],
            ['Monitora TDPF(s)', 'Finaliza Monitoramento de TDPF'], 
            ['Mostra TDPFs Monitorados', 'Supervisão'],            
            ['Menu Principal']]
    #userId = update.effective_user.id  
    #bot = update.effective_user.bot     
    if update.effective_user.is_bot:
        return #não atendemos bots                
    update.message.reply_text("Menu TDPF:", reply_markup=ReplyKeyboardMarkup(menu, one_time_keyboard=True))  
    return  

def mostraMenuSupervisao(update, context):
    global pendencias
    menu = [['Mostra TDPFs Supervisionados'],
            ['Envia Atividades (e-Mail) - Superv.'], 
            ['Recuperação Espontaneidade - Superv.'],
            ['Menu TDPF - Monitoramento'],
            ['Menu Principal']]   
    if update.effective_user.is_bot:
        return #não atendemos bots                
    update.message.reply_text("Menu Supervisão:", reply_markup=ReplyKeyboardMarkup(menu, one_time_keyboard=True))  
    return     

def mostraMenuCienciasAtividades(update, context):
    global pendencias
    menu = [['Informa Data de Ciência', 'Anula Data de Ciência Informada'], 
            ['Informa Ativ. e Prazo', 'Anula Atividade'], 
            ['Informa Horas Atividade', 'Informa Término de Ativ.'], 
            ['Ciências e Atividades - Email', 'Menu TDPF - Monitoramento'],
            ['Menu Principal']]   
    if update.effective_user.is_bot:
        return #não atendemos bots                
    update.message.reply_text("Menu Espontaneidade e Atividades:", reply_markup=ReplyKeyboardMarkup(menu, one_time_keyboard=True))  
    return             

def mostraMenuPrincipal(update, context):
    global pendencias
    #userId = update.effective_user.id     
    menu = [['Cadastros'], ['TDPF - Monitoramento']]    
    #bot = update.effective_user.bot     
    if update.effective_user.is_bot:
        return #não atendemos bots  
    update.message.reply_text("Menu Principal:", reply_markup=ReplyKeyboardMarkup(menu, one_time_keyboard=True))
    return  

def tipoOpcao1(userId):
    conn = conecta()
    if not conn: 
        return "Erro na conexão"
    cursor = conn.cursor(buffered=True)
    #comando = "Select CPF, Saida from Usuarios Where idTelegram=?"
    comando = "Select CPF, Saida from Usuarios Where Ativo='S' and idTelegram="+str(userId)
    #cursor.execute(comando, float(userId))
    try:
        #cursor.execute(comando, float(userId))
        cursor.execute(comando)
        row = cursor.fetchone()
        conn.close()
        if not row:
            return "Registra Usuário"
        else:
            if row[1]==None:
                return "Desativa Usuário"
            else:
                return "Reativa Usuário"
    except:
        conn.close()
        return "Erro na consulta (tipoOpcao)"            


def opcaoUsuario(update, context): #Cadastra usuário
    global pendencias
    if update.effective_user.is_bot:
        return #não atendemos bots      
    userId = update.effective_user.id
    bot = update.effective_user.bot     
    eliminaPendencia(userId)     
    msgOpcao1 = tipoOpcao1(userId)
    conn = None
    if msgOpcao1=="Registra Usuário" or msgOpcao1=="Reativa Usuário":       
        pendencias[userId] = 'registra' #usuário agora tem uma pendência de informação
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora, numa única mensagem, seu *CPF e o código de registro (chave)* (separe as informações com espaço):"  
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    else:
        conn = conecta()
        if not conn: 
            response_message = "Erro na conexão (8)."            
            enviaMsgBot(bot, userId, text=response_message)      
            return             
        cursor = conn.cursor(buffered=True)        
        if msgOpcao1=="Desativa Usuário":
            dataAtual = datetime.now().date()
            comando = "Update Usuarios Set Saida=%s Where idTelegram=%s"
            try:    
                cursor.execute(comando, (dataAtual, userId))
                conn.commit()
                response_message = "Usuário desativado."  
                enviaMsgBot(bot, userId, text=response_message)                                 
            except:
                conn.rollback()
                response_message = "Erro ao atualizar tabelas(8). Tente novamente mais tarde."            
                enviaMsgBot(bot, userId, text=response_message)
            mostraMenuPrincipal(update, context)  
    if conn:
        conn.close()                                                    
    return

def solicitaChaveRegistro(update, context): #envia chave de registro para o e-mail institucional do AFRFB (um envio a cada 24h)   
    global pendencias    
    userId = update.effective_user.id  
    bot = update.effective_user.bot       
    if update.effective_user.is_bot:
        return #não atendemos bots  
    eliminaPendencia(userId)                          
    pendencias[userId] = 'envioChave' #usuário agora tem uma pendência de informação (atividade)
    response_message = "Envie /menu para ver o menu principal.\nEnvie agora, numa única mensagem, o *nº do CPF (11 dígitos)* do usuário (fiscal) p/ o qual a chave será enviada:"  
    enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    return    

def verificaUsuario(userId, bot): #verifica se o usuário está cadastrado e ativo no serviço
    global textoRetorno
    conn = conecta()
    if not conn:
        response_message = "Erro na conexão (7)"
        enviaMsgBot(bot, userId, text=response_message)
        return False
    cursor = conn.cursor(buffered=True)
    comando = "Select CPF from Usuarios Where Saida Is Null and BloqueiaTelegram='N' and idTelegram=%s and Ativo='S'"
    try:
        cursor.execute(comando, (userId,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            enviaMsgBot(bot, userId, text="Usuário não está registrado no serviço, comunicação via Telegram está bloqueada ou usuário está inativo/desabilitado. "+textoRetorno)  
            return False
        else:
            return True
    except:
        try:
            conn.close()
        except:
            pass    
        enviaMsgBot(bot, userId, text="Erro na consulta (8).")        
        return False                

def opcaoInformaAtividade(update, context): #informa atividade relativa a TDPF e data de vencimento ou prazo em dias
    global pendencias   
    if update.effective_user.is_bot:
        return #não atendemos bots      
    userId = update.effective_user.id  
    bot = update.effective_user.bot  
    eliminaPendencia(userId)             
    achou = verificaUsuario(userId, bot)       
    if achou:                      
        pendencias[userId] = 'atividade' #usuário agora tem uma pendência de informação (atividade)
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora, numa única mensagem, o *nº do TDPF (16 dígitos), a data de vencimento (dd/mm/aaaa) ou o prazo de vencimento em dias (contados de hoje) e a data de início da atividade (dd/mm/aaaa)* - separe as informações com espaço:"  
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    else:
        mostraMenuPrincipal(update, context)
    return    

def exibeAtividadesEmAndamento(bot, userId, conn): #exibe atividades em andamento de TDPFs não encerrados para anulação, informação do término ou de horas dispendidas
    consulta = """
                Select Atividades.Codigo, TDPFS.Numero, Atividades.Atividade, Atividades.Inicio, Atividades.Vencimento, Atividades.Horas
                from Atividades, TDPFS, Alocacoes, Usuarios, Fiscais
                Where Usuarios.idTelegram=%s and Usuarios.CPF=Fiscais.CPF and Fiscais.Codigo=Alocacoes.Fiscal and Alocacoes.Desalocacao Is Null and 
                Alocacoes.TDPF=TDPFS.Codigo and TDPFS.Encerramento Is Null and TDPFS.Codigo=Atividades.TDPF and Atividades.Termino Is Null
                Order by Atividades.Inicio, Atividades.Vencimento
                """    
    cursor = conn.cursor(buffered=True)
    try:
        cursor.execute(consulta, (userId,))
        linhas = cursor.fetchall()
    except:    
        enviaMsgBot(bot, userId, text="Erro na consulta das atividades em andamento")
        return 0 
    i = 0
    msg = ""
    for linha in linhas:
        i+=1
        msg = msg+"\n"+str(i)+") Código "+str(linha[0])+"; TDPF "+formataTDPF(linha[1])+"; "+linha[2]+"; Início "+datetime.strftime(linha[3], "%d/%m/%Y")+"; Vencimento "+datetime.strftime(linha[4], "%d/%m/%Y")+"; Horas "+str(linha[5])
        if i%15==0:
            enviaMsgBot(bot, userId, text="Atividades em Andamento:"+msg)
            msg = ""
    if msg!="":
        enviaMsgBot(bot, userId, text="Atividades em Andamento:"+msg)
    if i==0:
        enviaMsgBot(bot, userId, text="Usuário não possui atividades cadastradas em andamento para seus TDPFs.")            
    return i       

def opcaoAnulaAtividade(update, context): #anula informação de atividade
    global pendencias  
    if update.effective_user.is_bot:
        return #não atendemos bots         
    userId = update.effective_user.id  
    bot = update.effective_user.bot    
    eliminaPendencia(userId)  
    achou = verificaUsuario(userId, bot)       
    if achou:    
        conn = conecta()
        if not conn:
            enviaMsgBot(bot, userId, text="Erro ao tentar conectar ao Banco de Dados - opcaoAnulaAtividade")
            mostraMenuPrincipal(update, context)
            return
        if exibeAtividadesEmAndamento(bot, userId, conn)==0:
            conn.close()
            mostraMenuPrincipal(update, context) 
            return                
        pendencias[userId] = 'anulaAtividade'  #usuário agora tem uma pendência de informação   
        response_message = "Envie /menu para ver o menu principal.\nEnvie o *nº do TDPF (16 dígitos, sem espaços) e, opcionalmente, o código da atividade* a ser excluída (se o código não for informado, será excluída a última cadastrada p/ o TDPF) - separe as informa;óes (TDPF e código) com espaço."
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
        conn.close()
    else:
        mostraMenuPrincipal(update, context)         
    return

#'Informa Término de Atividade'
def opcaoInformaTerminoAtividade(update, context): 
    global pendencias 
    if update.effective_user.is_bot:
        return #não atendemos bots      
    userId = update.effective_user.id  
    bot = update.effective_user.bot 
    eliminaPendencia(userId)             
    achou = verificaUsuario(userId, bot)       
    if achou:    
        conn = conecta()
        if not conn:
            enviaMsgBot(bot, userId, text="Erro ao tentar conectar ao Banco de Dados - opcaoInformaTerminoAtividade")
            mostraMenuPrincipal(update, context)
            return
        if exibeAtividadesEmAndamento(bot, userId, conn)==0:
            conn.close()
            mostraMenuPrincipal(update, context) 
            return
        pendencias[userId] = 'informaTerminoAtividade' #usuário agora tem uma pendência de informação (término atividade)
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora, numa única mensagem, *o código da atividade, a data de seu término (dd/mm/aaaa) e quantidade de horas dispendidas (número inteiro)* até o momento - separe as informações com espaço:"  
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
        conn.close()
    else:
        mostraMenuPrincipal(update, context)   
    return

#'Informa horas - usuário pode informar horas gastas a qualquer momento e depois ir alterando (exceto após o término)
def opcaoInformaHorasAtividade(update, context): 
    global pendencias 
    if update.effective_user.is_bot:
        return #não atendemos bots      
    userId = update.effective_user.id  
    bot = update.effective_user.bot 
    eliminaPendencia(userId)             
    achou = verificaUsuario(userId, bot)       
    if achou:    
        conn = conecta()
        if not conn:
            enviaMsgBot(bot, userId, text="Erro ao tentar conectar ao Banco de Dados - opcaoInformaHorasAtividade")
            mostraMenuPrincipal(update, context)
            return
        if exibeAtividadesEmAndamento(bot, userId, conn)==0:
            conn.close()
            mostraMenuPrincipal(update, context) 
            return
        pendencias[userId] = 'informaHorasAtividade' #usuário agora tem uma pendência de informação (horas atividade)
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora, numa única mensagem, *o código da atividade e a quantidade de horas dispendidas* até o momento (número inteiro) - separe as informações com espaço:"  
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
        conn.close()
    else:
        mostraMenuPrincipal(update, context)   
    return    

def opcaoEnviaCienciasAtividades(update, context): #Envia para o e-mail do usuário relatório de ciências e atividades de TDPFs em andamento nos quais esteja alocado
    global pendencias
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id  
    bot = update.effective_user.bot 
    eliminaPendencia(userId)              
    achou = verificaUsuario(userId, bot)       
    if not achou: 
        mostraMenuPrincipal(update, context)
        return                     
    conn = conecta()
    if not conn:
        enviaMsgBot(bot, userId, text="Erro ao criar conexão ao Banco de Dados - opcaoEnviaCienciasAtividades")
        mostraMenuPrincipal(update, context)
        return
    consulta = """
                Select TDPFS.Numero, TDPFS.Emissao, TDPFS.Nome, Usuarios.email, TDPFS.Codigo, TDPFS.Tipo, TDPFS.TDPFPrincipal
                From TDPFS, Usuarios, Alocacoes, Fiscais
                Where Usuarios.idTelegram=%s and Usuarios.Saida Is Null and Usuarios.Adesao Is Not Null and Usuarios.CPF=Fiscais.CPF and Fiscais.Codigo=Alocacoes.Fiscal
                and Alocacoes.Desalocacao Is Null and TDPFS.Codigo=Alocacoes.TDPF and TDPFS.Encerramento Is Null and (TDPFS.Tipo='F' or TDPFS.Tipo='D')
                Order by TDPFS.Numero
                """
    cursor = conn.cursor(buffered=True)           
    cursor.execute(consulta, (userId,))
    linhas = cursor.fetchall()
    bAchou = True
    if linhas==None:
        bAchou = False
    if len(linhas)==0:
        bAchou = False  
    if not bAchou:
        enviaMsgBot(bot, userId, text="Não há TDPFs em andamento em que o usuário esteja alocado")
        mostraMenuPrincipal(update, context)
        conn.close()
        return  
    email = linhas[0][3]   
    if email==None or email=="": #email vazio
        enviaMsgBot(bot, userId, text="Email do usuário não foi informado - não haverá envio.")
        mostraMenuPrincipal(update, context)  
        conn.close() 
        return 
    if not "@rfb.gov.br" in email: 
        enviaMsgBot(bot, userId, text="Email do usuário é inválido - aparenta não ser institucional - não haverá envio.")
        mostraMenuPrincipal(update, context)  
        conn.close() 
        return                                            
    book = Workbook()
    sheet2 = book.active 
    sheet2.title = "Ciências"    
    sheet1 = book.create_sheet(title="Atividades") 
    sheet1.cell(row=1, column=1).value = "TDPF"
    sheet1.cell(row=1, column=2).value = "Data Emissão"
    sheet1.cell(row=1, column=3).value = "Nome Fiscalizado"
    sheet1.cell(row=1, column=4).value = "Atividade"
    sheet1.cell(row=1, column=5).value = "Data de Início"            
    sheet1.cell(row=1, column=6).value = "Vencimento"
    sheet1.cell(row=1, column=7).value = "Término"
    sheet1.cell(row=1, column=8).value = "Horas"
    sheet1.cell(row=1, column=9).value = "Observacoes"
    sheet1.cell(row=1, column=10).value = "Tipo do Procedimento"  
    larguras = [19, 13, 32, 27, 13, 13, 13, 8, 60, 18]
    for col in range(len(larguras)):
        sheet1.column_dimensions[get_column_letter(col+1)].width = larguras[col]   
        currentCell = sheet1.cell(row=1, column=col+1)
        currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)                 
    sheet2.cell(row=1, column=1).value = "TDPF"
    sheet2.cell(row=1, column=2).value = "Data Emissão"
    sheet2.cell(row=1, column=3).value = "Nome Fiscalizado"
    sheet2.cell(row=1, column=4).value = "Data de Ciência"
    sheet2.cell(row=1, column=5).value = "Documento"     
    sheet2.cell(row=1, column=6).value = "Vencimento da Intimação"             
    sheet2.cell(row=1, column=7).value = "60 dias da Ciência"  
    sheet2.cell(row=1, column=8).value = "Tipo do Procedimento"  
    larguras = [19, 13, 32, 14, 25, 18, 16, 18]
    for col in range(len(larguras)):
        sheet2.column_dimensions[get_column_letter(col+1)].width = larguras[col]  
        currentCell = sheet2.cell(row=1, column=col+1)
        currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)             
    i = 1
    j = 1
    for linha in linhas:                             
        tdpf = linha[0]
        chaveTdpf = linha[4]
        tipo = linha[5]
        if tipo=="F":
            tipoE = "Fiscalização"
        elif linha[6]!=None:
            tipoE = "Diligência Vinculada"
        else:
            tipoE = "Diligência"
        emissao = linha[1]
        if emissao==None:
            emissao = ""
        else:
            emissao = emissao.date()    
        fiscalizado = linha[2]
        if fiscalizado==None:
            fiscalizado = ""                
        consulta = "Select Atividade, Inicio, Vencimento, Termino, Horas, Observacoes from Atividades Where TDPF=%s Order By Inicio"
        cursor.execute(consulta,(chaveTdpf,))
        rows = cursor.fetchall()                  
        for row in rows:
            currentCell = sheet1.cell(row=i+1, column=1)
            currentCell.value = formataTDPF(tdpf)
            currentCell.alignment = Alignment(horizontal='center')             
            currentCell = sheet1.cell(row=i+1, column=2)
            currentCell.value = emissao
            currentCell.alignment = Alignment(horizontal='center')             
            sheet1.cell(row=i+1, column=3).value = fiscalizado                           
            sheet1.cell(row=i+1, column=4).value = row[0]
            sheet1.cell(row=i+1, column=10).value = tipoE
            if row[1]!=None:
                sheet1.cell(row=i+1, column=5).value = row[1].date()
                currentCell = sheet1.cell(row=i+1, column=5)
                currentCell.alignment = Alignment(horizontal='center')                 
            if row[2]!=None:                              
                sheet1.cell(row=i+1, column=6).value = row[2].date()
                currentCell = sheet1.cell(row=i+1, column=6)
                currentCell.alignment = Alignment(horizontal='center')                  
            if row[3]!=None:                              
                sheet1.cell(row=i+1, column=7).value = row[3].date()
            if row[4]!=None:
                sheet1.cell(row=i+1, column=8).value = row[4]            
            if row[5]!=None:
                sheet1.cell(row=i+1, column=9).value = row[5]                    
            if row[2]!=None and row[3]==None:
                cor = None
                if row[2].date()==datetime.now().date(): #se está vencendo hoje, fica azul
                    cor = Font(color="0000FF")
                if row[2].date()<datetime.now().date(): #se atividade está vencida, fica vermelha
                    cor = Font(color="FF0000") 
                if cor!=None:
                    for col in range(9):
                        sheet1.cell(row=i+1, column=col+1).font = cor
            i+=1  
        consulta = "Select Data, Documento, Vencimento from Ciencias Where TDPF=%s and Data Is Not Null Order By Data"
        cursor.execute(consulta,(chaveTdpf,))
        rows = cursor.fetchall()      
        totalRows = len(rows) 
        rowAtual = 1           
        for row in rows: 
            currentCell = sheet2.cell(row=j+1, column=1)
            currentCell.value = formataTDPF(tdpf)
            currentCell.alignment = Alignment(horizontal='center')             
            currentCell = sheet2.cell(row=j+1, column=2)
            currentCell.value = emissao
            currentCell.alignment = Alignment(horizontal='center')             
            sheet2.cell(row=j+1, column=3).value = fiscalizado                           
            currentCell = sheet2.cell(row=j+1, column=4)
            currentCell.value = row[0].date() #data da ciência
            currentCell.alignment = Alignment(horizontal='center')             
            if row[1]!=None: #documento
                sheet2.cell(row=j+1, column=5).value = row[1]                
            if row[2]!=None: #vencimento da intimação
                sheet2.cell(row=j+1, column=6).value = row[2].date()
                currentCell = sheet2.cell(row=j+1, column=6)                
                currentCell.alignment = Alignment(horizontal='center')  
            sheet2.cell(row=j+1, column=8).value = tipoE
            if tipo=="F":                                    
                diaEspont = (row[0]+timedelta(days=60)).date()
                sheet2.cell(row=j+1, column=7).value = diaEspont #dia que recupera a espontaneidade
                cor = None
                if diaEspont<=(datetime.now()+timedelta(days=15)).date() and rowAtual==totalRows: #se faltar 15 dias ou menos para recuperar a espontaneidade, 
                                                                                                #a linha fica azul, mas só se for na última ciência
                    cor = Font(color="0000FF")
                if diaEspont<datetime.now().date() and rowAtual==totalRows: #se a espontaneidade já foi recuperada, a linha fica vermelha na última ciência
                    cor = Font(color="FF0000") 
                elif diaEspont<datetime.now().date(): #se a espontaneidade tiver sido recuperada em ciência que não é a última, fica na cor roxa
                    #temos que verificar em relação à ciência seguinte (há uma próxima ciência)
                    if (rows[rowAtual][0]-row[0]).days>60: #entre a ciência atual e a subsequente, decorreram mais de 60 dias
                        cor = Font(color="800080")
                if cor!=None:
                    for col in range(7):
                        sheet2.cell(row=j+1, column=col+1).font = cor   
            rowAtual+=1         
            j+=1
    if i>1 or j>1:              
        nomeArq = "CiencAtiv_"+str(userId)+"_"+datetime.now().strftime("%Y_%m_%d_%H_%M_%S")+".xlsx"
        book.save(nomeArq)   
        message = "Prezado(a),\n\nConforme solicitado, enviamos, em anexo, planilha com relação das ciências e das atividades de TDPFs em andamento sob sua responsabilidade.\n\nAtenciosamente,\n\nDisav/Cofis\n\nAmbiente: "+ambiente    
        resultado = enviaEmail(email, message, "Relação de Ciências e Atividades - TDPFs", nomeArq)
        if resultado!=3:
            msg = "Erro no envio de email - opcaoEnviaCienciasAtividades - "+str(resultado)
            logging.info(msg + " - "+email)
            enviaMsgBot(bot, userId, text=msg)
        else:
            enviaMsgBot(bot, userId, text="E-mail enviado.")   
        os.remove(nomeArq)
    else:    
        enviaMsgBot(bot, userId, text="Não há atividades e ciências informadas relativamente aos TDPFs em andamento sob sua responsabilidade.")
    conn.close()
    mostraMenuPrincipal(update, context)         
    return   

def opcaoInformaCiencia(update, context): #Informa ciência de TDPF
    global pendencias
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id  
    bot = update.effective_user.bot 
    eliminaPendencia(userId)              
    achou = verificaUsuario(userId, bot)       
    if achou:                      
        pendencias[userId] = 'ciencia' #usuário agora tem uma pendência de informação (ciência)
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora, numa única mensagem, *o nº do TDPF (16 dígitos), a data de ciência (dd/mm/aaaa) e, opcionalmente, a data de vencimento da intimação (ou prazo em dias corridos*; será aplicada a regra de que não começa nem termina em sábado/domingo) válida para fins de perda da espontaneidade tributária relativa ao respectivo procedimento fiscal - separe as informações com espaço:"  
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    else:
        mostraMenuPrincipal(update, context) 
    return
        
def opcaoPrazos(update, context): #Informa prazos para receber avisos
    global pendencias, textoRetorno  
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id 
    bot = update.effective_user.bot     
    eliminaPendencia(userId)     
    comando = "Select d1, d2, d3, Saida from Usuarios Where idTelegram=%s and Ativo='S'"
    saida = None
    conn = conecta()
    if not conn:
        response_message = "Erro na conexão (5)"
        enviaMsgBot(bot, userId, text=response_message)
        return     
    try:
        cursor = conn.cursor(buffered=True)    
        cursor.execute(comando, (userId,))
        achou = False
        row = cursor.fetchone()
        if row:
            achou = True
            d1 = row[0]
            d2 = row[1]
            d3 = row[2]
            saida = row[3] 
    except:
        conn.close()
        response_message = "Erro na consulta. (5)"
        enviaMsgBot(bot, userId, text=response_message)
        return             
    if not achou:
        response_message = "Usuário não está registrado no serviço ou está desabilitado. "+textoRetorno
    else:    
        if saida!=None:
            response_message = "Usuário está inativo no serviço. "+textoRetorno            
        else:    
            pendencias[userId] = 'prazos'   #usuário agora tem uma pendência de informação
            response_message = "Envie /menu para ver o menu principal.\nPrazos vigentes para receber alertas: *{} (d1), {} (d2) e {} (d3) dias* antes de o contribuinte readquirir a espontaneidade.\nEnvie agora, numa única mensagem, *três quantidades de dias (1 a 50) distintas* antes de o contribuinte readquirir a espontaneidade tributária em que você deseja receber alertas (separe as informações com espaço):".format(d1, d2, d3)
    enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    conn.close()  
    return

def opcaoAnulaCiencia(update, context): #Anula ciência de TDPF
    global pendencias    
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id  
    bot = update.effective_user.bot    
    eliminaPendencia(userId)  
    achou = verificaUsuario(userId, bot)       
    if achou:            
        pendencias[userId] = 'anulaCiencia'  #usuário agora tem uma pendência de informação   
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora *o nº do TDPF (16 dígitos)* para o qual você deseja anular a última ciência informada que impedia a recuperação da espontaneidade (retornará para a anterior):"
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    else:
        mostraMenuPrincipal(update, context)         
    return

def opcaoFinalizaAvisos(update, context): #Finaliza avisos para um certo TDPF
    global pendencias    
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id
    bot = update.effective_user.bot      
    eliminaPendencia(userId)  
    achou = verificaUsuario(userId, bot)       
    if achou:   
        pendencias[userId] = 'fim'     #usuário agora tem uma pendência de informação
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora *o nº do TDPF (16 dígitos) ou a palavra TODOS* para finalizar alertas/monitoramento de um ou de todos os TDPFs:"
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    else:
        mostraMenuPrincipal(update, context)         
    return

def opcaoAcompanhaTDPFs(update, context): #acompanha um TDPF ou todos os TDPFs em que estiver alocado ou em que for supervisor
    global pendencias    
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id
    bot = update.effective_user.bot     
    eliminaPendencia(userId)  
    achou = verificaUsuario(userId, bot)       
    if achou:   
        pendencias[userId] = 'acompanha' #usuário agora tem uma pendência de informação
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora o *nº do TDPF (16 dígitos) ou a palavra TODOS* (incluirá apenas fiscalizações) para receber alertas relativos ao TDPF informado ou a todos em que estiver alocado e/ou que for supervisor:"  
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    else:
        mostraMenuPrincipal(update, context)         
    return
        
def montaListaTDPFs(userId, tipo=1): #monta lista de TDPFs com diversas informações que dependem se é fiscal alocado (tipo=1) ou supervisor (tipo=2)
    conn = conecta()
    if not conn:
        return  ["Erro na conexão"]
    try:
        cursor = conn.cursor(buffered=True)
        comando = "Select CPF from Usuarios Where idTelegram=%s and Saida Is Null and Ativo='S'"
        cursor.execute(comando, (userId,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return None
        cpf = row[0]
        if tipo==1:
        #seleciona monitoramentos ativos (incluído pelo usuário, ainda alocado nele e não encerrado)
            comando = '''Select TDPFS.Numero, TDPFS.Vencimento, TDPFS.Codigo, Alocacoes.Supervisor, TDPFS.Tipo from CadastroTDPFs, Alocacoes, Fiscais, TDPFS
                        Where Fiscais.CPF=%s and CadastroTDPFs.Fiscal=Fiscais.Codigo and CadastroTDPFs.Fim Is Null and CadastroTDPFs.Fiscal=Alocacoes.Fiscal and 
                        CadastroTDPFs.TDPF=Alocacoes.TDPF and CadastroTDPFs.TDPF=TDPFS.Codigo and Alocacoes.Desalocacao Is Null and 
                        TDPFS.Encerramento Is Null Order By TDPFS.Numero'''
        elif tipo==2:
        #seleciona todos os TDPFs dos quais o usuário é supervisor
            comando = """ 
                        Select TDPFS.Numero, TDPFS.Vencimento, TDPFS.Codigo, TDPFS.Tipo from TDPFS, Supervisores, Fiscais Where Encerramento Is Null and 
                        Supervisores.Equipe=TDPFS.Grupo and Supervisores.Fim Is Null and Fiscais.CPF=%s and Fiscais.Codigo=Supervisores.Fiscal 
                        and (TDPFS.Tipo='F' or TDPFS.Tipo='D')
                        Order by TDPFS.Numero
                        """
            #comando = '''Select Alocacoes.TDPF as tdpf, Vencimento from Alocacoes, TDPFS Where Desalocacao Is Null and Encerramento Is Null 
            #            and CPF=%s and TDPF=Numero and Supervisor='S' Order by TDPF'''
        else:
            conn.close()
            return None  
        cursor.execute(comando, (cpf,))
        listaAux = cursor.fetchall()
        if not listaAux:
            conn.close()
            return None
        if len(listaAux)==0:
            conn.close()
            return None            
        result = []
        for linha in listaAux:
            tdpf = linha[0]
            chaveTdpf = linha[2]
            vencimento = linha[1]
            if tipo==1:
                tipoP = linha[4]
            else:
                tipoP = linha[3]    
            if vencimento:
                vencimento = vencimento.date()
                vctoTDPFNum = (vencimento-datetime.today().date()).days
                vctoTDPF = str(vctoTDPFNum)
                if vctoTDPFNum<0:
                    vctoTDPF = vctoTDPF + " (vencido)"
            else:
                vctoTDPF = "ND"
            comando = "Select Data, Documento, Vencimento from Ciencias Where TDPF=%s order by Data DESC"
            #logging.info(comando)
            cursor.execute(comando, (chaveTdpf,))            
            cienciaReg = cursor.fetchone() #busca a data de ciência mais recente (DESC acima)
            if tipo==2: #verificamos se o TDPF está sendo monitorado
                comando = "Select Codigo as codigo from CadastroTDPFs Where TDPF=%s and Fim Is Null"
                cursor.execute(comando, (chaveTdpf,))
                monitoradoReg = cursor.fetchone()
                if monitoradoReg:
                    monitorado = "SIM"
                else:
                    monitorado = "NÃO"       
            tdpfForm = formataTDPF(tdpf)
            documento = ""
            vencimentoCiencia = None
            if cienciaReg:
                if len(cienciaReg)>0: 
                    ciencia = cienciaReg[0] #obtem a data de ciência mais recente
                    documento = cienciaReg[1]
                    vencimentoCiencia = cienciaReg[2]
                else:
                    ciencia = None    
            else:
                ciencia = None
            if tipo==1: #fiscal alocado
                atividades = []
                comando = "Select Codigo, Atividade, Inicio, Vencimento, Horas from Atividades Where TDPF=%s and Termino Is Null order by Inicio"   #somente as atividade em andamento
                cursor.execute(comando, (chaveTdpf,))
                regAtividades = cursor.fetchall()
                for regAtividade in regAtividades:
                    lista = []
                    lista.append(regAtividade[0])
                    lista.append(regAtividade[1])
                    lista.append(regAtividade[2])
                    lista.append(regAtividade[3])
                    lista.append(regAtividade[4])
                    atividades.append(lista)            
                registro = [tdpfForm, linha[3], ciencia, documento, vencimentoCiencia, vctoTDPF, atividades, tipoP]
            else: #supervisor
                registro = [tdpfForm, monitorado, ciencia, documento, vctoTDPF, tipoP]
            result.append(registro)       
        if len(result)>0:
            #logging.info(result)
            conn.close()
            return result
        else:
            conn.close()
            return None
    except:
        conn.close()
        return ["Erro na consulta (9). Tente novamente mais tarde."]
            
        
def opcaoMostraTDPFs(update, context): #Relação de TDPFs monitorados, prazos e atividades do fiscal alocado
    global pendencias, ambiente   
    if update.effective_user.is_bot:
        return #não atendemos bots        
    userId = update.effective_user.id
    bot = update.effective_user.bot      
    eliminaPendencia(userId) 
    achou = verificaUsuario(userId, bot)       
    if not achou: 
        return        
    lista = montaListaTDPFs(userId, 1) #relaciona TDPFs para um usuário comum (NÃO na qualidade de supervisor)
    atividades = []
    if lista==None:
        response_message = "Você não monitora nenhum TDPF ou nenhum deles possui data de ciência informada neste serviço."        
    else:
        if len(lista)==1 and type(lista[0]) is str: #só retornou uma mensagem de erro
            response_message = lista[0]
            enviaMsgBot(bot, userId, text=response_message) 
            return
        i = 1
        msg = ""
        for item in lista:
            tdpf = item[0]
            supervisor = item[1]
            if not supervisor:
                supervisor = "NÃO"
            if supervisor=="S":
                supervisor = "SIM"
            ciencia = item[2]
            documento = item[3]
            vencimentoCiencia = item[4]
            vctoTDPF = item[5]
            if documento==None:
                documento = "ND"
            if vctoTDPF==None:
                vctoTDPF = "ND"    
            atividades.append([tdpf, item[6]]) #item[6] é uma lista de atividades
            tipoP = item[7]
            if ciencia:
                delta = ciencia.date() + timedelta(days=60)-datetime.today().date()
                dias = delta.days
                if vencimentoCiencia==None:
                    parteF = "; f) ND"
                else:
                    vencimentoTxt = vencimentoCiencia.strftime("%d/%m/%Y")
                    if vencimentoCiencia.date()>datetime.now().date():
                        parteF = "; f) "+vencimentoTxt
                    elif vencimentoCiencia.date()==datetime.now().date():
                        parteF = "; f) "+vencimentoTxt +" (HOJE)"
                    else:
                        parteF = "; f) "+vencimentoTxt +" (EXPIRADO)"
                if tipoP=='F': #prazo para espontaneidade só se aplica a fiscalizações
                    if dias<0:
                        dias = "*d) "+str(dias)+" (recuperada)*; e) "+documento+parteF+"; g) "+vctoTDPF
                    else:
                        dias = "*d) "+ str(dias) + "*; e) "+documento+parteF+"; g) "+vctoTDPF    
                else:
                    dias = "d) NA; e) "+documento+parteF+"; g) "+vctoTDPF    
                msg = msg+"\n\n"+str(i)+"a) *"+tdpf+"* ("+tipoP+"); b) "+supervisor+";\nc) "+ciencia.strftime('%d/%m/%Y')+"; "+dias
            else:
                msg = msg+"\n\n"+str(i)+"a) *"+tdpf+"* ("+tipoP+"); b) "+supervisor+"\nc) ND; d) ND; e) ND; f) ND; g) "+vctoTDPF  
            if i%15==0:   #há limite de tamanho de msg - enviamos 15 TDPFs por vez, no máximo
                response_message = "TDPFs Monitorados Por Você:\na) TDPF; b) Supervisor; c) Data da última ciência; d) Dias restantes p/ recuperação da espontaneidade; e) Documento; f) Vencimento do Prazo da Intimação; g) Dias restantes para o vencto. do TDPF:"
                response_message = response_message+msg  
                if ambiente!="PRODUÇÃO":
                    response_message = response_message+"\n\nAmbiente: "+ambiente   
                enviaMsgBot(bot, userId, text=response_message)               
                msg = ""                       
            i+=1                 
        if msg!="":
            response_message = "TDPFs Monitorados Por Você:\na) TDPF; b) Supervisor; c) Data da última ciência; d) Dias restantes p/ recuperação da espontaneidade; e) Documento; f) Vencimento do Prazo da Intimação; g) Dias restantes para o vencto. do TDPF:"
            response_message = response_message+msg
            response_message = response_message + "\n\nVencimento do TDPF pode ser inferior ao do Ação Fiscal, pois as informações do serviço se baseiam"
            response_message = response_message + " na data de distribuição, que pode ter ocorrido antes da emissão do TDPF."             
            if ambiente!="PRODUÇÃO":
                response_message = response_message+"\n\nAmbiente: "+ambiente
            response_message = limpaMarkdown(response_message)
            enviaMsgBot(bot, userId, text=response_message, parse_mode= 'MarkdownV2')                
        response_message = ""         
        msg = ""
        i = 1
        for atividade in atividades:
            logging.info(atividade)
            for registro in atividade[1]:
                horas = registro[4]
                if horas==None:
                    horas = 0
                msg = msg + "\n\n"+str(i)+"a) *"+atividade[0]+"*; b) "+str(registro[0])+"; c) "+registro[1]+"; d) "+registro[2].strftime('%d/%m/%Y')+"; e) "+registro[3].strftime('%d/%m/%Y')+"; f) "+str(horas)
                i+=1
        if msg!="":
            response_message = "Lista de atividades em andamento dos TDPFs Monitorados:\na)TDPF; b) Código; c) Descrição; d) Início; e) Vencimento; f) Horas" + msg            
            if ambiente!="PRODUÇÃO":
                response_message = response_message+"\n\nAmbiente: "+ambiente			
    if response_message!="": 
        response_message = limpaMarkdown(response_message)
        enviaMsgBot(bot, userId, text=response_message, parse_mode= 'MarkdownV2')                
    mostraMenuPrincipal(update, context)
    return

def opcaoMostraSupervisionados(update, context): #acompanha um TDPF ou todos os TDPFs em que estiver alocado ou em que for supervisor
    global pendencias    
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id
    bot = update.effective_user.bot     
    eliminaPendencia(userId)  
    achou = verificaUsuario(userId, bot)       
    if achou:   
        pendencias[userId] = 'mostraSuperv' #usuário agora tem uma pendência de informação
        response_message = "Envie /menu para ver o menu principal.\nDeseja que envie também e-mail (SIM/S ou NÃO/N)?"  
        enviaMsgBot(bot, userId, text=response_message)
    else:
        mostraMenuPrincipal(update, context)         
    return

def mostraSupervisionados(update, context): #Relação de TDPFs supervisionados pelo usuário
    global pendencias, textoRetorno, ambiente
    userId = update.message.from_user.id   
    bot = update.effective_user.bot
    if update.effective_user.is_bot:
        return #não atendemos bots       
    msg = update.message.text    
    parametros = getParametros(msg)
    if len(parametros)!=1:
        response_message = "Envie somente SIM, S, NÃO ou N."
        response_message = response_message+textoRetorno  
        enviaMsgBot(bot, userId, text=response_message)  
        return  
    resposta = parametros[0].upper()
    if not resposta in ["SIM", "S", "NÃO", "N", "NAO"]:
        response_message = "Envie somente SIM, S, NÃO ou N."
        response_message = response_message+textoRetorno  
        enviaMsgBot(bot, userId, text=response_message)
        return          
    resposta = resposta[:1]                 
    eliminaPendencia(userId)       
    lista = montaListaTDPFs(userId, 2)
    if lista==None:
        response_message = "Você não supervisiona nenhum TDPF." 
        enviaMsgBot(bot, userId, text=response_message)
        mostraMenuPrincipal(update, context)  
        return             
    if len(lista)==1 and type(lista[0]) is str:
        response_message = lista[0]
        enviaMsgBot(bot, userId, text=response_message) 
        mostraMenuPrincipal(update, context)         
        return 
    if resposta=='S':
        book = Workbook()
        sheet = book.active  
        sheet.cell(row=1, column=1).value = "TDPF"
        sheet.cell(row=1, column=2).value = "Data Emissão"
        sheet.cell(row=1, column=3).value = "Nome Fiscalizado"
        sheet.cell(row=1, column=4).value = "Auditor-Fiscal"
        sheet.cell(row=1, column=5).value = "DCC"
        sheet.cell(row=1, column=6).value = "Monitorado"
        sheet.cell(row=1, column=7).value = "Última Ciência"
        sheet.cell(row=1, column=8).value = "Dias p/ Recuperação da Espontaneidade"   
        sheet.cell(row=1, column=9).value = "Documento que Efetivou a Ciência"            
        sheet.cell(row=1, column=10).value = "Dias p/ Vencimento do TDPF"
        sheet.cell(row=1, column=11).value = "Horas Alocadas (TODOS os AFRFBs)"
        sheet.cell(row=1, column=12).value = "Tipo do Procedimento"
        sheet.row_dimensions[1].height = 42    
        larguras = [18, 13, 32, 32, 25, 11, 13, 14.5, 25, 13, 13.5, 16]
        for col in range(len(larguras)):
            sheet.column_dimensions[get_column_letter(col+1)].width = larguras[col]  
            currentCell = sheet.cell(row=1, column=col+1)
            currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)                     
    i = 1
    msg = ""
    conn = conecta()  
    if conn:
        cursor = conn.cursor(buffered=True)       
    for item in lista:
        tdpf = item[0]
        monitorado = item[1]
        ciencia = item[2]
        documento = item[3]
        vctoTDPF = item[4]
        if vctoTDPF==None:
            vctoTDPF = "ND"
        if documento==None:
            documento = ""
        tipoP = item[5]
        email = ""
        consulta = """
                    Select TDPFS.Emissao, TDPFS.Nome, Fiscais.Nome, TDPFS.Codigo, TDPFS.DCC From TDPFS, Alocacoes, Fiscais 
                    Where TDPFS.Numero=%s and TDPFS.Codigo=Alocacoes.TDPF and Alocacoes.Desalocacao Is Null and
                    Alocacoes.Supervisor='N' and Alocacoes.Fiscal=Fiscais.Codigo
                    """
        fiscal = ""
        fiscalizado = ""
        emissao = None   
        chaveTdpf = None                    
        if conn:
            cursor.execute(consulta, (getAlgarismos(tdpf),))
            linhas = cursor.fetchall()
            for linha in linhas:
                emissao = linha[0]
                if emissao!=None:
                    emissao = emissao.date()
                else:
                    emissao = ""    
                fiscalizado = linha[1]
                fiscal = linha[2]
                chaveTdpf = linha[3]
                dcc = linha[4]
                if dcc==None:
                    dcc = ""
                else:
                    dcc = dcc[:5]+"."+dcc[5:11]+"/"+dcc[11:15]+"-"+dcc[-2:]
                break
        if resposta=='S':
            consulta = "Select email from Usuarios Where idTelegram=%s"  
            cursor.execute(consulta, (userId,))   
            linhas = cursor.fetchall()
            for linha in linhas:
                email = linha[0]
                break          
            currentCell = sheet.cell(row=i+1, column=1)
            currentCell.value = tdpf
            currentCell.alignment = Alignment(horizontal='center') 
            currentCell = sheet.cell(row=i+1, column=2)
            currentCell.value = emissao
            currentCell.alignment = Alignment(horizontal='center')             
            sheet.cell(row=i+1, column=3).value = fiscalizado
            sheet.cell(row=i+1, column=4).value = fiscal
            currentCell = sheet.cell(row=i+1, column=5)
            currentCell.value = dcc
            currentCell.alignment = Alignment(horizontal='center')             
            currentCell = sheet.cell(row=i+1, column=6)
            currentCell.value = monitorado
            currentCell.alignment = Alignment(horizontal='center')              
            sheet.cell(row=i+1, column=10).value = getAlgarismos(vctoTDPF)
            sheet.cell(row=i+1, column=12).value = "Fiscalização" if tipoP=="F" else "Diligência"
            consulta = "Select SUM(Horas) from Alocacoes Where TDPF=%s"
            if conn:
                cursor.execute(consulta,(chaveTdpf,))
                horasAloc = cursor.fetchone()
                if horasAloc:
                    horas = horasAloc[0]
                else:
                    horas = 0    
                sheet.cell(row=i+1, column=11).value = horas
        if ciencia:
            delta = ciencia.date() + timedelta(days=60)-datetime.today().date()
            dias = delta.days
            if resposta=='S':
                currentCell = sheet.cell(row=i+1, column=7)
                currentCell.value = ciencia.date()
                currentCell.alignment = Alignment(horizontal='center')   
                sheet.cell(row=i+1, column=9).value = documento
                if tipoP=='F':              
                    sheet.cell(row=i+1, column=8).value = int(dias)
                    cor = None
                    if dias<=15: #se faltar 15 dias ou menos para recuperar a espontaneidade, a linha fica azul
                        cor = Font(color="0000FF")
                    if dias<0: #se a espontaneidade já foi recuperada, a linha fica vermelha
                        cor = Font(color="FF0000") 
                    if cor!=None:
                        for col in range(6,9):
                            sheet.cell(row=i+1, column=col+1).font = cor 
            if tipoP=='F':                
                if dias<0:
                    dias = " d) "+str(dias)+" (recuperada); e) "+vctoTDPF
                else:
                    dias = " d) "+ str(dias) + "; e) "+vctoTDPF
            else:
                dias = " d) NA; e) "+vctoTDPF
            msg = msg+"\n\n"+str(i)+"a) *"+tdpf+"*("+fiscal.strip().split()[0]+"-"+tipoP+"); b) "+monitorado+"; c) "+ciencia.strftime('%d/%m/%Y')+";"+dias
        else:
            msg = msg+"\n\n"+str(i)+"a) "+tdpf+"; b) "+monitorado+";c) ND; d) ND; e) "+vctoTDPF              
        if (i % 15) == 0: #há limite de tamanho de msg - enviamos 15 TDPFs por vez, no máximo
            response_message = "TDPFs Supervisionados Por Você:\na) TDPF; b) Monitorado Por Algum Fiscal; c) Data da última ciência; d) Dias restantes p/ recuperação da espontaneidade; e) Dias restantes para o vencto. do TDPF:" 
            response_message = response_message + msg  
            response_message = response_message + "\n\nVencimento do TDPF pode ser inferior ao do Ação Fiscal, pois as informações do serviço se baseiam"
            response_message = response_message + " na data de distribuição, que pode ter ocorrido antes da emissão do TDPF."            
            if ambiente!="PRODUÇÃO":
                response_message = response_message+"\n\nAmbiente: "+ambiente
            response_message = response_message.replace(".", "\.").replace("_", "\_").replace("[", "\[").replace("]", "\]").replace(")", "\)").replace("(", "\(").replace("-","\-")#.replace("*", "\\*")        
            enviaMsgBot(bot, userId, text=response_message, parse_mode= 'MarkdownV2')                  	                                  
            msg = ""
        i+=1  
    if conn:
        conn.close()                   
    if msg!="":    
        response_message = "TDPFs Supervisionados Por Você:\na) TDPF; b) Monitorado Por Algum Fiscal; c) Data da última ciência; d) Dias restantes p/ recuperação da espontaneidade; e) Dias restantes para o vencto. do TDPF:"  
        response_message = response_message+msg
        if ambiente!="PRODUÇÃO":
            response_message = response_message+"\n\nAmbiente: "+ambiente		
            response_message = limpaMarkdown(response_message)
            enviaMsgBot(bot, userId, text=response_message, parse_mode= 'MarkdownV2')   
    if resposta=='S':
        if email!="" and "@rfb.gov.br" in email:
            nomeArq = "Sup_"+str(userId)+"_"+datetime.now().strftime("%Y_%m_%d_%H_%M_%S")+".xlsx"
            book.save(nomeArq)   
            message = "Prezado(a),\n\nConforme solicitado, enviamos, em anexo, planilha com relação dos TDPFs sob sua supervisão.\n\nAtenciosamente,\n\nDisav/Cofis\n\nAmbiente: "+ambiente    
            resultado = enviaEmail(email, message, "Relação de TDPFs Supervisionados", nomeArq)
            if resultado!=3:
                msg = "Erro no envio de email - mostraSupervisionados - "+str(resultado)
                logging.info(msg + " - "+email)
                enviaMsgBot(bot, userId, text=msg)
            else:
                enviaMsgBot(bot, userId, text="E-mail enviado.")   
            os.remove(nomeArq)  
        conn.close()    
    mostraMenuPrincipal(update, context)
    return

def opcaoEnviaAtividades(update, context): #envia relação de atividades de TDPFs em andamento para o supervisor
    global textoRetorno
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id
    bot = update.effective_user.bot  
    eliminaPendencia(userId)          
    achou = verificaUsuario(userId, bot)       
    if achou:   
        conn = conecta()
        if not conn:
            enviaMsgBot(bot, userId, text="Erro ao criar conexão ao Banco de Dados - opcaoEnviaAtividades")
            mostraMenuPrincipal(update, context)
            return
        consulta = """
                   Select TDPFS.Numero, TDPFS.Emissao, TDPFS.Nome, Usuarios.email, TDPFS.Codigo, TDPFS.Tipo
                   From TDPFS, Usuarios, Supervisores, Fiscais
                   Where Usuarios.idTelegram=%s and Usuarios.Saida Is Null and Usuarios.Adesao Is Not Null and TDPFS.Encerramento Is Null 
                   and Usuarios.CPF=Fiscais.CPF and Fiscais.Codigo=Supervisores.Fiscal and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo
                   and (TDPFS.Tipo='F' or TDPFS.Tipo='D')
                   Order by TDPFS.Numero
                   """
        cursor = conn.cursor(buffered=True)           
        cursor.execute(consulta, (userId,))
        linhas = cursor.fetchall()
        bAchou = True
        if linhas==None:
            bAchou = False
        if len(linhas)==0:
            bAchou = False  
        if not bAchou:
            enviaMsgBot(bot, userId, text="Não há TDPFs sob sua supervisão ou usuário está inativo no serviço - opcaoEnviaAtividades")
            mostraMenuPrincipal(update, context)
            conn.close()
            return  
        email = linhas[0][3]   
        if email==None or email=="": #email vazio
            enviaMsgBot(bot, userId, text="Email do usuário não foi informado - não haverá envio.")
            mostraMenuPrincipal(update, context)  
            conn.close() 
            return 
        if not "@rfb.gov.br" in email: 
            enviaMsgBot(bot, userId, text="Email do usuário é inválido - aparenta não ser institucional - não haverá envio.")
            mostraMenuPrincipal(update, context)  
            conn.close() 
            return                                            
        book = Workbook()
        sheet = book.active  
        sheet.cell(row=1, column=1).value = "TDPF"
        sheet.cell(row=1, column=2).value = "Data Emissão"
        sheet.cell(row=1, column=3).value = "Nome Fiscalizado"
        sheet.cell(row=1, column=4).value = "Auditor-Fiscal"
        sheet.cell(row=1, column=5).value = "Atividade"
        sheet.cell(row=1, column=6).value = "Data de Início"            
        sheet.cell(row=1, column=7).value = "Vencimento"
        sheet.cell(row=1, column=8).value = "Término"
        sheet.cell(row=1, column=9).value = "Horas"
        sheet.cell(row=1, column=10).value = "Tipo do Procedimento"
        larguras = [19, 13, 32, 32, 27, 14, 13, 13, 10, 18]
        for col in range(len(larguras)):
            sheet.column_dimensions[get_column_letter(col+1)].width = larguras[col]  
            currentCell = sheet.cell(row=1, column=col+1)
            currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)         
        i = 1
        for linha in linhas:  
            chaveTdpf = linha[4]                           
            tdpf = linha[0]
            emissao = linha[1]
            if emissao==None:
                emissao = ""
            else:
                emissao = emissao.date()    
            fiscalizado = linha[2]
            if fiscalizado==None:
                fiscalizado = ""  
            tipoP = linha[5]              
            consulta = "Select Atividade, Inicio, Vencimento, Termino, Horas from Atividades Where TDPF=%s Order by Inicio"
            cursor.execute(consulta,(chaveTdpf,))
            rows = cursor.fetchall()
            if not rows:
                continue
            if len(rows)==0:
                continue 
            fiscal = ""
            consulta = "Select Fiscais.Nome From Alocacoes, Fiscais Where Alocacoes.Fiscal=Fiscais.Codigo and Alocacoes.Desalocacao Is Null and Alocacoes.TDPF=%s"
            cursor.execute(consulta,(chaveTdpf,))
            rowFisc = cursor.fetchone()
            if rowFisc:
                if len(rowFisc)>0:
                    fiscal = rowFisc[0]  
                    #print(fiscal)                       
            for row in rows:
                currentCell = sheet.cell(row=i+1, column=1)
                currentCell.value = formataTDPF(tdpf)
                currentCell.alignment = Alignment(horizontal='center')                                             
                currentCell = sheet.cell(row=i+1, column=2)
                currentCell.value = emissao
                currentCell.alignment = Alignment(horizontal='center')                
                sheet.cell(row=i+1, column=3).value = fiscalizado 
                sheet.cell(row=i+1, column=4).value = fiscal
                sheet.cell(row=i+1, column=5).value = row[0]                             
                sheet.cell(row=i+1, column=10).value = "Fiscalização" if tipoP=='F' else "Diligência"
                if row[1]!=None:
                    sheet.cell(row=i+1, column=6).value = row[1].date()
                if row[2]!=None:                              
                    sheet.cell(row=i+1, column=7).value = row[2].date()
                if row[3]!=None:                              
                    sheet.cell(row=i+1, column=8).value = row[3].date() 
                if row[4]!=None:                              
                    sheet.cell(row=i+1, column=9).value = row[4]
                for col in range(6, 9):
                    currentCell = sheet.cell(row=i+1, column=col)
                    currentCell.alignment = Alignment(horizontal='center')                    
                if row[2]!=None and row[3]==None:
                    cor = None
                    if row[2].date()==datetime.now().date(): #se a atividade está vencendo, a linha fica azul
                        cor = Font(color="0000FF")
                    if row[2].date()<datetime.now().date(): #se a atividade está vencida, a linha fica vermelha
                        cor = Font(color="FF0000") 
                    if cor!=None:
                        for col in range(9):
                            sheet.cell(row=i+1, column=col+1).font = cor                                     
                i+=1          
        if i>1:
            nomeArq = "SupAtiv_"+str(userId)+"_"+datetime.now().strftime("%Y_%m_%d_%H_%M_%S")+".xlsx"
            book.save(nomeArq)   
            message = "Prezado(a),\n\nConforme solicitado, enviamos, em anexo, planilha com relação das atividades de TDPFs em andamento sob sua supervisão.\n\nAtenciosamente,\n\nDisav/Cofis\n\nAmbiente: "+ambiente    
            resultado = enviaEmail(email, message, "Relação de Atividades - TDPFs Supervisionados", nomeArq)
            if resultado!=3:
                msg = "Erro no envio de email - opcaoEnviaAtividades - "+str(resultado)
                logging.info(msg + " - "+email)
                enviaMsgBot(bot, userId, text=msg)
            else:
                enviaMsgBot(bot, userId, text="E-mail enviado.")   
            os.remove(nomeArq)
        else:    
            enviaMsgBot(bot, userId, text="Não há atividades relativamente aos TDPFs em andamento sob sua supervisão.")
        conn.close()
    mostraMenuPrincipal(update, context)         
    return   

def opcaoSupervisorEspontaneidade(update, context): #exibição de TDPFS que estejam vencendo hoje ou em X dias para o supervisor
    global pendencias    
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id
    bot = update.effective_user.bot     
    eliminaPendencia(userId)  
    achou = verificaUsuario(userId, bot)       
    if achou:   
        pendencias[userId] = 'supervisorEspontaneidade' #usuário agora tem uma pendência de informação
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora um intervalo de dias (um ou dois dígitos para cada) em que haverá a recuperação da espontaneidade tributária - TDPFs de sua EQUIPE:"  
        enviaMsgBot(bot, userId, text=response_message)
    else:
        mostraMenuPrincipal(update, context)         
    return

def mostraSupervisorEspontaneidade(update, context): #exibe os TDPFs da equipe que estão recuperando a espontaneidade em até X dias
    global pendencias, textoRetorno
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id
    bot = update.effective_user.bot       
    eliminaPendencia(userId) 
    achou = verificaUsuario(userId, bot)       
    if not achou: 
        mostraMenuPrincipal(update, context)         
        return
    msg = update.message.text  
    prazo = [0, 0]    
    parametros = getParametros(msg)  
    if len(parametros)==1 and parametros[0].isdigit:
        try:
            prazo[0] = int(parametros[0])
            prazo[1] = prazo[0]
        except:    
            response_message = "Envie somente dois números inteiros (separados por espaço)."
            response_message = response_message+textoRetorno  
            enviaMsgBot(bot, userId, text=response_message)  
            return
    elif len(parametros)==3 and parametros[0].isdigit and parametros[1].upper()=="A" and parametros[2].isdigit():
        try:
            prazo[0] = int(parametros[0])
            prazo[1] = int(parametros[2])
        except:    
            response_message = "Envie somente dois números inteiros (separados por espaço) (2)."
            response_message = response_message+textoRetorno  
            enviaMsgBot(bot, userId, text=response_message)  
            return
    elif len(parametros)!=2:
        response_message = "Envie somente dois números inteiros (separados por espaço) (3)."
        response_message = response_message+textoRetorno  
        enviaMsgBot(bot, userId, text=response_message)  
        return
    else: 
        resposta1 = parametros[0]
        resposta2 = parametros[1]
        if resposta1.isdigit() and resposta2.isdigit():
            try:
                prazo[0] = int(resposta1)
                prazo[1] = int(resposta2)
            except:
                response_message = "Envie somente dois números inteiros (separados por espaço) (4)."
                response_message = response_message+textoRetorno  
                enviaMsgBot(bot, userId, text=response_message)  
                return 
        else:
            response_message = "Envie somente dois números inteiros (separados por espaço) (5)."
            response_message = response_message+textoRetorno  
            enviaMsgBot(bot, userId, text=response_message)  
            return  
    prazo.sort()  
    if prazo[0]>99 or prazo[1]>99 or prazo[0]<0 or prazo[1]<0:
        response_message = "Envie somente dois números inteiros positivos (separados por espaço) de no máximo dois dígitos cada."
        response_message = response_message+textoRetorno  
        enviaMsgBot(bot, userId, text=response_message)  
        return                           
    conn = conecta()
    if not conn:
        enviaMsgBot(bot, userId, text="Erro na conexão - mostraSupervisorEspontaneidade")
        return        
    cursor = conn.cursor(buffered=True)
    consulta = """
                Select TDPFS.Numero, TDPFS.Codigo
                From TDPFS, Usuarios, Supervisores, Fiscais
                Where Usuarios.idTelegram=%s and Usuarios.Saida Is Null and TDPFS.Encerramento Is Null and Usuarios.CPF=Fiscais.CPF and Fiscais.Codigo=Supervisores.Fiscal and 
                Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and TDPFS.Encerramento Is Null and TDPFS.Tipo='F'
                """
    cursor.execute(consulta, (userId,))
    linhas = cursor.fetchall()
    msg = ""
    i = 1
    for linha in linhas:
        chaveTdpf = linha[1]
        tdpf = linha[0]
        consulta = "Select Fiscais.Nome from Alocacoes, Fiscais Where Alocacoes.TDPF=%s and Alocacoes.Fiscal=Fiscais.Codigo and Alocacoes.Desalocacao Is Null"
        cursor.execute(consulta, (chaveTdpf,))
        linhaFiscal = cursor.fetchone()
        if linhaFiscal:
            fiscalPrimNome = linhaFiscal[0].strip().split()[0]
        else:
            fiscalPrimNome = "ND"
        consulta = "Select Ciencias.Data, Ciencias.Documento From Ciencias Where Ciencias.TDPF=%s Order By Ciencias.Data DESC"
        cursor.execute(consulta, (chaveTdpf,))
        ciencia = cursor.fetchone() #busca a última
        if ciencia:
            if ciencia[0]!=None:
                dataCiencia = ciencia[0].date()
                prazoRestante = (dataCiencia+timedelta(days=60)-datetime.now().date()).days 
                if prazoRestante>=prazo[0] and prazoRestante<=prazo[1]:
                    if ciencia[1]==None:
                        documento = "ND"
                    else:
                        documento = ciencia[1]
                    consultaPostagem = "Select Data from ControlePostal Where TDPF=%s and Data>=%s"
                    cursor.execute(consultaPostagem, (chaveTdpf, dataCiencia))
                    regPostal = cursor.fetchone()
                    if regPostal:
                        postagem = "; Há termo emitido em "+regPostal[0].strftime("%d/%m/%Y")+" e postado" #indica que houve uma postagem em data igual ou posterior à da última ciência
                    else:
                        postagem = ""
                    if msg!="":
                        msg = msg + "\n"
                    msg = msg +"\n"+str(i)+") *TDPF: "+formataTDPF(tdpf)+"* ("+fiscalPrimNome+"); Documento: "+documento+"; Ciência: "+dataCiencia.strftime("%d/%m/%Y")+"; Recupera em "+str(prazoRestante)+" dias"+postagem
                    i+=1
    if msg!="":
        response_message = "Relação de TDPFs cuja recuperação da espontaneidade tributária ocorrerá em "+str(prazo[0])+" a "+str(prazo[1])+" dias:"+msg
    else:
        response_message = "Não haverá recuperação da espontaneidade tributária para nenhum TDPF neste intervalo ou usuário não é supervisor."    
    response_message = limpaMarkdown(response_message)            
    enviaMsgBot(bot, userId, text=response_message, parse_mode= 'MarkdownV2')                          
    conn.close()
    mostraMenuPrincipal(update, context)  
    return
    
def opcaoEMail(update, context): #cadastra e-mail para o recebimento de avisos
    global pendencias
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id
    bot = update.effective_user.bot       
    eliminaPendencia(userId) 
    achou = verificaUsuario(userId, bot)       
    if not achou: 
        mostraMenuPrincipal(update, context)         
        return     
    conn = conecta()
    if not conn:
        enviaMsgBot(bot, userId, text="Erro na conexão - opcaoEmail")
        return        
    cursor = conn.cursor(buffered=True)
    comando = "Select CPF, email from Usuarios Where Saida Is Null and idTelegram=%s and Ativo='S'" 
    try:
        cursor.execute(comando, (userId,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            enviaMsgBot(bot, userId, text="Usuário não está registrado no serviço ou está inativo/desabilitado. "+textoRetorno)
            return
        email = row[1]
    except:
        enviaMsgBot(bot, userId, text="Erro na consulta (10).")  
        conn.close()      
        return False     
    pendencias[userId] = 'email'     #usuário agora tem uma pendência de informação
    if email!=None and email!="":
        response_message = "Envie /menu para ver o menu principal.\nEmail atualmente cadastrado - "+email+". Informe seu *novo nome de usuário do endereço de e-mail institucional ou a palavra NULO* para descadastrar o atual (exemplo - se seu e-mail é fulano@rfb.gov.br, envie fulano):"
    else:    
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora seu *nome de usuário do endereço de e-mail institucional* no qual você também receberá alertas (exemplo - se seu e-mail é fulano@rfb.gov.br, envie fulano):"
    enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')  
    return  

def opcaoRelacionaPostagens(update, context): #relaciona postagens do usuário no prazo em dias informado por ele (máximo de 90 dias)
    global pendencias   
    if update.effective_user.is_bot:
        return #não atendemos bots       
    userId = update.effective_user.id
    bot = update.effective_user.bot     
    eliminaPendencia(userId)  
    achou = verificaUsuario(userId, bot)       
    if achou:   
        pendencias[userId] = 'relacionaPostagens' #usuário agora tem uma pendência de informação
        response_message = "Envie /menu para ver o menu principal.\nEnvie agora uma *quantidade de dias (um ou dois dígitos; máximo de 90)* que serão consideradas para pesquisar no passado, por data de envio, suas postagens e, separado por espaço, o *indicador se considera (S) ou não (N)* sua eventual qualidade de supervisor na pesquisa:"
        enviaMsgBot(bot, userId, text=limpaMarkdown(response_message), parse_mode= 'MarkdownV2')
    else:
        mostraMenuPrincipal(update, context)         
    return    


############################# Handlers #########################################
def error_handler(_, update, error):
    #try:
    #    print("aqui")
    #    raise error
    #except:
    logging.info("Falha ao enviar mensagem - Erro: "+str(error)) #com ou sem o try acima, dá o erro e interrompe o envio das mensagens (quando há várias)
    return

def botTelegram():
    global updater, token, senhaAvisoUrgente
    logging.info('Tentando iniciar o Updater Telegram')
    updater = Updater(token, use_context=True) 
    updater.dispatcher.add_error_handler(error_handler)
    updater.dispatcher.add_handler(CommandHandler('start', start))
    #updater.dispatcher.add_handler(CallbackQueryHandler(menuTDPF))    
    updater.dispatcher.add_handler(CommandHandler('menu', start))
    updater.dispatcher.add_handler(CommandHandler('retorna', start)) 
    if len(senhaAvisoUrgente)>=6:
        updater.dispatcher.add_handler(MessageHandler(Filters.regex('DisparaAvisoUrgente'+senhaAvisoUrgente), disparaAvisoUrgente))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Menu Principal'), mostraMenuPrincipal))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Cadastros'), mostraMenuCadastro))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('TDPF - Monitoramento'), mostraMenuTDPF))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Menu TDPF - Monitoramento'), mostraMenuTDPF))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Espontaneidade e Atividades'), mostraMenuCienciasAtividades))  
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Relaciona Postagens'), opcaoRelacionaPostagens))  
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Prazos Para Receber Avisos'), opcaoPrazos))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Registra Usuário'), opcaoUsuario))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Solicita Chave de Registro e do ContÁgil'), solicitaChaveRegistro))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Desativa Usuário'), opcaoUsuario))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Reativa Usuário'), opcaoUsuario))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Cadastra/Exclui e-Mail'), opcaoEMail))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Informa Data de Ciência'), opcaoInformaCiencia))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Anula Data de Ciência Informada'), opcaoAnulaCiencia))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Informa Ativ. e Prazo'), opcaoInformaAtividade))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Anula Atividade'), opcaoAnulaAtividade))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Finaliza Monitoramento de TDPF'), opcaoFinalizaAvisos))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Mostra TDPFs Monitorados'), opcaoMostraTDPFs))     
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Monitora TDPF\(s\)'), opcaoAcompanhaTDPFs))   
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Supervisão'), mostraMenuSupervisao))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Mostra TDPFs Supervisionados'), opcaoMostraSupervisionados)) 
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Envia Atividades \(e-Mail\) - Superv.'), opcaoEnviaAtividades))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Recuperação Espontaneidade - Superv.'), opcaoSupervisorEspontaneidade))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Informa Término de Ativ.'), opcaoInformaTerminoAtividade))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Informa Horas Atividade'), opcaoInformaHorasAtividade))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Ciências e Atividades - Email'), opcaoEnviaCienciasAtividades))
    updater.dispatcher.add_handler(MessageHandler(Filters.regex('Inicio'), mostraMenuPrincipal))          
    updater.dispatcher.add_handler(MessageHandler(Filters.all, unknown)) 
    
    updater.start_polling()
    logging.info('Serviço iniciado - '+datetime.now().strftime('%d/%m/%Y %H:%M'))
    #enviaMsgUpdater(842552685, "Serviço iniciado")    
    #updater.idle()    #não é necessário pq o programa vai ficar rodando em loop infinito
################################################################################

def disparaAvisoUrgente(update, context): #avisos urgentes da Cofis disparados por comando no Bot (exige comando especifico e senha - DisparaAvisoUrgente + senhaAvisoUrgente)
    global updater, ambiente
    if update.effective_user.is_bot:
        return #não atendemos bots    
    conn = conecta()
    userId = update.effective_user.id
    if not conn:
        enviaMsgUpdater(userId, "Erro na conexão ao banco de dados")
        return
    logging.info("Acionado o disparo de mensagens URGENTES - "+datetime.now().strftime('%d/%m/%Y %H:%M'))
    cursor = conn.cursor()
    dataAtual = datetime.today()
    cursor.execute('Select Mensagem from AvisosUrgentes Where DataEnvio Is Null')    
    mensagens = cursor.fetchall()
    if not mensagens:
        msgErro = "Não há mensagem a ser enviada"
        logging.info(msgErro) 
        enviaMsgUpdater(userId, msgErro)
        return #não há mensagens a serem enviadas
    msgCofis = ""
    for mensagem in mensagens:
        if msgCofis!="":
            msgCofis = msgCofis+";\n"
        msgCofis = msgCofis+mensagem[0]
    if msgCofis!="":
        msgCofis = "Mensagem URGENTE Cofis:\n"+msgCofis
        if ambiente!="PRODUÇÃO":
            msgCofis = msgCofis + "\nAmbiente: "+ambiente
    else:
        msgErro = "Não há mensagem a ser enviada (2)"
        logging.info(msgErro) 
        enviaMsgUpdater(userId, msgErro)        
        return
    comando = "Select idTelegram from Usuarios Where Saida Is Null and idTelegram Is Not Null and BloqueiaTelegram='N' and and idTelegram<>0 and Adesao Is Not Null and Ativo='S'"
    cursor.execute(comando)
    usuarios = cursor.fetchall()
    totalMsg = 0
    msgDisparadas = 0
    for usuario in usuarios: #percorremos os usuários ativos Telegram
        enviaMsgUpdater(usuario[0], text=msgCofis)   
        totalMsg+=1
        msgDisparadas+=1
        if msgDisparadas>=30:
            msgDisparadas = 0
            time.sleep(1) #a cada 30 mensagens, dormimos um segundo (limitação do Bot é 30 por seg - TESTE) 
    msg = "Total de usuários para os quais foi enviada a mensagem (AvisoUrgente) no ambiente "+ambiente+": "+str(totalMsg)
    logging.info(msg) 
    enviaMsgUpdater(userId, msg)     
    enviaMsgUpdater(userId, "Mensagem que foi enviada para cada usuário:\n'"+msgCofis+"'")  
    try:
        comando = "Update AvisosUrgentes Set DataEnvio=%s Where DataEnvio Is Null"
        cursor.execute(comando, (dataAtual,))
        conn.commit()
    except:
        msgErro = "Erro ao atualizar a tabela de AvisosUrgentes - datas de envio ficaram em branco. Cuidado para não reenviar."
        logging.info(msgErro)
        enviaMsgUpdater(userId, msgErro)         
        conn.rollback()
    return

def disparaMensagens(): #avisos diários (produção) ou de hora em hora (teste) contendo os alertas e mensagens da Cofis (não inclui alertas de juntadas, nem de prorrogação pendente de assinatura)
    global updater, termina, ambiente
		
    try:
        server = smtplib.SMTP('INETRFOC.RFOC.SRF: 25') #servidor de email Notes
    except:
        logging.info("Erro na criação do servidor SMTP (disparaMensagens")
        server = None        
    conn = conecta()
    if not conn:
        logging.info("Conexão ao BD falhou (disparaMensagens")
        return
    logging.info("Acionado o disparo de mensagens - "+datetime.now().strftime('%d/%m/%Y %H:%M'))
    print("Acionado o disparo de mensagens - "+datetime.now().strftime('%d/%m/%Y %H:%M'))
    cursor = conn.cursor(buffered=True)
    dataAtual = datetime.today().date()
    cursor.execute('Select Mensagem from MensagensCofis Where Data=%s', (dataAtual,))    
    mensagens = cursor.fetchall()
    msgCofis = ""
    for mensagem in mensagens:
        if msgCofis!="":
            msgCofis = msgCofis+";\n"
        msgCofis = msgCofis+mensagem[0]
    if msgCofis!="":
        msgCofis = "Mensagens Cofis:\n"+msgCofis+"."   #todos os usuários receberão essa mensagem na data informada pela Cofis, com ou sem alertas do dia 
    comando = "Select idTelegram, CPF, d1, d2, d3, email from Usuarios Where Adesao Is Not Null and BloqueiaTelegram='N' and Saida Is Null and idTelegram Is Not Null and idTelegram<>0 and Ativo='S'"
    cursor.execute(comando)
    usuarios = cursor.fetchall()
    totalMsg = 0
    msgDisparadas = 0
    tdpfsAvisadosUpdate = set()
    tdpfsAvisadosInsert = set()
    tdpfsCienciasPendentes = set() #guarda os registros de tdpfs com ciências pendentes há 30 dias (aviso se repete de 15 em 15 dias) para avisar o supervisor
    cabecalho = "Alertas do dia (TDPF | Dias Restantes):" 
    rodape = "\n\n(a) P/ recuperação da espontaneidade tributária."
    rodape += "\n(b) P/ vencimento da atividade."            
    rodape += "\n(c) P/ vencimento do TDPF no Ação Fiscal - pode ser inferior ao verificado no Ação Fiscal, pois as informações do serviço se baseiam"
    rodape += " na data de distribuição, que pode ter ocorrido antes da emissão do TDPF."
    rodape += "\n(d) Vence HOJE o prazo do contribuinte para atendimento da última intimação." 
    rodape += "\n(P) Indica que há postagem em data igual ou posterior à da última ciência."   
    rodape += "\n(S) Indica que você é supervisor (titular ou substituto) da equipe do TDPF." 
    for usuario in usuarios: #percorremos os usuários ativos Telegram
        if termina: #programa foi informado de que é para encerrar (quit)
            if server:
                server.quit()
            return        
        listaUsuario = "" #lista de TDPF com espontaneidade, atividade ou o próprio TDPF vencendo (bot)
        listaUsuarioEmail = "" #lista de TDPF com espontaneidade, atividade ou o próprio TDPF vencendo (e-mail)
        cpf = usuario[1]
        d1 = usuario[2]
        d2 = usuario[3]
        d3 = usuario[4]
        email = usuario[5]
        #selecionamos os TDPFs do usuário em andamento e monitorados (ativos) pelo serviço
        comando = """
                Select TDPFS.Numero, Supervisor, TDPFS.Codigo, TDPFS.Nome, TDPFS.Tipo
                from CadastroTDPFs, TDPFS, Alocacoes, Fiscais
                Where Fiscais.CPF=%s and CadastroTDPFs.Fiscal=Fiscais.Codigo and CadastroTDPFs.TDPF=TDPFS.Codigo and 
                CadastroTDPFs.TDPF=Alocacoes.TDPF and  
                CadastroTDPFs.Fiscal=Alocacoes.Fiscal and TDPFS.Encerramento Is Null and 
                CadastroTDPFs.Fim Is Null and Alocacoes.Desalocacao Is Null
                Order by TDPFS.Numero
                """
        cursor.execute(comando, (cpf,))        
        fiscalizacoes = cursor.fetchall()
        comandoCiencias = "Select Data, Vencimento from Ciencias Where TDPF=%s Order By Data DESC"
        comandoAtividades = "Select Atividade, Vencimento, Inicio from Atividades Where TDPF=%s and Vencimento>=%s and Termino Is Null Order by Inicio, Vencimento"
        if fiscalizacoes:
            for fiscalizacao in fiscalizacoes: #percorremos os TDPFs MONITORADOS do usuário
                if termina: #foi solicitado o término do bot
                    return  
                chaveTdpf = fiscalizacao[2]          
                tdpf = fiscalizacao[0] 
                tdpfFormatado = formataTDPF(tdpf)
                supervisor = fiscalizacao[1]
                nome = fiscalizacao[3].strip()[:75]
                tipoP = fiscalizacao[4]
                cursor.execute(comandoCiencias, (chaveTdpf,)) #buscamos as ciências do TDPF
                ciencias = cursor.fetchall() #buscamos todas por questões técnicas do mysqlconnector
                bPrazo = False        #indica se há aviso de proximidade do prazo de recuperaçao da espontaneidade
                if len(ciencias)>0:   
                    cienciaReg = ciencias[0]    
                    dataCiencia = cienciaReg[0].date()  #só é necessária a última (selecionamos por ordem descendente)  
                    if tipoP=='F':    #só interessa a fiscalização para avisar sobre recuperação de espontaneidade
                        prazoRestante = (dataCiencia+timedelta(days=60)-dataAtual).days      
                        if prazoRestante==d1 or prazoRestante==d2 or prazoRestante==d3:
                            if len(listaUsuario)==0:
                                listaUsuario = cabecalho  
                                listaUsuarioEmail = cabecalho  
                            listaUsuarioEmail += "\n\n"+tdpfFormatado                 
                            if supervisor=='S':    
                                tdpfFormatado2 = tdpfFormatado + '* (S)'   
                                listaUsuarioEmail += " (S)"                     
                            else:
                                tdpfFormatado2 = tdpfFormatado + '*'
                            listaUsuario += "\n\n*"+tdpfFormatado2+ " | "+str(prazoRestante)+" (a)"
                            listaUsuarioEmail += " - "+nome+" | "+str(prazoRestante)+" (a)"  #no email, vai o nome da empresa  
                            #verificamos se há postagem relativa ao TDPF na data da última ciência deste ou após
                            consultaPostagem = "Select Codigo from ControlePostal Where TDPF=%s and Data>=%s"
                            cursor.execute(consultaPostagem, (chaveTdpf, dataCiencia))
                            regPostal = cursor.fetchone()
                            if regPostal:
                                listaUsuario += " (P)" #indica que houve uma postagem em data igual ou posterior à da última ciência
                                listaUsuarioEmail += " (P)"
                            bPrazo = True
                    vencimentoIntimacao = cienciaReg[1]
                    if vencimentoIntimacao!=None:
                        if vencimentoIntimacao.date()==datetime.now().date(): #a intimação vence hoje
                            if not bPrazo: #não há aviso de prazo da recuperação da espontaneidade
                                if len(listaUsuario)==0:
                                    listaUsuario = cabecalho        
                                    listaUsuarioEmail = cabecalho  
                                listaUsuarioEmail += "\n\n"+tdpfFormatado                                                               
                                if supervisor=='S':    
                                    tdpfFormatado2 = tdpfFormatado + '* (S)('+tipoP+")"
                                    listaUsuarioEmail += " (S)"
                                else:
                                    tdpfFormatado2 = tdpfFormatado + '* ('+tipoP+")"
                                listaUsuario += "\n\n*"+tdpfFormatado2+" (d)"    
                                listaUsuarioEmail += '('+tipoP+") - "+nome+" (d)"    
                            else: #há aviso de prazo da recuperação da espontaneidade (bastante improvável, mas posssível)
                                listaUsuario += " (d)"  
                                listaUsuarioEmail += " (d)"                          
           
                #buscamos as atividades do TDPF    
                cursor.execute(comandoAtividades, (chaveTdpf, dataAtual))
                atividades = cursor.fetchall()
                for atividade in atividades:
                    prazoRestante = (atividade[1].date()-dataAtual).days
                    if prazoRestante==0 or prazoRestante==d3: #para atividade, alertamos só no d3 (o menor) e no dia do vencimento (prazo restante == 0)
                        if len(listaUsuario)==0:
                            listaUsuario = cabecalho
                            listaUsuarioEmail = cabecalho
                        listaUsuario += "\n\n*"+tdpfFormatado+"* ("+tipoP+") | "+str(prazoRestante)+" (b)"
                        listaUsuarioEmail += "\n\n"+tdpfFormatado+" ("+tipoP+") - "+nome+" | "+str(prazoRestante)+" (b)"
                        #listaUsuario = listaUsuario+"\nAtividade: "+atividade[0]+"; Início: "+atividade[2].strftime("%d/%m/%Y")
                        listaUsuario += "\nAtividade: "+atividade[0]+"; Início: "+atividade[2].strftime("%d/%m/%Y")+"; Vencimento: "+atividade[1].strftime("%d/%m/%Y")
                        listaUsuarioEmail += "\nAtividade: "+atividade[0]+"; Início: "+atividade[2].strftime("%d/%m/%Y")+"; Vencimento: "+atividade[1].strftime("%d/%m/%Y")

        #selecionamos as datas de vencimento dos TDPFs-F ou D em que o usuário está alocado, mesmo que não monitorados
        comando = """
                Select TDPFS.Numero, TDPFS.Vencimento, Supervisor, TDPFS.Codigo, Fiscais.Codigo, TDPFS.Nome
                from TDPFS, Alocacoes, Fiscais
                Where Fiscais.CPF=%s and Alocacoes.Fiscal=Fiscais.Codigo and TDPFS.Codigo=Alocacoes.TDPF and TDPFS.Encerramento Is Null and 
                Alocacoes.Desalocacao Is Null and TDPFS.Tipo in ('F', 'D')
                """        
        cursor.execute(comando, (cpf,))     
        comandoVencimento = "Select Codigo, Data from AvisosVencimento Where TDPF=%s and Fiscal=%s"                   
        tdpfUsuarios = cursor.fetchall()
        for tdpfUsuario in tdpfUsuarios: #percorremos os TDPFs do usuário para ver suas datas de vencimento (TODOS em andamento no qual o usuário esteja atualmente alocado)
            vencimento = tdpfUsuario[1]
            tdpf = tdpfUsuario[0]
            tdpfFormatado = formataTDPF(tdpf)
            supervisor = tdpfUsuario[2]
            chaveTdpf = tdpfUsuario[3]
            chaveFiscal = tdpfUsuario[4]
            nome = tdpfUsuario[5].strip()[:75]
            if vencimento: #não deve ser nulo, mas garantimos ...
                vencimento = vencimento.date()
                prazoVenctoTDPF = (vencimento-dataAtual).days
                if not (1<=prazoVenctoTDPF<=15): #se não estiver próximo do vencimento (1 a 15 dias), prosseguimos (pulamos)
                    continue 
                cursor.execute(comandoVencimento, (chaveTdpf, chaveFiscal))  #buscamos as datas de aviso para o CPF e TDPF  
                avisos = cursor.fetchall()
                if len(avisos)==0: #não há avisos para o tdpf (este cpf)
                    avisou = None
                else:
                    avisou = avisos[0][1].date() #só retorna uma linha para cada tdpf/cpf      
                    codigo = avisos[0][0] #chave primária do registro        
                #não avisamos do vencimento de TDPF recentemente avisado (prazo: 7 dias - depende da carga)
                if not avisou:
                    podeAvisar = True
                    tdpfsAvisadosInsert.add(str(chaveTdpf).rjust(15,"0")+str(chaveFiscal).rjust(12,"0"))     
                else:
                    if (avisou+timedelta(days=7))<dataAtual: #vai depender da periodicidade da extração e carga - aqui estamos considerando 7 dias
                        podeAvisar = True
                        tdpfsAvisadosUpdate.add(codigo)                        
                    else:
                        podeAvisar = False                      
                if podeAvisar: #verificar estes prazos quando for colocar em produção
                    if len(listaUsuario)==0:
                        listaUsuario = cabecalho
                        listaUsuarioEmail = cabecalho
                    if supervisor=='S':    
                        tdpfFormatado2 = tdpfFormatado + ' (S)'
                    else:
                        tdpfFormatado2 = tdpfFormatado                         
                    listaUsuario += "\n\n*"+tdpfFormatado2+ "* | "+str(prazoVenctoTDPF)+" (c)"                 
                    listaUsuarioEmail += "\n\n"+tdpfFormatado2+" - "+nome+" | "+str(prazoVenctoTDPF)+" (c)" 
                    
        tdpfUsuarios = None
        if len(listaUsuario)>0:
            if len(listaUsuario)>0:
                listaUsuario += rodape
                listaUsuarioEmail += rodape      

        #avisamos o fiscal (monitora TDPF)/supervisor de suas postagens enviadas há 21/28 dias ou na mesma situação há 7/14 dias e NÃO Entregues
        #avisamos também ambas pessoas que não ocorreu movimentação para a postagem em 7 dias após seu envio
        #avisamos também o fiscal que monitora o TDPF que sua correspondência foi entregue ontem ao destinatário 
        #se a correspondência será devolvida por fato verificado ONTEM, também avisamos ao fiscal que monitora o TDPF
        #nestes dois últimos casos, não pego a data do dia, pois podemos rodar essa rotina de manhã e os correios atualizarem à tarde, p ex
        comando = """
                    Select TDPFS.Numero, TDPFS.Nome, Documento, DataEnvio, DataSituacao, SituacaoAtual
                    from TDPFS, ControlePostal, Fiscais, Alocacoes, CadastroTDPFs
                    Where Fiscais.CPF=%s and Alocacoes.Fiscal=Fiscais.Codigo and TDPFS.Codigo=Alocacoes.TDPF and TDPFS.Encerramento Is Null and 
                    Alocacoes.Desalocacao Is Null and ControlePostal.TDPF=TDPFS.Codigo and CadastroTDPFs.Fiscal=Fiscais.Codigo and CadastroTDPFs.TDPF=TDPFS.Codigo and
                    (((DataEnvio=cast((now() - interval 21 day) as date) or DataEnvio=cast((now() - interval 28 day) as date) or 
                    DataSituacao=cast((now() - interval 7 day) as date) or DataSituacao=cast((now() - interval 14 day) as date)) and 
                    Upper(SituacaoAtual) Not Like '%ENTREGUE%') or (DataSituacao Is Null and DataEnvio=cast((now() - interval 7 day) as date))) and
                    DataRecebimento Is Null
                    UNION
                    Select TDPFS.Numero, TDPFS.Nome, Documento, DataEnvio, DataSituacao, SituacaoAtual
                    from TDPFS, ControlePostal, Fiscais, Alocacoes, CadastroTDPFs
                    Where Fiscais.CPF=%s and Alocacoes.Fiscal=Fiscais.Codigo and TDPFS.Codigo=Alocacoes.TDPF and TDPFS.Encerramento Is Null and 
                    Alocacoes.Desalocacao Is Null and ControlePostal.TDPF=TDPFS.Codigo and CadastroTDPFs.Fiscal=Fiscais.Codigo and CadastroTDPFs.TDPF=TDPFS.Codigo and
                    DataSituacao=cast((now() - interval 1 day) as date) and Upper(SituacaoAtual) Like '%ENTREGUE AO DESTINATÁRIO%' and
                    DataRecebimento Is Null              
                    UNION
                    Select TDPFS.Numero, TDPFS.Nome, Documento, DataEnvio, DataSituacao, SituacaoAtual
                    from TDPFS, ControlePostal, Fiscais, Alocacoes, CadastroTDPFs
                    Where Fiscais.CPF=%s and Alocacoes.Fiscal=Fiscais.Codigo and TDPFS.Codigo=Alocacoes.TDPF and TDPFS.Encerramento Is Null and 
                    Alocacoes.Desalocacao Is Null and ControlePostal.TDPF=TDPFS.Codigo and CadastroTDPFs.Fiscal=Fiscais.Codigo and CadastroTDPFs.TDPF=TDPFS.Codigo and
                    DataSituacao=cast((now() - interval 1 day) as date) and Upper(SituacaoAtual) Like '%SERÁ DEVOLVIDO AO REMETENTE%' and
                    DataRecebimento Is Null                 
                    UNION                    
                    Select TDPFS.Numero, TDPFS.Nome, Documento, DataEnvio, DataSituacao, SituacaoAtual
                    from TDPFS, ControlePostal, Fiscais, Supervisores
                    Where Fiscais.CPF=%s and Supervisores.Fiscal=Fiscais.Codigo and TDPFS.Grupo=Supervisores.Equipe and  
                    Supervisores.Fim Is Null and TDPFS.Encerramento Is Null and ControlePostal.TDPF=TDPFS.Codigo and 
                    (((DataEnvio=cast((now() - interval 21 day) as date) or DataEnvio=cast((now() - interval 28 day) as date) or 
                    DataSituacao=cast((now() - interval 7 day) as date) or DataSituacao=cast((now() - interval 14 day) as date)) and 
                    Upper(SituacaoAtual) Not Like '%ENTREGUE%') or (DataSituacao Is Null and DataEnvio=cast((now() - interval 7 day) as date))) and
                    DataRecebimento Is Null
                """     
        bAlertaPostagem = False   
        cursor.execute(comando, (cpf, cpf, cpf, cpf))    
        linhas = cursor.fetchall()
        for linha in linhas:
            tdpf = linha[0]
            tdpfFormatado = formataTDPF(tdpf)  
            nome = linha[1][:100].strip()
            documento = linha[2].strip()
            envio = linha[3].strftime("%d/%m/%Y")            
            dataSituacao = linha[4]
            bOntem = False
            if dataSituacao==None:
                dataSituacao = "ND"
            else:
                if dataSituacao.date()==datetime.now().date()-timedelta(days=1): #é uma entrega ontem
                    bOntem = True
                dataSituacao = dataSituacao.strftime("%d/%m/%Y")     
            situacao = linha[5]       
            if situacao==None:
                situacao = "ND"
            else:
                situacao = situacao.strip()
            if len(listaUsuario)==0:
                listaUsuario = "Alertas do dia - Postagens (TDPF | Documento | Dt Envio | Dt Situação | Situação):" 
                listaUsuarioEmail = "Alertas do dia - Postagens (TDPF | Documento | Dt Envio | Dt Situação | Situação):" 
            elif bAlertaPostagem==False: #só antes do primeiro TDPF
                listaUsuario += "\n\nAlertas do dia - Postagens (TDPF | Documento | Dt Envio | Dt Situação | Situação):" 
                listaUsuarioEmail += "\n\nAlertas do dia - Postagens (TDPF | Documento | Dt Envio | Dt Situação | Situação):" 
            listaUsuario += "\n\n*"+tdpfFormatado+ "* | "+documento[:50].strip()+" | "+envio+" | "+dataSituacao
            listaUsuarioEmail += "\n\n"+tdpfFormatado+" - "+nome+" | "+documento+" | "+envio+" | "+dataSituacao
            if bOntem:
                if "ENTREGUE AO DESTINATÁRIO" in situacao.upper():
                    listaUsuario += " | (E) "+situacao[:50].strip()
                    listaUsuarioEmail += " | (E) "+situacao
                else: #objeto será devolvido ao remetente
                    listaUsuario += " | (D) "+situacao[:50].strip()
                    listaUsuarioEmail += " | (D) "+situacao   
            elif dataSituacao=="ND" and linha[3].date()==datetime.now().date()-timedelta(days=7): #correspondência sem qualquer movimentação 7 dias após o envio                  
                listaUsuario += " | (M)"
                listaUsuarioEmail += " | (M)"
            else:                
                listaUsuario += " | (A) "+situacao[:50].strip()
                listaUsuarioEmail += " | (A) "+situacao
            bAlertaPostagem = True            
        if bAlertaPostagem: #se houve alerta de postagem, colocamos uma explicação
            listaUsuario += "\n\nA - Postagens enviadas há 21 ou 28 dias ou na mesma situação há 7 ou 14 (atraso), e não entregues"
            listaUsuarioEmail += "\n\nA - Postagens enviadas há 21 ou 28 dias ou na mesma situação há 7 ou 14 (atraso), e não entregues"
            listaUsuario +="\nE - Entregue ao destinatário ONTEM"
            listaUsuarioEmail +="\nE - Entregue ao destinatário ONTEM"
            listaUsuario +="\nD - Correspondência será devolvida por causa de fato verificado ONTEM"
            listaUsuarioEmail +="\nD - Correspondência será devolvida por causa de fato verificado ONTEM"
            listaUsuario +="\nM - Correspondência sem qualquer Movimentação 7 dias após o envio"
            listaUsuarioEmail +="\nD - Correspondência sem qualquer Movimentação 7 dias após o envio"
        if len(listaUsuario)>0 or msgCofis!="":
            if msgCofis!="":
                if len(listaUsuario)>0:
                    listaUsuario += "\n\n"
                    listaUsuarioEmail += "\n\n"
                listaUsuario += msgCofis
                listaUsuarioEmail += msgCofis              
            if ambiente!="PRODUÇÃO":
                listaUsuario += "\n\nAmbiente: "+ambiente
                listaUsuarioEmail += "\n\nAmbiente: "+ambiente
            logging.info("Disparando mensagem para "+cpf)
            if ambiente=="HOMOLOGAÇÃO" and cpf!="53363833172": #só envio mensagens para mim no ambiente de homologação
                print(listaUsuarioEmail)
            else:
                enviaMsgUpdater(usuario[0], text=limpaMarkdown(listaUsuario), parse_mode= 'MarkdownV2')  
                #enviamos e-mail também, se houver um na tabela
                if email!=None and server!=None:
                    email = email.strip()
                    if email!="":
                        # create message object instance
                        msg = MIMEMultipart()               
                        # setup the parameters of the e-mail message
                        msg['From'] = "botespontaneidade@rfb.gov.br"
                        msg['Subject'] = "BotEspontaneidade - Avisos e Alertas do Dia"                    
                        msg['To'] = email			
                        # add in the message body
                        msg.attach(MIMEText(listaUsuarioEmail, 'plain'))				
                        # send the message via the server.
                        try:
                            server.sendmail(msg['From'], msg['To'], msg.as_string())
                        except:
                            logging.info("Erro no envio de email com os avisos do dia - "+email)	                        
                        #print(listaUsuarioEmail)
            #logging.info(listaUsuario)            
            totalMsg+=1
            msgDisparadas+=1
            if msgDisparadas>=30:
                msgDisparadas = 0
                time.sleep(1) #a cada 30 mensagens, dormimos um segundo (limitação do Bot é 30 por seg - TESTE) 
                
        #verificamos se o usuário possui TDPF com processo integrado sem ciência informada há 30 dias - avisamos ele e o supervisor (independe de monitoramento) 
        #  (não vai e-mail)
        #o aviso é repetido a cada 15 dias, mas limitado a 60 dias após a integração
        consulta = """
                    Select TDPFS.Numero, AvisosCiencia.Processo, AvisosCiencia.Integracao, AvisosCiencia.Codigo
                    From TDPFS, Alocacoes, AvisosCiencia, Fiscais
                    Where TDPFS.Codigo=AvisosCiencia.TDPF and TDPFS.Codigo=Alocacoes.TDPF and Alocacoes.Desalocacao Is Null 
                    and Alocacoes.Fiscal=Fiscais.Codigo and Fiscais.CPF=%s and AvisosCiencia.Finalizado Is Null and
                    (AvisosCiencia.Integracao=cast((now() - interval 30 day) as date) or 
                    (cast(AvisosCiencia.Aviso as date)=cast((now() - interval 15 day) as date) and AvisosCiencia.Integracao<=cast((now() - interval 60 day) as date)) or 
                    (AvisosCiencia.Aviso Is Null and AvisosCiencia.Integracao<cast((now() - interval 30 day) as date)))
                    """
        cursor.execute(consulta, (cpf,))
        rowsCienciasPendentes = cursor.fetchall()
        listaCienciasPendentes = ""
        i = 0
        for cienciaPendente in rowsCienciasPendentes:
            i+=1
            tdpf = cienciaPendente[0]
            processo = cienciaPendente[1]
            integracao = cienciaPendente[2].date().strftime("%d/%m/%Y")
            tdpfsCienciasPendentes.add(cienciaPendente[3])
            listaCienciasPendentes = listaCienciasPendentes+"\n"+str(i)+"*"+formataTDPF(tdpf)+"* | "+formataDCC(processo)+" | "+integracao+"\n"

        if i>0:
            listaCienciasPendentes = "Processos sem registro de ciência há mais de 30 dias da integração.\nTDPF | Processo | Data da Integração:\n"+listaCienciasPendentes
            if ambiente!="PRODUÇÃO":
                listaCienciasPendentes = listaCienciasPendentes+"\nAmbiente: "+ambiente
            logging.info("Disparando mensagem para "+cpf+" - Ciências Pendentes")
            enviaMsgUpdater(usuario[0], text=limpaMarkdown(listaCienciasPendentes), parse_mode= 'MarkdownV2')  
            totalMsg+=1
            msgDisparadas+=1
            if msgDisparadas>=30:
                msgDisparadas = 0
                time.sleep(1) #a cada 30 mensagens, dormimos um segundo (limitação do Bot é 30 por seg - TESTE)                         

    if server:
        server.quit()  

    logging.info("Total de mensagens disparadas: "+str(totalMsg))    

    #se avisamos do vencimento do TDPF, colocamos a data na tabela para não avisarmos novamente
    #todo dia (espera 7 dias)
    lista = []
    for tdpfCpf in tdpfsAvisadosInsert:
        tupla = (int(tdpfCpf[:15]), int(tdpfCpf[15:]), dataAtual) #ao adicionar, o código do TDPF ocupou 15 posições e a chave do fiscal, 12
        lista.append(tupla)
    if len(lista)>0:
        logging.info("Inserção:")
        logging.info(lista)
        comando = "Insert Into AvisosVencimento (TDPF, Fiscal, Data) Values (%s, %s, %s)"
        try:
            cursor.executemany(comando, lista)    
            conn.commit()
        except:
            logging.info("Erro ao tentar inserir as datas de aviso na tabela AvisosVencimento.")
            conn.rollback()   
    lista = []
    for codigo in tdpfsAvisadosUpdate:
        tupla = (dataAtual, codigo)
        lista.append(tupla)                         
    if len(lista)>0:   
        logging.info("Atualização:")
        logging.info(lista)         
        comando = "Update AvisosVencimento Set Data=%s Where Codigo=%s" 
        try:
            cursor.executemany(comando, lista)
            conn.commit()
        except:
            logging.info("Erro ao tentar atualizar as datas de aviso na tabela AvisosVencimento.")
            conn.rollback()

    #buscamos todos os TDPFs para os quais não foram registradas nenhuma data de ciência após 30 dias (exatos) de sua distribuição
    # e avisamos o supervisor - apenas fiscalização e diligência
    comando = """
              Select Distinctrow Fiscais.CPF, TDPFS.Numero, Usuarios.idTelegram, Supervisores.Equipe, TDPFS.Codigo, TDPFS.Emissao, Usuarios.email, TDPFS.Nome, Usuarios.BloqueiaTelegram, TDPFS.Tipo
              From TDPFS, Usuarios, Supervisores, Fiscais
              Where Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and TDPFS.Emissao=cast((now() - interval 30 day) as date) 
              and TDPFS.Encerramento Is Null and Supervisores.Fiscal=Fiscais.Codigo and Fiscais.CPF=Usuarios.CPF and Usuarios.idTelegram Is Not Null and 
              Usuarios.idTelegram<>0 and Usuarios.Adesao Is Not Null and Usuarios.Saida Is Null and 
              TDPFS.Codigo not in (Select TDPF from Ciencias Where Data Is Not Null) and (TDPFS.Tipo='F' or TDPFS.Tipo='D')
              Order By Fiscais.CPF, Supervisores.Equipe, TDPFS.Numero
              """
    consultaFiscal = "Select Fiscais.Nome From TDPFS, Alocacoes, Fiscais Where TDPFS.Codigo=%s and TDPFS.Codigo=Alocacoes.TDPF and Alocacoes.Fiscal=Fiscais.Codigo and Alocacoes.Desalocacao Is Null Order By Alocacoes.Alocacao"
    consultaPostagem = "Select Codigo from ControlePostal Where TDPF=%s" #verificamos se há postagem para aquele TDPF
    cursor.execute(comando)
    linhas = cursor.fetchall() 
    msg = ""  
    cpfAnt = "" 
    userId = 0 
    cabecalho = "TDPFs sem informação de início (ciência) do procedimento fiscal há 30 dias:"
    equipe = ""
    texto = "Sr. Usuário,\n\nEnviamos, em anexo, planilha com relação de TDPFs sob sua supervisão sem ciência registrada há 30 dias da emissão.\n\nAtenciosamente,\n\nCofis/Disav"
    if ambiente!="PRODUÇÃO":
        texto = texto+"\n\nAmbiente: "+ambiente      
    assunto = "TDPFs sem Ciência Registrada Há 30 Dias da Emissão"
    for linha in linhas:
        if cpfAnt=="":
            i = 1
            cpfAnt = linha[0]
            email = linha[6]
            bloqueio = linha[8]
            tipoP = linha[9]
            book = Workbook()
            sheet = book.active
            sheet.cell(row=1, column=1).value='Nº Ordem'
            sheet.cell(row=1, column=2).value='Equipe'
            sheet.cell(row=1, column=3).value='Nº TDPF'
            sheet.cell(row=1, column=4).value='Nome'
            sheet.cell(row=1, column=5).value='Fiscal'
            sheet.cell(row=1, column=6).value='Emissão'
            sheet.cell(row=1, column=7).value='Tipo do Procedimento'
            sheet.cell(row=1, column=8).value='Postagem'
            larguras = [13, 20, 20, 45, 35, 15, 18, 13]
            for col in range(len(larguras)):
                sheet.column_dimensions[get_column_letter(col+1)].width = larguras[col]
                currentCell = sheet.cell(row=1, column=col+1)
                currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                currentCell.font = Font(bold=True)            
        if linha[0]!=cpfAnt:
            if msg!="":
                msg = cabecalho+msg
                if ambiente!="PRODUÇÃO":
                    msg = msg+"\n"+"Ambiente: "+ambiente    
                if bloqueio=='N':            
                    enviaMsgUpdater(userId, text=limpaMarkdown(msg), parse_mode= 'MarkdownV2')
                if not email in [None, ""] and server:
                    nomeArq = "Ciencia30_"+str(userId)+"_"+datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")[:-3]+".xlsx"
                    book.save(nomeArq)
                    if enviaEmail(email, texto, assunto, nomeArq)!=3:
                        logging.info("Erro ao enviar email de TDPFs sem ciência registrada após 30 dias - "+email)
                    os.remove(nomeArq)
                i = 1
                book = Workbook()
                sheet = book.active
                sheet.cell(row=1, column=1).value='Nº Ordem'
                sheet.cell(row=1, column=2).value='Equipe'
                sheet.cell(row=1, column=3).value='Nº TDPF'
                sheet.cell(row=1, column=4).value='Nome'
                sheet.cell(row=1, column=5).value='Fiscal'
                sheet.cell(row=1, column=6).value='Emissão'
                sheet.cell(row=1, column=7).value='Tipo do Procedimento'
                sheet.cell(row=1, column=8).value='Postagem'
                for col in range(len(larguras)):
                    sheet.column_dimensions[get_column_letter(col+1)].width = larguras[col]
                    currentCell = sheet.cell(row=1, column=col+1)
                    currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    currentCell.font = Font(bold=True)
                msg = ""
            equipe = ""    
            cpfAnt = linha[0] 
            email = linha[6]  
            bloqueio = linha[8] 
        if linha[3]!=equipe:
            equipe = linha[3]
            msg = msg+"\nEquipe " + formataEquipe(equipe) +":"
        chaveTdpf = linha[4]    
        tipoP = linha[9]
        cursor.execute(consultaFiscal, (chaveTdpf,))              
        fiscal = cursor.fetchone()
        if fiscal==None:
            nomeFiscal = "ND"
        else:
            nomeFiscal = fiscal[0].strip()
        cursor.execute(consultaPostagem, (chaveTdpf, ))
        row = cursor.fetchone()
        postagemMsg = ""
        if row:
            postagem = "Sim"
            postagemMsg = " (P)"
        else:
            postagem = "Não"
        msg = msg + "\n  *"+formataTDPF(linha[1])+"* ("+nomeFiscal.split()[0]+")"+postagemMsg+"\n"
        userId = linha[2]  
        i+=1
        sheet.cell(row=i, column=1).value = i-1
        sheet.cell(row=i, column=2).value = formataEquipe(equipe)
        sheet.cell(row=i, column=3).value = formataTDPF(linha[1])
        sheet.cell(row=i, column=4).value = linha[7]
        sheet.cell(row=i, column=5).value = nomeFiscal
        sheet.cell(row=i, column=6).value = linha[5].strftime("%d/%m/%Y")
        sheet.cell(row=i, column=7).value = "Fiscalização" if tipoP=="F" else "Diligência"
        sheet.cell(row=i, column=8).value = postagem
        for col in range(len(larguras)):
            if not col in [3, 4]:
                currentCell = sheet.cell(row=i, column=col+1)
                currentCell.alignment = Alignment(horizontal='center', vertical='center')#, wrap_text=True)
        
    if msg!="": #envia a mensagem para o último usuário
        msg = cabecalho+msg
        if ambiente!="PRODUÇÃO":
            msg = msg+"\n"+"Ambiente: "+ambiente  
        if bloqueio=='N':      
            enviaMsgUpdater(userId, text=limpaMarkdown(msg), parse_mode= 'MarkdownV2')
        if not email in [None, ""] and server:
            nomeArq = "Ciencia30_"+str(userId)+"_"+datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")[:-3]+".xlsx"
            book.save(nomeArq)
            if enviaEmail(email, texto, assunto, nomeArq)!=3:
                logging.info("Erro ao enviar email de TDPFs sem ciência registrada após 30 dias - "+email)
            os.remove(nomeArq)        

    #avisamos os supervisores dos processos integrados e sem registro de ciência há mais de 30 dias (o aviso se repete a cada 15 dias, limitado a 60 dias da integração)  
    #(não vai e-mail)
    consulta = """
                Select Distinctrow TDPFS.Numero, AvisosCiencia.Processo, AvisosCiencia.Integracao, Supervisores.Equipe, Usuarios.idTelegram, AvisosCiencia.Codigo, TDPFS.Codigo
                From TDPFS, AvisosCiencia, Fiscais, Supervisores, Usuarios
                Where TDPFS.Codigo=AvisosCiencia.TDPF and TDPFS.Grupo=Supervisores.Equipe and Supervisores.Fim Is Null 
                and Supervisores.Fiscal=Fiscais.Codigo and AvisosCiencia.Finalizado Is Null and Usuarios.CPF=Fiscais.CPF and 
                Usuarios.idTelegram Is Not Null and Usuarios.idTelegram<>0 and Usuarios.Adesao Is Not Null and BloqueiaTelegram='N' and
                (AvisosCiencia.Integracao=cast((now() - interval 30 day) as date) or 
                (cast(AvisosCiencia.Aviso as date)=cast((now() - interval 15 day) as date) and AvisosCiencia.Integracao<=cast((now() - interval 60 day) as date))
                or (AvisosCiencia.Aviso Is Null and AvisosCiencia.Integracao<cast((now() - interval 30 day) as date)))
                Order By Supervisores.Equipe, Usuarios.idTelegram, TDPFS.Numero, AvisosCiencia.Processo            
                """   
    cursor.execute(consulta)
    linhas = cursor.fetchall()
    msg = ""  
    cpfAnt = "" 
    userId = 0 
    cabecalho = "Processos sem registro de ciência há mais de 30 dias da integração(\*).\nTDPF | Processo | Data da Integração:\n"
    equipe = ""     
    for linha in linhas:
        if userId==0:
            userId = linha[4]
        if linha[4]!=userId:
            if msg!="":
                msg = cabecalho+msg
                if ambiente!="PRODUÇÃO":
                    msg = msg+"\n"+"Ambiente: "+ambiente                
                enviaMsgUpdater(userId, text=limpaMarkdown(msg), parse_mode= 'MarkdownV2')
                msg = ""
            equipe = ""    
            userId = linha[4]    
        if linha[3]!=equipe:
            equipe = linha[3]
            msg = msg+"\nEquipe " + formataEquipe(equipe) + ":"    
        chaveTdpf = linha[6]    
        cursor.execute(consultaFiscal, (chaveTdpf,))              
        fiscal = cursor.fetchone()
        if fiscal==None:
            nomeFiscal = "ND"
        else:
            nomeFiscal = fiscal[0].strip()
        processo = linha[1]
        integracao = linha[2].strftime("%d/%m/%Y")
        tdpfsCienciasPendentes.add(linha[5]) 
        msg = msg + "\n *"+formataTDPF(linha[0])+"* ("+nomeFiscal.split()[0]+") | "+formataDCC(processo)+" | "+integracao+"\n"

    if msg!="": #envia a mensagem para o último usuário
        msg = cabecalho+msg
        if ambiente!="PRODUÇÃO":
            msg = msg+"\n"+"Ambiente: "+ambiente        
        enviaMsgUpdater(userId, text=limpaMarkdown(msg), parse_mode= 'MarkdownV2')  

    agora = datetime.now()
    lista = []
    for codigo in tdpfsCienciasPendentes:
        tupla = (agora, codigo)
        lista.append(tupla)
    #atualiza a data de aviso para todos os tdpfs/processos para os quais foram enviados avisos (30 dias da integração ou 15 dias do último aviso e não finalizados)
    atualiza = "Update AvisosCiencia Set Aviso=%s Where Codigo=%s"
    cursor.executemany(atualiza, lista)
    try:
        conn.commit()
    except:
        conn.rollback()
        logging.info("Erro ao tentar atualizar as datas de aviso na tabela AvisosCiencia.")       

    #buscamos todos os TDPFs que estão recuperando a espontaneidade em 15 dias exatos e avisamos o supervisor (apenas Fiscalização)
    comando = """
              Select Distinctrow Fiscais.CPF, TDPFS.Numero, Usuarios.idTelegram, Supervisores.Equipe, TDPFS.Codigo, Usuarios.email, TDPFS.Nome, Usuarios.BloqueiaTelegram
              From TDPFS, Usuarios, Supervisores, Fiscais 
              Where Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and TDPFS.Encerramento Is Null and
              Supervisores.Fiscal=Fiscais.Codigo and Fiscais.CPF=Usuarios.CPF and Usuarios.idTelegram Is Not Null and Usuarios.idTelegram<>0 and Usuarios.Adesao Is Not Null and 
              Usuarios.Saida Is Null and TDPFS.Codigo in (Select TDPF from Ciencias Where Data Is Not Null) and TDPFS.Tipo='F'
              Order By Fiscais.CPF, Supervisores.Equipe, TDPFS.Numero
              """          
    cursor.execute(comando)
    linhas = cursor.fetchall() 
    msg = ""  
    cpfAnt = "" 
    userId = 0 
    consulta = "Select Data, Documento, Vencimento from Ciencias Where TDPF=%s Order By Data DESC"    
    cabecalho = "TDPFs em andamento com espontaneidade tributária sendo recuperada em 15 dias (ciência em "+(dataAtual-timedelta(days=45)).strftime("%d/%m/%Y")+"):"
    equipe = ""
    texto = "Sr. Usuário,\n\nEnviamos, em anexo, planilha com relação de TDPFs sob sua supervisão cuja espontaneidade o contribuinte recuperará em 15 dias.\n\nAtenciosamente,\n\nCofis/Disav"
    if ambiente!="PRODUÇÃO":
        texto = texto+"\n\nAmbiente: "+ambiente    
    assunto = "TDPFs - Recuperação de Espontaneidade em 15 Dias"    
    bloqueio = 'N'   
    for linha in linhas:
        tdpf = linha[1]               
        if cpfAnt=="":
            cpfAnt = linha[0]
            i = 1
            email = linha[5]
            bloqueio = linha[7]
            book = Workbook()
            sheet = book.active
            sheet.cell(row=1, column=1).value='Nº Ordem'
            sheet.cell(row=1, column=2).value='Equipe'
            sheet.cell(row=1, column=3).value='Nº TDPF'
            sheet.cell(row=1, column=4).value='Nome'
            sheet.cell(row=1, column=5).value='Fiscal'
            sheet.cell(row=1, column=6).value='Última Ciência'
            sheet.cell(row=1, column=7).value='Documento'
            sheet.cell(row=1, column=8).value='Vencimento Intimação'
            sheet.cell(row=1, column=9).value='Último Termo Emitido'
            sheet.cell(row=1, column=10).value='Postagem do Últ Termo'
            larguras = [13, 20, 20, 45, 35, 18, 35, 20, 20, 21]
            for col in range(len(larguras)):
                sheet.column_dimensions[get_column_letter(col+1)].width = larguras[col]
                currentCell = sheet.cell(row=1, column=col+1)
                currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                currentCell.font = Font(bold=True)             
        if linha[0]!=cpfAnt:
            if msg!="":
                msg = cabecalho+msg
                if ambiente!="PRODUÇÃO":
                    msg = msg+"\n"+"Ambiente: "+ambiente  
                if bloqueio=='N':              
                    enviaMsgUpdater(userId, text=limpaMarkdown(msg), parse_mode= 'MarkdownV2')
                msg = ""
                if not email in [None, ""] and server:
                    nomeArq = "Ciencia15_"+str(userId)+"_"+datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")[:-3]+".xlsx"
                    book.save(nomeArq)
                    if enviaEmail(email, texto, assunto, nomeArq)!=3:
                        logging.info("Erro ao enviar email de TDPFs recuperando espontaneidade em 15 dias - "+email)
                    os.remove(nomeArq)                  
                i = 1
                book = Workbook()
                sheet = book.active
                sheet.cell(row=1, column=1).value='Nº Ordem'
                sheet.cell(row=1, column=2).value='Equipe'
                sheet.cell(row=1, column=3).value='Nº TDPF'
                sheet.cell(row=1, column=4).value='Nome'
                sheet.cell(row=1, column=5).value='Fiscal'
                sheet.cell(row=1, column=6).value='Última Ciência'
                sheet.cell(row=1, column=7).value='Documento'
                sheet.cell(row=1, column=8).value='Vencimento Intimação'
                sheet.cell(row=1, column=9).value='Último Termo Emitido'
                sheet.cell(row=1, column=10).value='Postagem do Últ Termo'                
                for col in range(len(larguras)):
                    sheet.column_dimensions[get_column_letter(col+1)].width = larguras[col]
                    currentCell = sheet.cell(row=1, column=col+1)
                    currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    currentCell.font = Font(bold=True)  
            email = linha[5]    
            bloqueio = linha[7]            
            equipe = ""    
            cpfAnt = linha[0]    
        if linha[3]!=equipe:
            equipe = linha[3]
        chaveTdpf = linha[4]
        cursor.execute(consultaFiscal, (chaveTdpf,))              
        fiscal = cursor.fetchone()
        if fiscal==None:
            nomeFiscal = "ND"
        else:
            nomeFiscal = fiscal[0].strip()
        cursor.execute(consulta, (chaveTdpf,))
        cienciaReg = cursor.fetchone() #buscamos a última data de ciência do TDPF
        if cienciaReg:
            if len(cienciaReg)>0:
                dataCiencia = cienciaReg[0].date()
                prazoRestante = (dataCiencia+timedelta(days=60)-dataAtual).days                              
                if prazoRestante==15:
                    #verificamos se há postagem relativa ao TDPF na data da última ciência deste ou após
                    consultaPostagem = "Select Data, DataEnvio from ControlePostal Where TDPF=%s and Data>=%s"
                    cursor.execute(consultaPostagem, (chaveTdpf, dataCiencia))
                    regPostal = cursor.fetchone()
                    if regPostal:
                        postagem = " (P)" #indica que houve uma postagem em data igual ou posterior à da última ciência
                        dataEmissao = regPostal[0]
                        dataEnvio = regPostal[1]
                    else:
                        postagem = ""
                        dataEmissao = None
                        dataEnvio = None
                    i+=1                    
                    if msg=="":
                        msg = msg+"\nEquipe " + formataEquipe(equipe) + ":"                                
                    msg = msg + "\n  *"+formataTDPF(tdpf)+"* ("+nomeFiscal.split()[0]+")"
                    if cienciaReg[1]: #documento informado
                        msg = msg +" - "+cienciaReg[1]+postagem+"\n"
                        sheet.cell(row=i, column=7).value = cienciaReg[1]         
                    else:
                        msg = msg + postagem+"\n"
                    sheet.cell(row=i, column=1).value = i-1
                    sheet.cell(row=i, column=2).value = formataEquipe(equipe)
                    sheet.cell(row=i, column=3).value = formataTDPF(linha[1])
                    sheet.cell(row=i, column=4).value = linha[6]
                    sheet.cell(row=i, column=5).value = nomeFiscal
                    sheet.cell(row=i, column=6).value = cienciaReg[0].strftime("%d/%m/%Y")
                    if cienciaReg[2]!=None: #vencimento
                        sheet.cell(row=i, column=8).value = cienciaReg[2].strftime("%d/%m/%Y")
                    if dataEmissao!=None:
                        sheet.cell(row=i, column=9).value = dataEmissao.strftime("%d/%m/%Y")                        
                        sheet.cell(row=i, column=10).value = dataEnvio.strftime("%d/%m/%Y")                        
                    for col in range(len(larguras)):
                        if not col in [3, 4]:
                            currentCell = sheet.cell(row=i, column=col+1)
                            currentCell.alignment = Alignment(horizontal='center', vertical='center') #, wrap_text=True)                        
        userId = linha[2]               
    if msg!="":
        msg = cabecalho+msg
        if ambiente!="PRODUÇÃO":
            msg = msg+"\n"+"Ambiente: "+ambiente
        if bloqueio=='N':
            enviaMsgUpdater(userId, text=limpaMarkdown(msg), parse_mode= 'MarkdownV2')
        if not email in [None, ""] and server:
            nomeArq = "Ciencia15_"+str(userId)+"_"+datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")[:-3]+".xlsx"
            book.save(nomeArq)
            if enviaEmail(email, texto, assunto, nomeArq)!=3:
                logging.info("Erro ao enviar email de TDPFs recuperando espontaneidade em 15 dias - "+email)
            os.remove(nomeArq)          
    conn.close()  
    print("Envio de mensagens finalizado ", datetime.now())            
    return

def formataDCC(dcc):
    if dcc==None:
        return ""
    if len(dcc)<17:
        return dcc
    return dcc[:5]+"."+dcc[5:11]+"/"+dcc[11:15]+"-"+dcc[15:]

def disparaMsgJuntada(): #alerta de juntada para fiscais alocados (não vai para o e-mail)
    global updater, termina, ambiente
    conn = conecta()
    if not conn:
        logging.info("Conexão ao BD falhou (disparaMsgJuntada")
        return
    logging.info("Acionado o disparo de mensagens Solic Juntada - "+datetime.now().strftime('%d/%m/%Y %H:%M'))
    cursor = conn.cursor(buffered=True)
    dataAtual = datetime.today().date()
    comando = "Select idTelegram, CPF from Usuarios Where Adesao Is Not Null and Saida Is Null and idTelegram Is Not Null and idTelegram<>0 and BloqueiaTelegram='N' and Ativo='S'"
    cursor.execute(comando)
    usuarios = cursor.fetchall()
    totalMsg = 0
    msgDisparadas = 0
    juntadasAvisadas = set()
    cabecalho = "Alerta de Solicitação de Juntada Pendente de Análise (TDPF | DCC):\n\n" 
    for usuario in usuarios: #percorremos os usuários ativos Telegram
        if termina: #programa foi informado de que é para encerrar (quit)
            return
        #logging.info("Verificando disparo de Aviso de Juntada para "+usuario[1])
        
        listaUsuario = "" #lista de TDPF com espontaneidade, atividade ou o próprio TDPF vencendo
        userId = usuario[0]
        cpf = usuario[1]
        #selecionamos os TDPFs do usuário em andamento e monitorados (ativos) pelo serviço
        comando = """
                Select TDPFS.Numero, TDPFS.DCC, Juntadas.Solicitacao, Juntadas.Codigo
                from CadastroTDPFs, TDPFS, Alocacoes, Fiscais, Juntadas
                Where Fiscais.CPF=%s and CadastroTDPFs.Fiscal=Fiscais.Codigo and CadastroTDPFs.TDPF=TDPFS.Codigo and 
                CadastroTDPFs.TDPF=Alocacoes.TDPF and  
                CadastroTDPFs.Fiscal=Alocacoes.Fiscal and TDPFS.Encerramento Is Null and TDPFS.DCC Is Not Null and TDPFS.DCC<>'' and
                CadastroTDPFs.Fim Is Null and Alocacoes.Desalocacao Is Null and TDPFS.Codigo=Juntadas.TDPF and Juntadas.Solicitacao Is Not Null and 
                (Juntadas.Solicitacao>Juntadas.Aviso or Juntadas.Aviso Is Null)
                Order by TDPFS.Numero
                """
        cursor.execute(comando, (cpf,))        
        fiscalizacoes = cursor.fetchall()  
        for fiscalizacao in fiscalizacoes:
            if listaUsuario!="":
                listaUsuario = listaUsuario + "\n"
            listaUsuario = "*"+formataTDPF(fiscalizacao[0])+"* |\n"+formataDCC(fiscalizacao[1])+"\n"
            juntadasAvisadas.add(fiscalizacao[3])
        if listaUsuario!="":
            listaUsuario = cabecalho + listaUsuario
            if ambiente!="PRODUÇÃO":
                listaUsuario = listaUsuario + "\n\nAmbiente: "+ambiente
            enviaMsgUpdater(userId, text=limpaMarkdown(listaUsuario), parse_mode= 'MarkdownV2')  
            totalMsg+=1
            msgDisparadas+=1
            if msgDisparadas>=30:
                msgDisparadas = 0
                time.sleep(1) #a cada 30 mensagens, dormimos um segundo (limitação do Bot é 30 por seg - TESTE)             
    lista = []
    for codigo in juntadasAvisadas:
        tupla = (dataAtual, codigo)
        lista.append(tupla)                         
    if len(lista)>0:   
        logging.info("Atualização dos Avisos de Pendência de Solicitação de Juntada:")
        logging.info(lista)         
        comando = "Update Juntadas Set Aviso=%s Where Codigo=%s" 
        try:
            cursor.executemany(comando, lista)
            conn.commit()
        except:
            logging.info("Erro ao tentar atualizar as datas de aviso na tabela Juntadas.")
            conn.rollback()   
    logging.info("Total de mensagens disparadas (juntadas): "+str(totalMsg)) 
    conn.close()
    return

def disparaMsgProrrogacao(): #avisa sobre prorrogação pendente de assinatura ou de registro no RHAF - apenas fiscais alocados (não vai para e-mail)
    global updater, termina, ambiente, dirLog
    #verificamos o dia e a hora da última pesquisa    
    try:
        f = open(dirLog+"AvisoProrrogacao.txt",'r')
        ultimaDataHora = f.read()
        ultimaDataHora = datetime.strptime(ultimaDataHora, "%d/%m/%Y %H:%M") #esta data serve para não mandar msg sobre evento já verificado
        f.close()
    except:
        ultimaDataHora =  datetime.strptime("01/01/1900 00:00", "%d/%m/%Y %H:%M")  
    conn = conecta()
    if not conn:
        logging.info("Conexão ao BD falhou (disparaMsgProrrogacao")
        return
    logging.info("Acionado o disparo de mensagens Prorrogação - "+datetime.now().strftime('%d/%m/%Y %H:%M'))
    cursor = conn.cursor(buffered=True)
    comando = """Select idTelegram, Fiscais.Codigo, Usuarios.CPF from Usuarios, Fiscais
                 Where Adesao Is Not Null and Saida Is Null and idTelegram Is Not Null and idTelegram<>0 and Usuarios.CPF=Fiscais.CPF and Usuarios.BloqueiaTelegram='N' and Ativo='S' """
    cursor.execute(comando)
    usuarios = cursor.fetchall()
    totalMsg = 0
    msgDisparadas = 0    
    cabecalho = "Alerta de Prorrogação Pendente (TDPF | Pendência):\n\n" 
    consulta1 = """Select Distinctrow TDPFS.Numero From Prorrogacoes, AssinaturaFiscal, TDPFS 
                   Where Prorrogacoes.TDPF=TDPFS.Codigo and Prorrogacoes.Codigo=AssinaturaFiscal.Prorrogacao and Prorrogacoes.Data>%s and
                   AssinaturaFiscal.Fiscal=%s and AssinaturaFiscal.DataAssinatura Is Null"""    
    consulta2 = """Select TDPFS.Numero, Prorrogacoes.Codigo, AssinaturaFiscal.Codigo From Prorrogacoes, AssinaturaFiscal, TDPFS 
                   Where Prorrogacoes.DataAssinatura Is Not Null and Prorrogacoes.DataAssinatura>%s and Prorrogacoes.TDPF=TDPFS.Codigo and AssinaturaFiscal.Fiscal=%s
                   and Prorrogacoes.Codigo=AssinaturaFiscal.Prorrogacao and Prorrogacoes.RegistroRHAF Is Null and AssinaturaFiscal.DataAssinatura Is Not Null"""                    
    for usuario in usuarios: #percorremos os usuários ativos Telegram
        if termina: #programa foi informado de que é para encerrar (quit)
            return
        #logging.info("Verificando disparo de Aviso de Prorrogação para "+usuario[2])
        
        listaUsuario = "" #lista de TDPF com espontaneidade, atividade ou o próprio TDPF vencendo
        userId = usuario[0]
        chaveFiscal = usuario[1]
        cursor.execute(consulta1, (ultimaDataHora, chaveFiscal))
        linhas = cursor.fetchall()
        for linha in linhas: #avisamos de assinaturas pendentes
            listaUsuario = listaUsuario + "*"+formataTDPF(linha[0])+"* | Assinatura\n\n"
        cursor.execute(consulta2, (ultimaDataHora, chaveFiscal))
        linhas = cursor.fetchall()
        for linha in linhas: #avisamos de pendência de registro no RHAF - apenas fiscal que criou a prorrogação
            tdpfFormatado = formataTDPF(linha[0])
            prorrogacao = linha[1]
            codigoAssinatura = linha[2]
            #temos que verificar se este é o fiscal que criou a prorrogação, ou seja, o que tem menor código na tabela assinatura fiscal dentre os da prorrogação corrente
            consultaFiscalResp = "Select MIN(Codigo) from AssinaturaFiscal Where AssinaturaFiscal.Prorrogacao=%s"
            cursor.execute(consultaFiscalResp, (prorrogacao, ))
            menorCodigoAssinatura = cursor.fetchone()
            if codigoAssinatura==menorCodigoAssinatura[0]: #é a menor, então é o fiscal que criou a prorrogação - receberá o aviso
                listaUsuario = listaUsuario + "*"+tdpfFormatado+"* | Registro no RHAF"
        if listaUsuario!="":
            listaUsuario = cabecalho + listaUsuario
            if ambiente!="PRODUÇÃO":
                listaUsuario = listaUsuario + "\n\nAmbiente: "+ambiente            
            enviaMsgUpdater(userId, text=limpaMarkdown(listaUsuario), parse_mode= 'MarkdownV2')  
            totalMsg+=1
            msgDisparadas+=1
            if msgDisparadas>=30:
                msgDisparadas = 0
                time.sleep(1) #a cada 30 mensagens, dormimos um segundo (limitação do Bot é 30 por seg - TESTE)  
    try:
        f = open(dirLog+"AvisoProrrogacao.txt",'w')
        f.write(datetime.now().strftime("%d/%m/%Y %H:%M")) #atualizamos a data para não informarmos sobre pendências já verificadas
        f.close()                  
    except:
        print("Falhou a criação do arquivo com o registro da data e hora do envio de mensagens - PRORROGAÇÕES")
        logging.info("Falhou a criação do arquivo com o registro da data e hora do envio de mensagens - PRORROGAÇÕES")
    logging.info("Total de mensagens disparadas (prorrogações): "+str(totalMsg)) 
    conn.close()
    return                      

def disparador():
    global termina, dirLog, sistema, ambiente, diaAtual
    logging.info("Disparador (thread) iniciado ...")
    while not termina:
        schedule.run_pending() 
        logging.info("Disparador (thread) indo 'dormir'")
        time.sleep(60*60) #dorme por 1 h
        dia = datetime.now().date()
        if ambiente!='TESTE' and diaAtual!=dia:
            diaAtual = dia
            #a cada dia inicia um arquivo de log diferente
            logging.basicConfig(filename=dirLog+datetime.now().strftime('%Y-%m-%d %H_%M')+' Bot'+sistema+'.log', format='%(asctime)s - %(message)s', level=logging.INFO, force=True)       
    return 

def conecta():
    global MYSQL_DATABASE, MYSQL_USER, MYSQL_PASSWORD, hostSrv
    try:
        #logging.info("BD: "+MYSQL_DATABASE)
        #logging.info(MYSQL_PASSWORD)
        #logging.info("User: "+MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        return conn                         
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD - veja Log")
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return None

sistema = sys.platform.upper()
if "WIN32" in sistema or "WIN64" in sistema or "WINDOWS" in sistema:
    hostSrv = 'localhost'
    dirLog = 'log\\'
else:
    hostSrv = 'mysqlsrv'
    dirLog = '/Log/'  

#print(datetime.now().strftime('%Y-%m-%d %H_%M')+' BotLog'+sistema+'.log')
    
logging.basicConfig(filename=dirLog+datetime.now().strftime('%Y-%m-%d %H_%M')+' Bot'+sistema+'.log', format='%(asctime)s - %(message)s', level=logging.INFO)

MYSQL_ROOT_PASSWORD = os.getenv("MYSQL_ROOT_PASSWORD", "EXAMPLE")
MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234")
token = os.getenv("TOKEN", "ERRO")
ambiente = os.getenv("AMBIENTE", "TESTE")
senhaAvisoUrgente = os.getenv("AVISOURGENTE", "760714")
logging.info("Ambiente: "+ambiente)
if token=="ERRO":
    logging.error("Token do Bot Telegram não foi fornecido.")
    print("Token não informado")
    sys.exit(1)

logging.info("Conectando ao servidor de banco de dados ...")
conn = conecta() #não será utilizada - apenas conectamos para ver se está ok
if not conn:
    sys.exit(1)
logging.info("Conexão efetuada com sucesso ao MySql!")
logging.info("BD: "+MYSQL_DATABASE)
#logging.info(MYSQL_PASSWORD)
logging.info("User: "+MYSQL_USER)
conn.close() #cada função criará sua conexão (aqui é só um teste para inicializarmos o Bot)
#dias de antecedência padrão para avisar
d1padrao = 30
d2padrao = 20
d3padrao = 5

atividadeTxt = {} #guarda o tdpf e o prazo de uma atividade, pendente a informação do texto de sua descrição (id: [tdpf, data])

cienciaTxt = {} #guarda o tdpf e a data de ciência, pendente a informação do texto da descrição do documento (id: [tdpf, data])

pendencias = {} #indica que próxima função deve ser chamada para analisar entrada de dados

#encaminha a pendência para a função adequada para tratar o input do usuário
dispatch = { 'registra': registra, 'ciencia': ciencia, 'prazos': prazos, 'acompanha': acompanha,
             'anulaCiencia': anulaCiencia, 'fim': fim, 'email': cadastraEMail,
             'atividade': atividade, 'anulaAtividade': anulaAtividade, 
             'atividadeTexto': atividadeTexto, 'cienciaTexto': cienciaTexto, 'envioChave': envioChave, 
             'mostraSuperv': mostraSupervisionados, 'informaTerminoAtividade': terminoAtividade,
             'supervisorEspontaneidade': mostraSupervisorEspontaneidade, 'informaHorasAtividade': informaHorasAtividade, 
             'relacionaPostagens': relacionaPostagens}
textoRetorno = "\nEnvie /menu para retornar ao menu principal"
updater = None #para ser acessível ao disparador de mensagens
#schedule.every().day.at("07:30").do(disparaMensagens)
if ambiente=="TESTE":
    schedule.every(60).minutes.do(disparaMensagens) #deixamos enviar msgs a cada 1 h no ambiente de testes
    schedule.every(60).minutes.do(disparaMsgJuntada)
    schedule.every(60).minutes.do(disparaMsgProrrogacao)
else:
	schedule.every().day.at("07:30").do(disparaMensagens) #uma vez por dia - produção     
if ambiente!="TESTE":
    schedule.every().day.at("15:00").do(disparaMsgJuntada) #não funciona colocando no else acima (dá erro de identação)
    schedule.every().day.at("10:00").do(disparaMsgProrrogacao) 
    schedule.every().day.at("15:45").do(disparaMsgProrrogacao)
termina = False
diaAtual = datetime.now().date() #será utilizado para criar um arquivo de Log p/ cada dia
botTelegram()
threadDisparador = threading.Thread(target=disparador, daemon=True) #encerra thread quando sair do programa sem esperá-la
threadDisparador.start()
print("Serviço iniciado [", datetime.now(), "]")
if ambiente=="PRODUÇÃO" and datetime.now().strftime("%d/%m/%Y")=="15/06/2021":
    disparaMensagens()
#disparaMsgJuntada
#disparaMsgProrrogacao()
while not termina:
    entrada = input("Digite QUIT para terminar o serviço BOT: ")
    if entrada:
        if entrada.strip().upper()=="QUIT":
            termina = True
updater.stop() 
schedule.clear() 
       