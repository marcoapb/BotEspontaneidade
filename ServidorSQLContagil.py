# -*- coding: utf-8 -*-
"""
Created on Tue Jul 21 17:27:16 2020

@author: 53363833172
"""

from __future__ import unicode_literals
from datetime import datetime, timedelta
import time
import sys
import os
import logging   
import mysql.connector
from mysql.connector import errorcode
import re
import socket
import threading
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_OAEP
from Crypto.Cipher import AES
from random import randint
import calendar
import schedule #para mandar e-mail com pontos do trimestre do fiscal e a média de sua equipe
import requests #para pesquisar situação de correspondências nos correios
import pandas as pd #para converter o html em tabelas (consulta correios)

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.utils import formatdate
from email import encoders
import smtplib

import zlib #para compactar as mensagens enviadas

from openpyxl import Workbook
#from openpyxl.styles import colors
from openpyxl.styles import Font #, Color
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

#função para o envio de e-mail - copiado do Bot_Telegram.py
def enviaEmail(email, texto, assunto, arquivo=None):
    try:
        #server = smtplib.SMTP('INETRFOC.RFOC.SRF: 25') #servidor de email Notes
        #pass
        server = smtplib.SMTP('exchangerfoc.rfoc.srf: 25')
    except:
        return 1	
    # create message object instance
    msg = MIMEMultipart()
    # setup the parameters of the message
    msg['From'] = "botespontaneidade@rfb.gov.br"
    msg['To'] = email
    msg['Subject'] = assunto
    # add in the message body
    msg.attach(MIMEText(texto, 'plain'))  
    if arquivo!=None and arquivo!="":  
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(arquivo, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename='+arquivo)
        msg.attach(part)                 
    # send the message via the server.
    try:
        server.sendmail(msg['From'], msg['To'], msg.as_string())
        pass
    except:
        return 2
    server.quit()  
    return 3 #sucesso
##########################################################################################

def descriptografa(msgCripto, addr, c):
    global chavesCripto
    try:
        chaveCriptoAES = str.encode(chavesCripto[segmentoIP(addr)][1].decrypt(msgCripto[:256]).decode("utf-8"))
        nonce = msgCripto[256:272]
        tag = msgCripto[272:288]
        cipher = AES.new(chaveCriptoAES, AES.MODE_EAX, nonce=nonce)
        decrypted = cipher.decrypt(msgCripto[288:]).decode("utf-8")
        cipher.verify(tag)
    except:
        #se não deu certo, esperamos um tempo para ver se chega o restante da mensagem (pode ter vindo só uma parte)
        try:
            msg = c.recv(2048)
            if msg==None or msg==b"":
                return "000000000A", None
        except: #time out
            return "000000000A", None #só uma mensagem dummy para demonstrar o erro quando não for possível descriptografar (não chegou o restante)
        #chegou - tentamos descriptografar agora com o restante da mensagem
        try:
            msgCripto += msg
            chaveCriptoAES = str.encode(chavesCripto[segmentoIP(addr)][1].decrypt(msgCripto[:256]).decode("utf-8"))
            nonce = msgCripto[256:272]
            tag = msgCripto[272:288]
            cipher = AES.new(chaveCriptoAES, AES.MODE_EAX, nonce=nonce)
            decrypted = cipher.decrypt(msgCripto[288:]).decode("utf-8")
            cipher.verify(tag)
        except:
            return "000000000A", None #só uma mensagem dummy para demonstrar o erro quando não for possível descriptografar
    #decrypted = decryptor.decrypt(chaveCripto).decode("utf-8")
    return decrypted, chaveCriptoAES    

def getAlgarismos(texto): #retorna apenas os algarismos de uma string
    limpo = ""
    for car in texto:
        if car.isdigit():
            limpo = limpo + car
    return limpo 

def formataTDPF(tdpf):
    if tdpf==None:
        return None
    tipo = str(type(tdpf)).upper()
    if not 'STR' in tipo and not 'UNICODE' in tipo:
        return None
    if len(tdpf)<16:
        return tdpf
    return tdpf[:7]+"."+tdpf[7:11]+"."+tdpf[11:] 

def formataEquipe(equipe):
    if equipe==None:
        return ""
    tipo = str(type(equipe)).upper()
    if not 'STR' in tipo and not 'UNICODE' in tipo:
        return ""       
    if len(equipe)<14:
        return equipe  
    return equipe[:7]+"."+equipe[7:11]+"."+equipe[11:]     

#verifica se um CPF é válido
def validaCPF(cpfPar):
#The MIT License (MIT) Copyright (c) 2015 Derek Willian Stavis
    if cpfPar==None:
        return False
    if not 'STR' in str(type(cpfPar)).upper() and not "UNICODE" in str(type(cpfPar)).upper():
        return False
    cpf = getAlgarismos(cpfPar)
    if len(cpf)!=11:
        return False

    if cpf in [s * 11 for s in [str(n) for n in range(10)]]:
        return False

    calc = lambda t: int(t[1]) * (t[0] + 2)
    d1 = (sum(map(calc, enumerate(reversed(cpf[:-2])))) * 10) % 11
    d2 = (sum(map(calc, enumerate(reversed(cpf[:-1])))) * 10) % 11
    if d1==10:
        d1 = 0
    if d2==10:
        d2 = 0    
    return str(d1) == cpf[-2] and str(d2) == cpf[-1]

#transforma uma data string de dd/mm/yyyy para yyyy/mm/dd para fins de consulta, inclusão ou atualização no BD SQL
#se o BD esperar a data em outro formato, basta alterarmos aqui
def converteAMD(data):
    return data[6:]+"/"+data[3:5]+"/"+data[:2]


def verificaEMail(email): #valida o e-mail se o usuário informou um completo
    regex1 = '^[a-zA-Z0-9]+[\._]?[a-zA-Z0-9\.\-]+[@]\w+[.]\w{2,3}$'
    regex2 = '^[a-zA-Z0-9]+[\._]?[a-zA-Z0-9\.\-]+[@]\w+[.]\w+[.]\w{2,3}$'  

    if(re.search(regex1,email)):  
        return True   
    elif(re.search(regex2,email)):  
        return True
    else:  
        return False

def dataTexto(data):
    if data==None:
        return "00/00/0000"
    else:
        try:
            return data.strftime("%d/%m/%Y")     
        except:
            return "00/00/0000"

def digitoVerificadorModulo11(codigo, dv1QuandoModulo0):
    soma = 0
    peso = 2
    i = len(codigo)-1
    while i>=0:
        soma += int(codigo[i:i+1]) * peso
        peso+=1
        i-=1
    modulo = soma % 11
    if (11 - modulo)==10:
        return "0"
    if (11 - modulo)==11:
        return "1" if dv1QuandoModulo0 else "0"
    return str(11 - modulo)
                        
def verificaDVDCC(dcc): #só 17 dígitos
    primeiroDV = dcc[-2:-1]
    segundoDV = dcc[-1:]
    dccSemDV = dcc[:15]
    primDVCalculado = digitoVerificadorModulo11(dccSemDV, True)
    if primDVCalculado==primeiroDV:
        segDVCalculado = digitoVerificadorModulo11(dccSemDV + primDVCalculado, True)
        if segDVCalculado==segundoDV:
            return True
    return False

def conecta():
    global MYSQL_DATABASE, MYSQL_USER, MYSQL_PASSWORD, hostSrv
    try:
        #logging.info("BD: "+MYSQL_DATABASE)
        #logging.info(MYSQL_PASSWORD)
        #logging.info("User: "+MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        return conn                         
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD - veja Log")
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return None   

def conectaRaw():
    global MYSQL_DATABASE, MYSQL_USER, MYSQL_PASSWORD, hostSrv
    try:
        #logging.info("BD: "+MYSQL_DATABASE)
        #logging.info(MYSQL_PASSWORD)
        #logging.info("User: "+MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE, raw=True)
        return conn                         
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD - veja Log")
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return None              

def enviaRespostaSemFechar(resposta, c, bCriptografa=False, chaveCriptoAES=None, bComprime=False):
    #logging.info(resposta)  
    resposta = resposta.encode('utf-8')
    if bComprime:   
        resposta = zlib.compress(resposta)
        compressao = b"S"
    else:
        compressao = b'N' 
    tam = len(resposta)+1+1+(32 if (bCriptografa and chaveCriptoAES!=None) else 0) #inclui a informação de se está criptografada e comprimido ou não, mais o nonce e o tag no caso de criptografar
    if tam>999999: #tamanho máximo de uma resposta
        print("Mensagem ultrapassou o tamanho de 999999 bytes - ", c.getpeername())
        print("Resposta: ", resposta[:2])
        logging.error("Mensagem ultrapassou o tamanho de 99999 bytes - "+str(c.getpeername())+" - Resposta: "+resposta[:2])
        return
    else:
        tam = str(tam).rjust(6,"0").encode('utf-8')    
    if bCriptografa and chaveCriptoAES!=None:
        cipher = AES.new(chaveCriptoAES, AES.MODE_EAX)   
        nonce = cipher.nonce #16 bytes
        resposta, tag = cipher.encrypt_and_digest(resposta) #tag - 16 bytes
        resposta = b"S"+compressao+nonce+tag+resposta
    else:
        resposta = b"N"+compressao+resposta #primeiro S/N refere-se à criptografia; segundo, à compressão
    try:    
        c.sendall(tam+resposta) 
    except:
        logging.info("Erro ao enviar a resposta - exceção - "+str(resposta))
    return

def enviaResposta(resposta, c, bCriptografa=False, chaveCriptoAES=None, bComprime=False):
    enviaRespostaSemFechar(resposta, c, bCriptografa, chaveCriptoAES, bComprime)
    c.close()        
    return

def ultimoDiaMes(mes, ano): #retorna o último dia do mês
    if mes in (1, 3, 5, 7, 8, 10, 12):
        ultimo_dia = 31
    elif mes == 2:
        # verifica se é ano bissexto
        if (ano % 4 == 0) and (ano % 100 != 0 or ano % 400 == 0):
            ultimo_dia = 29
        else:
            ultimo_dia = 28
    else:
        ultimo_dia = 30
    return ultimo_dia        

def isDate(data): #verifica se a string é uma data válida
    if data==None:
        return False
    if len(data)!=10: #exige no formato dd/mm/aaaa
        return False
    # faz o split e transforma em números
    try:
        dia, mes, ano = map(int, data.split('/'))
    except:
        return False
    # mês ou ano inválido (só considera do ano 1 em diante), retorna False
    if mes < 1 or mes > 12 or ano <= 0:
        return False
    ultimo_dia = ultimoDiaMes(mes, ano)
    # verifica se o dia é válido
    if dia < 1 or dia > ultimo_dia:
        return False
    return True

def atualizaTentativas(codigo, tentativas, conn): #incrementa o numero de tentativas em caso de erro
    if tentativas==None:
        tentativas = 1
    else:
        tentativas+=1
    cursor = conn.cursor(buffered=True)
    comando = "Update Usuarios Set Tentativas=%s Where Codigo=%s"
    try:
        cursor.execute(comando, (tentativas, codigo))
        conn.commit()   #FALTA CONTROLE DE ERROS
    except:
        conn.rollback() #não há o que fazer
    return

def zeraTentativas(codigo, conn): #zera as tentativas de utilizar a senha em caso de acerto
    cursor = conn.cursor(buffered=True)
    comando = "Update Usuarios Set Tentativas=0 Where Codigo=%s"
    try:
        cursor.execute(comando, (codigo,))
        conn.commit()    #FALTA CONTROLE DE ERROS
    except:
        conn.rollback() #não há o que fazer
    return 

def eliminaTags(texto):
    tags = 0
    if texto==None:
        return ""
    textoLimpo = ""
    for j in range(len(texto)):
        if texto[j:j+1]=="<":
            tags+=1
            continue
        if texto[j:j+1]==">":
            tags-=1
            continue
        if tags==0:
            textoLimpo = textoLimpo + texto[j:j+1]
    return textoLimpo

def verificaAlocacao(conn, cpf, tdpf): #verifica se o fiscal (cpf) está alocado ao TDPF em andamento - retorna tb o nome do fiscalizado
    cursor = conn.cursor(buffered=True)
    comando = """Select TDPFS.Nome
                 from TDPFS, Alocacoes, Fiscais 
                 Where TDPFS.Numero=%s and Fiscais.CPF=%s and Fiscais.Codigo=Alocacoes.Fiscal and TDPFS.Codigo=Alocacoes.TDPF and 
                 TDPFS.Encerramento Is Null and Alocacoes.Desalocacao Is Null"""
    cursor.execute(comando, (tdpf, cpf))
    row = cursor.fetchone()  
    if not row:
        return False, None
    if len(row)==0:
        return False, None
    return True, row[0]

def verificaSupervisao(conn, cpf, tdpf): #verifica se o fiscal (cpf) é supervisor da equipe do fiscal responsável pelo TDPF - retorna tb o nome do fiscalizado
    cursor = conn.cursor(buffered=True)
    comando = """Select TDPFS.Nome
                 from TDPFS, Supervisores, Fiscais 
                 Where TDPFS.Numero=%s and TDPFS.Grupo=Supervisores.Equipe and Fiscais.CPF=%s and Fiscais.Codigo=Supervisores.Fiscal and Supervisores.Fim Is Null"""
    cursor.execute(comando, (tdpf, cpf))
    row = cursor.fetchone()  
    if not row:
        return False, None
    if len(row)==0:
        return False, None
    return True, row[0] 

def tdpfMonitoradoCPF(conn, tdpf, cpf): #verifica se o TDPF está sendo monitorado e se tal monitoramento está ativo
    cursor = conn.cursor(buffered=True)
    comando = """Select CadastroTDPFs.Codigo, CadastroTDPFs.Fim from CadastroTDPFs, Fiscais, TDPFS 
                 Where TDPFS.Numero=%s and Fiscais.CPF=%s and CadastroTDPFs.TDPF=TDPFS.Codigo and CadastroTDPFs.Fiscal=Fiscais.Codigo"""
    cursor.execute(comando, (tdpf, cpf))
    row = cursor.fetchone()
    tdpfMonitorado = False
    monitoramentoAtivo = None
    chave = None
    if row:
        if len(row)>0:
            tdpfMonitorado = True  
            chave = row[0]
            fim = row[1]
            if fim==None:
                monitoramentoAtivo = True
            else:
                monitoramentoAtivo = False
    return tdpfMonitorado, monitoramentoAtivo, chave #indica se o tdpf está sendo monitorado; se True, o segundo retorno indica se tal monitoramento está ativo e o terceiro a chave do registro     

def segmentoIP(addr):
    global ambiente
    return sum(map(int, addr.split("."))) % 10 #10 segmentos (chaves)

def geraChaves(addr): #gera a chave (par) do usuário, coloca no dicionário de validades e retorna a chave pública para ser encaminhada a ele
    global chavesCripto
    keyPair = RSA.generate(2048)
    decryptor = PKCS1_OAEP.new(keyPair)        
    chavesCripto[segmentoIP(addr)] = [keyPair, decryptor, datetime.now()+timedelta(hours=1)] #validade de 1h para este conjunto de chaves
    #pubKey = keyPair.publickey()
    return #pubKey
    #print(f"Public key:  (n={hex(pubKey.n)}, e={hex(pubKey.e)})")
    #pubKeyPEM = pubKey.exportKey()   

def inicializaChaves():
    global chavesCripto, ambiente   
    for i in range(10): #geramos 10 chaves para dificultar qualquer ação maliciosa - a atribuição é feita de acordo com o IP
        keyPair = RSA.generate(2048)
        decryptor = PKCS1_OAEP.new(keyPair)        
        chavesCripto[i] = [keyPair, decryptor, datetime.now()+timedelta(hours=1)] #cada chave tem validade de uma hora
    return

def estaoChavesValidas(addr):
    global chavesCripto
    if chavesCripto==None:
        return False #não há chave a ser revalidada
    if chavesCripto[segmentoIP(addr)][2]<datetime.now():
        return False
    else:
        return True

def buscaTipoOrgao(orgao, cursor):
    if orgao==0 or orgao==None:
        return "L", ""
    comando = "Select Tipo, Orgao from Orgaos Where Codigo=%s"
    cursor.execute(comando, (orgao,))
    row = cursor.fetchone()
    if row==None or len(row)==0:
        return "I", ""
    else:
        orgaoResp = row[0]
        if orgaoResp=="" or orgaoResp==None:
            orgaoResp = "L"
        return orgaoResp, row[1].strip()

def trataMsgRecebida(msgRecebida, c, addr): #c é o socket estabelecido com o cliente que será utilizado para a resposta
    global chavesCripto, ambiente, CPF1, CPF2, CPF3, CPF4, CPF5, SENHADCCS, dataExtracao, ultimaVerificacao
    tamChave = 6 #tamanho da chave do ContÁgil (chave de registro) 
    tamMsg = 100
    tamNome = 100
    try: #tentamos obter mensagens descriptografadas - solicitação de chave criptográfica pública ou solicitação de envio de chave/senha para o e-mail
        msgOrigem = msgRecebida.decode("utf-8")     
        if len(msgOrigem)==13:
            if msgOrigem[:2]=="00": #está fazendo uma requisição de chave pública (msg sem criptografia)
                cpf = msgOrigem[2:]
                if not validaCPF(cpf):
                    resposta = "97CPF INVÁLIDO"
                    enviaResposta(resposta, c)        
                    return                
                if not estaoChavesValidas(addr): #meio difícil de ocorrer aqui, pq tem lá no servidor, mas ...
                    geraChaves(addr) 
                versaoScript = "200" #versão mínima do Script (X.XX, sem o ponto; colocar o zero ao final, se for o caso) - só informa na mensagem abaixo (não restringe nas requisições)
                enviaRespostaSemFechar("0000"+chavesCripto[segmentoIP(addr)][2].strftime("%d/%m/%Y %H:%M:%S")+versaoScript+dataExtracao.strftime("%d/%m/%Y %H:%M"), c) #envia a validade da chave e a versão mínima exigida do Script
                try:
                    msg = c.recv(1024).decode("utf-8") #só aguarda um 00
                    if msg=="00":
                        #print(len(chavesCripto[segmentoIP(addr)][0].publickey().export_key()))
                        c.sendall(chavesCripto[segmentoIP(addr)][0].publickey().export_key()) #envia a chave pública (tamanho = 450)
                except:
                    logging.info("Não enviou o flag para receber a chave "+msgOrigem[2:])
                c.close()
                return
            elif msgOrigem[:2]=="26": #está fazendo uma requisição de envio de chave/senha para o e-mail institucional vinculado ao CPF
                cpf = msgOrigem[2:]
                if not validaCPF(cpf):
                    resposta = "97CPF INVÁLIDO"
                    enviaResposta(resposta, c)        
                    return                      
                conn = conecta()
                if not conn:
                    resposta = "97ERRO NA CONEXÃO COM O BANCO DE DADOS" #erro de conexão ou de BD
                    enviaResposta(resposta, c) 
                    return   
                cursor = conn.cursor(buffered=True)                                   
                comando = "Select Codigo, email, DataEnvio, Ativo from Usuarios Where CPF=%s"
                cursor.execute(comando, (cpf,))
                row = cursor.fetchone()
                if not row or len(row)==0:
                    resposta = "97CPF NÃO CONSTA DA BASE DE DADOS DO SERVIÇO"
                    enviaResposta(resposta, c) 
                    return  
                codigoReg = row[0]
                email = row[1]
                dataEnvio = row[2]  
                ativo = row[3]
                if ativo!='S':
                    resposta = "26USUÁRIO NÃO ESTÁ ATIVO NA BASE DE DADOS - ENTRE EM CONTATO COM A COFIS/DISAV"
                    enviaResposta(resposta, c) 
                    return                     
                if email==None:
                    resposta = "26USUÁRIO NÃO TEM EMAIL CADASTRADO NA BASE - CONTACTE botespontaneidade@rfb.gov.br" 
                    enviaResposta(resposta, c) 
                    return 
                email = email.strip()
                if email[-11:].upper()!='@RFB.GOV.BR':
                    resposta = "26EMAIL DO USUÁRIO CADASTRADO NA BASE NÁO É INSTITUCIONAL - CONTACTE botespontaneidade@rfb.gov.br"
                    enviaResposta(resposta, c) 
                    return  
                if dataEnvio!=None:     
                    if dataEnvio.date()>=datetime.now().date():
                        resposta = "26SOMENTE É PERMITIDA UMA REQUISIÇÃO DESTA POR DIA"
                        enviaResposta(resposta, c) 
                        return 
                chave = randint(100000, 1000000) #a chave é um número inteiro de seis dígitos
                message = "Prezado(a),\n\nSua chave SIGILOSA de registro no Bot Espontaneidade (Telegram) é "+str(chave)+"\n\nEsta chave é utilizada também para acesso via ContÁgil no Script AlertasFiscalização e tem validade de 30 dias.\n\nAtenciosamente,\n\nDisav/Cofis\n\nAmbiente: "+ambiente 
                assunto = "Chave de Registro - Bot Espontaneidade"
                sucesso = enviaEmail(email, message, assunto)
                if sucesso==3:
                    comando = "Update Usuarios Set Chave=%s, ValidadeChave=%s, Tentativas=%s, DataEnvio=%s Where Codigo=%s"
                    validade = datetime.today().date()+timedelta(days=30) #a chave tem validade de 30 dias
                    try:
                        cursor.execute(comando, (chave, validade, 0, datetime.today().date(), codigoReg))
                        conn.commit()
                        resposta = "CHAVE FOI ENVIADA PARA O E-MAIL INSTITUCIONAL E ATUALIZADA NA BASE DE DADOS."
                    except:
                        conn.rollback()    
                        resposta = "ERRO AO ATUALIZAR A TABELA DE USUÁRIOS - A CHAVE ENVIADA PARA O E-MAIL NÃO É VALIDA"
                else:
                    resposta = "ERRO AO TENTAR ENVIAR O E-MAIL"
                    if ambiente!="PRODUÇÃO":
                        print(email)                    
                        print(message)
                enviaResposta("26"+resposta, c) 
                return                     
    except:
        pass #não conseguiu decodificar a msgRecebida

    if not estaoChavesValidas(addr):
        resposta = "99REQUISIÇÃO RECUSADA - CHAVE CRIPTOGRÁFICA VENCIDA" #código de erro na mensagem recebida
        enviaResposta(resposta, c)
        return        

    msgRecebida, chaveCriptoAES = descriptografa(msgRecebida, addr, c)
    if msgRecebida=="000000000A": #houve falha na descriptografia
        resposta = "99REQUISIÇÃO RECUSADA - CRIPTOGRAFIA INVÁLIDA - REINICIE O SCRIPT" #código de erro na mensagem recebida
        enviaResposta(resposta, c)
        return   
    c.settimeout(10)
    #todas as mensagens tem, no mínimo, um código, um cpf para acesso e uma chave deste - total: 19 caracteres (chave de 6 dígitos)
    if len(msgRecebida)<(13+tamChave):
        resposta = "99REQUISIÇÃO INVÁLIDA (A)" #código de erro na mensagem recebida
        enviaResposta(resposta, c)
        return 
    
    codigoStr = msgRecebida[:2]
    cpf = msgRecebida[2:13]  
    chaveContagil = msgRecebida[13:(13+tamChave)]      

    logging.info(codigoStr+" - "+cpf)
    #logging.info(chaveContagil)     
   
    if not codigoStr.isdigit():
        resposta = "99REQUISIÇÃO INVÁLIDA (B)"
        enviaResposta(resposta, c)       
        return  
    try:
        codigo = int(codigoStr)             
    except:
        resposta = "99REQUISIÇÃO INVÁLIDA (B1)"
        enviaResposta(resposta, c)       
        return         
    
    if cpf=="12345678909": #este CPF flag não vem por aqui - ele é esperado dentro de alguns lugares neste procedimento para mandar mais informações da requisição
        resposta = "97CPF INVÁLIDO PARA ESTA REQUISIÇÃO"
        enviaResposta(resposta, c)        
        return
    
    if not validaCPF(cpf):
        resposta = "97CPF INVÁLIDO"
        enviaResposta(resposta, c)        
        return  

        
    if codigo<1 or (codigo>52 and not codigo in [60, 61, 62, 63, 64, 65]): #número de requisições válidas
        resposta = "99CÓD DA REQUISIÇÃO É INVÁLIDO (C)" 
        enviaResposta(resposta, c)          
        return     
             
    if not chaveContagil.isdigit():
        resposta = "90CHAVE DE ACESSO VIA CONTAGIL NÃO É NUMÉRICA"
        enviaResposta(resposta, c)        
        return     
    #dbpath = "C:\\Users\\marco\\Downloads\\"
    #db = "BotTelegramCofisDisaf.accdb"
    #driver = "{Microsoft Access Driver (*.mdb, *.accdb)}"
    #conn = pyodbc.connect("DRIVER={};DBQ={}".format(driver, dbpath+db))  #estabelece conexão com o BD
    conn = conecta()
    if not conn:
        resposta = "97ERRO NA CONEXÃO COM O BANCO DE DADOS" #erro de conexão ou de BD
        enviaResposta(resposta, c) 
        return            
    cursor = conn.cursor(buffered=True)    
    if datetime.now()>ultimaVerificacao+timedelta(hours=1): #buscamos de hora em hora a última data de extração dos dados
        cursor.execute("Select Data from Extracoes Order By Data DESC")
        row = cursor.fetchone() #data de extração dos dados do Ação Fiscal, via DW ou Receita Data
        if row:
            dataExtracao = row[0]
        else:
            dataExtracao = datetime.strptime("01/01/2021", "%d/%m/%Y")
        ultimaVerificacao = datetime.now()


    if codigo==1: #status do usuário 
        comando = "Select Codigo, CPF, Adesao, Saida, d1, d2, d3, email, Chave, ValidadeChave, Tentativas, Orgao, Ativo from Usuarios Where CPF=%s"
        cursor.execute(comando, (cpf,))
        row = cursor.fetchone() 
        if row==None or len(row)==0:
            resposta = "0104" #01 - status; 04 - não consta na base (não foram carregados)
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if row[12]!='S':
            resposta = "0105" #01 - status; 05 - usuário desabilitado
            enviaResposta(resposta, c) 
            conn.close()
            return                        
        ativo = False
        registrado = False
        if row[3]==None and row[2]!=None: #tem um registro ativo
            ativo = True
        if row[8]!=None: #usuário tem chave cadastrada (está registrado)
            registrado = True            
            tentativas = row[10]
            if tentativas==None:
                tentativas = 0
            chaveBD = row[8]
            if tentativas>=3:
                result = "4"
            elif chaveBD==int(chaveContagil):
                if row[9]==None:
                    result = "2"
                else:  
                    logging.info(row[9].date())                        
                    if datetime.today().date()<=row[9].date():
                        result = "1"+row[9].strftime("%d/%m/%Y") #chave dentro da validade
                        zeraTentativas(row[0], conn)
                    else:
                        result = "2" #chave fora da validade
            else: #chave digitada não confere 
                result = "3"
                atualizaTentativas(row[0], tentativas, conn)
        orgao = row[11]
        tipoOrgaoUsuario, nomeOrgao = buscaTipoOrgao(orgao, cursor)        
        if ativo:
            resposta = "0101"+tipoOrgaoUsuario+result #01 - status; 01 - ativo
        elif registrado:
            resposta = "0102"+tipoOrgaoUsuario+result #01 - status; 02 - inativo no Bot (não tem senha ou saiu)
        else:
            resposta - "0103"+tipoOrgaoUsuario #01 - status; 03 - não registrado (nem tem senha)
        enviaResposta(resposta, c) 
        conn.close()
        return    
    
    if not codigo in [18, 19] and ambiente=="PRODUÇÃO": #fazemos o log no ambiente de produção, exceto para envio/pedido de entrada do diário da fiscalização,
                                         #solicitação de status e solicitação de chave pública (acima - cód = 00) e de status (acima - cód = 01)
        comando = "Insert Into Log (IP, Requisicao, Mensagem, Data) Values (%s, %s, %s, %s)"
        try:
            cursor.execute(comando, (c.getpeername()[0], codigo, msgRecebida[2:], datetime.now()))
            conn.commit()
        except:
            logging.info("Falhou o log - IP: "+c.getpeername()[0]+"; Msg: "+msgRecebida)
            conn.rollback()    
    
    #validamos a chave do contágil ligada àquele CPF (registro ativo) - serviços de 2 em diante
    comando = "Select Codigo, Chave, ValidadeChave, Tentativas, email, d1, d2, d3, Orgao, Adesao, Saida, BloqueiaTelegram, Ativo, Cadastrador from Usuarios Where CPF=%s"            
    cursor.execute(comando, (cpf,))
    row = cursor.fetchone()   
    if not row or len(row)==0: #o usuário está inativo
        resposta = "90USUÁRIO NÃO ENCONTRADO NA BASE DE DADOS"
        enviaResposta(resposta, c)  
        conn.close()
        return
    statusBloqueio = row[11]
    orgaoUsuario =  row[8]
    email = row[4]
    ativo = row[12]
    if ativo!='S':
        resposta = "90USUÁRIO DESABILITADO - CONTACTE A COFIS/DISAV"
        enviaResposta(resposta, c)  
        conn.close()
        return 
    cadastrador = row[13]        
    tipoOrgaoUsuario, nomeOrgao = buscaTipoOrgao(orgaoUsuario, cursor) 
    if (row[9]==None or row[10]!=None) and (tipoOrgaoUsuario=="L" or not codigo in [7, 13, 24, 28, 29, 31, 32, 41, 42, 49, 51, 52, 60, 61, 62, 63, 64, 65]): #adesão nula ou inatividade
                                                                                                                      #só permitimos acesso para usuários nacionais e regionais nas 
                                                                                                                      #requisições listadas
        if row[9]==None:
            resposta = "90USUÁRIO NÃO SE REGISTROU NO BOT TELEGRAM OU OPÇÃO NÃO DISPONÍVEL PARA O TIPO DE USUÁRIO"
        else:
            resposta = "90USUÁRIO ESTÁ INATIVO NO BOT TELEGRAM OU OPÇÃO NÃO DISPONÍVEL PARA O TIPO DE USUÁRIO"
        enviaResposta(resposta, c)  
        conn.close()
        return           
    tentativas = row[3]
    if tentativas==None:
        tentativas = 0         
    if row[1]==None:
        resposta = "90CHAVE DE ACESSO VIA CONTÁGIL NÃO FOI GERADA"       
        enviaResposta(resposta, c) 
        conn.close()
        return        
    if row[1]!=int(chaveContagil):
        resposta = "90CHAVE DE ACESSO VIA CONTÁGIL É INVÁLIDA OU INCORRETA"
        atualizaTentativas(row[0], tentativas, conn)        
        enviaResposta(resposta, c)
        conn.close()
        return
    if row[2]==None:
        resposta = "90CHAVE DE ACESSO VIA CONTÁGIL SEM VALIDADE - GERE OUTRA"
        enviaResposta(resposta, c) 
        conn.close()
        return        
    if datetime.today().date()>row[2].date():
        resposta = "90CHAVE DE ACESSO VIA CONTÁGIL ESTÁ EXPIRADA - GERE OUTRA"
        enviaResposta(resposta, c)  
        conn.close()
        return    
    if tentativas>=3:
        resposta = "90CHAVE DE ACESSO VIA CONTÁGIL ESTÁ EXPIRADA - TENTATIVAS EXCEDIDAS - GERE OUTRA"
        enviaResposta(resposta, c) 
        conn.close()
        return               
    
    zeraTentativas(row[0], conn) #como a chave está correta, zera o nº de tentativas
        
    #para todas as funções abaixo, temos que verificar se o cpf está cadastrado e ativo <- JÁ FOI FEITO ACIMA
    #comando = "Select Codigo, CPF, Adesao, Saida, d1, d2, d3, email, Chave, ValidadeChave, Tentativas from Usuarios Where Saida Is Null and Adesao Is Not Null and CPF=%s"        
    #cursor.execute(comando, (cpf,))
    #row = cursor.fetchone()
    #if not row:
    #    resposta = "99CPF NÃO ENCONTRADO OU INATIVO NO SERVIÇO"
    #    enviaResposta(resposta, c)
    #    conn.close()
    #    return 

    #if len(row)==0:
    #    resposta = "99CPF NÃO ENCONTRADO OU INATIVO NO SERVIÇO"
    #    enviaResposta(resposta, c)
    #    conn.close()
    #    return           

    if codigo==22: #troca de senha
        if len(msgRecebida)!=(13+2*tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (22A)"
            enviaResposta(resposta, c)  
            conn.close()
            return 
        novaChave = msgRecebida[-6:]
        if not novaChave.isdigit():
            resposta = "22NCHAVE DEVE SER NUMÉRICA (A)"
            enviaResposta(resposta, c)  
            conn.close()
            return
        if novaChave==chaveContagil:
            resposta = "22NVALIDAÇÃO FEITA PELO SCRIPT FOI BURLADA"
            enviaResposta(resposta, c)  
            conn.close()
            return 
        try:
            chaveNum = int(novaChave)
        except:
            resposta = "22NCHAVE DEVE SER NUMÉRICA (B)"
            enviaResposta(resposta, c)  
            conn.close()
            return
        if chaveNum<100000 or chaveNum>999999:
            resposta = "22NCHAVE DEVE CONTER 6 ALGARISMOS, SENDO O PRIMEIRO DIFERENTE DE ZERO"
            enviaResposta(resposta, c)  
            conn.close()
            return 
        if (novaChave in '12345678909876543210'):  
            resposta = "22NCHAVE NÃO PODE SER UMA SEQUÊNCIA DE ALGARISMOS"
            enviaResposta(resposta, c)  
            conn.close()
            return   
        if (novaChave in ['111111', '222222', '333333', '444444', '555555', '666666', '777777', '888888', '999999']):  
            resposta = "22NCHAVE NÃO PODE SER COMPOSTA DOS MESMOS ALGARISMOS"
            enviaResposta(resposta, c)  
            conn.close()
            return                                
        comando = "Update Usuarios Set Chave=%s, ValidadeChave=%s Where CPF=%s"
        try:
            validade = datetime.now().date()+timedelta(days=30)
            cursor.execute(comando, (chaveNum, validade, cpf))
            resposta = "22STROCA DE SENHA EFETUADA - CHAVE VÁLIDA ATÉ "+validade.strftime("%d/%m/%Y")
            conn.commit()
        except:
            resposta = "22NERRO AO ATUALIZAR TABELA"   
            conn.rollback()       
        enviaResposta(resposta, c)  
        conn.close()
        return 


    #obtemos a chave do fiscal para ser utilizada em todas as requisições  que não sejam a troca de senha, nem status - acima  
    comando = "Select Fiscais.Codigo, Fiscais.Nome From Fiscais Where Fiscais.CPF=%s"          
    cursor.execute(comando, (cpf,))
    rowFiscal = cursor.fetchone()
    if (not rowFiscal or len(rowFiscal)==0) and not codigo in [7, 13, 24, 28, 29, 31, 32, 41, 42, 49, 51, 52, 60, 61, 62, 63, 64, 65]: #estes códigos podem ser utilizados por usuários Cofis (nacionais) ou Difis (regionais)
        resposta = "97CPF NÃO FOI LOCALIZADO NA TABELA DE FISCAIS/SUPERVISORES"
        enviaResposta(resposta, c)  
        conn.close()
        return 
    elif rowFiscal!=None:
        chaveFiscal = rowFiscal[0] #<---
        nomeFiscal = rowFiscal[1]
    else:
        chaveFiscal = 0
        nomeFiscal = ""

    if codigo in [2, 3, 4, 5, 14, 15, 16, 17, 18, 19, 20, 21, 23, 24, 27, 28, 30, 31, 33, 34, 35, 36, 38, 40, 44, 45, 46, 49]: 
    #verificações COMUNS relativas ao TDPF - TDPF existe, em andamento (p/ alguns), cpf está alocado nele ou é supervisor/regional/nacional, dependendo do caso
        if len(msgRecebida)<(29+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (D)"
            enviaResposta(resposta, c)  
            conn.close()
            return 
        pontosTdpf = 0 
        if codigo!=46:                 
            tdpf = msgRecebida[(tamChave+13):(tamChave+29)] #obtemos o TDPF que será utilizado em todas as requisições acima elencadas
        else:
            if msgRecebida[tamChave+13:tamChave+14]=="T": #para solicitação de informações sobre controle postal, há que se ter esse indicador
                tdpf = msgRecebida[(tamChave+14):(tamChave+30)]
            else:
                tdpf = None
        if tdpf!=None: #se for None/Null, indica que é uma solicitação de informações sobre controle postal com base em período e não em TDPF (ver if acima)
            if not tdpf.isdigit():
                resposta = "99REQUISIÇÃO INVÁLIDA - TDPF DEVE SER NUMÉRICO"
                enviaResposta(resposta, c)  
                conn.close()
                return            
            comando = """Select Codigo, Encerramento, Nome, Emissao, Vencimento, Grupo, NI, DCC, Porte, Acompanhamento, TrimestrePrevisto, Tipo, 
                         TDPFPrincipal, FAPE, Pontos, DataPontos from TDPFS Where Numero=%s"""
            cursor.execute(comando, (tdpf,))
            row = cursor.fetchone()
            if row:
                chaveTdpf = row[0] #<-- chave do TDPF que será utilizado nas requisições aqui tratadas
                encerramento = row[1]
                emissao = row[3] #<-- data de emissão fica disponível
                nome = row[2] #<-- nome do fiscalizado tb
                if nome==None or nome=="":
                    nome = "NOME INDISPONÍVEL"
                nome = nome[:tamNome].ljust(tamNome)
                vencimento = row[4]
                tipoProc = row[11] #esta variável será útil no tratamento das requisições
                chaveTdpfPrincipal = row[12]
                fape = row[13]
                pontosTdpf = row[14]
                if pontosTdpf==None:
                    pontosTdpf = 0
                dataPontos = row[15]
                tipoProc = tipoProc if (tipoProc!='D' or chaveTdpfPrincipal==None) else 'V' #V = diligência vinculada              
                if encerramento!=None and not codigo in [4, 15, 19, 21, 23, 24, 27, 28, 30, 31, 34, 46, 49]: #nestes códigos podemos listar ciências, atividades, entradas do diário, informar DCC,
                                                                                    #incluir/listar pontuação de TDPFs encerrados, incluir trimestre de encerramento previsto,
                                                                                    # mostrar fiscais alocados ou informar se é supervisor do tdpf (mesmo encerrado), dados do TDPF tb
                                                                                    #listar prorrogações e listar controle postal 
                                                                                    #solicitar fatores de pontuação (Serpro)
                    msg = "TDPF encerrado"
                    msg = msg.ljust(tamMsg)
                    resposta = codigoStr+(("N"+msg+nome) if (2<=codigo<=5) else msg) #o código 27 estava aqui antes de colocar na lista acima
                    if codigo in [33, 35, 36, 38, 40, 45]:
                        resposta = codigoStr+"E"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return 
                if not tipoProc in ["F", "D", "V"] and codigo in [2, 3, 4, 5, 14, 15, 16, 17, 18, 19, 20, 44, 45, 46]: #estas requisições valem apenas para Fiscalizações e Diligências
                    msg = "TDPF NÃO É DE DILIGÊNCIA OU DE FISCALIZAÇÃO"
                    if codigo==15:
                        resposta = codigoStr+"000"
                    elif codigo==18:
                        resposta = codigoStr+"88"+msg
                    elif codigo==19:
                        resposta = codigoStr+"AAA"+msg 
                    elif codigo==20:
                        resposta = codigoStr+"99"+msg
                    elif codigo in [44, 45, 46]:
                        resposta = codigoStr+"I"
                    else:
                        resposta = codigoStr+(("I"+msg) if 2<=codigo<=5 else msg)   
                    enviaResposta(resposta, c) 
                    conn.close()
                    return                                                             
                if tipoProc!="F" and codigo in [23, 27, 33, 34, 35, 36, 38, 40]: #só permitimos informar parâmetros para pontos, trimestre da meta, prorrogação (em geral) de Fiscalizações
                    resposta = codigoStr+"I" #I - indica tb que o TDPF não é do tipo fiscalização (podendo ser inexistente, conforme abaixo)
                    enviaResposta(resposta, c) 
                    conn.close()
                    return  
                if fape=='S':
                    tipoProc += 'P'
                else:
                    tipoProc += ' '                      
            else: 
                msg = "TDPF NÃO foi localizado ou foi encerrado há muito tempo e não colocado na base deste serviço"
                msg = msg.ljust(tamMsg)          
                if codigo in [30, 31, 33, 35, 36, 38, 40, 44, 45, 46, 49]:
                    resposta = codigoStr+"I"
                else:
                    resposta = codigoStr+(("I"+msg) if (2<=codigo<=5 or codigo in [23, 24, 27, 28]) else msg)
                enviaResposta(resposta, c) 
                conn.close()
                return  
            rowTdpf = row #para utilizarmos na requisição 31            
            comando = "Select Alocacoes.Desalocacao from Alocacoes Where Alocacoes.Fiscal=%s and Alocacoes.TDPF=%s"
            cursor.execute(comando, (chaveFiscal, chaveTdpf))
            row = cursor.fetchone()    
            bSupervisor = False 
            if codigo in [2, 4, 15, 21, 23, 24, 27, 28, 31, 34, 35, 36, 44, 45, 46, 49]: #supervisor pode informar e relacionar ciências (2 e 4), atividades (15) de TDPF, 
                                                    #incluir ou lista pontuação (23, 24), incluir DCC (21), informar trimestre da meta (27) ou listar fiscais alocados (28)
                                                    #ou dados do TDPF (31) e apagar/assinar prorrogação, listar prorrogações,
                                                    #incluir e excluir postagem e listar controle postal
                if tipoOrgaoUsuario=="N" and codigo in [24, 28, 31, 49]: #nestes códigos, usuário Cofis pode fazer consulta (#para os fins destas requisições, é supervisor)
                    bSupervisor = True
                elif tipoOrgaoUsuario=="R" and codigo in [24, 28, 31, 49]: #nestes códigos, usuário Difis pode fazer consulta (#para os fins destas requisições, é supervisor)
                    comando = "Select TDPFS.Numero from TDPFS Where TDPFS.Codigo=%s and TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s)"
                    cursor.execute(comando, (chaveTdpf, orgaoUsuario))
                    rowSuperv = cursor.fetchone()
                    if rowSuperv:
                        bSupervisor = True
                if not bSupervisor:
                    bSupervisor, _ = verificaSupervisao(conn, cpf, tdpf)   
            if codigo==30: #a requisição é uma pergunta se o cpf é do supervisor do tdpf ou está alocado nele    
                bSupervisor, _ = verificaSupervisao(conn, cpf, tdpf)       
                if row!=None: #fiscal esteve ou está alocado do tdpf
                    if row[0]==None: #fiscal está ainda alocado ao tdpf
                        alocado = "S"
                    else:
                        alocado = "N"                    
                else:
                    alocado = "N"
                if bSupervisor:
                    supervisor = "S"
                else:
                    supervisor = "N"           
                resposta = "30"+alocado+supervisor
                enviaResposta(resposta, c)   
                conn.close()
                return  
            if not bSupervisor and codigo==27: #só supervisor pode incluir trimestre da meta 
                msg = "Usuário NÃO é supervisor do TDPF"
                resposta = "27"+msg.ljust(tamMsg)            
                enviaResposta(resposta, c)   
                conn.close()
                return              
            achou = False
            if row and not bSupervisor:
                achou = True
                if row[0]!=None and not codigo in [28, 31]: #fiscal desalocado pode consultar TDPFs em que participou (28 - lista fiscais alocados no TDPF; 31 - dados do TDPF; no script não terá como num primeiro momento)
                    msg = "CPF NÃO está mais alocado ao TDPF ou não é supervisor, em requisições em que isso seria relevante"
                    msg = msg.ljust(tamMsg)                
                    resposta = codigoStr+(("N"+msg+nome) if (2<=codigo<=5 or codigo==23) else msg)
                    enviaResposta(resposta, c)  
                    conn.close()
                    return                         
            if not achou and not bSupervisor:
                msg = "CPF NÃO está alocado ao TDPF ou não é supervisor, em requisições em que isso seria relevante"
                msg = msg.ljust(tamMsg)   
                if codigo in [31, 33, 34, 35, 36, 38, 40, 44, 45, 46, 49]:
                    resposta = codigoStr+"N"
                else:
                    resposta = codigoStr+(("N"+msg+nome) if (2<=codigo<=5 or codigo in [23, 24]) else msg)     
                enviaResposta(resposta, c)   
                conn.close()
                return    

    if codigo==31: #pediu as informações do registro do tdpf na tabela TDPFS - todas as verificações foram feitas acima e o resultado está em rowTdpf
        grupo = rowTdpf[5]
        if grupo==None:
            grupo = ""
        grupo = grupo.ljust(25)
        emissao = dataTexto(rowTdpf[3])
        encerramento = rowTdpf[1]
        encerramento=dataTexto(encerramento)
        nome = rowTdpf[2]
        if nome==None:
            nome = ""
        nome = nome.ljust(150)
        niFiscalizado = rowTdpf[6]
        if niFiscalizado==None:
            niFiscalizado = ""
        niFiscalizado = niFiscalizado.ljust(18)  
        vencimento = dataTexto(rowTdpf[4])
        dcc = rowTdpf[7]
        if dcc==None:
            dcc = ""
        dcc = dcc.ljust(17)
        porte = rowTdpf[8]
        if porte==None:
            porte = ""
        porte = porte.ljust(3)
        acompanhamento = rowTdpf[9]
        if acompanhamento==None:
            acompanhamento = " "
        trimestrePrevisto = rowTdpf[10]
        if trimestrePrevisto==None:
            trimestrePrevisto = "      "
        #procuramos o supervisor atual do grupo do TDPF
        cpfSupervSub = " ".ljust(11)
        nomeSupervSub = " ".ljust(100)        
        comando = "Select Supervisores.Codigo, Fiscais.CPF, Fiscais.Nome from Fiscais, Supervisores Where Supervisores.Fiscal=Fiscais.Codigo and Supervisores.Fim Is Null and Supervisores.Equipe=%s and Titular Is Null"
        cursor.execute(comando, (grupo.strip(),))
        row = cursor.fetchone()
        if not row:
            cpfSuperv = " ".ljust(11)
            nomeSuperv = " ".ljust(100)
        else:
            cpfSuperv = row[1]
            nomeSuperv = row[2][:100].ljust(100)
            supervCod = row[0] 
            #procura o supervisor substituto (se houver)           
            comando = "Select Fiscais.CPF, Fiscais.Nome from Fiscais, Supervisores Where Supervisores.Fiscal=Fiscais.Codigo and Supervisores.Fim Is Null and Supervisores.Titular=%s"
            cursor.execute(comando, (supervCod, ))
            rowSubst = cursor.fetchone()
            if rowSubst:
                cpfSupervSub = rowSubst[0]
                nomeSupervSub = rowSubst[1][:100].ljust(100) 
        tdpfPrincipal = "0".rjust(16,"0")                 
        if chaveTdpfPrincipal!=None:  
            cursor.execute("Select TDPFS.Numero From TDPFS Where Codigo=%s", (chaveTdpfPrincipal,))
            linha = cursor.fetchone()
            if linha:
                tdpfPrincipal = linha[0]
        resposta = "31S"+ grupo+emissao+encerramento+nome+niFiscalizado+vencimento+dcc+porte+acompanhamento+trimestrePrevisto+cpfSuperv+nomeSuperv+cpfSupervSub+nomeSupervSub+tipoProc+tdpfPrincipal
        enviaResposta(resposta, c, True, chaveCriptoAES)   
        conn.close()
        return
           
    if codigo==2: #informa data de ciência relativa a TDPF 
        try:   #deve enviar imediatamente a descrição do documento que efetivou a ciência (sem criptografia)
            mensagemRec = c.recv(512) #.decode('utf-8') #chegou a requisicao
        except:
            c.close()
            logging.info("Erro de time out 2 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
            conn.close()
            return         
        if len(msgRecebida)!=(49+tamChave): #inclui o tdpf, a data da intimação e a data de seu vencimento
            resposta = "99REQUISIÇÃO INVÁLIDA (2A)"
            enviaResposta(resposta, c) 
            conn.close()
            return      
        
        data = msgRecebida[-20:-10] 
        if not isDate(data):
            msg = "Data de ciência inválida"
            msg = msg.ljust(tamMsg)             
            resposta = "02N"+msg+nome 
            enviaResposta(resposta, c) 
            conn.close()
            return
        try:
            dataObj = datetime.strptime(data, '%d/%m/%Y')
        except:
            msg = "Data de ciência inválida (2)"
            msg = msg.ljust(tamMsg)             
            resposta = "02N"+msg+nome 
            enviaResposta(resposta, c) 
            conn.close()
            return            
        if datetime.today().date()<dataObj.date(): #data não pode ser futura
            msg = "Data de ciência não pode ser futura"
            msg = msg.ljust(tamMsg)             
            resposta = "02N"+msg+nome 
            enviaResposta(resposta, c)  
            conn.close()
            return
        if dataObj.date()<emissao.date(): #ciência não pode ser inferior à data de emissão (obtida nas verificacões gerais)
            msg = "Data de ciência não pode ser inferior à de emissão ("+emissao.strftime("%d/%m/%Y")+")"
            msg = msg.ljust(tamMsg)             
            resposta = "02N"+msg+nome 
            enviaResposta(resposta, c)  
            conn.close()
            return   
        # estou permitindo informar data de ciência anterior à última por questões de registro - manter a crítica depois de um certo tempo                     
        #comando = "Select Data from Ciencias Where TDPF=%s and Data>=%s Order by Data DESC"
        #cursor.execute(comando, (chaveTdpf, dataObj.date()))
        #row = cursor.fetchone()
        #if row:
        #    msg = "Data de ciência informada DEVE ser posterior à ultima informada para o TDPF ("+row[0].strftime('%d/%m/%Y')+")"
        #    msg = msg.ljust(tamMsg)             
        #    resposta = "02N"+msg+nome
        #    enviaResposta(resposta, c)  
        #    conn.close()
        #    return  

        vencimento = msgRecebida[-10:] 
        if vencimento=="00/00/0000":
            vencimentoObj = None
        else:
            if not isDate(vencimento):
                msg = "Data de vencimento da intimação inválida"
                msg = msg.ljust(tamMsg)             
                resposta = "02N"+msg+nome 
                enviaResposta(resposta, c) 
                conn.close()
                return
            try:
                vencimentoObj = datetime.strptime(vencimento, '%d/%m/%Y').date()
            except:
                msg = "Data de vencimento da intimação inválida (2)"
                msg = msg.ljust(tamMsg)             
                resposta = "02N"+msg+nome 
                enviaResposta(resposta, c) 
                conn.close()
                return            
            if vencimentoObj<=dataObj.date(): #vencimento da intimação não pode ser inferior à data de ciência
                msg = "Data de vencimento da intimação deve ser posterior à de ciência."
                msg = msg.ljust(tamMsg)             
                resposta = "02N"+msg+nome 
                enviaResposta(resposta, c)  
                conn.close()
                return 

        requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c) 
        if requisicao=="000000000A": #não foi possível descriptografar
            resposta = "99REQUISIÇÃO INVÁLIDA - NÃO FOI POSSÍVEL DESCRIPTOGRAFAR O DOCUMENTO (2A1)"
            enviaResposta(resposta, c) 
            conn.close()
            return            
        codReq = ""
        if len(requisicao)>=2:
            codReq = requisicao[:2]
        if len(requisicao)!=72 or codReq!="02":
            msg = "Documento não foi informado ou requisição inválida."
            msg = msg.ljust(tamMsg)             
            resposta = "02N"+msg+nome
            enviaResposta(resposta, c)  
            conn.close()
            return
        documento = requisicao[2:].strip().upper()    
        if len(documento)<3:
            msg = "Documento não foi informado ou tem menos de 3 caracteres."
            msg = msg.ljust(tamMsg)             
            resposta = "02N"+msg+nome
            enviaResposta(resposta, c)  
            conn.close()
            return   
        if len(documento)>70: #isso aqui é só uma garantia extrema
            msg = "Documento tem mais de 70 caracteres."
            msg = msg.ljust(tamMsg)             
            resposta = "02N"+msg+nome
            enviaResposta(resposta, c)  
            conn.close()
            return                                          
        tdpfMonitorado, monitoramentoAtivo, chave = tdpfMonitoradoCPF(conn, tdpf, cpf)
        try:
            comando = "Insert into Ciencias (TDPF, Data, Documento, Vencimento) Values (%s, %s, %s, %s)"
            cursor.execute(comando, (chaveTdpf, dataObj.date(), documento, vencimentoObj))
            msg = "Ciência registrada para o TDPF"
            if tdpfMonitorado and monitoramentoAtivo==False: #monitoramento do tdpf estava desativado - ativa
                msg = "Monitoramento deste TDPF foi reativado e a ciência foi registrada."
                comando = "Update CadastroTDPFs Set Fim=Null Where Codigo=%s"
                cursor.execute(comando, (chave,))                         
            elif not tdpfMonitorado: #tdpf não estava sendo monitorado - inclui ele
                comando = "Insert into CadastroTDPFs (Fiscal, TDPF, Inicio) Values (%s, %s, %s)"
                cursor.execute(comando, (chaveFiscal, chaveTdpf, datetime.today().date()))
            conn.commit()
            msg = msg.ljust(tamMsg) 
            resposta = "02S"+msg+nome          
        except:
            conn.rollback()
            msg = "Erro ao atualizar as tabelas."  
            msg = msg.ljust(tamMsg)             
            resposta = "02N"+msg+nome
        enviaResposta(resposta, c)  
        conn.close()
        return            
        
    if codigo==3: #anula ciência relativa ao TDPF
        comando = "Select Codigo, TDPF, Data from Ciencias Where TDPF=%s Order by Data DESC"
        cursor.execute(comando, (chaveTdpf,))
        rows = cursor.fetchall()
        if rows==None or len(rows)==0:
            msg  = "Não há data de ciência informada para o TDPF" #Não havia data de ciência para o TDPF
            msg = msg.ljust(tamMsg)             
            resposta = "03N"+msg+nome
            enviaResposta(resposta, c, True, chaveCriptoAES)     
            conn.close()
            return   
        #print(rows)         
        if len(rows)==1:
            dataAnt = "Nenhuma Data Vigente" #não haverá data anterior
        else: #2 ou mais linhas
            dataAnt = "Data agora em vigor: "+rows[1][2].strftime('%d/%m/%Y')        
        chave = rows[0][0]    
        try:
            comando = "Delete from Ciencias Where Codigo=%s"
            cursor.execute(comando, (chave,))
            conn.commit() 
            msg = dataAnt.ljust(tamMsg)            
            resposta = "03S"+msg+nome
            enviaResposta(resposta, c, True, chaveCriptoAES)  
            conn.close()
            return            
        except:
            conn.rollback()
            msg = "Erro na atualização das tabelas. Tente novamente mais tarde."        
            msg = msg.ljust(tamMsg)             
            resposta = "03N"+msg+nome
            enviaResposta(resposta, c, True, chaveCriptoAES)      
            conn.close()
            return            
        
    if codigo==4: #relaciona ciências de um tdpf
        comando = "Select Data, Documento, Vencimento from Ciencias Where TDPF=%s Order by Data"
        cursor.execute(comando, (chaveTdpf,))
        rows = cursor.fetchall()
        if len(rows)==0:
            msg = "Não há nenhuma ciência registrada para o TDPF."
            msg = msg.ljust(tamMsg)
            resposta = "04N"+msg+nome
            enviaResposta(resposta, c, True, chaveCriptoAES)
            conn.close()
            return            
        nn = len(rows)
        if nn<10:
            nn = "0"+str(nn)
        elif nn>99:
            nn = "99"
        else:
            nn = str(nn)
        datas = ""
        i = 0
        for row in rows:
            documento = row[1]
            if not documento:
                documento = ""
            documento = documento.ljust(70)
            vencimento = dataTexto(row[2])
            datas = datas + row[0].strftime('%d/%m/%Y')+documento+vencimento
            i+=1
            if i>=30: #limite de 30 datas/documentos
                break
        resposta = "04S"+nn+nome+tipoProc+datas
        enviaResposta(resposta, c, True, chaveCriptoAES)  
        conn.close()
        return        
    
    if codigo==5: #finaliza alertas de um tdpf
        dataAtual = datetime.today().date()
        comando = "Update CadastroTDPFs Set Fim=%s Where TDPF=%s and Fiscal=%s"
        try:
            cursor.execute(comando, (dataAtual, chaveTdpf, chaveFiscal))
            conn.commit()
            msg = "TDPF não será mais objeto de alerta"
            msg = msg.ljust(tamMsg)
            resposta = "05S"+msg+nome
        except:
            conn.rollback()
            msg = "Erro na atualização das tabelas. Tente novamente mais tarde."        
            msg = msg.ljust(tamMsg)             
            resposta = "05N"+msg+nome            
        enviaResposta(resposta, c, True, chaveCriptoAES) 
        conn.close()
        return 
        
    if codigo==6: #mostra lista de tdpfs ativos e últimas ciências 
        if len(msgRecebida)>(13+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (6A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        comando = """Select TDPFS.Codigo, TDPFS.Numero, TDPFS.Nome, TDPFS.Vencimento, TDPFS.Emissao, Tipo, TDPFS.TDPFPrincipal, FAPE
                     from CadastroTDPFs, Alocacoes, TDPFS 
                     Where CadastroTDPFs.Fiscal=%s and CadastroTDPFs.Fim Is Null and CadastroTDPFs.Fiscal=Alocacoes.Fiscal
                     and CadastroTDPFs.TDPF=Alocacoes.TDPF and CadastroTDPFs.TDPF=TDPFS.Codigo and Alocacoes.Desalocacao Is Null 
                     and TDPFS.Encerramento Is Null"""       
        cursor.execute(comando, (chaveFiscal,))
        rows = cursor.fetchall()
        if rows==None or len(rows)==0:
            tam = 0
        else:    
            tam = len(rows)
        if tam==0:
            resposta = "0600"
            enviaResposta(resposta, c)  
            conn.close()
            return             
        if tam>=100: #limite de 99 tdpfs
            nn = "99"
            tam = 99
        else:
            nn = str(tam).rjust(2,"0")
        registro = "" 
        resposta = "06"+nn
        i = 0
        total = 0
        for row in rows:
            chaveTdpf = row[0]
            tdpf = row[1]
            nome = row[2]
            if nome==None:
                nome = ""            
            nome = nome[:tamNome].ljust(tamNome)
            vencimento = dataTexto(row[3])
            emissao = dataTexto(row[4]) 
            tipoProc = row[5]  
            tdpfPrincipal = row[6]
            fape = row[7]
            tipoProc = tipoProc if (tipoProc!='D' or tdpfPrincipal==None) else 'V' #V = diligência vinculada    
            if fape=='S':
                tipoProc += 'P'
            else:
                tipoProc += ' '               
            comando = "Select Data, Documento, Vencimento from Ciencias Where TDPF=%s order by Data DESC"
            cursor.execute(comando, (chaveTdpf,))
            cienciaReg = cursor.fetchone() #busca a data de ciência mais recente (DESC acima)
            documento = ""
            documento = documento.ljust(70)             
            if cienciaReg: 
                if cienciaReg[1]!=None:
                    documento = cienciaReg[1].ljust(70)               
                ciencia = cienciaReg[0] #obtem a data de ciência mais recente
                if ciencia!=None:
                    cienciaStr = ciencia.strftime('%d/%m/%Y')                    
                    registro = registro + tdpf + tipoProc + nome + emissao + vencimento + cienciaStr + documento + dataTexto(cienciaReg[2])  
                else:
                    registro = registro + tdpf + tipoProc + nome + emissao + vencimento + "00/00/0000" + documento + "00/00/0000"        
            else:
                registro = registro + tdpf + tipoProc + nome + emissao + vencimento + "00/00/0000" + documento + "00/00/0000"
            i+=1
            total+=1
            if i==5 or total==tam: #de cinco em cinco ou no último registro, enviamos
                enviaRespostaSemFechar(resposta+registro, c, True, chaveCriptoAES)
                resposta = "06"
                registro = ""
                i = 0
                if total==tam:
                    c.close()
                    break #percorreu os registros ou 99 deles, que é o limite
                if total<tam: #ainda não chegou ao final - aguardamos a requisição da continuação
                    try:
                        mensagemRec = c.recv(1024) #.decode('utf-8') #chegou a requisicao
                        requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                        if requisicao!="0612345678909":
                            resposta = "99REQUISIÇÃO INVÁLIDA (6B)"
                            enviaResposta(resposta, c)    
                            conn.close()
                            return
                    except:
                        c.close()
                        logging.info("Erro de time out 6 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                        conn.close()
                        return
                    
    if codigo==7: #requer e-mail cadastrado
        if len(msgRecebida)>(13+tamChave): #despreza
            resposta = "99REQUISIÇÃO INVÁLIDA (7)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        #a variável row foi buscada na pequena rotina antes do código validador previo de 2 a 5
        #email = row[4]
        if email==None or email=="":
            resposta = "07N"
            enviaResposta(resposta, c)  
            conn.close()
            return
        email = email[:50].ljust(50)
        resposta = "07S"+email
        enviaResposta(resposta, c)    
        conn.close()
        return

    if codigo==8: #cadastra ou substitui e-mail        
        if len(msgRecebida)!=(63+tamChave): 
            resposta = "99REQUISIÇÃO INVÁLIDA (8)"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        email = msgRecebida[(13+tamChave):].strip()
        if "@" in email:
            if not verificaEMail(email):
                msg = "Email inválido"
                msg = msg.ljust(100)            
                resposta = "08N"+msg
                enviaResposta(resposta, c)
                conn.close()
                return
            elif email[-11:].upper()!="@RFB.GOV.BR":
                msg = "Email não é institucional"
                msg = msg.ljust(100)            
                resposta = "08N"+msg
                enviaResposta(resposta, c)
                conn.close()
                return                       
        else:
            email = email + "@rfb.gov.br"
        if len(email)<15:
            msg = "Email inválido (nome de usuário curto)"
            msg = msg.ljust(100)            
            resposta = "08N"+msg
            enviaResposta(resposta, c)
            conn.close()
            return                
        comando = "Update Usuarios Set email=%s Where CPF=%s and Saida Is Null"
        try:
            cursor.execute(comando, (email, cpf))
            conn.commit()
            msg = "Email atualizado"
            msg = msg.ljust(100)
            resposta = "08S"+msg
        except:
            msg = "Erro na atualização de tabelas (2)"
            msg = msg.ljust(100)            
            resposta = "08N"+msg
        enviaResposta(resposta, c)
        conn.close()
        return
     
    if codigo==9: #apaga e-mail
        if len(msgRecebida)>(13+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (9)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if statusBloqueio=='S': #comunicação via Telegram está bloqueada, não podemos permitir a exclusão do e-mail (statusBloqueio foi obtido acima)
            msg = "Email não pode ser excluído porque a comunicação via Telegram está bloqueada".ljust(100)
            resposta = "09N"+msg
        else:
            comando = "Update Usuarios Set email=Null Where CPF=%s and Saida Is Null"
            cursor.execute(comando, (cpf,))
            try:
                conn.commit()
                msg = "Email excluido"
                msg = msg.ljust(100)
                resposta = "09S"+msg
            except:
                conn.rollback()
                msg = "Erro na atualização da tabela"
                msg = msg.ljust(100)            
                resposta = "09N"+msg
        enviaResposta(resposta, c)
        conn.close()
        return
    
    if codigo==10: #solicita prazos vigentes
        if len(msgRecebida)!=(13+tamChave): #despreza
            resposta = "99REQUISIÇÃO INVÁLIDA (10)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        #a variável row foi buscada na pequena rotina antes do código validador previo de 2 a 5
        d1 = str(row[5]).rjust(2,"0")
        d2 = str(row[6]).rjust(2,"0")
        d3 = str(row[7]).rjust(2,"0")
        resposta = "10"+d1+d2+d3
        enviaResposta(resposta, c)
        conn.close()
        return
    
    if codigo==11: #altera prazos
        if len(msgRecebida)!=(19+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (11)"
            enviaResposta(resposta, c) 
            conn.close()
            return                    
        d1 = msgRecebida[-6:][:2]
        d2 = msgRecebida[-4:][:2]
        d3 = msgRecebida[-2:]
        logging.info(d1)
        logging.info(d2)
        logging.info(d3)
        logging.info("---------")
        if not d1.isdigit() or not d2.isdigit() or not d3.isdigit():
            msg = "Dias devem ser numéricos"
            msg = msg.ljust(100)
            resposta = "11N"+msg
            enviaResposta(resposta, c)  
            conn.close()
            return

        d = [int(d1), int(d2), int(d3)]
        erro = False
        for dia in d:
            if dia<1 or dia>50:
                erro = True
                break
        if erro:
            msg = "Dias devem estar na faixa de 1 a 50"
            msg = msg.ljust(100)
            resposta = "11N"+msg
            enviaResposta(resposta, c)  
            conn.close()
            return

        d.sort(reverse = True)
        d1 = d[0]
        d2 = d[1]
        d3 = d[2]
        comando = "Update Usuarios set d1=%s, d2=%s, d3=%s where CPF=%s and Saida Is Null"
        try:
            cursor.execute(comando, (d1, d2, d3, cpf))
            conn.commit()
            msg = "Alteração efetuada"
            msg = msg.ljust(100)
            resposta = "11S"+msg        
        except:
            conn.rollback()
            msg = "Erro na atualização da tabela"
            msg = msg.ljust(100)
            resposta = "11N"+msg
        enviaResposta(resposta, c)     
        conn.close()
        return            

    if codigo==12: #Relação de TDPFs alocados ao CPF - indica se está sendo monitorado (se estiver em andamento) e se o cpf é supervisor
        if len(msgRecebida)>(15+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (12A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        encerrados = msgRecebida[-2:-1]        
        if encerrados=="N":
            comando = """Select TDPFS.Numero, Alocacoes.Supervisor, TDPFS.Nome, TDPFS.Codigo, TDPFS.DCC, TDPFS.Encerramento, TDPFS.NI, TDPFS.Tipo, TDPFS.TDPFPrincipal, FAPE from Alocacoes, TDPFS 
                         Where Alocacoes.Fiscal=%s and Alocacoes.Desalocacao Is Null and Alocacoes.TDPF=TDPFS.Codigo and 
                         TDPFS.Encerramento Is Null"""
            orderBy = " Order by TDPFS.Numero"
        elif encerrados=="S":
            comando = """Select TDPFS.Numero, Alocacoes.Supervisor, TDPFS.Nome, TDPFS.Codigo, TDPFS.DCC, TDPFS.Encerramento, TDPFS.NI, TDPFS.Tipo, TDPFS.TDPFPrincipal, FAPE from Alocacoes, TDPFS 
                         Where Alocacoes.Fiscal=%s and Alocacoes.Desalocacao Is Null and Alocacoes.TDPF=TDPFS.Codigo and
                         TDPFS.Encerramento Is Not Null""" 
            orderBy = " Order by TDPFS.Encerramento DESC, TDPFS.Numero ASC"
        else:
            resposta = "99INDICADOR DE ENCERRAMENTO INVÁLIDO (12B)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        tipoTdpfs = msgRecebida[-1:]
        if not tipoTdpfs in ['D', 'F', 'I', 'L', 'R', 'T']:
            resposta = "99INDICADOR DE TIPO INVÁLIDO (12C)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        if tipoTdpfs!='T':
            comando += " and TDPFS.Tipo=%s"
            comando += orderBy                         
            cursor.execute(comando, (chaveFiscal, tipoTdpfs))
        else:
            comando += orderBy
            cursor.execute(comando, (chaveFiscal, ))
        rows = cursor.fetchall()
        tam = len(rows)
        if tam==0:
            resposta = "1200"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if tam>=100: #limite de 99 tdpfs
            nn = "99"
            tam = 99
        else:
            nn = str(tam).rjust(2,"0")         
        registro = "" 
        resposta = "12"+nn
        i = 0
        total = 0            
        for row in rows:
            tdpf = row[0]
            nome = row[2]
            supervisor = row[1]
            chaveTdpf = row[3]
            if nome==None or nome=="":
                nome = "ND"
            if supervisor==None or supervisor=="":
                supervisor = "N"
            dcc = row[4]
            if dcc==None:
                dcc = ""
            dcc = dcc.ljust(17)
            encerramento = dataTexto(row[5])
            niFiscalizado = row[6]
            if niFiscalizado==None:
                niFiscalizado = ""
            niFiscalizado = niFiscalizado.ljust(18)  
            tipoProc = row[7]
            tdpfPrincipal = row[8]
            tipoProc = tipoProc if (tipoProc!='D' or tdpfPrincipal==None) else 'V' #V = diligência vinculada
            fape = row[9]
            if fape=='S':
                tipoProc += 'P'
            else:
                tipoProc += ' '             
            nome = nome[:tamNome].ljust(tamNome)  
            registro = registro + tdpf + tipoProc + nome + niFiscalizado
            if encerrados=="N":             
                comando = "Select Inicio, Fim from CadastroTDPFs Where Fiscal=%s and TDPF=%s"
                cursor.execute(comando, (chaveFiscal, chaveTdpf))
                linha = cursor.fetchone()
                if linha:
                    if linha[1]==None:
                        registro = registro+"S"
                    else:
                        registro = registro+"N"
                else:
                    registro = registro+"N"
            registro = registro + supervisor + dcc + ("" if encerrados=="N" else encerramento)
            i+=1
            total+=1
            if i==5 or total==tam: #de cinco em cinco ou no último registro, enviamos
                enviaRespostaSemFechar(resposta+registro, c, True, chaveCriptoAES)
                resposta = "12"
                registro = ""
                i = 0
                if total==tam:
                    c.close()
                    return #percorreu os registros ou 99 deles, que é o limite
                if total<tam: #ainda não chegou ao final - aguardamos a requisição da continuação
                    try:
                        mensagemRec = c.recv(1024) #.decode('utf-8') #chegou a requisicao
                        requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                        if requisicao!="1212345678909":
                            resposta = "99REQUISIÇÃO INVÁLIDA (12C)"
                            enviaResposta(resposta, c) 
                            conn.close()
                            return
                    except:
                        c.close()
                        conn.close()
                        logging.info("Erro de time out 12 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                        return 

    if codigo==13: #mostra lista de tdpfs ativos e últimas ciências sob supervisão do CPF (tipo Orgao R ou N contam como supervisores) - semelhante ao código 6
        qtdeRegistros = 300 #qtde de registros de tdpfs que enviamos por vez - se alterar aqui, tem que alterar no script e vice-versa
        if len(msgRecebida)!=(19+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (13A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        encerrados = msgRecebida[-6:-5]
        if not encerrados in ["S", "N"]:
            resposta = "99INDICADOR DE ENCERRAMENTO INVÁLIDO (13B)"
            enviaResposta(resposta, c) 
            conn.close()
            return                   
        regInicial = msgRecebida[-5:]
        if not regInicial.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA (13C)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        try:
            regInicial = int(regInicial)            
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA (13D)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        c.settimeout(20)            
        if regInicial>0: #se foi informado o registro, devemos buscar a partir dele, limitado a CINCO
            offsetReg = "Limit "+str(qtdeRegistros)+" Offset "+str(regInicial-1)             
        else: #caso contrário, buscamos todos para informar a quantidade total que existe, mas só enviamos 5 (conforme if ao final do for abaixo)
             offsetReg = "Limit "+str(qtdeRegistros)+" Offset 0"
        logging.info("Offset: "+offsetReg) 
        comando = """Select TDPFS.Numero, TDPFS.Nome, TDPFS.Vencimento, TDPFS.Emissao, TDPFS.Codigo, TDPFS.DCC, TDPFS.Encerramento, TDPFS.Porte, 
                     TDPFS.Acompanhamento, TDPFS.TrimestrePrevisto, TDPFS.Grupo, TDPFS.Tipo, TDPFS.TDPFPrincipal, FAPE, Pontos """
        if tipoOrgaoUsuario=="L": #esta variável e orgaoUsuario vem de rotina comum de validação do usuário
            if encerrados=="N": 
                comando = comando + """from TDPFS, Supervisores 
                                       Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and TDPFS.Encerramento Is Null 
                                       Order by TDPFS.Numero """+offsetReg
                if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                    consulta = """Select Count(TDPFS.Numero)
                                  from TDPFS, Supervisores 
                                  Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and TDPFS.Encerramento Is Null"""
            else:
                comando = comando + """from TDPFS, Supervisores 
                                       Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and TDPFS.Encerramento Is Not Null 
                                       Order by TDPFS.Encerramento DESC, TDPFS.Numero ASC """+offsetReg            
                if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                    consulta = """Select Count(TDPFS.Numero)
                                  from TDPFS, Supervisores 
                                  Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and TDPFS.Encerramento Is Not Null"""
        elif tipoOrgaoUsuario=="R":
            if encerrados=="N": 
                if chaveFiscal==None or chaveFiscal==0: 
                    comando = comando + """from TDPFS
                                           Where TDPFS.Encerramento Is Null and TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s)
                                           Order by TDPFS.Numero """+offsetReg
                else:
                    comando = comando + """from TDPFS
                                           Where TDPFS.Encerramento Is Null and 
                                           (TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) or 
                                           TDPFS.Grupo in (Select Equipe from Supervisores Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null))
                                           Order by TDPFS.Numero """+offsetReg

                if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                    if chaveFiscal==None or chaveFiscal==0: 
                        consulta = """Select Count(TDPFS.Numero)
                                      from TDPFS
                                      Where TDPFS.Encerramento Is Null and TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) """
                    else:
                        consulta = """Select Count(TDPFS.Numero)
                                      from TDPFS
                                      Where TDPFS.Encerramento Is Null and (TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) or 
                                      TDPFS.Grupo in (Select Equipe from Supervisores Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null)) """                        
            else: #encerrados
                if chaveFiscal==None or chaveFiscal==0: 
                    comando = comando + """from TDPFS
                                           Where TDPFS.Encerramento Is Not Null and TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s)
                                           Order by TDPFS.Encerramento DESC, TDPFS.Numero """+offsetReg
                else:
                    comando = comando + """from TDPFS
                                           Where TDPFS.Encerramento Is Not Null and 
                                           (TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) or 
                                           TDPFS.Grupo in (Select Equipe from Supervisores Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null))
                                           Order by TDPFS.Encerramento DESC, TDPFS.Numero """+offsetReg

                if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                    if chaveFiscal==None or chaveFiscal==0: 
                        consulta = """Select Count(TDPFS.Numero)
                                      from TDPFS
                                      Where TDPFS.Encerramento Is Not Null and TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) """
                    else:
                        consulta = """Select Count(TDPFS.Numero)
                                      from TDPFS
                                      Where TDPFS.Encerramento Is Not Null and (TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) or 
                                      TDPFS.Grupo in (Select Equipe from Supervisores Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null)) """             
        elif tipoOrgaoUsuario=="N":
            if encerrados=="N": 
                comando = comando + """from TDPFS
                                       Where TDPFS.Encerramento Is Null  
                                       Order by TDPFS.Numero """+offsetReg
                if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                    consulta = """Select Count(TDPFS.Numero)
                                  from TDPFS
                                  Where TDPFS.Encerramento Is Null """
            else:
                comando = comando + """from TDPFS
                                       Where TDPFS.Encerramento Is Not Null  
                                       Order by TDPFS.Encerramento DESC, TDPFS.Numero ASC """+offsetReg            
                if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                    consulta = """Select Count(TDPFS.Numero)
                                from TDPFS
                                Where TDPFS.Encerramento Is Not Null """             

        if regInicial==0:
            if tipoOrgaoUsuario=="L": 
                cursor.execute(consulta, (chaveFiscal,))
            elif tipoOrgaoUsuario=="R":
                if chaveFiscal==None or chaveFiscal==0:
                    cursor.execute(consulta, (orgaoUsuario,))
                else:
                    cursor.execute(consulta, (orgaoUsuario, chaveFiscal))
            else:
                cursor.execute(consulta)
            totalReg = cursor.fetchone()
            #print(str(totalReg[0]))
            if totalReg:
                tam = totalReg[0]
            else:
                tam = 0
            if tam==0:
                resposta = "1300000" #13+qtde TDPFs
                enviaResposta(resposta, c) 
                conn.close()
                return             
            if tam>=100000: #limite de  tdpfs
                nnnnn = "99999"
                tam = 99999
            else:
                nnnnn = str(tam).rjust(5, "0")
            #resposta = "13"+nnnn #código da resposta e qtde de TDPFs que serão retornados
        #else:
        #    resposta = "13"
        if tipoOrgaoUsuario=="L":
            cursor.execute(comando, (chaveFiscal,))
        elif tipoOrgaoUsuario=="R":
            if chaveFiscal==None or chaveFiscal==0:
                cursor.execute(comando, (orgaoUsuario,))  
            else:
                cursor.execute(comando, (orgaoUsuario, chaveFiscal))  
        else:
            cursor.execute(comando)                      
        rows = cursor.fetchall()   
        if regInicial>0:
            tam = len(rows)     
        registro = "" 
        i = 0
        total = 0         
        for row in rows:
            tdpf = row[0]          
            nome = row[1]
            if nome==None:
                nome = ""   
            nome = nome[:tamNome].ljust(tamNome)  
            vencimento = dataTexto(row[2])                      
            emissao = dataTexto(row[3])  
            chaveTdpf = row[4]  
            tipoProc = row[11]
            tdpfPrincipal = row[12]
            tipoProc = tipoProc if (tipoProc!='D' or tdpfPrincipal==None) else 'V' #V = diligência vinculada
            fape = row[13]  
            if fape=='S':
                tipoProc += 'P'
            else:
                tipoProc += ' '  
            pontos = row[14]
            if pontos==None:
                pontos = 0
            pontos = str(int(pontos*100)).rjust(8, "0")                        
            dcc = row[5]
            if dcc==None:
                dcc = ""
            dcc = dcc.ljust(17) 
            encerramento = dataTexto(row[6])
            porte = row[7]
            acompanhamento = row[8]
            trimestre = row[9]
            if trimestre==None:
                trimestre = " ".ljust(6)
            elif len(trimestre)!=6:
                trimestre = " ".ljust(6)
            if porte==None or porte=="":
                porte = "ND "
            if acompanhamento==None or acompanhamento=="":
                acompanhamento = "N"
            equipe = row[10]
            if equipe==None:
                equipe =""
            equipe = equipe.ljust(25)
            #busca a primeira ciência e a última
            comando = "Select Data, Documento from Ciencias Where TDPF=%s order by Data"
            cursor.execute(comando, (chaveTdpf,))
            cienciaRegs = cursor.fetchall() #busca a data de ciência mais recente (DESC acima)
            if cienciaRegs: 
                primCiencia = dataTexto(cienciaRegs[0][0]) #obtem a primeira data de ciência
                ultCiencia =  dataTexto(cienciaRegs[len(cienciaRegs)-1][0]) #obtém a última data de ciência (mais recente)
                documento = cienciaRegs[len(cienciaRegs)-1][1] #documento da última data de ciência
                if documento==None:
                    documento = ""
                documento = documento.ljust(70)                                   
                registro = registro + tdpf + tipoProc + nome + emissao + vencimento + dcc + primCiencia + ultCiencia + documento 
            else:
                registro = registro + tdpf + tipoProc + nome + emissao + vencimento + dcc + "00/00/0000" + "00/00/0000" + " ".ljust(70) #provavelmente nenhum fiscal iniciou monitoramento
            #verifica se o TDPF está sendo monitorado 
            if encerrados=="N":
                comando = "Select * from CadastroTDPFs Where TDPF=%s and Fim Is Null"   
                cursor.execute(comando, (chaveTdpf,))
                linha = cursor.fetchone()
                if linha:
                    registro = registro+"S"
                else:
                    registro = registro+"N"
            #busca o fiscal alocado há mais tempo no TDPF  
            comando = """Select Fiscais.Nome, Alocacoes.Alocacao 
                         from Fiscais, Alocacoes 
                         Where Alocacoes.Desalocacao Is Null and Alocacoes.Fiscal=Fiscais.Codigo and Alocacoes.TDPF=%s Order by Alocacoes.Alocacao"""
            cursor.execute(comando, (chaveTdpf,))
            linha = cursor.fetchone()
            if linha:
                nomeFiscal = linha[0]
                if nomeFiscal==None:
                    nomeFiscal = "NÃO ENCONTRADO NO REGISTRO"
            else:
                nomeFiscal = "NÃO ENCONTRADO"
            nomeFiscal = nomeFiscal[:100].ljust(100)   
            registro = registro+nomeFiscal+pontos+porte+acompanhamento+trimestre+equipe
            registro = registro + ("" if encerrados=="N" else encerramento)
            #logging.info(registro)
            total+=1
            i+=1
            if i%qtdeRegistros==0 or total==tam: #de qtdeRegistros em qtdeRegistros ou no último registro enviamos a mensagem
                if regInicial==0:
                    #tamanhoReg = str(len(nnnnn+registro)).rjust(5, "0")
                    resposta = "13"+nnnnn #tamanhoReg+nnnnn                  
                else:
                    #tamanhoReg = str(len(registro)).rjust(5, "0")
                    resposta = "13"#+tamanhoReg 
                #enviaResposta(resposta+registro, c, True, chaveCriptoAES)
                enviaResposta(resposta+registro, c, True, chaveCriptoAES, True) #criptografa e comprime (ambos os Trues)
                return 
                
    if codigo==14: #inclui atividade em um tdpf
        if len(msgRecebida)!=(113+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (14A)"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        atividade = msgRecebida[-84:-34].strip()     
        if len(atividade)<4:
            resposta = "99REQUISIÇÃO INVÁLIDA - ATIVIDADE - DESCRIÇÃO CURTA (14C)"
            enviaResposta(resposta, c) 
            conn.close()
            return                  
        inicio = msgRecebida[-34:-24]
        try:
            inicio = datetime.strptime(inicio, "%d/%m/%Y")
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE INÍCIO INVÁLIDA (14D)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        if inicio.date()>datetime.today().date():  
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE INÍCIO FUTURA (14D1)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if inicio.date()<emissao.date(): #ciência não pode ser inferior à data de emissão (obtida nas verificacões gerais)
            resposta = "14Data de início da atividade não pode ser inferior à de emissão do TDPF ("+emissao.strftime("%d/%m/%Y")+")"
            enviaResposta(resposta, c)  
            conn.close()
            return                                               
        vencimento = msgRecebida[-24:-14]
        try:
            vencimento = datetime.strptime(vencimento,"%d/%m/%Y")
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE VENCIMENTO INVÁLIDA (14E)"
            enviaResposta(resposta, c) 
            conn.close()
            return        
        if inicio>vencimento: #foi pedido para retirar essa crítica or vencimento.date()<datetime.today().date():
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE VENCIMENTO ANTERIOR À DE INÍCIO (14F)" #OU PASSADA (14F)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        termino = msgRecebida[-14:-4]
        horas = msgRecebida[-4:-1]
        if termino!="00/00/0000":
            try:
                terminoAux = datetime.strptime(termino, "%d/%m/%Y")
            except:
                resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE TÉRMINO INVÁLIDA (14G)"
                enviaResposta(resposta, c) 
                conn.close()
                return        
            if inicio>terminoAux or terminoAux.date()>datetime.today().date():
                resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE TÉRMINO ANTERIOR À DE INÍCIO OU FUTURA (14H)"
                enviaResposta(resposta, c) 
                conn.close()
                return
        else:
            terminoAux = None
        if not horas.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - HORAS INVÁLIDAS (14I)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        try:
            horas = int(horas)   
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - HORAS INVÁLIDAS (14J)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if horas<0:
            resposta = "99REQUISIÇÃO INVÁLIDA - HORAS INVÁLIDAS (14K)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        haObservacoes = msgRecebida[-1:]
        if haObservacoes=="S":
            enviaRespostaSemFechar("1400", c) #envia este para que o cliente envie as observações
            try:
                mensagemRec = c.recv(1024) #.decode('utf-8') #chegou a requisicao criptografada
                requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                if requisicao[:2]!="14":
                    resposta = "99REQUISIÇÃO INVÁLIDA (14L)"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return
                observacoes = requisicao[2:102].strip()
            except:
                c.close()
                conn.close()
                logging.info("Erro de time out 14 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                return            
        else:
            observacoes = ""
        #já foi verificado se está alocado
        #verificamos se o tdpf existe e se é de responsabilidade do usuário e está em andamento
        #if not verificaAlocacao(conn, cpf, tdpf):
        #    resposta = "14CPF NÃO ESTÁ ALOCADO OU NÃO É SUPERVISOR OU TDPF ENCERRADO"
        #    enviaResposta(resposta, c) 
        #    conn.close()
        #    return 
        #podemos incluir a atividade 
        tdpfMonitorado, monitoramentoAtivo, chave = tdpfMonitoradoCPF(conn, tdpf, cpf) #tdpf e cpf vêm da rotina comum às requisições 2-5 e 14-17
        try:
            comando = "Insert Into Atividades (TDPF, Atividade, Inicio, Vencimento, Termino, Horas, Observacoes) Values (%s, %s, %s, %s, %s, %s, %s)"           
            cursor.execute(comando, (chaveTdpf, atividade, inicio, vencimento, terminoAux, horas, observacoes))  #chaveTdpf vem da rotina comum às requisições 2-5 e 14-17
            resposta = "14REGISTRO INCLUÍDO"            
            if tdpfMonitorado and monitoramentoAtivo==False: #monitoramento do tdpf estava desativado - ativa
                comando = "Update CadastroTDPFs Set Fim=Null Where Codigo=%s"
                cursor.execute(comando, (chave,))  
                resposta = resposta + " - MONITORAMENTO REATIVADO"                       
            elif not tdpfMonitorado: #tdpf não estava sendo monitorado - inclui ele
                comando = "Insert into CadastroTDPFs (Fiscal, TDPF, Inicio) Values (%s, %s, %s)"
                cursor.execute(comando, (chaveFiscal, chaveTdpf, datetime.today().date())) #chaveFiscal vem para todas as funções; chaveTdpf vem da rotina comum às requisições 2-5 e 14-17
                resposta = resposta + " - MONITORAMENTO INICIADO"  
            else:
                resposta = resposta + " - TDPF MONITORADO"                  
            conn.commit()
        except:
            conn.rollback()
            resposta = "14ERRO NA INCLUSÃO DO REGISTRO"
        enviaResposta(resposta, c)  
        conn.close()
        return             


    if codigo==15: #mostra lista de atividades de um tdpf - cpf deve estar alocado ou ser supervisor (já foi verificado)
        #if len(msgRecebida)>(13+tamChave): #despreza
        #    pass
        if len(msgRecebida)!=(32+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (15A)"
            enviaResposta(resposta, c) 
            conn.close()
            return         
        regInicial = msgRecebida[-3:]
        if not regInicial.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA (15B)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        try:
            regInicial = int(regInicial)            
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA (15C)"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if regInicial>0: #se foi informado o registro, devemos buscar a partir dele, limitado a dez
            offsetReg = "Limit 10 Offset "+str(regInicial-1)             
        else: #caso contrário, buscamos todos para informar a quantidade que há 
             offsetReg = ""  
        comando = "Select Codigo, Atividade, Vencimento, Inicio, Termino, Horas, Observacoes from Atividades Where TDPF=%s Order by Inicio "+offsetReg
        cursor.execute(comando, (chaveTdpf,))
        rows = cursor.fetchall()
        if rows:
            tam = len(rows)
        else:
            tam = 0           
        if tam==0:
            resposta = "15000" 
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if tam>=1000: #limite de 999 atividades
            nnn = "999"
            tam = 999
        else:
            nnn= str(tam).rjust(3, "0")
        registro = "" 
        if regInicial==0:
            resposta = "15"+nnn #código da resposta e qtde de TDPFs que serão retornados
        else:
            resposta = "15"     
        total = 0
        i = 0           
        for row in rows:
            codigoAtiv = str(row[0]).rjust(10, "0")
            atividade = row[1].ljust(50)
            vencimentoAtiv = dataTexto(row[2])
            inicio = dataTexto(row[3])
            termino = dataTexto(row[4])  
            horas = row[5]    
            if horas==None:
                horas = 0                            
            horas = str(horas)[:3].rjust(3, "0")
            observacoes = row[6] if row[6]!=None else ""
            registro = registro + codigoAtiv + atividade + inicio + termino + vencimentoAtiv + horas + observacoes.ljust(100)
            total+=1
            i+=1
            if i%10==0 or total==tam: #de 10 em 10 ou no último registro enviamos a mensagem
                enviaResposta(resposta+registro, c, True, chaveCriptoAES)
                return      

    if codigo==16: #apaga atividade em um tdpf
        if len(msgRecebida)!=(39+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (16A)"
            enviaResposta(resposta, c) 
            conn.close()
            return         
        codAtividade = msgRecebida[-10:].strip()     
        if not codAtividade.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓD ATIVIDADE NÃO NUMÉRICO (16C)"
            enviaResposta(resposta, c) 
            conn.close()
            return                  
        try:
            codAtividade = int(codAtividade)
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓD ATIVIDADE INVÁLIDO (16D)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        if codAtividade<0:  
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓD ATIVIDADE NEGATIVO (16E)"
            enviaResposta(resposta, c) 
            conn.close()
            return                                  
        #já foi verificado
        #verificamos se o tdpf existe e se é de responsabilidade do usuário e está em andamento
        #if not verificaAlocacao(conn, cpf, tdpf):
        #    resposta = "16CPF NÃO ESTÁ ALOCADO OU TDPF ENCERRADO OU INEXISTENTE"
        #    enviaResposta(resposta, c) 
        #    conn.close()
        #    return
        #verificamos se a atividade existe e é do TDPF
        comando = "Select Codigo, TDPF From Atividades Where Codigo=%s" 
        cursor.execute(comando, (codAtividade,))
        row = cursor.fetchone()
        bAchou = False
        if row:
            if len(row)>0:
                bAchou = True
        if not bAchou:
            resposta = "99REQUISIÇÃO INVÁLIDA - ATIVIDADE NÃO ENCONTRADA(16G)"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if row[1]!=chaveTdpf:
            resposta = "99REQUISIÇÃO INVÁLIDA - ATIVIDADE NÃO PERTENCE AO TDPF (16H)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        tdpfMonitorado, monitoramentoAtivo, chave = tdpfMonitoradoCPF(conn, tdpf, cpf)  
        if not tdpfMonitorado or monitoramentoAtivo==False:
            resposta = "16TDPF NÃO ESTÁ SENDO MONITORADO PELO USUÁRIO - EXCLUSÃO NÃO PERMITIDA"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        #podemos excluir a atividade 
        comando = "Delete from Atividades Where Codigo=%s"
        try:
            cursor.execute(comando, (codAtividade,))
            conn.commit()
            resposta = "16REGISTRO EXCLUÍDO"
        except:
            conn.rollback()
            resposta = "16ERRO NA EXCLUSÃO DO REGISTRO"
        enviaResposta(resposta, c)  
        conn.close()
        return             

    if codigo==17: #altera atividade de um tdpf
        if len(msgRecebida)!=(123+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (17A)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        codAtividade = msgRecebida[-94:-84].strip()     
        if not codAtividade.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓD ATIVIDADE NÃO NUMÉRICO (17C)"
            enviaResposta(resposta, c) 
            conn.close()
            return                  
        try:
            codAtividade = int(codAtividade)
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓD ATIVIDADE INVÁLIDO (17D)"
            enviaResposta(resposta, c) 
            conn.close()
            return                   
        atividade = msgRecebida[-84:-34].strip()     
        if len(atividade)<4:
            resposta = "99REQUISIÇÃO INVÁLIDA - ATIVIDADE - DESCRIÇÃO CURTA (17E)"
            enviaResposta(resposta, c) 
            conn.close()
            return                  
        inicio = msgRecebida[-34:-24]
        try:
            inicio = datetime.strptime(inicio, "%d/%m/%Y")
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE INÍCIO INVÁLIDA (17F)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        if inicio.date()>datetime.today().date():  
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE INÍCIO FUTURA (17G)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if inicio.date()<emissao.date(): #ciência não pode ser inferior à data de emissão (obtida nas verificacões gerais)
            resposta = "17Data de início da atividade não pode ser inferior à de emissão do TDPF ("+emissao.strftime("%d/%m/%Y")+")"
            enviaResposta(resposta, c)  
            conn.close()
            return                                            
        vencimento = msgRecebida[-24:-14]
        try:
            vencimento = datetime.strptime(vencimento,"%d/%m/%Y")
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE VENCIMENTO INVÁLIDA (17H)"
            enviaResposta(resposta, c) 
            conn.close()
            return        
        if inicio>vencimento: #foi pedido para retirar essa crítica or vencimento.date()<datetime.today().date():
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE VENCIMENTO ANTERIOR À DE INÍCIO (17I)" #OU PASSADA (17I)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        termino = msgRecebida[-14:-4]
        if termino!="00/00/0000":
            try:
                terminoAux = datetime.strptime(termino, "%d/%m/%Y")
            except:
                resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE TÉRMINO INVÁLIDA (17J)"
                enviaResposta(resposta, c) 
                conn.close()
                return        
            if inicio>terminoAux or terminoAux.date()>datetime.today().date():
                resposta = "99REQUISIÇÃO INVÁLIDA - DATA DE TÉRMINO ANTERIOR À DE INÍCIO OU FUTURA (17K)"
                enviaResposta(resposta, c) 
                conn.close()
                return
        else:
            terminoAux = None
        horas = msgRecebida[-4:-1]
        if not horas.isdigit() and horas!="AAA":
            resposta = "99REQUISIÇÃO INVÁLIDA - HORAS INVÁLIDAS (17L)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        try:
            horas = int(horas)   
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - HORAS INVÁLIDAS (17M)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if horas<0:
            resposta = "99REQUISIÇÃO INVÁLIDA - HORAS INVÁLIDAS (17N)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        haObservacoes = msgRecebida[-1:]
        if haObservacoes=="S":
            enviaRespostaSemFechar("1700", c) #envia este para que o cliente envie as observações
            try:
                mensagemRec = c.recv(1024) #.decode('utf-8') #chegou a requisicao criptografada
                requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                if requisicao[:2]!="17":
                    resposta = "99REQUISIÇÃO INVÁLIDA (17O)"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return
                observacoes = requisicao[2:102].strip()
            except:
                c.close()
                conn.close()
                logging.info("Erro de time out 17 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                return            
        else:
            observacoes = ""
        tdpfMonitorado, monitoramentoAtivo, chave = tdpfMonitoradoCPF(conn, tdpf, cpf)  
        if not tdpfMonitorado or monitoramentoAtivo==False:
            resposta = "17TDPF NÃO ESTÁ SENDO MONITORADO PELO USUÁRIO - ALTERAÇÃO NÃO PERMITIDA"
            enviaResposta(resposta, c) 
            conn.close()
            return                                             
        #já foi verificado
        #verificamos se o tdpf existe e se é de responsabilidade do usuário e está em andamento
        #if not verificaAlocacao(conn, cpf, tdpf):
        #    resposta = "17CPF NÃO ESTÁ ALOCADO AO TDPF OU TDPF ENCERRADO/INEXISTENTE"
        #    enviaResposta(resposta, c) 
        #    conn.close()
        #    return 
        #podemos alterar a atividade 
        try:
            comando = "Update Atividades Set TDPF=%s, Atividade=%s, Inicio=%s, Vencimento=%s, Termino=%s, Horas=%s, Observacoes=%s Where Codigo=%s"           
            cursor.execute(comando, (chaveTdpf, atividade, inicio, vencimento, terminoAux, horas, observacoes, codAtividade))                
            conn.commit()
            resposta = "17ALTERAÇÃO EFETIVADA"
        except:
            conn.rollback()
            resposta = "17ERRO NA ALTERAÇÃO"
        enviaResposta(resposta, c)  
        conn.close()
        return             

    if codigo==18: #inclui entrada em diário da fiscalização
        if len(msgRecebida)!=(38+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (18A)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        primEnvio = msgRecebida[-9:-7]
        numPartes = msgRecebida[-7:-5]
        extensao = msgRecebida[-5:].strip()
        try:
            primEnvio = int(primEnvio) 
            if primEnvio!=0:       
                resposta = "99REQUISIÇÃO INVÁLIDA - CÓDIGO INVÁLIDO DO PRIMEIRO ENVIO (18C)"
                enviaResposta(resposta, c) 
                conn.close()
                return
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓDIGO INVÁLIDO DO PRIMEIRO ENVIO (18D)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        try:
            numPartes = int(numPartes) 
            if primEnvio!=0 or numPartes>64:    #no máximo 64 partes (para não passar, no total, de 65535)   
                resposta = "99REQUISIÇÃO INVÁLIDA - NÚMERO DE PARTES INVÁLIDA (18E)"
                enviaResposta(resposta, c) 
                conn.close()
                return
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - NÚMERO DE PARTES INVÁLIDA (18F)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if len(extensao)==0:
            resposta = "99REQUISIÇÃO INVÁLIDA - EXTENSÃO VAZIA (18F1)"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        #recebemos os dados da criptografia simétrica 
        try:    #informações sobre a criptografia simétrica (chave 16 caracteres criptografada com certificado digital [712], nonce [16] e tag [16] = 744)         
            infoCripto = c.recv(1024) #chegou a requisicao criptografada simetricamente 
            tamEfetivo = len(infoCripto)
            tentativas = 0
            while tamEfetivo<744: #tamanho da mensagem com os dados
                mensagemRec = infoCripto + c.recv(1024) #chegou a requisicao criptografada simetricamente
                tamEfetivo = len(infoCripto)
                tentativas+=1
                if tentativas>15:
                    logging.info("Tentativas de recebimento dos dados da chave criptográfica simétricaforam excedidas - CPF "+cpf)   
                    c.close()
                    conn.close()
                    return
        except:
            c.close()
            conn.close()
            logging.info("Erro de time out 18 (A) - provavelmente cliente não respondeu no prazo. Abandonando operação.")   
            return     
        resposta = "1801OK"
        respostaErro = "1888"
        entrada = None
        logging.info(msgRecebida)
        c.settimeout(15)  
        for parte in range(1, numPartes+1):    
            enviaRespostaSemFechar(resposta, c)   
            logging.info("Enviou")              
            try:              
                mensagemRec = c.recv(4096) #chegou a requisicao criptografada simetricamente 
                tam = int(mensagemRec[:5].decode('utf-8'))
                mensagemRec = mensagemRec[5:]
                tamEfetivo = len(mensagemRec)
                tentativas = 0
                while tamEfetivo<tam:
                    mensagemRec = mensagemRec + c.recv(4096) #chegou a requisicao criptografada simetricamente
                    tamEfetivo = len(mensagemRec)
                    tentativas+=1
                    if tentativas>30:
                        logging.info("Tentativas de recebimento de parte de uma entrada foram excedidas - CPF "+cpf)   
                        c.close()
                        conn.close()
                        return
            except:
                c.close()
                conn.close()
                logging.info("Erro de time out 18 (B) - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                return
            if parte==1:
                entrada = mensagemRec
            else:                 
                entrada = entrada + mensagemRec    
        if entrada:                              
            comando = "Insert into DiarioFiscalizacao (Fiscal, TDPF, Data, Entrada, Extensao) Values (%s, %s, %s, %s, %s)"
            try:
                logging.info("Inserindo ...")
                cursor.execute(comando, (chaveFiscal, chaveTdpf, datetime.today().date(), infoCripto+entrada, extensao)) #chaveFiscal vem para todas as funções; chaveTdpf vem da rotina comum às requisições 2-5 e 14-20
                conn.commit()            
                consulta = "Select Codigo from DiarioFiscalizacao Where Fiscal=%s and TDPF=%s Order by Codigo DESC" #obtém a última entrada deste fiscal/tdpf
                cursor.execute(consulta, (chaveFiscal, chaveTdpf))
                registro = cursor.fetchone()
                if registro!=None:
                    chave = registro[0]
                    enviaResposta("1899"+str(chave).rjust(10,"0"), c)  #enviamos a chave do registro incluído para facilitar a atualização e sua exclusão posteriormente
                else:
                    enviaResposta(respostaErro+"INCLUSÃO DA ENTRADA OU CONSULTA FALHOU (18K)", c)                              
            except:
                conn.rollback()
                enviaResposta(respostaErro+"INCLUSÃO DA ENTRADA FALHOU (18L)", c) 
        conn.close()
        return             

    if codigo==19: #solicita entradas do diário da fiscalização
        if len(msgRecebida)!=(31+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (19A)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        primEnvio = msgRecebida[-2:]
        if primEnvio!="00":
            resposta = "99REQUISIÇÃO INVÁLIDA (19B)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        consulta = "Select Codigo, Data, Extensao from DiarioFiscalizacao Where Fiscal=%s and TDPF=%s order by Codigo"
        cursor.execute(consulta, (chaveFiscal, chaveTdpf))
        entradas = cursor.fetchall()
        if entradas==None:
            numEntradas = 0
        else:
            numEntradas = len(entradas)
        if numEntradas==0:
            enviaResposta("19000NÃO HÁ ENTRADAS CADASTRADAS PARA ESTE TDPF/CPF", c)
            conn.close()
            return
        enviaRespostaSemFechar("19"+str(numEntradas).rjust(3, "0")+"ENTRADAS DISPONÍVEIS", c)
        connRaw = conectaRaw() #para recuperar as entradas (varbinary)
        if connRaw==None:
            resposta = "97ERRO DE CONEXÃO AO BD P/ RECUPERAR ENTRADAS (RAW)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        cursorRaw = connRaw.cursor(buffered=True)
        consulta = "Select Entrada from DiarioFiscalizacao Where Fiscal=%s and TDPF=%s order by Codigo"                        
        cursorRaw.execute(consulta, (chaveFiscal, chaveTdpf))
        entradasRaw = cursorRaw.fetchall()
        j = 0
        tamParte = 4000 #tamanho de cada parte da entrada enviada separadamente
        for entrada in entradas:
            c.settimeout(20)
            mensagemRec = c.recv(1024).decode('utf-8') #chegou a requisicao sem criptografia            
            if mensagemRec[:2]!="19" or mensagemRec[2:4]!="11":
                enviaResposta("99REQUISIÇÃO INVÁLIDA - AGUARDANDO PEDIDOS DE CONTINUAÇÃO (19C)", c) 
                conn.close()
                return             
            codReg = entrada[0]   
            data = entrada[1].strftime("%d/%m/%Y") 
            extensao = entrada[2]
            if extensao==None:
                extensao = ""
            extensao = extensao.ljust(5)
            texto = entradasRaw[j][0]   
            j+=1              
            totalPartes = len(texto) // tamParte #tamParte caracteres do texto são enviados de cada vez
            if (len(texto) % tamParte) > 0:
                totalPartes+=1
            enviaRespostaSemFechar("19"+str(codReg).rjust(10, "0")+data+extensao+str(totalPartes).rjust(2, "0"), c)
            for i in range(totalPartes): #para cada entrada (codReg), fazemos o envio das partes de 300 caracteres cada
                try:
                    mensagemRec = c.recv(1024).decode('utf-8') #chegou a requisicao sem criptografia
                    if mensagemRec==None:
                        mensagemRec = "NULO"
                except:
                    c.close()
                    conn.close()
                    connRaw.close()
                    logging.info("Erro de time out 19 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return                     
                if mensagemRec[:2]!="19" or mensagemRec[2:4]!="11":
                    enviaResposta("99REQUISIÇÃO INVÁLIDA - AGUARDANDO PEDIDOS DE CONTINUAÇÃO (19D)", c) 
                    conn.close()
                    connRaw.close()
                    return 
                resposta = texto[i*tamParte:(i*tamParte+tamParte)]
                c.sendall(str(len(resposta)).rjust(5,"0").encode('utf-8')+resposta) #enviamos o tamanho da resposta (5 primeiros) e o diário criptografado
        c.close()
        conn.close()
        connRaw.close()
        return

    if codigo==20: #apaga entrada de um diário
        if len(msgRecebida)!=(39+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (20A)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        #print(msgRecebida)      
        codEntrada = msgRecebida[-10:].strip()     
        if not codEntrada.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓD ENTRADA NÃO NUMÉRICO (20B)"
            enviaResposta(resposta, c) 
            conn.close()
            return                  
        try:
            codEntrada = int(codEntrada)
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓD ENTRADA INVÁLIDO (20C)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        if codEntrada<0:  
            resposta = "99REQUISIÇÃO INVÁLIDA - CÓD ENTRADA NEGATIVO (20D)"
            enviaResposta(resposta, c) 
            conn.close()
            return                                  
        #já foi verificado
        #verificamos se o tdpf existe e se é de responsabilidade do usuário e está em andamento
        #if not verificaAlocacao(conn, cpf, tdpf):
        #    resposta = "16CPF NÃO ESTÁ ALOCADO OU TDPF ENCERRADO OU INEXISTENTE"
        #    enviaResposta(resposta, c) 
        #    conn.close()
        #    return
        #verificamos se a atividade existe e é do TDPF
        comando = "Select Codigo, TDPF From DiarioFiscalizacao Where Codigo=%s" 
        cursor.execute(comando, (codEntrada,))
        row = cursor.fetchone()
        bAchou = False
        if row:
            if len(row)>0:
                bAchou = True
        if not bAchou:
            resposta = "99REQUISIÇÃO INVÁLIDA - ENTRADA NÃO ENCONTRADA(20E)"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if row[1]!=chaveTdpf:
            resposta = "99REQUISIÇÃO INVÁLIDA - ENTRADA NÃO PERTENCE AO TDPF (20F)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        #tdpfMonitorado, monitoramentoAtivo, chave = tdpfMonitoradoCPF(conn, tdpf, cpf)  <-- nesta funcionalidade, é irrelevante o monitoramento
        #if not tdpfMonitorado or monitoramentoAtivo==False:
        #    resposta = "20TDPF NÃO ESTÁ SENDO MONITORADO PELO USUÁRIO - EXCLUSÃO NÃO PERMITIDA"
        #    enviaResposta(resposta, c) 
        #    conn.close()
        #    return             
        #podemos excluir a entrada
        comando = "Delete from DiarioFiscalizacao Where Codigo=%s"
        try:
            cursor.execute(comando, (codEntrada,))
            conn.commit()
            resposta = "2000REGISTRO EXCLUÍDO"
        except:
            conn.rollback()
            resposta = "2099ERRO NA EXCLUSÃO DO REGISTRO"
        enviaResposta(resposta, c)  
        conn.close()
        return             

    if codigo==21: #inclui informação do DCC        
        if len(msgRecebida)!=(46+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (21A)"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        dcc = msgRecebida[-17:].strip()  
        if not dcc.isdigit() and len(dcc)!=0:
            resposta = "99REQUISIÇÃO INVÁLIDA - DCC DEVE SER NUMÉRICO (21B)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if len(dcc)!=17 and len(dcc)!=0: #o dcc pode estar em branco
            resposta = "99REQUISIÇÃO INVÁLIDA - DCC DEVE TER 17 DÍGITOS (21C)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if len(dcc)==17:
            if not verificaDVDCC(dcc):             
                resposta = "99REQUISIÇÃO INVÁLIDA - DCC INVÁLIDO (21D)"
                enviaResposta(resposta, c) 
                conn.close()
                return
            #verificamos se OUTRO tdpf utilizou o número do DCC
            consulta = "Select TDPFS.Codigo, TDPFS.Numero From TDPFS Where DCC=%s"       
            cursor.execute(consulta, (dcc,))
            rows = cursor.fetchall()
            bAchou = False
            for row in rows:
                if row[0]!=chaveTdpf:
                    bAchou = True
                    break
            if bAchou:
                resposta = "21Nº DO DCC ESTÁ EM USO POR OUTRO TDPF - "+row[1]
                enviaResposta(resposta, c) 
                conn.close()
                return         
        #chaveTdpf vêm da rotina comum às requisições 2-5 e 14-21, onde tb foi verifica se o usuário está alocado ou é supervisor
        if tipoProc[:1] in ['F', 'D']:
            tdpfMonitorado, monitoramentoAtivo, chave = tdpfMonitoradoCPF(conn, tdpf, cpf)
            bMonitoravel = True
        else:
            bMonitoravel = False
        try:
            if dcc=="":
                dcc = None
            comando = "Update TDPFS Set DCC=%s Where TDPFS.Codigo=%s"           
            cursor.execute(comando, (dcc, chaveTdpf)) 
            if bMonitoravel:
                if tdpfMonitorado and monitoramentoAtivo==False: #monitoramento do tdpf estava desativado - ativa
                    comando = "Update CadastroTDPFs Set Fim=Null Where Codigo=%s"
                    cursor.execute(comando, (chave,)) 
                    msgMonitoramento = "MONITORAMENTO DO TDPF REATIVADO"                        
                elif not tdpfMonitorado: #tdpf não estava sendo monitorado - inclui ele
                    comando = "Insert into CadastroTDPFs (Fiscal, TDPF, Inicio) Values (%s, %s, %s)"
                    cursor.execute(comando, (chaveFiscal, chaveTdpf, datetime.today().date()))  
                    msgMonitoramento = "MONITORAMENTO DO TDPF INICIADO" 
                else:
                    msgMonitoramento = "TDPF MONITORADO"
            else:
                msgMonitoramento = "MONITORAMENTO ND"                       
            resposta = "21INFORMAÇÃO REGISTRADA - "+msgMonitoramento                               
            conn.commit()
        except:
            conn.rollback()
            resposta = "21ERRO NO REGISTRO DA INFORMAÇÃO"
        enviaResposta(resposta, c)  
        conn.close()
        return   

    if codigo==23: #inclui/atualiza/exclui parâmetros de pontuação do TDPF - DESATIVADA
        resposta = "99REQUISIÇÃO DESATIVADA (23A1)"
        enviaResposta(resposta, c) 
        conn.close()
        return          

    if codigo==24: #recupera parâmetros de pontuação do TDPF informados pelo usuário, calcula e informa a quantidade de PONTOS dele; ou recupera apenas os parâmetros internos utilizados para cálculo
        if len(msgRecebida)!=(29+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (24A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if encerramento==None:
            encerramento = "A"
        else:
            encerramento = "S"
        resposta = "24"+encerramento+str(int(pontosTdpf*100)).rjust(8,"0")+dataTexto(dataPontos)
        enviaResposta(resposta, c)
        conn.close()
        return             

    if codigo==25: #Relação de TDPFs (diligencia ou fiscalização) alocados ao CPF ou de que este seja supervisor sem ciência entre XX e YY dias após a emissão
        if len(msgRecebida)!=(17+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (25A)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        inicial = msgRecebida[-4:-2]
        final = msgRecebida[-2:]
        try:
            inicial = int(inicial)
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - QTDE DIAS INICIAL INVÁLIDA (25B)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        try:
            final = int(final)
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - QTDE DIAS FINAL INVÁLIDA (25C)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        #if tipoOrgaoUsuario=="L": #considera TDPFs supervisionados e nos quais o fiscal está alocado
        comando = """
                    Select Distinctrow TDPFS.Numero, TDPFS.Nome, TDPFS.Emissao, TDPFS.Vencimento, TDPFS.Codigo, TDPFS.Tipo, TDPFS.TDPFPrincipal, FAPE
                    From TDPFS, Supervisores
                    Where (Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and Supervisores.Fiscal=%s) 
                    and TDPFS.Emissao<=cast((now() - interval %s day) as date) and TDPFS.Emissao>=cast((now() - interval %s day) as date) 
                    and TDPFS.Encerramento Is Null and TDPFS.Codigo not in (Select TDPF from Ciencias Where Data Is Not Null)
                    and TDPFS.Tipo in ('F', 'D')
                    Union
                    Select Distinctrow TDPFS.Numero, TDPFS.Nome, TDPFS.Emissao, TDPFS.Vencimento, TDPFS.Codigo, TDPFS.Tipo, TDPFS.TDPFPrincipal, FAPE
                    From TDPFS, Alocacoes              
                    Where (Alocacoes.Fiscal=%s and Alocacoes.Desalocacao Is Null and Alocacoes.TDPF=TDPFS.Codigo)
                    and TDPFS.Emissao<=cast((now() - interval %s day) as date) and TDPFS.Emissao>=cast((now() - interval %s day) as date) 
                    and TDPFS.Encerramento Is Null and TDPFS.Codigo not in (Select TDPF from Ciencias Where Data Is Not Null)  
                    and TDPFS.Tipo in ('F', 'D')              
                    Order By Numero
                    """                                                       
        cursor.execute(comando, (chaveFiscal, inicial, final, chaveFiscal, inicial, final))
        #elif tipoOrgaoUsuario=="R": #considera TDPFs do órgão e nos quais o fiscal está alocado, se ele estiver na base de fiscais
        #    comando = """
        #              Select Distinctrow TDPFS.Numero, TDPFS.Nome, TDPFS.Emissao, TDPFS.Vencimento, TDPFS.Codigo
        #              From TDPFS
        #              Where TDPFS.Emissao<=cast((now() - interval %s day) as date) and TDPFS.Emissao>=cast((now() - interval %s day) as date) 
        #              and TDPFS.Encerramento Is Null and TDPFS.Codigo not in (Select TDPF from Ciencias Where Data Is Not Null)
        #              and TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) """

        #    if chaveFiscal==0 or chaveFiscal==None:
        #        cursor.execute(comando, (inicial, final, orgaoUsuario))  
        #    else:
        #        comando = comando + """Union
        #                               Select Distinctrow TDPFS.Numero, TDPFS.Nome, TDPFS.Emissao, TDPFS.Vencimento, TDPFS.Codigo
        #                               From TDPFS, Alocacoes              
        #                               Where (Alocacoes.Fiscal=%s and Alocacoes.Desalocacao Is Null and Alocacoes.TDPF=TDPFS.Codigo)
        #                               and TDPFS.Emissao<=cast((now() - interval %s day) as date) and TDPFS.Emissao>=cast((now() - interval %s day) as date) 
        #                               and TDPFS.Encerramento Is Null and TDPFS.Codigo not in (Select TDPF from Ciencias Where Data Is Not Null)                
        #                               Order By Numero                    
        #                            """ 
        #        cursor.execute(comando, (inicial, final, orgaoUsuario, chaveFiscal, inicial, final))
        #else: #órgão nacional - seleciona todos os TDPFs
        #    comando = """
        #              Select Distinctrow TDPFS.Numero, TDPFS.Nome, TDPFS.Emissao, TDPFS.Vencimento, TDPFS.Codigo
        #              From TDPFS
        #              Where TDPFS.Emissao<=cast((now() - interval %s day) as date) and TDPFS.Emissao>=cast((now() - interval %s day) as date) 
        #              and TDPFS.Encerramento Is Null and TDPFS.Codigo not in (Select TDPF from Ciencias Where Data Is Not Null)                   
        #              """ 
        #    cursor.execute(comando, (inicial, final))                       
        rows = cursor.fetchall()
        tam = len(rows)
        if tam==0:
            resposta = "2500"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if tam>=100: #limite de 99 tdpfs
            nn = "99"
            tam = 99
        else:
            nn = str(tam).rjust(2,"0")          
        registro = "" 
        resposta = "25"+nn
        i = 0
        total = 0            
        for row in rows:
            tdpf = row[0]
            nome = row[1]
            emissao = dataTexto(row[2])
            vencimento = dataTexto(row[3])          
            if nome==None or nome=="":
                nome = "ND"  
            nome = nome[:tamNome].ljust(tamNome) 
            #obtem o fiscal há mais tempo alocado ao TDPF
            consulta = "Select Fiscais.Nome From Fiscais, Alocacoes Where Alocacoes.TDPF=%s and Alocacoes.Desalocacao Is Null and Alocacoes.Fiscal=Fiscais.Codigo Order By Alocacoes.Alocacao"
            cursor.execute(consulta, (row[4],)) #passa como parâmetro a chave primária da tabela TDPFS
            fiscaisRow = cursor.fetchone()
            nomeFiscal = None
            if fiscaisRow:
                nomeFiscal = fiscaisRow[0]
            if nomeFiscal==None:
                nomeFiscal = "ND"
            else:
                nomeFiscal = nomeFiscal.split()[0][:20] #manda o primeiro nome do fiscal tb (limitado a 20 caracteres)
            nomeFiscal = nomeFiscal.ljust(20) 
            tipoProc = row[5]
            tdpfPrincipal = row[6]
            tipoProc = tipoProc if (tipoProc!='D' or tdpfPrincipal==None) else 'V' #V = diligência vinculada
            fape = row[7]
            if fape=='S':
                tipoProc += 'P'
            else:
                tipoProc += ' '             
            registro = registro + tdpf + tipoProc + nome + emissao + vencimento + nomeFiscal
            i+=1
            total+=1
            if i==5 or total==tam: #de cinco em cinco ou no último registro, enviamos
                enviaRespostaSemFechar(resposta+registro, c, True, chaveCriptoAES)
                resposta = "25"
                registro = ""
                i = 0
                if total==tam:
                    c.close()
                    return #percorreu os registros ou 99 deles, que é o limite
                if total<tam: #ainda não chegou ao final - aguardamos a requisição da continuação
                    try:
                        mensagemRec = c.recv(1024) #.decode('utf-8') #chegou a requisicao
                        requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                        if requisicao!="2512345678909":
                            resposta = "99REQUISIÇÃO INVÁLIDA (25D)"
                            enviaResposta(resposta, c) 
                            conn.close()
                            return
                    except:
                        c.close()
                        conn.close()
                        logging.info("Erro de time out 25 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                        return 

    if codigo==27: #Supervisor informa trimestre previsto para encerramento (meta) do TDPF
        trimestreDict = {1:"1", 2:"1", 3:"1", 4:"2", 5:"2", 6:"2", 7:"3", 8:"3", 9:"3", 10:"4", 11:"4", 12:"4"}
        if len(msgRecebida)!=(35+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (27A)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        trimestre = msgRecebida[-6:]
        ano = trimestre[:4]
        trimDigit = trimestre[-1:]
        if not ano.isdigit() or trimestre[4:5]!="/" or not trimDigit.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE INVÁLIDO (27B)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if trimDigit=='0' or trimDigit>'4':      
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE INVÁLIDO (27C)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        trimAtual = str(datetime.now().year)+"/"+trimestreDict[datetime.now().month]           
        if trimAtual>trimestre:
            resposta = "27NTRIMESTRE INFORMADO NÃO PODE SER ANTERIOR AO ATUAL"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        #lá em cima, na verificação comum a diversas funcionalidades, já foi verificado o TDPF, se está em andamento e se o CPF é do supervisor
        comando = "Update TDPFS Set TrimestrePrevisto=%s Where TDPFS.Codigo=%s"
        try:
            cursor.execute(comando, (trimestre, chaveTdpf))      
            conn.commit()
            resposta = "STRIMESTRE ATUALIZADO COM SUCESSO"
        except:
            conn.rollback()
            resposta = "NNÃO FOI POSSÍVEL ATUALIZAR A TABELA DE TDPFS"
        resposta = "27"+resposta
        enviaResposta(resposta, c) 
        conn.close()
        return        

    if codigo==28: #envia lista de fiscais alocados a TDPF (nome fiscalizado, data de encerramento, alocacação, desalocação, horas e nome do fiscal)
        if len(msgRecebida)!=(29+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (28A)"
            enviaResposta(resposta, c) 
            conn.close()
            return            
        comando = """Select TDPFS.Nome, TDPFS.Encerramento, Alocacoes.Alocacao, Alocacoes.Desalocacao, Alocacoes.Horas, Fiscais.Nome, Fiscais.CPF
                     From TDPFS, Alocacoes, Fiscais
                     Where Alocacoes.TDPF=TDPFS.Codigo and TDPFS.Codigo=%s and Alocacoes.Fiscal=Fiscais.Codigo"""
        cursor.execute(comando, (chaveTdpf,))
        rows = cursor.fetchall()
        if not rows or len(rows)==0:
            comando = "Select TDPFS.Nome, TDPFS.Encerramento from TDPFS Where TDPFS.Codigo=%s"
            cursor.execute(comando, (chaveTdpf,))
            row = cursor.fetchone()
            data = dataTexto(row[1])
            nome = row[0][:100].ljust(100)
            resposta = "28S"+nome+data+"00"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        row = rows[0]
        data = dataTexto(row[1])
        nome = row[0][:100].ljust(100)   
        nFiscais = len(rows)
        if nFiscais>=100:
            nFiscais = 99
        nFiscais = str(nFiscais).rjust(2,"0")  
        respostaInicio = "28S"+nome+data+nFiscais     
        registro = "" 
        i = 0                  
        for row in rows:
            alocacao = dataTexto(row[2])
            desalocacao = dataTexto(row[3])
            horas = row[4]
            if horas==None:
                horas = 0
            horas = str(horas).rjust(4, "0")
            nomeFiscal = row[5][:100]
            nomeFiscal = nomeFiscal.ljust(100)
            cpfFiscal = row[6]
            registro = registro + cpfFiscal + nomeFiscal + alocacao + desalocacao + horas
            i+=1
            if i%10==0 or i==len(rows): #de 10 em 10 registros ou no último enviamos
                resposta = respostaInicio + registro
                registro = ""
                if i==len(rows):
                    enviaResposta(resposta, c, True, chaveCriptoAES)
                    conn.close()
                    return  
                else:
                    enviaRespostaSemFechar(resposta, c, True, chaveCriptoAES)
                    respostaInicio = "28"    
                    try:
                        mensagemRec = c.recv(512)
                        requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                        if requisicao!="2812345678909":
                            resposta = "99REQUISIÇÃO INVÁLIDA (28B)"
                            enviaResposta(resposta, c) 
                            conn.close()
                            return
                    except:
                        c.close()
                        conn.close()
                        logging.info("Erro de time out 28 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                        return 
    #utilizado no código abaixo e no 42
    mesTrimIni = {"1": "01", "2": "04", "3": "07", "4": "10"}
    mesTrimFim = {"1": "03", "2": "06", "3": "09", "4": "12"}

    if codigo==29: #relaciona TDPFs encerrados em um período ou com previsão de encerramento nele sob supervisão do CPF
        qtdeRegistros = 200 #qtde de registros de tdpfs que enviamos por vez (se mudar aqui, tem que alterar o script)
        if len(msgRecebida)!=(25+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (29A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        trimInicial = msgRecebida[-12:-6]
        trimFinal = msgRecebida[-6:]
        anoInicial = trimInicial[:4]
        anoFinal = trimFinal[:4]
        trimIni = trimInicial[5:]
        trimFim = trimFinal[5:]
        if not anoInicial.isdigit() or not anoFinal.isdigit() or trimInicial[4:5]!="/" or trimFinal[4:5]!="/" or not trimIni.isdigit() or not trimFim.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE INVÁLIDO (29B)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if trimInicial>trimFinal:           
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE INICIAL POSTERIOR AO FINAL (29C)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        if anoInicial<"2021":
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE INICIAL NÃO PODE SER ANTERIOR A 2021 (29D)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        if trimFim>"4":
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE NÃO PODE SER SUPERIOR A 4 (29E)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        mesIni = mesTrimIni[trimIni]
        mesFim = mesTrimFim[trimFim]
        dataInicial = datetime.strptime("01/"+mesIni+"/"+anoInicial, "%d/%m/%Y")
        dataFinal = datetime.strptime(str(calendar.monthrange(int(anoFinal), int(mesFim))[1])+"/"+mesFim+"/"+anoFinal, "%d/%m/%Y")
        if dataFinal.date().year>datetime.now().date().year:
            resposta = "99REQUISIÇÃO INVÁLIDA -  ANO FINAL NÃO PODE SER FUTURO (29F)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if tipoOrgaoUsuario=="L":           
            comando = """Select TDPFS.Codigo, TDPFS.Numero, TDPFS.Encerramento, TDPFS.TrimestrePrevisto, TDPFS.Grupo, TDPFS.Emissao, Fiscais.Nome, Alocacoes.Horas, Pontos
                        From TDPFS, Supervisores, Fiscais, Alocacoes
                        Where Supervisores.Fiscal=%s and Supervisores.Equipe=TDPFS.Grupo and Supervisores.Fim Is Null and Alocacoes.TDPF=TDPFS.Codigo and Alocacoes.Fiscal=Fiscais.Codigo
                        and ((Encerramento Is Null and TrimestrePrevisto>=%s and TrimestrePrevisto<=%s) or (Encerramento Is Not Null and Encerramento>=%s and Encerramento<=%s))
                        and TDPFS.Tipo='F'
                        Order by TDPFS.Grupo, TDPFS.Numero"""
            cursor.execute(comando, (chaveFiscal, trimInicial, trimFinal, dataInicial, dataFinal))
        elif tipoOrgaoUsuario=="R":
            comando = """Select TDPFS.Codigo, TDPFS.Numero, TDPFS.Encerramento, TDPFS.TrimestrePrevisto, TDPFS.Grupo, TDPFS.Emissao, Fiscais.Nome, Alocacoes.Horas, Pontos
                        From TDPFS, Fiscais, Alocacoes
                        Where TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) and Alocacoes.TDPF=TDPFS.Codigo and Alocacoes.Fiscal=Fiscais.Codigo
                        and ((Encerramento Is Null and TrimestrePrevisto>=%s and TrimestrePrevisto<=%s) or (Encerramento Is Not Null and Encerramento>=%s and Encerramento<=%s))
                        and TDPFS.Tipo='F'
                        Order by TDPFS.Grupo, TDPFS.Numero"""
            cursor.execute(comando, (orgaoUsuario, trimInicial, trimFinal, dataInicial, dataFinal))  
        else: #órgão nacional
            comando = """Select TDPFS.Codigo, TDPFS.Numero, TDPFS.Encerramento, TDPFS.TrimestrePrevisto, TDPFS.Grupo, TDPFS.Emissao, Fiscais.Nome, Alocacoes.Horas, Pontos
                        From TDPFS, Fiscais, Alocacoes
                        Where ((Encerramento Is Null and TrimestrePrevisto>=%s and TrimestrePrevisto<=%s) or 
                        (Encerramento Is Not Null and Encerramento>=%s and Encerramento<=%s)) and Alocacoes.TDPF=TDPFS.Codigo and Alocacoes.Fiscal=Fiscais.Codigo
                        and TDPFS.Tipo='F'
                        Order by TDPFS.Grupo, TDPFS.Numero"""
            cursor.execute(comando, (trimInicial, trimFinal, dataInicial, dataFinal))                               
        rows = cursor.fetchall()
        if not rows or len(rows)==0:
            resposta = "29"+"000"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        i = 0
        tam = len(rows)
        if tam>999:
            tam = 999
        #print("Nº TDPFs supervisionados:  ",tam)
        respostaIn = "29"+str(tam).rjust(3,"0")
        registro = ""
        for row in rows:
            chaveTdpf = row[0]
            tdpf = row[1]
            encerramento = dataTexto(row[2])
            trimestre = row[3]
            grupo = row[4].ljust(25)
            if trimestre==None:
                trimestre = "0000/0" 
            emissao = dataTexto(row[5]) 
            nomeFiscal = row[6][:100].ljust(100)
            horas = row[7]
            if horas==None:
                horas = 0
            horas = str(horas).rjust(4, "0")
            consulta = "Select Data from Ciencias Where TDPF=%s Order By Data"
            cursor.execute(consulta, (chaveTdpf,))
            rowCiencia = cursor.fetchone()
            if rowCiencia==None:
                primCiencia = None
            else:
                primCiencia = rowCiencia[0]
            primCiencia = dataTexto(primCiencia)
            i+=1
            pontos = row[8]
            if pontos==None:
                pontos = 0
            pontos = str(int(pontos*100)).rjust(8, "0")
            registro = registro + tdpf + encerramento + trimestre + grupo + emissao + primCiencia + pontos + nomeFiscal + horas         
            if i%qtdeRegistros==0 or i==tam:
                resposta = respostaIn + registro
                registro = ""
                if i==tam:
                    enviaResposta(resposta, c, True, chaveCriptoAES, True) #criptografa e comprime 
                    conn.close()
                    return 
                enviaRespostaSemFechar(resposta, c, True, chaveCriptoAES, True ) #criptografa e comprime
                respostaIn = "29"  
                try:
                    mensagemRec = c.recv(512)
                    requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                    if requisicao!="2912345678909":
                        resposta = "99REQUISIÇÃO INVÁLIDA (29G)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 29 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return

    if codigo==32: #envia mensagens da Cofis do dia           
        if len(msgRecebida)!=(13+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (32A)"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        comando = "Select Mensagem from MensagensCofis Where Data=%s"
        cursor.execute(comando, (datetime.now().date(),))
        rows = cursor.fetchall()
        if len(rows)==0 or rows==None:
            resposta = "32N"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        i = 0         
        mensagens = ""
        for row in rows:
            i+=1
            mensagens = mensagens + row[0].ljust(200)
            if i==9:
                break
        resposta = "32S"+str(i)+mensagens
        enviaResposta(resposta, c) 
        conn.close()
        return         

    if codigo==33: #inclui prorrogação de um TDPF
        requisicao = ""
        try: #aguarda o restante da mensagem
            mensagemRec = c.recv(2048)
            restante, chaveCriptoAES = descriptografa(mensagemRec, addr, c)   
        except:
            c.close()
            conn.close()
            logging.info("Erro de time out 33 - cliente não mandou o complemento da mensagem no prazo. Abandonando operação.")
            return
        if restante[:2]!="33" or restante[2:13]!=cpf:
            resposta = "99REQUISIÇÃO COMPLEMENTAR INVÁLIDA (33A)"
            enviaResposta(resposta, c) 
            conn.close()
            return     
        enviaRespostaSemFechar("33", c) #indicador para enviar os fundamentos            
        try: #aguarda o restante da mensagem
            fundamentos = c.recv(8192).decode("utf-8")  #não há criptografia
            if fundamentos[:2]!="33" or fundamentos[2:13]!=cpf:
                resposta = "99REQUISIÇÃO COMPLEMENTAR INVÁLIDA (33A2)"
                enviaResposta(resposta, c) 
                conn.close()
                return 
            tamanhoFund = int(fundamentos[13:17])
            fundamentos = fundamentos[17:]
            while len(fundamentos)<tamanhoFund:
                fundamentos = fundamentos + c.recv(8192).decode("utf-8")   #não há criptografia
        except:
            c.close()
            conn.close()
            logging.info("Erro de time out 33 - cliente não mandou o complemento da mensagem no prazo. Abandonando operação.")
            return                         
        msgRecebida = msgRecebida + restante[13:]
        if len(msgRecebida)!=(281+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (33A1)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        cpfSuperv = msgRecebida[29+tamChave:40+tamChave]
        assunto = msgRecebida[40+tamChave:140+tamChave].strip()
        numero = msgRecebida[140+tamChave:142+tamChave]
        motivo = msgRecebida[142+tamChave:145+tamChave]
        documento = msgRecebida[145+tamChave:245+tamChave].strip()
        tipo = msgRecebida[245+tamChave:247+tamChave]
        nFiscais = msgRecebida[247+tamChave:248+tamChave]
        if not validaCPF(cpfSuperv):
            resposta = "99REQUISIÇÃO INVÁLIDA - CPF SUPERVISOR INVÁLIDO (33B1)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if  cpfSuperv==cpf:  
            resposta = "99REQUISIÇÃO INVÁLIDA - CPF REQUISITANTE É IGUAL AO DO SUPERVISOR (33B2)"
            enviaResposta(resposta, c) 
            conn.close()
            return                      
        if len(assunto)<10:
            resposta = "99REQUISIÇÃO INVÁLIDA - ASSUNTO (33B)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if not numero.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - NÚMERO INVÁLIDO (33B3)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        try:
            numero = int(numero)   
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - NÚMERO INVÁLIDO (33B4)"
            enviaResposta(resposta, c) 
            conn.close()
            return                
        if not motivo.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - MOTIVO INVÁLIDO (33B5)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        try:
            motivo = int(motivo)   
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - MOTIVO INVÁLIDO (33B6)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        if motivo==0:
            resposta = "99REQUISIÇÃO INVÁLIDA - MOTIVO INVÁLIDO (33B7)"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if len(documento)<9:
            resposta = "99REQUISIÇÃO INVÁLIDA - DOCUMENTO (33C)"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if not tipo.isdigit():   
            resposta = "99REQUISIÇÃO INVÁLIDA - TIPO (33D)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if not nFiscais.isdigit():   
            resposta = "99REQUISIÇÃO INVÁLIDA - Nº DE FISCAIS (33E)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        nFiscais = int(nFiscais)
        if nFiscais>4:
            resposta = "99REQUISIÇÃO INVÁLIDA - Nº DE FISCAIS EXCEDIDO (33E1)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        fundamentos = fundamentos.strip()
        tamFund = len(eliminaTags(fundamentos))
        if tamFund<50 or tamFund>5000:
            resposta = "99REQUISIÇÃO INVÁLIDA - TAMANHO INVÁLIDO DA FUNDAMENTAÇÃO (33K)"
            enviaResposta(resposta, c) 
            conn.close()
            return                       
        fiscais = []
        if nFiscais>1:        
            regFiscais = msgRecebida[-33:]           
            for i in range(nFiscais-1):
                cpfFiscal = regFiscais[i*11:(i*11)+11]
                if cpfFiscal==cpfSuperv:
                    resposta = "99REQUISIÇÃO INVÁLIDA - FISCAL ALOCADO NÃO PODE SER SUPERVISOR (33B3)"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return                     
                if not cpfFiscal in fiscais:
                    fiscais.append(cpfFiscal)      
        consulta = "Select Codigo from Prorrogacoes Where TDPF=%s and Data=%s" 
        cursor.execute(consulta, (chaveTdpf, datetime.now().date())) 
        row = cursor.fetchone()
        if row:
            if row[0]!=None:
                resposta = "33J"
                enviaResposta(resposta, c) 
                conn.close()
                return
        comando = """
                  Select Prorrogacoes.Codigo from Prorrogacoes, AssinaturaFiscal Where TDPF=%s and AssinaturaFiscal.Prorrogacao=Prorrogacoes.Codigo and 
                  (Prorrogacoes.DataAssinatura Is Null or AssinaturaFiscal.DataAssinatura Is Null or Prorrogacoes.RegistroRHAF Is Null)
                  """
        cursor.execute(comando, (chaveTdpf,))
        row = cursor.fetchone()
        if row: #há assinatura ou registro no RHAF pendente em prorrogação já incluída
            resposta = "33J"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        #verificamos se falta muito tempo para o TDPF vencer
        comando = "Select Vencimento from TDPFS Where TDPFS.Codigo=%s"
        cursor.execute(comando, (chaveTdpf, ))
        row = cursor.fetchone()
        if row:
            if row[0].date()>(datetime.now()+timedelta(days=30)).date():
                resposta = "33T" #falta muito Tempo para o TDPF vencer
                enviaResposta(resposta, c) 
                conn.close()
                return 
        else:
            resposta = "99REQUISIÇÃO INVÁLIDA - TDPF NÃO EXISTE MAIS NA BASE DE DADOS (33B4)"
            enviaResposta(resposta, c) 
            conn.close()
            return                              
        #verificamos se o cpfSuperv é supervisor titular ou substituto da equipe do TDPF
        comando = """Select Fiscais.Codigo, Fiscais.CPF, TDPFS.Grupo From Fiscais, TDPFS, Supervisores 
                      Where TDPFS.Codigo=%s and TDPFS.Grupo=Supervisores.Equipe and Supervisores.Fim Is Null and Supervisores.Fiscal=Fiscais.Codigo and Fiscais.CPF=%s"""
        cursor.execute(comando, (chaveTdpf, cpfSuperv))
        row = cursor.fetchone()
        if row==None:
            resposta = "33C"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        supervisor = row[0]
        comando = "Select Codigo, Numero from Prorrogacoes Where TDPF=%s Order by Numero DESC"                  
        cursor.execute(comando, (chaveTdpf, ))
        row = cursor.fetchone()
        if row:
            if row[1]>=numero: #número da prorrogação que recebemos deve ser maior do que o último registrado para o TDPF
                resposta = "33X"
                enviaResposta(resposta, c) 
                conn.close()
                return 
        comando = "Insert Into Prorrogacoes (TDPF, Assunto, Documento, Tipo, Data, Supervisor, Fundamentos, Numero, Motivo) Values (%s, %s, %s, %s, %s, %s, %s, %s, %s)"
        cursor.execute(comando, (chaveTdpf, assunto, documento, tipo, datetime.now().date(), supervisor, fundamentos, numero, motivo))
        cursor.execute(consulta, (chaveTdpf, datetime.now().date())) 
        row = cursor.fetchone()
        bErro = False
        if row==None:
            bErro = True
        elif len(row)==0:
            bErro = True
        if bErro:
            conn.rollback()
            resposta = "33F"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        chaveProrrogacao = row[0]   
        consulta = "Select Fiscais.Codigo From Fiscais, Alocacoes Where Fiscais.CPF=%s and Alocacoes.Fiscal=Fiscais.Codigo and Alocacoes.TDPF=%s and Alocacoes.Desalocacao Is Null"
        listaInclusao = [(chaveProrrogacao, chaveFiscal, datetime.now())] #incluímos o fiscal requisitante
        bErro = False
        for cpfFiscal in fiscais:
            cursor.execute(consulta, (cpfFiscal, chaveTdpf))
            rowFiscal = cursor.fetchone()
            if rowFiscal==None:
                bErro = True
            elif len(rowFiscal)==0:
                bErro = True
            if bErro:
                conn.rollback()
                resposta = "33O"
                enviaResposta(resposta, c) 
                conn.close()
                return 
            listaInclusao.append((chaveProrrogacao, rowFiscal[0], None))     #incluímos os demais fiscais que assinarão     
        comando = "Insert Into AssinaturaFiscal (Prorrogacao, Fiscal, DataAssinatura) Values (%s, %s, %s)"
        cursor.executemany(comando, listaInclusao)
        try:
            conn.commit()
            resposta = "33S"
        except:
            conn.rollback()
            resposta = "33F"
        enviaResposta(resposta, c) 
        if len(fiscais)==0: #só há o fiscal e o supervisor, então só falta o supervisor p/ assinar - mandamos um e-mail para ele avisando
            comando = "Select email from Fiscais, Usuarios Where Fiscais.CPF=Usuarios.CPF and Fiscais.Codigo=%s"
            cursor.execute(comando, (supervisor, )) #pesquisamos o e-mail
            row = cursor.fetchone()
            if row:
                email = row[0]
                tdpfFormatado = formataTDPF(tdpf)
                texto = "Sr. Chefe de Equipe,\n\nInformamos que a Prorrogação nº "+str(numero)+" do TDPF nº "+tdpfFormatado+" - "+nome.strip()+" está pendente de sua assinatura no script Alertas Fiscalização do ContÁgil.\n\nAtenciosamente,\n\nCofis/Disav"
                if ambiente!="PRODUÇÃO":
                    texto = texto + "\n\nAmbiente: "+ambiente
                resultado = enviaEmail(email, texto, "Prorrogação Pendente de Assinatura - TDPF nº "+tdpfFormatado)
                if resultado!=3:
                    print("Falhou o envio do e-mail avisando da prorrogação pendente para o supervisor "+email+" - "+str(resultado))
                    logging.info("Falhou o envio do e-mail avisando da prorrogação pendente para o supervisor "+email+" - "+str(resultado))    
                    if ambiente!="PRODUÇÃO":
                        print(texto)                    
        conn.close()
        return 

    if codigo==34: #consulta prorrogações de um TDPF
        if len(msgRecebida)!=(29+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (34A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        comando ="""Select Prorrogacoes.Codigo, Assunto, Documento, Tipo, Data, Fiscais.CPF, Fiscais.Nome, DataAssinatura, Fundamentos, Prorrogacoes.Numero, RegistroRHAF 
                    from Prorrogacoes, Fiscais Where TDPF=%s and Supervisor=Fiscais.Codigo Order by Data"""    
        cursor.execute(comando, (chaveTdpf,))           
        prorrogacoes = cursor.fetchall()
        if prorrogacoes==None:
            resposta = "34S00"
            enviaResposta(resposta, c)
            conn.close()
            return            
        elif len(prorrogacoes)==0:
            resposta = "34S00"
            enviaResposta(resposta, c)
            conn.close()
            return            
        else:
            i = 0
            tam = len(prorrogacoes)
            if tam>99:
                tam = 99
            consulta = """Select Fiscais.CPF, Fiscais.Nome, DataAssinatura from Fiscais, AssinaturaFiscal 
                          Where AssinaturaFiscal.Fiscal=Fiscais.Codigo and AssinaturaFiscal.Prorrogacao=%s"""
            for prorrogacao in prorrogacoes:
                i+=1
                chaveProrrogacao = prorrogacao[0]
                assunto = prorrogacao[1].ljust(100)
                documento = prorrogacao[2].ljust(100)
                tipo = prorrogacao[3].ljust(2)
                data = dataTexto(prorrogacao[4])
                cpfSupervisor = prorrogacao[5]
                nomeSupervisor = prorrogacao[6][:100].ljust(100)
                assSupervisor = dataTexto(prorrogacao[7])
                fundamentos = prorrogacao[8]
                if fundamentos==None:
                    fundamentos = "" #não deve acontecer, mas ...
                nTermo = prorrogacao[9]
                if nTermo==None:
                    nTermo = 0
                nTermo = str(nTermo).rjust(2, "0") 
                registroRhaf = dataTexto(prorrogacao[10])                   
                cursor.execute(consulta,(chaveProrrogacao,))
                fiscais = cursor.fetchall()
                registro = ""
                j = 0
                for fiscal in fiscais:
                    j+=1
                    registro = registro + fiscal[0] + fiscal[1][:100].ljust(100)+dataTexto(fiscal[2])
                    if j==4:
                        break #no máximo 4 fiscais (não deve ter mais, mas prefiro não arriscar)
                registro = registro.ljust(484)
                registro = assunto+documento+tipo+data+nTermo+cpfSupervisor+nomeSupervisor+assSupervisor+registroRhaf+registro
                if i==1:
                    resposta = "34S"+str(tam).rjust(2, "0") + registro
                else:
                    resposta = "34" + registro
                if i==tam:
                    enviaRespostaSemFechar(resposta, c)
                    if not espera34("1", c, conn, addr):
                        return                       
                    enviaResposta("34"+fundamentos, c)
                    conn.close()
                    return
                else:
                    enviaRespostaSemFechar(resposta, c)
                    if not espera34("2", c, conn, addr):
                        return                    
                    enviaRespostaSemFechar("34"+fundamentos, c)
                    if not espera34("3", c, conn, addr):
                        return                   

    if codigo==35: #exclui prorrogação (deve ser a última; deve estar com assinatura pendente; cpf do fiscal ou do supervisor)
        if len(msgRecebida)!=(39+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (35A)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        data = msgRecebida[-10:]
        if not isDate(data):    
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA INVÁLIDA (35B)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        data = datetime.strptime(data, "%d/%m/%Y")
        if data.date()>datetime.now().date():
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA FUTURA (35C)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        comando = "Select Codigo, DataAssinatura, Data, Supervisor From Prorrogacoes Where TDPF=%s Order by Data DESC" 
        cursor.execute(comando, (chaveTdpf,))
        row = cursor.fetchone()
        if not row:
            resposta = "35P"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        chaveProrrogacao = row[0] 
        if row[2].date()!=data.date(): #não é a última prorrogação ou não existe a data
            resposta = "35L"
            enviaResposta(resposta, c) 
            conn.close()
            return                                              
        if row[3]!=chaveFiscal and not bSupervisor: #usuário não é supervisor do TDPF incluído na prorrogação
            if row[1]!=None: #só é permitida a exclusão se houver assinatura pendente de alguém (e a do supervisor não está pendente)
                comando = "Select * from AssinaturaFiscal Where Prorrogacao=%s and DataAssinatura Is Null"
                cursor.execute(comando, (chaveProrrogacao,))
                assinaturas = cursor.fetchall()
                if len(assinaturas)==0: #usuário não é supervisor e não há assinaturas pendentes
                    resposta = "35U"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return   
        comando = "Delete From AssinaturaFiscal Where Prorrogacao=%s"
        cursor.execute(comando, (chaveProrrogacao,))
        comando = "Delete From Prorrogacoes Where Codigo=%s"
        cursor.execute(comando, (chaveProrrogacao,))
        try:
            conn.commit()
            resposta = "35S"
        except:
            conn.rollback()
            resposta = "35F"
        enviaResposta(resposta, c) 
        conn.close()
        return         

    if codigo==36: #informa assinatura na prorrogação
        if len(msgRecebida)!=(39+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (36A)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        dataDoc = msgRecebida[-10:]
        if not isDate(dataDoc):
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA INVÁLIDA (36B)"
            enviaResposta(resposta, c) 
            conn.close()
            return                     
        dataDoc = datetime.strptime(dataDoc, "%d/%m/%Y")
        if dataDoc.date()>datetime.now().date():
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA FUTURA (36C)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        comando = "Select Codigo, DataAssinatura, Supervisor, Numero from Prorrogacoes Where TDPF=%s and Data=%s"
        cursor.execute(comando, (chaveTdpf, dataDoc))
        row = cursor.fetchone()
        if not row:
            resposta = "36P"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        chaveProrrogacao = row[0]
        assSupervisor = row[1]
        supervisor = row[2]
        numero = row[3]
        if supervisor==chaveFiscal: #é o supervisor (efetivo ou substituto) - temos que ver se ele é o último a assinar e se sua assinatura está pendente
            comando = "Select * from AssinaturaFiscal Where DataAssinatura Is Null and Prorrogacao=%s"
            cursor.execute(comando, (chaveProrrogacao,))
            rowRestantes = cursor.fetchall()
            if rowRestantes:
                if len(rowRestantes)>0: #há assinaturas pendentes de outros fiscais
                    resposta = "36P"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return 
            if assSupervisor!=None: #supervisor já assinou
                resposta = "36P"
                enviaResposta(resposta, c) 
                conn.close()
                return                     
            comando = "Update Prorrogacoes Set DataAssinatura=%s Where Codigo=%s"
            cursor.execute(comando, (datetime.now(), chaveProrrogacao))
        else: #é um dos fiscais responsáveis
            comando = "Select Codigo, DataAssinatura from AssinaturaFiscal Where Fiscal=%s and Prorrogacao=%s"
            cursor.execute(comando, (chaveFiscal, chaveProrrogacao))   
            row = cursor.fetchone()
            if row:
                chaveRegistro = row[0]
                if row[1]!=None: #fiscal já assinou
                    resposta = "36P"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return
                comando = "Update AssinaturaFiscal Set DataAssinatura=%s Where Codigo=%s"  
                cursor.execute(comando, (datetime.now(), chaveRegistro))
                #verificamos se só falta o supervisor assinar e, caso afirmativo, mandamos um e-mail para ele
                comando = "Select Codigo from AssinaturaFiscal Where Prorrogacao=%s and DataAssinatura Is Null"
                cursor.execute(comando, (chaveProrrogacao, ))   
                row = cursor.fetchone()        
                if row==None: #todos os fiscais assinaram
                    comando = "Select email from Fiscais, Usuarios Where Fiscais.CPF=Usuarios.CPF and Fiscais.Codigo=%s"
                    cursor.execute(comando, (supervisor, ))
                    row = cursor.fetchone()
                    if row:
                        email = row[0]
                        tdpfFormatado = formataTDPF(tdpf)
                        texto = "Sr. Chefe de Equipe,\n\nInformamos que a Prorrogação nº "+str(numero)+" do TDPF nº "+tdpfFormatado+" - "+nome.strip()+" está pendente de sua assinatura no script Alertas Fiscalização do ContÁgil.\n\nAtenciosamente,\n\nCofis/Disav"
                        if ambiente!="PRODUÇÃO":
                            texto = texto + "\n\nAmbiente: "+ambiente                        
                        resultado = enviaEmail(email, texto, "Prorrogação Pendente de Assinatura - TDPF nº "+tdpfFormatado)
                        if resultado!=3:
                            print("Falhou o envio do e-mail avisando da prorrogação pendente para o supervisor "+email+" - "+str(resultado))
                            logging.info("Falhou o envio do e-mail avisando da prorrogação pendente para o supervisor "+email+" - "+str(resultado))
                            if ambiente!="PRODUÇÃO":
                                print(texto)
            else:
                resposta = "36P"
                enviaResposta(resposta, c) 
                conn.close()
                return
        try:
            conn.commit()
            resposta = "36S"
        except:
            conn.rollback()
            resposta = "36F"
        enviaResposta(resposta, c) 
        conn.close()
        return  

    if codigo==37: #retorna lista de TDPFs pendentes de assinatura pelo cpf do usuário
        #assinatura do supervisor deve ser a última
        comando = """
                  Select TDPFS.Numero, TDPFS.Nome, Prorrogacoes.Numero, Data from TDPFS, Prorrogacoes, AssinaturaFiscal
                  Where TDPFS.Codigo=Prorrogacoes.TDPF and Prorrogacoes.Supervisor=%s and Prorrogacoes.DataAssinatura Is Null and
                  AssinaturaFiscal.Prorrogacao=Prorrogacoes.Codigo and AssinaturaFiscal.DataAssinatura Is Not Null
                  Union
                  Select TDPFS.Numero, TDPFS.Nome, Prorrogacoes.Numero, Data from TDPFS, Prorrogacoes, AssinaturaFiscal
                  Where TDPFS.Codigo=Prorrogacoes.TDPF and AssinaturaFiscal.Prorrogacao=Prorrogacoes.Codigo and
                  AssinaturaFiscal.Fiscal=%s and AssinaturaFiscal.DataAssinatura Is Null
                  """                 
        listaTdpfs = []
        cursor.execute(comando, (chaveFiscal, chaveFiscal))   
        rows = cursor.fetchall()
        for row in rows:
            if not [row[0], row[1], row[2], row[3]] in listaTdpfs:
                listaTdpfs.append([row[0], row[1], row[2], row[3]]) 
        tam = len(listaTdpfs)
        if tam==0:
            resposta = "3700"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if tam>99:
            tam = 99
        registro = ""
        for i in range(tam):
            registro = registro + listaTdpfs[i][0]+listaTdpfs[i][1][:100].ljust(100)+str(listaTdpfs[i][2]).rjust(2, "0")+listaTdpfs[i][3].strftime("%d/%m/%Y")
        resposta = "37"+str(tam).rjust(2,"0")+registro
        enviaResposta(resposta, c) 
        conn.close()
        return 

    if codigo==38: #solicita próximo número de prorrogação para numerar o termo
        consulta = """Select Numero, DataAssinatura, RegistroRHAF from Prorrogacoes Where TDPF=%s Order by Numero DESC"""
        cursor.execute(consulta, (chaveTdpf, ))
        row = cursor.fetchone()
        if not row:
            resposta = "38S01"                       
        else:
            resposta = "38S"+str(row[0]+1).rjust(2, "0")
            if row[1] == None:
                resposta = resposta + "P" #a última prorrogação está pendente de assinatura
            elif row[2] == None:
                resposta = resposta + "R" #a última prorrogação está pendente de registro no RHAF
        consulta = "Select Vencimento, Emissao from TDPFS Where Codigo=%s"
        cursor.execute(consulta, (chaveTdpf, ))
        row = cursor.fetchone()                
        if len(resposta)==5: #não houve pendência - verificamos quando ocorre o vencimento
            if row[0].date()>(datetime.now()+timedelta(days=30)).date(): #vencimento do TDPF ocorre em mais de 30 dias da data atual - prorrogação não é possível
                resposta = resposta + "T" #falta muito (T)tempo para o vencimento     
            else:
                resposta = resposta + " "
        resposta = resposta + row[1].strftime("%d/%m/%Y")+row[0].strftime("%d/%m/%Y")        
        enviaResposta(resposta, c) 
        conn.close()
        return  

    if codigo==39: #solicita prorrogações pendentes de assinatura no RHAF (todos assinaram a prorrogação) - usuário deve ser fiscal alocado ou supervisor
        consulta = """
                    Select Distinctrow TDPFS.Numero, TDPFS.Nome, Prorrogacoes.Data, Prorrogacoes.Numero, Motivo, Fundamentos
                    from Prorrogacoes, TDPFS, AssinaturaFiscal, Fiscais, Alocacoes, Supervisores
                    Where Fiscais.CPF=%s and ((Alocacoes.Fiscal=Fiscais.Codigo and Alocacoes.Desalocacao Is Null and Alocacoes.TDPF=TDPFS.Codigo) or
                    (Supervisores.Fiscal=Fiscais.Codigo and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and Supervisores.Fiscal=Prorrogacoes.Supervisor)) and 
                    TDPFS.Codigo=Prorrogacoes.TDPF and AssinaturaFiscal.Prorrogacao=Prorrogacoes.Codigo and Prorrogacoes.DataAssinatura Is Not Null and RegistroRHAF Is Null and
                    AssinaturaFiscal.DataAssinatura Is Not Null Order By Prorrogacoes.Data, TDPFS.Numero
                   """        
        cursor.execute(consulta, (cpf,))
        resposta = ""
        rows = cursor.fetchall()
        if not rows:
            resposta = "39N"  
        elif len(rows)==0:
            resposta = "39N" 
        else:
            tam = len(rows) 
        if resposta=="39N":
            enviaResposta(resposta, c) 
            conn.close()
            return   
        if tam>99:
            tam = 99 
        resposta = "39S"+str(tam).rjust(2,"0")
        i = 0
        for row in rows:
            tdpf = row[0]
            nome = row[1]
            if nome==None:
                nome = ""
            nome = nome[:100].ljust(100)
            data = dataTexto(row[2])
            numero = row[3]
            motivo = row[4]
            if motivo==None:
                motivo = 0
            fundamentos = row[5]
            if fundamentos==None:
                fundamentos = ""
            resposta = resposta + tdpf + nome + data + str(numero).rjust(2, "0")+str(motivo).rjust(3, "0")
            enviaRespostaSemFechar(resposta, c, True, chaveCriptoAES) 
            resposta = "39"             
            try:
                mensagemRec = c.recv(256)
                requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                if requisicao!="3912345678909":
                    resposta = "99REQUISIÇÃO INVÁLIDA (39G)"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return
                if i==tam:
                    enviaResposta("39"+fundamentos.strip(), c)
                    conn.close()
                    return                      
                else:
                    enviaRespostaSemFechar("39"+fundamentos.strip(), c)
                    try:
                        mensagemRec = c.recv(256)
                        requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                        if requisicao!="3912345678909":
                            resposta = "99REQUISIÇÃO INVÁLIDA (39H)"
                            enviaResposta(resposta, c) 
                            conn.close()
                            return     
                    except:
                        c.close()
                        conn.close()
                        logging.info("Erro de time out 39A - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                        return                                           
            except:
                c.close()
                conn.close()
                logging.info("Erro de time out 39B - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                return  
        return   

    if codigo==40: #informa registro no RHAF relativamente a uma prorrogação
        if len(msgRecebida)!=(39+tamChave) and len(msgRecebida)!=(40+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (40A)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        bExclui = False
        if len(msgRecebida)==(40+tamChave):
            if msgRecebida[-1:]!="E":
                resposta = "99REQUISIÇÃO INVÁLIDA (40A1)"
                enviaResposta(resposta, c) 
                conn.close()
                return 
            msgRecebida = msgRecebida[:39+tamChave]
            bExclui = True
        dataDoc = msgRecebida[-10:]
        if not isDate(dataDoc):
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA INVÁLIDA (40B)"
            enviaResposta(resposta, c) 
            conn.close()
            return                     
        dataDoc = datetime.strptime(dataDoc, "%d/%m/%Y")
        if dataDoc.date()>datetime.now().date():
            resposta = "99REQUISIÇÃO INVÁLIDA - DATA FUTURA (40C)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        comando = "Select Codigo, DataAssinatura from Prorrogacoes Where TDPF=%s and Data=%s" #como supervisor assina por último, sua assinatura não pode estar nula
        cursor.execute(comando, (chaveTdpf, dataDoc))
        row = cursor.fetchone()
        if not row:
            resposta = "40P"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        chaveProrrogacao = row[0]
        if not bExclui: #inclui o registro no RHAF para a prorrogação
            comando = "Select * from AssinaturaFiscal Where Fiscal=%s and Prorrogacao=%s"
            cursor.execute(comando, (chaveFiscal, chaveProrrogacao)) #o fiscal atual deve constar da prorrogação
            rowFiscal = cursor.fetchone()
            if rowFiscal==None:
                resposta = "40N"
                enviaResposta(resposta, c) 
                conn.close()
                return
            assSupervisor = row[1]
            if assSupervisor==None: #supervisor já deve ter assinado (o último a assinar)
                resposta = "40P"
                enviaResposta(resposta, c) 
                conn.close()
                return     
            comando = "Update Prorrogacoes Set RegistroRHAF=%s Where Codigo=%s"
            cursor.execute(comando, (datetime.now(), chaveProrrogacao)) 
            if vencimento!=None: #esta data vem lá das verificações comuns 
                cursor.execute("Update TDPFS Set Vencimento=%s Where Codigo=%s", (vencimento+timedelta(days=120), chaveTdpf)) #atualizamos a data de vencimento do TDPF
        else: #devemos excluir o registro no RHAF para a prorrogação
            comando = "Update Prorrogacoes Set RegistroRHAF=Null Where Codigo=%s"
            cursor.execute(comando, (chaveProrrogacao, ))            
        try:
            conn.commit()
            resposta = "40S"
        except:
            conn.rollback()
            resposta = "40F"
        enviaResposta(resposta, c) 
        conn.close()
        return  

    if codigo==41: #informações gerenciais de um período (mm/aaaa a mm/aaaa)
        qtdeRegistros = 300 #qtde de registros de tdpfs que enviamos por vez - se alterar aqui, tem que alterar no script e vice-versa
        if len(msgRecebida)!=(32+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (41A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        perInicial = msgRecebida[13+tamChave:20+tamChave]
        perFinal = msgRecebida[20+tamChave:27+tamChave]
        perInicial = "01/"+perInicial
        try:
            perInicial = datetime.strptime(perInicial, "%d/%m/%Y").date()
            if perInicial>datetime.now().date():
                resposta = "99PERÍODO INICIAL DEVE SER PASSADO (41B)"
                enviaResposta(resposta, c) 
                conn.close()
                return                
        except:
            resposta = "99PERÍODO INICIAL INVÁLIDO (41C)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        if perFinal in ["00/0000", "       "]:
            perFinal = datetime.now().date()
        else:
            try:
                data = datetime.strptime("01/"+perFinal, "%d/%m/%Y")
                ultimoDia = ultimoDiaMes(int(perFinal[:2]), int(perFinal[-4:]))
                perFinal = datetime.strptime(str(ultimoDia).rjust(2,"0")+"/"+perFinal, "%d/%m/%Y").date()
                if perFinal<perInicial:
                    resposta = "99PERÍODO FINAL DEVE SER POSTERIOR AO INICIAL (41D)"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return                 
            except:
                resposta = "99PERÍODO FINAL INVÁLIDO (41E)"
                enviaResposta(resposta, c) 
                conn.close()
                return                                
        if ((perFinal.year - perInicial.year)*12+(perFinal.month-perInicial.month))>12: #máximo de 12 meses
            resposta = "99PERÍODO de 12 MESES NO MÁXIMO (41E1)"
            enviaResposta(resposta, c) 
            conn.close()
            return                
        regInicial = msgRecebida[-5:]
        if not regInicial.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - REGISTRO INICIAL (41F)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        try:
            regInicial = int(regInicial)            
        except:
            resposta = "99REQUISIÇÃO INVÁLIDA - REGISTRO INICIAL (41G)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        c.settimeout(30)            
        if regInicial>0: #se foi informado o registro, devemos buscar a partir dele
            offsetReg = "Limit "+str(qtdeRegistros)+" Offset "+str(regInicial-1)             
        else: #caso contrário, buscamos todos para informar a quantidade total que existe
             offsetReg = "Limit "+str(qtdeRegistros)+" Offset 0"
        logging.info("Offset: "+offsetReg) 
        comando = """Select TDPFS.Codigo, TDPFS.Numero, TDPFS.Nome, TDPFS.Emissao, TDPFS.Encerramento, TDPFS.TrimestrePrevisto, TDPFS.Grupo, TDPFS.Porte, 
                     TDPFS.Acompanhamento, TDPFS.CasoEspecial, TDPFS.Tipo, TDPFS.TDPFPrincipal, FAPE, Pontos, DataPontos, NI """
        if tipoOrgaoUsuario=="L": #esta variável e orgaoUsuario vem de rotina comum de validação do usuário
            comandoLocal = comando
            comando = comando + """ from TDPFS, Supervisores 
                                    Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo
                                    and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                                    and ((TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null))
                                    Order by TDPFS.Numero """+offsetReg
            comandoLocal += """ Alocacoes.CPF from TDPFS, Supervisores, Alocacoes
                                Where """
            if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                consulta = """Select Count(TDPFS.Numero)
                                from TDPFS, Supervisores 
                                Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo 
                                and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                                and ((TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null)) """
        elif tipoOrgaoUsuario=="R":
            if chaveFiscal==None or chaveFiscal==0: 
                comando = comando + """ from TDPFS
                                        Where TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                                        and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                                        and ((TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null))
                                        Order by TDPFS.Numero """+offsetReg
            else:
                comando = comando + """ from TDPFS
                                        Where (TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) or 
                                        TDPFS.Grupo in (Select Equipe from Supervisores Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null))
                                        and ((TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null))
                                        and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                                        Order by TDPFS.Numero """+offsetReg

            if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                if chaveFiscal==None or chaveFiscal==0: 
                    consulta = """ Select Count(TDPFS.Numero)
                                    from TDPFS
                                    Where TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) 
                                    and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                                    and ((TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null)) """
                else:
                    consulta = """Select Count(TDPFS.Numero)
                                    from TDPFS
                                    Where (TDPFS.Grupo in (Select Equipe from Jurisdicao Where Orgao=%s) or 
                                    TDPFS.Grupo in (Select Equipe from Supervisores Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null)) 
                                    and ((TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null)) 
                                    and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                               """                                    
        elif tipoOrgaoUsuario=="N":
            comando = comando + """from TDPFS
                                    Where (TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null)
                                    and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                                    Order by TDPFS.Numero """+offsetReg
            if regInicial==0: #contamos a quantidade de registros para informar na primeira consulta
                consulta = """Select Count(TDPFS.Numero)
                                from TDPFS
                                Where (TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null) 
                                and (TDPFS.Tipo!='D' or TDPFS.TDPFPrincipal Is Null)
                           """            
        if regInicial==0:
            if tipoOrgaoUsuario=="L": 
                cursor.execute(consulta, (chaveFiscal, perInicial, perFinal, perFinal))
            elif tipoOrgaoUsuario=="R":
                if chaveFiscal==None or chaveFiscal==0:
                    cursor.execute(consulta, (orgaoUsuario, perInicial, perFinal, perFinal))
                else:
                    cursor.execute(consulta, (orgaoUsuario, chaveFiscal, perInicial, perFinal, perFinal))
            else:
                cursor.execute(consulta, (perInicial, perFinal, perFinal))
            totalReg = cursor.fetchone()
            #print(str(totalReg[0]))
            if totalReg:
                tam = totalReg[0]
            else:
                tam = 0
            if tam==0:
                resposta = "4100000" #41+qtde TDPFs
                enviaResposta(resposta, c) 
                conn.close()
                return             
            if tam>=100000: #limite de  tdpfs
                nnnnn = "99999"
                tam = 99999
            else:
                nnnnn = str(tam).rjust(5, "0")
        if tipoOrgaoUsuario=="L":
            cursor.execute(comando, (chaveFiscal, perInicial, perFinal, perFinal))
        elif tipoOrgaoUsuario=="R":
            if chaveFiscal==None or chaveFiscal==0:
                cursor.execute(comando, (orgaoUsuario, perInicial, perFinal, perFinal))  
            else:
                cursor.execute(comando, (orgaoUsuario, chaveFiscal, perInicial, perFinal, perFinal))  
        else:
            cursor.execute(comando, (perInicial, perFinal, perFinal))                      
        rows = cursor.fetchall()   
        if regInicial>0:
            tam = len(rows)     
        registro = "" 
        i = 0
        total = 0        
        for row in rows:
            chaveTdpf = row[0]
            tdpf = row[1]          
            nome = row[2]
            if nome==None:
                nome = ""   
            nome = nome[:tamNome].ljust(tamNome)                       
            emissao = dataTexto(row[3])   
            tipoProc = row[10]
            tdpfPrincipal = row[11]
            tipoProc = tipoProc if (tipoProc!='D' or tdpfPrincipal==None) else 'V' #V = diligência vinculada
            encerramento = row[4]
            #pontos = None
            #if tipoProc=='F' or (encerramento!=None and tipoProc!='V'):
                #não calculamos pontos de TDPFs em andamento, exceto de fiscalizações, nem de diligências vinculadas, em qualquer hipótese
            pontos = row[13]
            if pontos==None:
                pontos = 0
            pontos = str(int(pontos*100)).rjust(8,"0")
            fape = row[12]
            if fape=='S':
                tipoProc += 'P'
            else:
                tipoProc += ' '   
            dataPontos = dataTexto(row[14])   
            ni = row[15]
            tipoPessoa = "ND"
            if ni!=None:
                if len(ni)==14:
                    tipoPessoa = "PF"
                else:
                    tipoPessoa = "PJ"           
            encerramento = dataTexto(encerramento)         
            trimestre = row[5]
            if trimestre==None:
                trimestre = " ".ljust(6)
            elif len(trimestre)!=6:
                trimestre = " ".ljust(6)
            equipe = row[6]
            nomeEquipe = ""            
            if equipe==None:
                equipe =""
            else:
                cursor.execute("Select Nome From Equipes Where Equipe=%s", (equipe.strip(),))
                equipeRow = cursor.fetchone()
                if equipeRow!=None:
                    nomeEquipe = equipeRow[0]
            equipe = equipe.ljust(25)
            nomeEquipe = nomeEquipe.ljust(50)
            porte = row[7]
            acompanhamento = row[8]
            if porte==None or porte=="":
                porte = "ND "
            if acompanhamento==None or acompanhamento=="":
                acompanhamento = "N"     
            casoEspecial = row[9]
            if casoEspecial==None or casoEspecial==0:
                casoEspecial = ""
                casoEspecialDesc = " "
            else:
                cursor.execute("Select CasoEspecial, Descricao from CasosEspeciais Where Codigo=%s", (casoEspecial, )) 
                linhaCaso = cursor.fetchone()
                if not linhaCaso:
                    casoEspecial = ""
                    casoEspecialDesc = " "    
                else:
                    casoEspecial = str(linhaCaso[0])
                    casoEspecialDesc = linhaCaso[1].strip()[:90]    
            casoEspecial = casoEspecial.rjust(15,"0")
            casoEspecialDesc = casoEspecialDesc.ljust(90)
            #busca a primeira ciência
            comando = "Select Data, Documento from Ciencias Where TDPF=%s order by Data"
            cursor.execute(comando, (chaveTdpf,))
            cienciaReg = cursor.fetchone() #busca a data de ciência mais antiga (PRIMEIRA)
            if cienciaReg: 
                primCiencia = dataTexto(cienciaReg[0]) 
            else:
                primCiencia = "00/00/0000"
            #busca os dados de programação
            tributos = set()          
            if tipoProc[:1]=='F':
                periodoMin = datetime.strptime("01/01/2100", "%d/%m/%Y")
                periodoMax = datetime.strptime("01/01/1900", "%d/%m/%Y")                  
                comando = "Select Tributos.Tributo, PeriodoInicial, PeriodoFinal from Operacoes, Tributos Where Operacoes.TDPF=%s and Operacoes.Tributo=Tributos.Codigo"
                cursor.execute(comando, (chaveTdpf,))
                linhas = cursor.fetchall()
                for linha in linhas:
                    if len(tributos)<7:
                        tributos.add(linha[0])
                    periodoMin = min(periodoMin, linha[1])
                    periodoMax = max(periodoMax, linha[2])
                periodoMin = periodoMin.strftime("%m/%Y")
                periodoMax = periodoMax.strftime("%m/%Y")     
            else:
                periodoMin = " ".rjust(7)              
                periodoMax = periodoMin
            tributosStr = ''.join([str(tributo).rjust(4,"0") for tributo in tributos]).ljust(28)
            #busca os fiscais com mais horas alocadas
            comando = """Select Fiscais.Nome, Alocacoes.Horas 
                         from Fiscais, Alocacoes 
                         Where Alocacoes.TDPF=%s and Alocacoes.Fiscal=Fiscais.Codigo Order by Alocacoes.Horas DESC"""
            cursor.execute(comando, (chaveTdpf,))
            linhas = cursor.fetchall()
            fiscais = [i[0] for i in linhas]
            horasL = [i[1] for i in linhas]
            horas = str(sum(horasL)).rjust(4, "0")
            for k in range(len(fiscais), 8, 1):
                fiscais.append(" ".rjust(50))
                horasL.append(0)
            regFiscais = ""
            for k in range(min(len(fiscais), 8)): #pega os dados dos primeiros 8 fiscais
                regFiscais = regFiscais + fiscais[k][:50].ljust(50) + str(horasL[k]).rjust(4, "0")
            registro = registro + tdpf + tipoProc + nome + emissao + trimestre + encerramento + primCiencia + porte + acompanhamento + tipoPessoa + tributosStr + \
                       periodoMin + periodoMax + casoEspecial + casoEspecialDesc + equipe + nomeEquipe 
            registro += dataPontos + pontos + horas 
            registro += regFiscais
            total+=1
            i+=1
            if i%qtdeRegistros==0 or total==tam: #de qtdeRegistros em qtdeRegistros ou no último registro enviamos a mensagem
                if regInicial==0:
                    resposta = "41"+nnnnn 
                else:
                    resposta = "41"
                enviaResposta(resposta+registro, c, True, chaveCriptoAES, True) #criptografa e comprime (ambos os Trues)
                return 

    if codigo==42: #solicita média de pontos
        if len(msgRecebida)!=(19+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (42A)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        trimInicial = msgRecebida[-2:-1]
        trimFinal = msgRecebida[-1:]
        ano = msgRecebida[-6:-2]
        if not ano.isdigit() or not trimInicial.isdigit() or not trimFinal.isdigit():
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE INVÁLIDO (42B)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if trimInicial>trimFinal:           
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE INICIAL POSTERIOR AO FINAL (42C)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        if ano<"2020":
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE INICIAL NÃO PODE SER ANTERIOR A 2020 (42D)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        if trimFinal>"4":
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE NÃO PODE SER SUPERIOR A 4 (42E)"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        mesIni = mesTrimIni[trimInicial]
        mesFim = mesTrimFim[trimFinal]
        dataInicial = datetime.strptime("01/"+mesIni+"/"+ano, "%d/%m/%Y")
        dataFinal = datetime.strptime(str(calendar.monthrange(int(ano), int(mesFim))[1])+"/"+mesFim+"/"+ano, "%d/%m/%Y") #último dia do mês   
        ano = int(ano)
        mesAtual = datetime.now().month
        anoAtual = datetime.now().year
        dictMesTrimestre = {1:1, 2:1, 3:1, 4:2, 5:2, 6:2, 7:3, 8:3, 9:3, 10:4, 11:4, 12:4}
        trimestreAtual = dictMesTrimestre[mesAtual]        
        if ano>anoAtual:
            resposta = "99REQUISIÇÃO INVÁLIDA - ANO NÃO PODE SER POSTERIOR AO ATUAL (42F)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        trimInicial = int(trimInicial)       
        trimFinal = int(trimFinal)  
        if ano==anoAtual and trimFinal>=trimestreAtual:
            resposta = "99REQUISIÇÃO INVÁLIDA - TRIMESTRE FINAL NÃO PODE SER IGUAL OU POSTERIOR AO ATUAL (42G)"
            enviaResposta(resposta, c) 
            conn.close()
            return                       
        registro = ""    
        dictMediaEquipes = dict()
        equipes = set()   #guarda as equipes para as quais foram calculados médias para não enviar no registro mais de uma vez
        regioes = set()   #guarda as regiões para as quais foram calculadas médias para não enviar no registro mais de uma vez

        #produzimos as méidas de todas as equipes, pois pode ser que o órgão do usuário abranja regiões diferentes ou o usuário supervisione 
        #equipes de regiões diferentes e também para possibilitar o cálculo da média da rf do usuário/supervisor e nacional, 
        #no caso de usuários regionais ou nacionais
        cursor.execute("""Select Sum(Pontos), Sum(PontosMalha), Count(Distinct Fiscal), PontosMetas.Equipe, Equipes.Equipe, Equipes.Nome 
                          From PontosMetas, Equipes, TipoEquipes 
                          Where Ano=%s and Trimestre>=%s and Trimestre<=%s and Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo and 
                          TipoEquipes.Descricao Like 'EXECUÇÃO %' and PontosMetas.Equipe=Equipes.Codigo 
                          Group By PontosMetas.Equipe Order By PontosMetas.Equipe""", (ano, trimInicial, trimFinal))
        linhas = cursor.fetchall()
        for linha in linhas: #média das equipes
            pontos = linha[0]+linha[1]
            fiscais = linha[2]
            chaveEquipe = linha[3]
            equipe = linha[4].strip()
            nomeEquipe = linha[5].strip()
            equipes.add(chaveEquipe)
            if fiscais>0:
                dictMediaEquipes[chaveEquipe] = [round(pontos/fiscais, 2), equipe, nomeEquipe]
        dictRegioes = dict() #guarda média de cada equipe e número delas de uma região fiscal
        for chaveEquipe in dictMediaEquipes: #média das regiões
            media = dictMediaEquipes[chaveEquipe][0]
            rf = dictMediaEquipes[chaveEquipe][1][:2]
            if dictRegioes.get(rf, -1)==-1:
                mediaEquipes = media
                nEquipes = 1
            else:
                mediaEquipes = dictRegioes[rf][0]+media
                nEquipes = dictRegioes[rf][1]+1
            dictRegioes[rf] = [mediaEquipes, nEquipes]
            #com as médias das equipes, obtemos as médias das regiões
            nEquipes = 0
            mediaEquipes = 0
            for rf in dictRegioes: #calcula a média nacional
                nEquipes += dictRegioes[rf][1]
                mediaEquipes += dictRegioes[rf][0]
            if nEquipes>0:
                mediaNacional = round(mediaEquipes/nEquipes,2)
            else:
                mediaNacional = 0
        if not chaveFiscal in [0, None]: #temos que obter os pontos do fiscal
            cursor.execute("Select Sum(Pontos), Sum(PontosMalha), Equipe From PontosMetas Where Fiscal=%s and Ano=%s and Trimestre>=%s and Trimestre<=%s Group By Equipe", (chaveFiscal, ano, trimInicial, trimFinal))
            resultFiscal = cursor.fetchone()
            if resultFiscal: #fiscal tem informações de pontos nem metas cadastradas
                pontosFiscal = resultFiscal[0]+resultFiscal[1]
                chaveEquipe = resultFiscal[2]
                registro += "F"+cpf+nomeFiscal[:100].ljust(100)+dictMediaEquipes[chaveEquipe][1].ljust(25)+str(int(pontosFiscal*100)).rjust(10, "0")
                if chaveEquipe in equipes:
                    registro += "E"+dictMediaEquipes[chaveEquipe][1].ljust(25)+str(int(dictMediaEquipes[chaveEquipe][0]*100)).rjust(10, "0")
                    #equipes.add(chaveEquipe)
            #verificamos qual a equipe do supervisor, se o usuário for um
            consulta = """Select Equipes.Codigo, Equipes.Equipe, Equipes.Nome From Equipes, Supervisores 
                          Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and Supervisores.Equipe=Equipes.Equipe"""
            cursor.execute(consulta, (chaveFiscal, ))
            #enviamos as médias das equipes supervisionadas e dos respectivos fiscais
            linhas = cursor.fetchall()
            for linha in linhas:
                chaveEquipe = linha[0]
                equipe = linha[1]
                #print(chaveEquipe, equipe, linha[2])
                #procuramos os fiscais da equipe, exceto o do usuário, contemplado acima
                consultaFiscais = """Select PontosMetas.Fiscal, Sum(Pontos), Sum(PontosMalha), Fiscais.CPF, Fiscais.Nome From PontosMetas, Fiscais 
                                     Where PontosMetas.Equipe=%s and PontosMetas.Fiscal!=%s and PontosMetas.Fiscal=Fiscais.Codigo and 
                                     Ano=%s and Trimestre>=%s and Trimestre<=%s Group By PontosMetas.Fiscal Order By Fiscais.Nome"""
                cursor.execute(consultaFiscais, (chaveEquipe, chaveFiscal, ano, trimInicial, trimFinal))
                rowPontos = cursor.fetchall()
                for rowPonto in rowPontos:
                    fiscal = rowPonto[0]
                    pontosFiscal = rowPonto[1]+rowPonto[2]
                    registro += "F"+rowPonto[3]+rowPonto[4][:100].ljust(100)+dictMediaEquipes[chaveEquipe][1].ljust(25)+str(int(pontosFiscal*100)).rjust(10, "0")
                if chaveEquipe in equipes:
                    #equipes.add(chaveEquipe)
                    registro += "E"+dictMediaEquipes[chaveEquipe][1].ljust(25)+str(int(dictMediaEquipes[chaveEquipe][0]*100)).rjust(10, "0")
            for equipe in equipes: #temos que enviar as médias das respectivas regiões dos fiscais/supervisores
                rf = dictMediaEquipes[chaveEquipe][1][:2]
                if not rf in regioes:
                    regioes.add(rf)
                    if dictRegioes[rf][1]>0:
                        mediaRegiao = dictRegioes[rf][0]/dictRegioes[rf][1]
                    registro += "R"+rf+str(int(mediaRegiao*100)).rjust(10, "0")
             
        if tipoOrgaoUsuario in ["R", "N"]: #usuário regional ou nacional
            if tipoOrgaoUsuario=="R": #temos que informar as médias de todas as equipes subordinadas ao órgão do usuário
                comando = """Select Distinctrow Equipes.Codigo, Equipes.Equipe
                             from Jurisdicao, Equipes, PontosMetas, TipoEquipes
                             Where Jurisdicao.Orgao=%s and Jurisdicao.Equipe=Equipes.Equipe and Equipes.Codigo=PontosMetas.Equipe and Ano=%s
                             and Equipes.Sistema=6 and TipoEquipes.Descricao Like 'EXECUÇÃO %'"""
                cursor.execute(comando, (orgaoUsuario, ano))
            else: #órgão nacional - informamos as médias de TODAS as equipes   
                comando = """Select Distinctrow Equipes.Codigo, Equipes.Equipe From Equipes, PontosMetas, TipoEquipes
                             Where Equipes.Codigo=PontosMetas.Equipe and Ano=%s and Equipes.Sistema=6 and TipoEquipes.Descricao Like 'EXECUÇÃO %'"""
                cursor.execute(comando, (ano, ))
            equipesReg = cursor.fetchall()
            for equipeReg in equipesReg:
                chaveEquipe = equipeReg[0]
                equipe = equipeReg[1].strip()
                if chaveEquipe in equipes:
                    #equipes.add(chaveEquipe) #para não recalcular pontos desta equipe mais de uma vez
                    mediaEquipe = int(dictMediaEquipes[chaveEquipe][0]*100)
                    registro += "E"+equipe.ljust(25)+str(mediaEquipe).rjust(10, "0")
                #enviamos as médias de todas as regiões das equipes
                rf = equipe[:2]
                if not rf in regioes:
                    regioes.add(rf)
                    if dictRegioes[rf][1]>0:
                        mediaRegiao = dictRegioes[rf][0]/dictRegioes[rf][1]
                    registro += "R"+rf+str(int(mediaRegiao*100)).rjust(10, "0")                    
            #enviamos tb a média nacional
            registro += "N"+str(int(mediaNacional*100)).rjust(10, "0")                                       
        tamMsg = len(registro)
        if tamMsg>999989:
            enviaRespostaSemFechar("42"+registro[:999989], c, False, None, True)
            totalEnviado = 999989
            while totalEnviado<tamMsg:
                try:
                    mensagemRec = c.recv(256)
                    if mensagemRec!="4212345678909":
                        resposta = "99REQUISIÇÃO INVÁLIDA (42H)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return     
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 42A - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return                 
                enviaRespostaSemFechar(registro[totalEnviado:(totalEnviado+999999)], c, False, None, True) #não criptografa, mas compacta
                totalEnviado += 999999
            try:
                mensagemRec = c.recv(256)
                if mensagemRec!="4212345678909":
                    resposta = "99REQUISIÇÃO INVÁLIDA (42I)"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return     
            except:
                c.close()
                conn.close()
                logging.info("Erro de time out 42B - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                return                  
            enviaResposta("TERMINOU", c)
        else:
            enviaResposta("42"+registro+"TERMINOU", c, False, None, True) #não criptografa, mas compacta
        conn.close()
        return
                     
    if codigo==43: #bloqueia ou desbloqueia comunicação via Telegram ou mostra o status desta comunicação
        if len(msgRecebida)!=(14+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (43A)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        tipo = msgRecebida[-1:]
        if not tipo in ["S", "B", "D"]: 
            resposta = "99REQUISIÇÃO INVÁLIDA - TIPO SUBSERVIÇO(43B)"
            enviaResposta(resposta, c) 
            conn.close()
            return        
        if tipo=="S":
            if statusBloqueio=="S":
                enviaResposta("43B", c)
            else:
                enviaResposta("43D", c)
        else:
            if tipo=="D":
                comando = "Update Usuarios Set BloqueiaTelegram='N' Where CPF=%s"
            else:
                if email in ["", None]: #email não pode estar vazio para bloquearmos a comunicação (email foi obtido ao início de trataMsgRecebida)
                    enviaResposta("43E", c)
                    conn.close()
                    return
                comando = "Update Usuarios Set BloqueiaTelegram='S' Where CPF=%s"
            cursor.execute(comando, (cpf, ))
            try:
                conn.commit()
                enviaResposta("43S", c)
            except:
                conn.rollback()
                enviaResposta("43N", c)
        conn.close()
        return

    if codigo==44: #informa termo emitido e respectivos dados de controle POSTAL            
        if len(msgRecebida)!=(13+tamChave+121):
            resposta = "99TAMANHO DA MENSAGEM INVÁLIDA (44B)"
            enviaResposta(resposta, c) 
            conn.close()
            return                             
        #criticamos os campos - já foi verificado se o TDPF existe, se ele está em andamento e se o CPF é de fiscal ou supervisor
        registroAtual = msgRecebida[13+tamChave:]
        documento = registroAtual[16:86].strip()
        if len(documento)<3:
            resposta = "44M"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        documento = documento.upper()                  
        erroData = True
        data = registroAtual[86:96]
        if isDate(data):
            try:
                data = datetime.strptime(data, "%d/%m/%Y")
                erroData = False
            except:
                pass
        if erroData:
            resposta = "44D"
            enviaResposta(resposta, c) 
            conn.close()
            return                  
        rastreamento = registroAtual[96:111].strip().upper()
        if len(rastreamento)<13:
            resposta = "44C"
            enviaResposta(resposta, c) 
            conn.close()
            return                  
        envio = registroAtual[111:121]            
        erroData = True
        if isDate(envio):
            try:
                envio = datetime.strptime(envio, "%d/%m/%Y")
                erroData = False
            except:
                pass
        if erroData:
            resposta = "44V"
            enviaResposta(resposta, c) 
            conn.close()
            return     
        if envio<data or envio.date()>datetime.now().date():
            resposta = "44V"
            enviaResposta(resposta, c) 
            conn.close()
            return                   
        if data<emissao: #data de emissão do termo anterior à de emissão do TDPF
            resposta = "44D"
            enviaResposta(resposta, c) 
            conn.close()
            return                   
        consulta = "Select Data from ControlePostal Where TDPF=%s Order By Data DESC" #buscamos a última data de emissão - a agora enviada não pode ser anterior
        cursor.execute(consulta, (chaveTdpf, ))
        linha = cursor.fetchone()
        if linha:
            if linha[0].date()>data.date(): #data de emissão desse termo não pode ser anterior ao do último enviado
                resposta = "44D"
                enviaResposta(resposta, c) 
                conn.close()
                return                       
        consulta = "Select Codigo from ControlePostal Where CodRastreamento=%s and DataEnvio>cast((now() - interval 90 day) as date)" 
        cursor.execute(consulta, (rastreamento, ))  
        linhaRastreamento = cursor.fetchone()                
        if linhaRastreamento!=None:
            resposta = "44C"+tdpf
            enviaResposta(resposta, c) 
            conn.close()
            return    
        cursor.execute("Insert Into ControlePostal (TDPF, Documento, Data, CodRastreamento, DataEnvio) Values (%s, %s, %s, %s, %s)", (chaveTdpf, documento, data, rastreamento, envio))
        try:
            conn.commit()
            resposta = "44S"           
        except:
            conn.rollback()
            resposta = "44F"
        enviaResposta(resposta, c) 
        conn.close()
        return

    if codigo==45: #exclui informação de postagem; já foi verificado ao início se TDPF existe, está em andamento e se cpf é de fiscal ou supervisor
        if len(msgRecebida)!=(45+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (45A)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        rastreamento = msgRecebida[-16:-1].strip()
        if len(rastreamento)<13:
            resposta = "99TAMANHO DO CÓDIGO DE RASTREAMENTO DEVE SER IGUAL OU SUPERIOR A 13 (45B)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        informaAR = msgRecebida[-1:]      
        if not informaAR in ['S', 'N']:
            resposta = "99INDICADOR DE INFORMAÇÃO DE AR INVÁLIDO (45C)"
            enviaResposta(resposta, c) 
            conn.close()
            return            
        comando = consulta = "Select Codigo from ControlePostal Where CodRastreamento=%s and TDPF=%s" #consultamos o código de rastreamento para o TDPF 
        cursor.execute(comando, (rastreamento, chaveTdpf))
        linha = cursor.fetchone()
        if not linha:
            resposta = "45P"
            enviaResposta(resposta, c) 
            conn.close()
            return         
        codigo = linha[0]
        if informaAR=='N': #apaga o registro
            cursor.execute("Delete From ControlePostal Where Codigo=%s", (codigo, ))
        else:
            cursor.execute("Update ControlePostal Set DataRecebimento=%s Where Codigo=%s", (datetime.now().date(), codigo))
        try:
            conn.commit()
            resposta = "45S"
        except:
            conn.rollback()    
            resposta = "45F"
        enviaResposta(resposta, c) 
        conn.close()
        return 

    if codigo==46: #solicita informações sobre postagens
        if not len(msgRecebida) in [30+tamChave, 34+tamChave]:
            resposta = "99REQUISIÇÃO INVÁLIDA (46A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        tipo = msgRecebida[tamChave+13:tamChave+14]
        if tipo=="T": #com base em TDPF - já foi verificado se o TDPF existe e se o CPF é de fiscal ou supervisor
            comando = "Select TDPFS.Numero, TDPFS.Nome, Documento, Data, CodRastreamento, DataEnvio, SituacaoAtual, DataSituacao From ControlePostal, TDPFS Where TDPFS.Codigo=%s and TDPFS.Codigo=ControlePostal.TDPF Order by Data"
            cursor.execute(comando, (chaveTdpf, ))
        elif tipo in ["D", "P"]: #por período/data
            dataIni = msgRecebida[-20:-10]
            dataFim = msgRecebida[-10:]
            try:
                dataIni = datetime.strptime(dataIni, "%d/%m/%Y")
                dataFim = datetime.strptime(dataFim, "%d/%m/%Y")
                if dataIni>dataFim:
                    resposta = "99DATA INICIAL POSTERIOR À FINAL (46B)"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return                     
                if dataIni.date()>datetime.now().date():
                    resposta = "99DATA INICIAL NÃO PODE SER FUTURA (46C)"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return                     
            except:
                resposta = "99PERÍODO INVÁLIDO (46D)"
                enviaResposta(resposta, c) 
                conn.close()
                return  
            comando = """Select TDPFS.Numero, TDPFS.Nome, Documento, Data, CodRastreamento, DataEnvio, SituacaoAtual, DataSituacao, DataRecebimento
                         From ControlePostal, TDPFS, Alocacoes 
                         Where Alocacoes.Fiscal=%s and Alocacoes.Desalocacao Is Null and Alocacoes.TDPF=TDPFS.Codigo and 
                         TDPFS.Codigo=ControlePostal.TDPF and Data>=%s and Data<=%s and TDPFS.Encerramento Is Null
                         UNION
                         Select TDPFS.Numero, TDPFS.Nome, Documento, Data, CodRastreamento, DataEnvio, SituacaoAtual, DataSituacao, DataRecebimento
                         From ControlePostal, TDPFS, Supervisores 
                         Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and Supervisores.Equipe=TDPFS.Grupo and 
                         TDPFS.Codigo=ControlePostal.TDPF and Data>=%s and Data<=%s and TDPFS.Encerramento Is Null
                         Order by Data """ 
            cursor.execute(comando, (chaveFiscal, dataIni, dataFim, chaveFiscal, dataIni, dataFim))
        else:
            resposta = "99REQUISIÇÃO INVÁLIDA (46D)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        linhas = cursor.fetchall()        
        if not linhas:           
            nnn = 0
        elif len(linhas)>999:
            nnn = 999
        else:
            nnn = len(linhas)
        resposta = "46"+str(nnn).rjust(3, "0")
        if nnn==0:
            enviaResposta(resposta, c) 
            conn.close()
            return 
        total = 0
        registro = ""
        for linha in linhas:
            total+=1
            tdpf = linha[0]
            nome = linha[1][:100].ljust(100)
            documento = linha[2].ljust(70)
            data = dataTexto(linha[3])
            rastreamento = linha[4].ljust(15)
            envio = dataTexto(linha[5])
            situacao = linha[6]
            if situacao==None:
                situacao = ""
            situacao = situacao.ljust(100)
            dataSituacao = dataTexto(linha[7])
            dataRecebimento = dataTexto(linha[8])
            registro += tdpf + nome + documento + data + rastreamento + envio + situacao + dataSituacao + dataRecebimento
            if total%50==0 or total==nnn: #enviamos de 50 em 50 registros ou no final
                if total<nnn:
                    enviaRespostaSemFechar(resposta+registro, c)
                    resposta = "46"
                    registro = ""
                    try:
                        mensagemRec = c.recv(256)
                        if mensagemRec!="4612345678909":
                            resposta = "99REQUISIÇÃO INVÁLIDA (46E)"
                            enviaResposta(resposta, c) 
                            conn.close()
                            return     
                    except:
                        c.close()
                        conn.close()
                        logging.info("Erro de time out 46A - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                        return                      
                else:
                    enviaResposta(resposta+registro, c) 
                    conn.close()
                    return  
        return  #não chega aqui, mas ...

    if codigo==47: #solicita informações de supervisão do CPF
        if len(msgRecebida)!=(13+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (47A)"
            enviaResposta(resposta, c) 
            conn.close()
            return     
        comando = "Select Supervisores.Codigo, Fiscais.Nome, Equipe, Titular from Fiscais, Supervisores Where Fiscais.CPF=%s and Fiscais.Codigo=Supervisores.Fiscal and Supervisores.Fim Is Null"                 
        cursor.execute(comando, (cpf, ))
        linhas = cursor.fetchall()
        resposta = "47"+str(min(99, len(linhas))).rjust(2, "0")
        i = 0
        for linha in linhas:
            equipe = linha[2].ljust(25)
            if linha[3]!=None: #cpf do usuário é substituto da equipe - devemos descobrir nome do titular
                consulta = "Select Fiscais.CPF, Fiscais.Nome From Fiscais, Supervisores Where Supervisores.Codigo=%s and Supervisores.Fiscal=Fiscais.Codigo"
                cursor.execute(consulta, (linha[3], ))                
                linhaTitular = cursor.fetchone()
                cpfTitular = linhaTitular[0]
                nomeTitular = linhaTitular[1][:50].ljust(50)
                cpfSubstituto = cpf
                nomeSubstituto = "" #nome do usuário não precisa ir
            else:
                cpfTitular = cpf
                nomeTitular = "" #nome do usuário não precisa ir
                consulta = "Select Fiscais.CPF, Fiscais.Nome From Fiscais, Supervisores Where Supervisores.Titular=%s and Supervisores.Fiscal=Fiscais.Codigo and Supervisores.Fim Is Null"
                cursor.execute(consulta, (linha[0], ))
                linhaSubstituto = cursor.fetchone() 
                if not linhaSubstituto: #não necessariamente há substituto
                    cpfSubstituto = "".rjust(11)
                    nomeSubstituto = "" #quando não há substituto, o nome não vai
                else:
                    cpfSubstituto = linhaSubstituto[0]
                    nomeSubstituto = linhaSubstituto[1][:50].ljust(50)
            resposta+= equipe+cpfTitular+nomeTitular+cpfSubstituto+nomeSubstituto
            i+=1
            if i==99: #no máximo 99 registros
                break
        enviaResposta(resposta, c) 
        conn.close()
        return         

    if codigo==48: #registra substituto da equipe
        if len(msgRecebida)!=(49+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (48A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        equipe = msgRecebida[-36:-11].strip()
        cpfSubstituto = msgRecebida[-11:]
        if cpf==cpfSubstituto:
            resposta = "99REQUISIÇÃO INVÁLIDA - CPFS DE TITULAR E SUBSTITUTO DEVEM SER DISTINTOS (48B)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        if (len(cpfSubstituto.strip())!=11 or not cpfSubstituto.isdigit()) and cpfSubstituto.strip()!="":
            resposta = "99REQUISIÇÃO INVÁLIDA - CPF DO SUBSTITUTO É INVÁLIDO (48C)"
            enviaResposta(resposta, c) 
            conn.close()
            return            
        comando = "Select * from Supervisores Where Equipe=%s" #antes, verificamos se a equipe existe para detalhar melhor eventual inconsistência na requisição
        cursor.execute(comando, (equipe, ))
        linha = cursor.fetchone()
        if not linha:
            resposta = "48E"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if len(linha)==0:   
            resposta = "48E"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        comando = "Select Supervisores.Codigo, Fiscais.Nome, Equipe from Fiscais, Supervisores Where Fiscais.CPF=%s and Fiscais.Codigo=Supervisores.Fiscal and Supervisores.Fim Is Null and Equipe=%s and Titular Is Null"                 
        cursor.execute(comando, (cpf, equipe)) #verificamos se o CPF do usuário é titular da equipe
        linha = cursor.fetchone()
        if not linha:
            resposta = "48C"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if len(linha)==0:   
            resposta = "48C"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if cpfSubstituto.strip()=="": #cpf substituto vazio significa que deve ser finalizada a substituição da equipe
            cursor.execute("Update Supervisores Set Fim=%s Where Equipe=%s and Titular Is Not Null", (datetime.now().date(), equipe))
            try:
                conn.commit()  
                resposta = "48S"
            except:
                conn.rollback()   
                resposta = "48F"  
            enviaResposta(resposta, c) 
            conn.close()
            return                          
        codigoRegTitular = linha[0]
        comando = "Select Fiscais.Codigo From Fiscais Where Fiscais.CPF=%s"
        cursor.execute(comando, (cpfSubstituto, ))
        linha = cursor.fetchone()
        if not linha:
            resposta = "48A"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if len(linha)==0:   
            resposta = "48A"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        codigoFiscalSubstituto = linha[0]
        #matamos qualquer substituto que exista, exceto o que estão mandando
        cursor.execute("Update Supervisores Set Fim=%s Where Equipe=%s and Fiscal!=%s and Titular Is Not Null", (datetime.now().date(), equipe, codigoFiscalSubstituto))
        #verificamos se ele já existe ou existiu como substituto da equipe
        comando = "Select Supervisores.Codigo, Fim, Titular From Supervisores Where Equipe=%s and Supervisores.Fiscal=%s"
        cursor.execute(comando, (equipe, codigoFiscalSubstituto))
        linhas = cursor.fetchall()
        for linha in linhas:
            if linha[1]==None and linha[2]==codigoRegTitular: #registro de substituição está ativo   
                resposta = "48J"
                enviaResposta(resposta, c) 
                conn.close()
                return 
            if linha[1]==None and linha[2]!=codigoRegTitular: #está como substituo ativo da equipe, mas de outro titular - não devia acontecer, mas reparamos aqui
                comando = "Update Supervisores Set Titular=%s Where Codigo=%s"
                cursor.execute(comando, (codigoRegTitular, linha[0]))  
                try:
                    conn.commit()  
                    resposta = "48S"
                except:
                    conn.rollback()   
                    resposta = "48F"  
                enviaResposta(resposta, c) 
                conn.close()
                return   
            if linha[1]!=None and linha[2]==codigoRegTitular: #está como substituto inativo do titular  
                comando = "Update Supervisores Set Fim=Null Where Codigo=%s"
                cursor.execute(comando, (linha[0], ))  
                try:
                    conn.commit()  
                    resposta = "48S"
                except:
                    conn.rollback()   
                    resposta = "48F"  
                enviaResposta(resposta, c) 
                conn.close()
                return  
        #se chegou aqui, é pq as situações acima não foram contempladas - vamos incluir o registro do substituto pq ele não existe para esta equipe/chefe
        comando = "Insert Into Supervisores (Equipe, Fiscal, Inicio, Titular) Values (%s, %s, %s, %s)" 
        cursor.execute(comando, (equipe, codigoFiscalSubstituto, datetime.now().date(), codigoRegTitular))   
        try:
            conn.commit()  
            resposta = "48S"
        except:
            conn.rollback()   
            resposta = "48F"  
        enviaResposta(resposta, c) 
        conn.close()
        return  

    if codigo==49: #solicita fatores de pontuação gerados pelo Serpro
        nRegistros = 15 #enviamos de 15 em 15 fatores para cada TDPF - se alterar aqui, alterar no script e vice-versa
        if len(msgRecebida)!=(29+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (49A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        consulta = "Select Count(Codigo) From Fatores Where TDPF=%s"
        cursor.execute(consulta, (chaveTdpf,))
        total = cursor.fetchone()[0]
        if total==0:
            resposta = "4900"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if total>99:
            total = 99
        nn = str(total).rjust(2,"0")
        consulta = "Select Sequencia, Descricao, Elementos, Percentual, Pontos From Fatores Where TDPF=%s Order By Sequencia"
        cursor.execute(consulta, (chaveTdpf,))
        linhas = cursor.fetchall()
        resposta = "49"+nn
        registro = ""
        for i in range(total):
            linha = linhas[i]
            sequencia = str(linha[0]).rjust(2,"0")
            descricao = linha[1][:200].ljust(200) #o campo no BD tem 300 caracteres - se necessário, ajustar aqui e no script
            elementos = linha[2]
            if elementos==None:
                elementos = 0
            elementos = str(int(elementos*100)).rjust(8,"0")
            percentual = linha[3]
            if percentual==None:
                percentual = 0
            percentual = str(int(percentual*100)).rjust(5,"0")  
            pontos = linha[4]
            if pontos==None:
                pontos = 0
            pontos = str(int(pontos*100)).rjust(8,"0")   
            registro+=sequencia+descricao+elementos+percentual+pontos
            if (i+1)%nRegistros==0 or (i+1)==total:
                resposta+=registro                   
                if (i+1)==total:
                    enviaResposta(resposta, c) 
                    conn.close()
                    return 
                enviaRespostaSemFechar(resposta, c)  
                resposta = "49"
                registro = ""
                try:
                    mensagemRec = c.recv(256)
                    requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                    if requisicao!="4912345678909":
                        resposta = "99REQUISIÇÃO INVÁLIDA (49B)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 49 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return  

    if codigo==50: #calcula os pontos de um fiscal supervisionado em cada equipe de que ele faça parte
        nRegistros = 10 #número de registros a ser enviado por vez (se alterar aqui, alterar script e vice-versa)
        if len(msgRecebida)!=(33+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (50A)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        dataInicio = msgRecebida[-20:-10]
        dataFim = msgRecebida[-10:]
        try:
            dataInicio = datetime.strptime(dataInicio, "%d/%m/%Y")
        except:
            resposta = "99DATA DE INÍCIO INVÁLIDA (50B)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        try:
            dataFim = datetime.strptime(dataFim, "%d/%m/%Y")
        except:
            resposta = "99DATA FINAL INVÁLIDA (50C)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        if dataInicio>dataFim:
            resposta = "99DATA INICIAL POSTERIOR À FINAL (50D)"
            enviaResposta(resposta, c) 
            conn.close()
            return                                   
        consultaExterna = """Select Distinctrow Alocacoes.Fiscal From Alocacoes, TDPFS, Supervisores
                            Where Supervisores.Fiscal=%s and Supervisores.Fim Is Null and
                            Supervisores.Equipe=TDPFS.Grupo and ((TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or 
                            (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null)) and
                            TDPFS.Codigo=Alocacoes.TDPF and Alocacoes.Horas Is Not Null and Alocacoes.Horas>0"""   
        consultaInterna = """Select Distinctrow Fiscais.Nome, TDPFS.Codigo, TDPFS.Numero, TDPFS.Grupo, TDPFS.Encerramento, TDPFS.Pontos, Alocacoes.Horas, Equipes.Nome
                            from Fiscais, TDPFS, Alocacoes, Equipes Where Fiscais.Codigo=%s and Alocacoes.Fiscal=Fiscais.Codigo and
                            Alocacoes.TDPF=TDPFS.Codigo and Alocacoes.Horas Is Not Null and Alocacoes.Horas>0 and
                            ((TDPFS.Encerramento>=%s  and TDPFS.Encerramento<=%s) or 
                            (TDPFS.Emissao<=%s and TDPFS.Encerramento Is Null)) and TDPFS.Grupo=Equipes.Equipe"""
        consultaHoras = "Select Sum(Alocacoes.Horas) from Alocacoes Where Alocacoes.TDPF=%s"
        cursor.execute(consultaExterna, (chaveFiscal, dataInicio, dataFim, dataFim))   
        linhasExt = cursor.fetchall()
        if linhasExt==None:
            tam = 0
        else:
            tam = len(linhasExt)
        if tam==0:
            resposta = "500000"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        dictPontos = {}   
        dictEquipes = {}             
        for linhaExt in linhasExt:
            fiscal = linhaExt[0]
            cursor.execute(consultaInterna, (fiscal, dataInicio, dataFim, dataFim))
            linhasInt = cursor.fetchall()
            for linhaInt in linhasInt:
                nomeFiscal = linhaInt[0][:100].ljust(100)
                chaveTdpf = linhaInt[1]
                equipe = linhaInt[3].strip()
                encerramento = linhaInt[4]
                if encerramento==None:
                    mes = "00/0000"
                else:
                    mes = encerramento.strftime("%m/%Y")
                pontos = linhaInt[5]
                horas = linhaInt[6]
                nomeEquipe = linhaInt[7][:50].ljust(50)
                cursor.execute(consultaHoras, (chaveTdpf,))
                totalHoras = cursor.fetchone()[0]
                if totalHoras==0 or totalHoras==None:
                    continue
                pontosFiscal = pontos * horas / totalHoras
                if dictPontos.get((equipe, nomeFiscal, mes), -1)==-1: #ainda não existe a entrada para a equipe/mês
                    dictPontos[(equipe, nomeFiscal, mes)]=pontosFiscal
                else:
                    dictPontos[(equipe, nomeFiscal, mes)]+=pontosFiscal
                if dictEquipes.get(equipe, "1")=="1": #não existe a entrada para a equipe       
                    dictEquipes[equipe] = nomeEquipe 
        tam = len(dictPontos)
        if tam>9999:
            tam = 9999
        if tam==0:
            resposta = "500000"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        nnnn = str(tam).rjust(4, "0")
        registro = ""
        resposta = "50"+nnnn
        i = 0
        for chave in dictPontos:
            i+=1
            equipe = chave[0]
            nomeFiscal = chave[1]
            mes = chave[2]
            pontosFiscal = str(int(dictPontos[chave]*100)).rjust(8, "0")
            registro += equipe.ljust(25)+dictEquipes[equipe]+nomeFiscal+mes+pontosFiscal
            if i%nRegistros==0 or i==tam:
                resposta += registro
                enviaRespostaSemFechar(resposta, c, True, chaveCriptoAES, True)
                if i==tam:
                    c.close()
                    conn.close()
                    return
                resposta = "50"
                registro = ""
                try:
                    mensagemRec = c.recv(256)
                    requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                    if requisicao!="5012345678909":
                        resposta = "99REQUISIÇÃO INVÁLIDA (50E)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 50 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return                

    if codigo==51: #envia os pontos acumulados de fiscais e respectiva meta e equipe
        dictTrimFim = {"1": "31/03/", "2": "30/06/", "3": "30/09/", "4": "31/12/"}
        dictTrimInicio = {"1": "01/01/", "2": "01/04/", "3": "01/07/", "4": "01/10/"}
        dictMesTrim = {1:1, 2:1, 3:1, 4:2, 5:2, 6:2, 7:3, 8:3, 9:3, 10:4, 11:4, 12:4}
        nRegistros = 10 #número de registros a ser enviado por vez (se alterar aqui, alterar script e vice-versa)
        if len(msgRecebida)!=(21+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (51A)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        solicitaEmail = msgRecebida[-1:]
        if not solicitaEmail in ["S", "N"]:
            resposta = "99SOLICITAÇÃO DE EMAIL INVÁLIDO (51A1)"
            enviaResposta(resposta, c) 
            conn.close()
            return              
        ambito = msgRecebida[-2:-1].upper()
        if not ambito in ['P', 'G']:
            resposta = "99ÂMBITO DA PESQUISA INVÁLIDO (51B)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        trimestreFinal = msgRecebida[-3:-2] 
        trimestreInicial = msgRecebida[-4:-3]
        ano = msgRecebida[-8:-4]   
        if not ano.isdigit() or not trimestreFinal in ["1", "2", "3", "4"] or not trimestreInicial in ["1", "2", "3", "4"]:
            resposta = "99PERÍODO INVÁLIDO (51C)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if int(ano)<2021 or int(ano)>datetime.now().year: #de 2021 em diante
            resposta = "99ANO DO TRIMESTRE INVÁLIDO (51D)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if datetime.now().year==int(ano):
            if dictMesTrim[datetime.now().month]<int(trimestreFinal):
                resposta = "99TRIMESTRE FINAL INVÁLIDO - SUPERIOR AO DO ANO ATUAL (51E)"
                enviaResposta(resposta, c) 
                conn.close()
                return                  
        if solicitaEmail=="S" and email in [None, ""]:
            resposta = "51USUÁRIO NÃO TEM E-MAIL CADASTRADO - CONSULTA CANCELADA"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if solicitaEmail=="S" and email[-len("@rfb.gov.br"):].upper()!="@RFB.GOV.BR":
            resposta = "51E-MAIL DO USUÁRIO NÃO É INSTITUCIONAL - CONSULTA CANCELADA"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if solicitaEmail=="S" and ambito!="G":
            resposta = "51ENVIO PARA E-MAIL É PERMITIDO APENAS PARA CONSULTAS GERENCIAIS - CONSULTA CANCELADA"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        arq = [x for x in os.listdir(".") if len(x) >= 16 and  x[-16:] == (cpf+".xlsx")]
        if len(arq)>0: #há um arquivo xlsx sendo montado para usuário
            resposta = "51AGUARDE O TÉRMINO DE SUA REQUISIÇÃO ANTERIOR - CONSULTA CANCELADA"
            enviaResposta(resposta, c) 
            conn.close()
            return             
        if solicitaEmail=="S" and ambito=="G":
            resposta = "51OK"
            enviaResposta(resposta, c)  
            nomeArq = "PONTOS-METAS"+datetime.now().strftime("%H_%M_%S")+"_"+cpf+".xlsx"                            
            book = Workbook()
            sheet = book.active 
            sheet.title = "Metas-Pontos"    
            sheet.cell(row=1, column=1).value = "Trimestre"
            sheet.cell(row=1, column=2).value = "Equipe"
            sheet.cell(row=1, column=3).value = "Nome Equipe"
            sheet.cell(row=1, column=4).value = "Fiscal"
            sheet.cell(row=1, column=5).value = "CPF"
            sheet.cell(row=1, column=6).value = "Pontos"  
            sheet.cell(row=1, column=7).value = "Pontos Malha" 
            sheet.cell(row=1, column=8).value = "Pontos TDPFs em Andamento"                       
            sheet.cell(row=1, column=9).value = "Meta"
            sheet.cell(row=1, column=10).value = "Meta Anual"
            sheet.cell(row=1, column=11).value = "Regra"
            larguras = [11, 19, 30, 50, 15, 13, 13, 20, 13, 13, 75]
            for col in range(len(larguras)):
                sheet.column_dimensions[get_column_letter(col+1)].width = larguras[col]   
                currentCell = sheet.cell(row=1, column=col+1)
                currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
                currentCell.font = Font(bold=True)  
            book.save(nomeArq)                                               
        dataFinal = datetime.strptime(dictTrimFim[trimestreFinal]+ano, "%d/%m/%Y")
        dataInicial = datetime.strptime(dictTrimInicio[trimestreInicial]+ano, "%d/%m/%Y")
        trimestreInicial = int(trimestreInicial)
        trimestreFinal = int(trimestreFinal)
        ano = int(ano)
        if ambito=="P": #nos restringimos ao fiscal solicitante
            cursor.execute("Select Nome, CPF From Fiscais Where Codigo=%s", (chaveFiscal,))
            rowFiscal = cursor.fetchone()
            if rowFiscal==None:
                nomeFiscal =""
                cpfFiscal = ""
            else:
                nomeFiscal = rowFiscal[0][:100]
                cpfFiscal = rowFiscal[1]
            nomeFiscal = nomeFiscal.ljust(100)
            cpfFiscal = cpfFiscal.ljust(11)                      
            #vemos a qual equipe o fiscal deve ser considerado como vinculado
            consultaFiscalEquipe = """Select FiscaisEquipes.Equipe, Equipes.Equipe, Equipes.Nome, FiscaisEquipes.Trimestre From FiscaisEquipes, Equipes 
                                      Where FiscaisEquipes.Fiscal=%s and FiscaisEquipes.Ano=%s and 
                                      FiscaisEquipes.Trimestre>=%s and FiscaisEquipes.Trimestre<=%s 
                                      and FiscaisEquipes.Equipe=Equipes.Codigo Order By FiscaisEquipes.Trimestre DESC, Equipes.Equipe"""
            cursor.execute(consultaFiscalEquipe, (chaveFiscal, ano, trimestreInicial, trimestreFinal))   
            rowEquipe = cursor.fetchone()  
            if rowEquipe: #fiscal teve equipe já processada e incluída na tabela FiscaisEquipes no último trimestre do período
                codEquipe = rowEquipe[0] #não precisava, mas vai ficando por aí ...
                equipe = rowEquipe[1]
                nomeEquipe = rowEquipe[2]     
            else:
                metaFiscal = None
                equipe = ""
                nomeEquipe = ""          
            resposta = "5100001"+equipe[:25].ljust(25)+nomeEquipe[:50].ljust(50)+nomeFiscal+cpfFiscal      
            dictPontos = {}                            
            for trim in range(trimestreInicial, trimestreFinal+1, 1):           
                cursor.execute("Select Pontos, PontosMalha, MetaFiscal, MetaAnual From PontosMetas Where Fiscal=%s and Ano=%s and Trimestre=%s", (chaveFiscal, ano, trim)) 
                linha = cursor.fetchone()
                if not linha:
                    pontosFiscal = 0
                    pontosMalha = 0
                    metaFiscal = 0
                    metaAnual = 0
                else:
                    pontosFiscal = linha[0]
                    pontosMalha = linha[1]
                    metaFiscal = linha[2]
                    metaAnual = linha[3]
                dictPontos[trim]=[pontosFiscal, pontosMalha, metaFiscal]
            consultaPontosEmAndamento = """Select Sum(PontosFiscais.Pontos) from PontosFiscais, TDPFS 
                                            Where PontosFiscais.Fiscal=%s and PontosFiscais.TDPF=TDPFS.Codigo and TDPFS.Encerramento Is Null"""
            cursor.execute(consultaPontosEmAndamento, (chaveFiscal,))
            linha = cursor.fetchone()
            pontosEmAndamento = None
            if linha:
                pontosEmAndamento = linha[0]    
            if pontosEmAndamento==None:
                pontosEmAndamento = 0            
            for trim in dictPontos:
                resposta+=str(trim)+str(int(dictPontos[trim][0]*100)).rjust(8,"0")+str(int(dictPontos[trim][1]*100)).rjust(8,"0")+str(int(dictPontos[trim][2]*100)).rjust(8, "0")
            resposta+=str(int(pontosEmAndamento*100)).rjust(8,"0")+str(int(metaAnual*100)).rjust(8, "0")
            enviaResposta(resposta, c) 
            conn.close()
            return        
        #âmbito = G - gerencial (equipes supervisionadas e/ou (região xor país))
        #primeiramente, consultamos as equipes supervisionadas pelo CPF do usuário
        consultaEquipes = "Select Distinctrow Supervisores.Equipe, Equipes.Codigo, Equipes.Nome From Supervisores, Equipes Where Fiscal=%s and Fim Is Null and Supervisores.Equipe=Equipes.Equipe"
        cursor.execute(consultaEquipes, (chaveFiscal,))
        rowsEquipes = cursor.fetchall()
        fiscais = set() #guarda os fiscais pesquisados, sem repetir
        equipes = set() #guarda as equipes (para o caso de o usuário ser regional ou nacional também) para não repetir  
        dictEquipes = {}   
        consulta = "Select Fiscal, Regra, Trimestre From FiscaisEquipes Where Equipe=%s and Ano=%s and Trimestre<=%s Order By Fiscal, Trimestre DESC"   
        consultaPontosEmAndamento = """Select Sum(PontosFiscais.Pontos) from PontosFiscais, TDPFS 
                                       Where PontosFiscais.Fiscal=%s and PontosFiscais.TDPF=TDPFS.Codigo and TDPFS.Encerramento Is Null"""        
        for row in rowsEquipes:
            equipe = row[0].strip()
            chaveEquipe = row[1]
            nomeEquipe = row[2]
            equipes.add(chaveEquipe) #os fiscais desta equipe já estão sendo considerados
            #selecionamos todos os fiscais daquela equipe para fins de metas - consultamos os fiscais da tabela vínculos
            dados = (chaveEquipe, ano, trimestreFinal)
            cursor.execute(consulta, dados)
            linhasFiscais = cursor.fetchall()
            for linha in linhasFiscais:
                fiscal = linha[0]
                if fiscal in fiscais:
                    continue
                regra = linha[1]
                fiscais.add(fiscal)              
                dictEquipes[fiscal] = [chaveEquipe, equipe, nomeEquipe, regra]           
        #fazemos a busca regional e nacional aqui de fiscais e equipes aqui, se for o caso para o tipo do usuário
        if tipoOrgaoUsuario=='R': #usuário regional
            selecaoEquipesRegiao = """Select Equipes.Equipe, Equipes.Codigo, Equipes.Nome From Jurisdicao, Equipes, TipoEquipes 
                                      Where Jurisdicao.Orgao=%s and Jurisdicao.Equipe=Equipes.Equipe and Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo 
                                      and TipoEquipes.Tipo In (3, 4)""" #somente equipes da fiscalização que executem fiscalização ou revisão
            cursor.execute(selecaoEquipesRegiao, (orgaoUsuario,))
        elif tipoOrgaoUsuario=='N':     
            selecaoEquipesPais = """Select Distinctrow Equipes.Equipe, Equipes.Codigo, Equipes.Nome From Equipes, TipoEquipes 
                                    Where Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo and TipoEquipes.Tipo In (3, 4)"""
            cursor.execute(selecaoEquipesPais,)
        if tipoOrgaoUsuario in ['N', 'R']:                           
            linhas = cursor.fetchall()
            for linha in linhas:
                equipe = linha[0].strip()
                chaveEquipe = linha[1]
                nomeEquipe = linha[2]
                if not chaveEquipe in equipes:
                    equipes.add(chaveEquipe) #os fiscais desta equipe estão sendo considerados
                    #selecionamos todos os fiscais daquela equipe para fins de metas, inclusive os que não as tem (tabela vínculos)
                    dados = (chaveEquipe, ano, trimestreFinal)
                    cursor.execute(consulta, dados)
                    linhasFiscais = cursor.fetchall()
                    for linha in linhasFiscais:
                        fiscal = linha[0]
                        regra = linha[1]
                        if not fiscal in fiscais:
                            fiscais.add(fiscal)                           
                            dictEquipes[fiscal] = [chaveEquipe, equipe, nomeEquipe, regra]                                                  
        #depois de termos adicionados todos os fiscais e respectivas equipes, tanto de supervisores quanto de usuários regionais/nacionais, pesquisamos seus pontos e suas metas
        registros = []   
        j = 2
        for fiscal in fiscais:     
            cursor.execute("Select Nome, CPF From Fiscais Where Codigo=%s", (fiscal,))
            rowFiscal = cursor.fetchone()
            if rowFiscal==None:
                nomeFiscal = ""
                cpfFiscal = ""
            else:
                nomeFiscal = rowFiscal[0][:100]
                cpfFiscal = rowFiscal[1]     
            nomeFiscal = nomeFiscal.ljust(100)
            equipe = dictEquipes[fiscal][1]
            nomeEquipe = dictEquipes[fiscal][2][:50]
            regra = dictEquipes[fiscal][3]  
            cursor.execute(consultaPontosEmAndamento, (fiscal,))
            linha = cursor.fetchone()
            pontosEmAndamento = None
            if linha:
                pontosEmAndamento = linha[0]    
            if pontosEmAndamento==None:
                pontosEmAndamento = 0                       
            regPontos = {}
            for trim in range(trimestreInicial, trimestreFinal+1, 1):    
                cursor.execute("""Select Pontos, PontosMalha, MetaFiscal, MetaAnual From PontosMetas Where Fiscal=%s and Ano=%s and Trimestre=%s""", (fiscal, ano, trim))
                linha = cursor.fetchone()
                if linha:
                    pontosFiscal = linha[0] if linha[0]!=None else 0
                    pontosMalha = linha[1] if linha[1]!=None else 0
                    metaFiscal = linha[2] if linha[2]!=None else 0
                    metaAnual = linha[3] if linha[2]!=None else 0
                else:
                    pontosFiscal = 0
                    pontosMalha = 0
                    metaFiscal = 0
                    metaAnual = 0                     
                regPontos[trim] = [pontosFiscal, pontosMalha, metaFiscal]                    
                if solicitaEmail=="S":   
                    sheet.cell(row=j, column=1).value = trim
                    sheet.cell(row=j, column=2).value = formataEquipe(equipe)
                    sheet.cell(row=j, column=3).value = nomeEquipe.strip()
                    sheet.cell(row=j, column=4).value = nomeFiscal.strip()
                    sheet.cell(row=j, column=5).value = cpfFiscal[:3]+"."+cpfFiscal[3:6]+"."+cpfFiscal[6:9]+"-"+cpfFiscal[9:]
                    sheet.cell(row=j, column=6).value = round(pontosFiscal, 2)
                    sheet.cell(row=j, column=7).value = round(pontosMalha, 2)
                    sheet.cell(row=j, column=8).value = round(pontosEmAndamento, 2)
                    sheet.cell(row=j, column=9).value = round(metaFiscal, 2)
                    sheet.cell(row=j, column=10).value = round(metaAnual, 2)
                    sheet.cell(row=j, column=11).value = regra  
                    j+=1 
            registros.append([equipe, nomeEquipe, nomeFiscal, cpfFiscal, regPontos, pontosEmAndamento, metaAnual])                     
        if solicitaEmail=="S":
            book.save(nomeArq)   
            message = "Sr. Usuário,\n\nConforme solicitado, estamos enviando a planilha em anexo com dados sobre pontuação de RPFs e metas relativas ao período em epígrafe.\n\n"
            message += "Atenciosamente,\n\nCofis/Disav"
            resultado = enviaEmail(email, message, "Pontos/Metas de "+dataInicial.strftime("%d/%m/%Y")+" a "+dataFinal.strftime("%d/%m/%Y"), nomeArq)
            if resultado!=3:
                msg = "Erro no envio de email - codigo 51 - "+str(resultado)
                logging.info(msg + " - "+email)
                print("Erro ao enviar a planilha com pontos e metas para o e-mail "+email)
            else:
                print("A planilha com pontos e metas foi enviada com sucesso para o e-mail "+email)
            os.remove(nomeArq)    
            conn.close()
            return                          
        tam = len(registros)
        if tam>99999:
            tam = 99999
        elif tam==0:
            resposta = "5100000"
            enviaResposta(resposta, c) 
            conn.close()
            return                 
        nnnnn = str(tam).rjust(5,"0")
        resposta = "51"+nnnnn
        i = 0
        for registro in registros:
            i +=1
            equipe = registro[0][:25].ljust(25)
            nomeEquipe = registro[1][:50].ljust(50)
            nomeFiscal = registro[2][:100].ljust(100)
            cpfFiscal = registro[3]
            regPontos = registro[4]
            pontosEmAndamento = registro[5]
            metaAnual = registro[6]
            resposta += equipe+nomeEquipe+nomeFiscal+cpfFiscal
            for trim in regPontos:
                resposta+=str(trim)+str(int(regPontos[trim][0]*100)).rjust(8,"0")+str(int(regPontos[trim][1]*100)).rjust(8,"0")+str(int(regPontos[trim][2]*100)).rjust(8, "0")
            resposta+=str(int(pontosEmAndamento*100)).rjust(8, "0") +str(int(metaAnual*100)).rjust(8, "0")            
            if i%nRegistros==0 or i==tam:
                enviaRespostaSemFechar(resposta, c, True, chaveCriptoAES, True)
                if i==tam:
                    c.close()
                    conn.close()
                    return
                resposta = "51"
                try:
                    c.settimeout(15)
                    mensagemRec = c.recv(256)
                    requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                    if requisicao!="5112345678909":
                        resposta = "99REQUISIÇÃO INVÁLIDA (51E)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 51 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return                 
        return

    if codigo==52: #retorna equipes jurisdicionadas por usuário regional (tabela jurisdição) ou nacional (equipes com RPFs em andamento)
        nRegistros = 25 #mudou aqui, muda no script
        if len(msgRecebida)!=(14+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (52A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        somenteTipoOrgao = msgRecebida[-1:]
        if not somenteTipoOrgao in ['S', 'N']:
            resposta = "99SOMENTE TIPO ÓRGÃO - DEFINIÇÃO INVÁLIDA (52A1)"
            enviaResposta(resposta, c) 
            conn.close()
            return
        if somenteTipoOrgao=="S":
            resposta = "52000"+tipoOrgaoUsuario+nomeOrgao[:25].ljust(25)
            enviaResposta(resposta, c)
            conn.close()
            return
        if tipoOrgaoUsuario=="L": #usuário local não tem órgãos jurisdicionados - já podemos responder
            enviaResposta("52000L"+nomeOrgao[:25].ljust(25), c) #varíavel nome órgão vem das verificações gerais ao início desta função
            conn.close()
            return             
        if not tipoOrgaoUsuario in ['R', 'N']:
            resposta = "99TIPO DE USUÁRIO INVÁLIDO - CONTACTE SUPORTE (52B)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if tipoOrgaoUsuario=="R": #usuário regional - listamos equipes jurisdicionadas, conform tabela Jurisdicao
            comando = "Select Distinctrow Jurisdicao.Equipe, Equipes.Nome From Jurisdicao, Equipes Where Jurisdicao.Orgao=%s and Jurisdicao.Equipe=Equipes.Equipe"
            cursor.execute(comando, (orgaoUsuario,))
        else: #Nacional - apenas equipes que constem em TDPFs em andamento ou encerrados no ano corrente
            comando = "Select Distinctrow Grupo, Equipes.Nome From TDPFS, Equipes Where (TDPFS.Encerramento Is Null or TDPFS.Encerramento>='"+str(datetime.now().year)+"-01-01') and TDPFS.Grupo=Equipes.Equipe"
            cursor.execute(comando)
        linhas = cursor.fetchall()
        tam = len(linhas)
        if tam>999: #limitamos a 999 registros
            tam = 999
        if tam==0:
            enviaResposta("52000"+tipoOrgaoUsuario, c) 
            conn.close()
            return          
        i = 0 
        if nomeOrgao==None: #não deve acontecer, mas ... - essa variável é obtida nas verificações gerais do usuário ao início desta função
            nomeOrgao = "ND"
        resposta = "52"+str(tam).rjust(3, "0")+tipoOrgaoUsuario+nomeOrgao[:25].ljust(25) #no primeiro envio, enviamos a quantidade de registros, o tipo do órgão e o nome do órgão do usuário 
        for linha in linhas:
            i+=1
            equipe = linha[0]
            nomeEquipe = linha[1]
            resposta+=equipe.ljust(25)+nomeEquipe.ljust(70)
            if i%nRegistros==0 or i==tam:
                enviaRespostaSemFechar(resposta, c, False, None, True) #não criptografamos, mas compactamos as informações
                if i==tam:
                    c.close()
                    conn.close()
                    return
                resposta="52"
                try:
                    c.settimeout(15)
                    mensagemRec = c.recv(256)
                    requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                    if requisicao!="5212345678909":
                        resposta = "99REQUISIÇÃO INVÁLIDA (52E)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 52 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return       
        return           


    if codigo==60: #Solicita DCCs vinculados aos TDPFs em andamento (somente usuários autorizados - CPF1 a CPF5 [variáveis de ambiente] - e demanda mais uma senha)
                   #somente são enviados DCCs de TDPFs que tenham usuários registrados e ativos no serviço
        if len(msgRecebida)!=(23+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (60A)"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        hora = datetime.now().hour
        if 8<hora<10 and datetime.today().weekday()==0: #às segundas, há restrição de horário por conta da carga dos TDPFs
            resposta = "60H"
            enviaResposta(resposta, c) 
            conn.close()
            return         
        if not cpf in [CPF1, CPF2, CPF3, CPF3, CPF4, CPF5]:
            resposta = "60N"
            enviaResposta(resposta, c) 
            conn.close()
            return
        senha = msgRecebida[-10:]
        horaSenha = str(int(int(datetime.now().strftime('%H%M'))/10)).rjust(3,"0")
        if senha[-10:-3]!=SENHADCCS or senha[-3:]!=horaSenha or SENHADCCS==None: #senha é ----- seguida da hora e a dezena dos minutos (ex.: são 11:35, a senha é -----113)
            resposta = "60N"
            enviaResposta(resposta, c) 
            conn.close()
            return
        comando = """Select Distinctrow DCC from TDPFS, Alocacoes, Fiscais, Usuarios 
                     Where DCC Is Not Null and DCC!='' and Encerramento Is Null and TDPFS.Codigo=Alocacoes.TDPF and Alocacoes.Desalocacao Is Null and
                     Alocacoes.Fiscal=Fiscais.Codigo and Fiscais.CPF=Usuarios.CPF and Usuarios.Adesao Is Not Null and Usuarios.Saida Is Null
                     and TDPFS.Tipo in ('F', 'D')"""
        cursor.execute(comando)
        rows = cursor.fetchall()
        tam = len(rows)
        tamStr = str(tam).rjust(5,"0")
        resposta = "60S"+tamStr
        if tam==0:
            enviaResposta(resposta, c) 
            conn.close()
            return            
        i = 0
        for row in rows:
            resposta = resposta + row[0]
            i+=1
            if i%50==0 or i==tam:
                if i==tam:
                    enviaResposta(resposta, c) 
                    conn.close()
                    return 
                enviaRespostaSemFechar(resposta, c) 
                resposta = "60"  
                try:
                    mensagemRec = c.recv(256)
                    requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                    if requisicao!="60"+cpf:
                        resposta = "99REQUISIÇÃO INVÁLIDA (60B)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 60 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return           

    if codigo==61: #Consulta usuários
        nRegistros = 10 #número de registros a ser enviado por vez (se alterar aqui, alterar script e vice-versa)
        if len(msgRecebida)!=(23+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (61A)"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        if cadastrador!='S':
            resposta = "61N"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        senha = msgRecebida[-10:]
        #horaSenha = str(int(int(datetime.now().strftime('%H%M'))/10)).rjust(3,"0")
        #print(senha[-10:-3], SENHADCCS)
        if senha[-10:-3]!=SENHADCCS or SENHADCCS==None:
            resposta = "61N"
            enviaResposta(resposta, c) 
            conn.close()
            return            
        comando = "Select Count(Codigo) from Usuarios"
        cursor.execute(comando)
        totalUsuarios = cursor.fetchone()[0]  
        if totalUsuarios==0:
            resposta = "6100000"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if totalUsuarios>99999:
            totalUsuarios = 99999
            print("IMPORTANTE: TOTAL DE USUÁRIOS SUPEROU 99999 - CÓDIGO 61")
            logging.info("IMPORTANTE: TOTAL DE USUÁRIOS SUPEROU 99999 - CÓDIGO 61")
        nnnnn = str(totalUsuarios).rjust(5,"0") 
        comando = "Select CPF, Adesao, Saida, email, Tentativas, ValidadeChave, DataEnvio, BloqueiaTelegram, Orgao from Usuarios Order by CPF" 
        consulta = "Select Orgao, Tipo from Orgaos Where Codigo=%s"        
        cursor.execute(comando)
        linhas = cursor.fetchall()
        i = 0
        resposta = "61"+nnnnn
        registro = ""
        for linha in linhas:
            i+=1
            cpfUsuario = linha[0]
            adesao = dataTexto(linha[1])
            saida = dataTexto(linha[2])
            email = linha[3]
            #print(cpfUsuario)
            if email==None:
                email = " "
            email = email[:100].ljust(100)
            tentativas = linha[4]
            if tentativas==None:
                tentativas = 0
            tentativas = str(tentativas)[:1]
            validadeChave = dataTexto(linha[5])
            dataEnvio = dataTexto(linha[6])
            bloqueio = linha[7]
            if bloqueio==None:
                bloqueio = 'N'
            codOrgao = linha[8]
            tipoOrgao = " "
            orgao = ""
            if not codOrgao in [0, None]:
                cursor.execute(consulta, (codOrgao,))
                row = cursor.fetchone()
                if row:
                    orgao = row[0]
                    tipoOrgao = row[1]
            else:
                codOrgao = 0
            codOrgaoStr = str(codOrgao).rjust(4,"0")
            orgao = orgao.ljust(25)
            registro += cpfUsuario+adesao+saida+email+tentativas+validadeChave+dataEnvio+bloqueio+orgao+codOrgaoStr+tipoOrgao
            if i%nRegistros==0 or i==totalUsuarios:
                resposta += registro
                if i<totalUsuarios:
                    enviaRespostaSemFechar(resposta, c, True, chaveCriptoAES)
                    resposta = "61"
                    registro = ""
                    try:
                        mensagemRec = c.recv(256)
                        requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                        if requisicao!="6112345678909":
                            resposta = "99REQUISIÇÃO INVÁLIDA (61B)"
                            enviaResposta(resposta, c) 
                            conn.close()
                            return
                    except:
                        c.close()
                        conn.close()
                        logging.info("Erro de time out 61 - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                        return                     
                else:
                    enviaResposta(resposta, c, True, chaveCriptoAES)
                    conn.close()
                    return

    if codigo==62: #Inclui, exclui ou altera usuário - exclui: na verdade, desabilita
        if not len(msgRecebida) in [(35+tamChave), (139+tamChave)]:
            resposta = "99REQUISIÇÃO INVÁLIDA (62A)"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        if cadastrador!='S':
            resposta = "62N"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        senha = msgRecebida[13+tamChave:23+tamChave]
        #horaSenha = str(int(int(datetime.now().strftime('%H%M'))/10)).rjust(3,"0")
        if senha[-10:-3]!=SENHADCCS or SENHADCCS==None:
            resposta = "62N"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        operacao = msgRecebida[23+tamChave:24+tamChave]
        if not operacao in ['I', 'E', 'A']:
            resposta = "99REQUISIÇÃO INVÁLIDA (62B)"
            enviaResposta(resposta, c) 
            conn.close()
            return      
        if operacao in ['I', 'A'] and len(msgRecebida)!=(139+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (62C)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        if operacao=='E' and len(msgRecebida)!=(35+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (62D)"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        cpfUsuario = msgRecebida[24+tamChave:35+tamChave]
        if not validaCPF(cpfUsuario):
            resposta = "62C"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        if operacao in ['I', 'A']:
            email = msgRecebida[35+tamChave:135+tamChave].strip()
            orgao = msgRecebida[-4:].strip()
            if not verificaEMail(email) or email.upper()[-11:]!="@RFB.GOV.BR":
                resposta = "62M"
                enviaResposta(resposta, c) 
                conn.close()
                return    
            if not orgao.isdigit():  
                resposta = "62O"
                enviaResposta(resposta, c) 
                conn.close()
                return  
            orgao = int(orgao)                     
            tipo, orgaoDesc = buscaTipoOrgao(orgao, cursor)
            if tipo=='I': #órgão inválido ou inexistente (0 = órgão local; tipos [L, R, N])
                resposta = "62O"
                enviaResposta(resposta, c) 
                conn.close()
                return  
            consulta = "Select CPF from Usuarios Where email=%s"
            cursor.execute(consulta, (email, ))
            linha = cursor.fetchone()
            if linha:
                if linha[0]!=cpfUsuario:
                    resposta = "62M"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return                                    
        consulta = "Select Codigo from Usuarios Where CPF=%s"
        cursor.execute(consulta, (cpfUsuario, ))
        linha = cursor.fetchone()
        if (linha==None and operacao in ['A', 'E']) or (linha!=None and operacao=='I'):
            resposta = "62I"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if linha:
            chaveUsuario = linha[0]
        if operacao=='E':
            cursor.execute("Update Usuarios Set Saida=%s, Ativo='N', Orgao=0, Chave=0, ValidadeChave=Null, email=Null Where Codigo=%s", (datetime.now(), chaveUsuario))  
        elif operacao=='A':
            cursor.execute("Update Usuarios Set Ativo='S', email=%s, Orgao=%s Where codigo=%s", (email, orgao, chaveUsuario))         
        else:
            cursor.execute("Insert Into Usuarios (CPF, email, Orgao) Values (%s, %s, %s)", (cpfUsuario, email, orgao))
        try:
            conn.commit()
            resposta = "62S"
            if operacao in ["I", "A"]: #enviamos os dados do órgão na inclusão e alteração em caso de sucesso
                resposta+=tipo+orgaoDesc.ljust(25)
        except:
            conn.rollback()
            resposta = "62E"
        enviaResposta(resposta, c) 
        conn.close()
        return  

    if codigo==63: #Consulta órgãos e jurisdições
        nRegistros = 25 #número de registros a ser enviado por vez (se alterar aqui, alterar script e vice-versa)
        if len(msgRecebida)!=(23+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (63A)"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        if cadastrador!='S':
            resposta = "63N"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        senha = msgRecebida[-10:]
        #horaSenha = str(int(int(datetime.now().strftime('%H%M'))/10)).rjust(3,"0")
        if senha[-10:-3]!=SENHADCCS or SENHADCCS==None:
            resposta = "63N"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        cursor.execute("Select Count(Codigo) From Orgaos")
        total = cursor.fetchone()[0] 
        if total==0 or total==None:
            resposta = "630000"
            enviaRespostaSemFechar(resposta, c)
        else:
            if total>9999:
                total = 9999
            nnnn = str(total).rjust(4,"0")
            cursor.execute("Select Codigo, Orgao, Tipo From Orgaos")
            linhas = cursor.fetchall()
            resposta = "63"+nnnn
            registro = ""
            i = 0
            for linha in linhas:
                i+=1
                codigoOrgao = str(linha[0]).rjust(4, "0")
                orgao = linha[1].ljust(25)
                tipoOrgao = linha[2]
                registro += codigoOrgao+orgao+tipoOrgao
                if i%nRegistros==0 or i==total:
                    resposta += registro
                    enviaRespostaSemFechar(resposta, c)
                    registro = ""
                    resposta = "63"
                    try:
                        mensagemRec = c.recv(256)
                        requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                        if requisicao!="6312345678909":
                            resposta = "99REQUISIÇÃO INVÁLIDA (63B)"
                            enviaResposta(resposta, c) 
                            conn.close()
                            return
                    except:
                        c.close()
                        conn.close()
                        logging.info("Erro de time out 63A - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                        return 
                    if i==total:
                        break
        cursor.execute("Select Distinctrow Grupo From TDPFS Where Encerramento Is Null and Grupo Not in (Select Distinctrow Equipe from Jurisdicao) Order By Grupo") 
        linhasESJ = cursor.fetchall() #equipes sem entradas na tabela de Jurisdição
        if not linhasESJ:
            totalESJur = 0
        else:
            totalESJur = len(linhasESJ)
        cursor.execute("Select Orgao, Equipe From Jurisdicao Order By Equipe")
        linhasEJur = cursor.fetchall() #equipes da tabela de Jurisdicao
        if not linhasEJur:
            totalEJur = 0
        else:
            totalEJur = len(linhasEJur)  
        if totalEJur>9999:
            totalEJur = 9999     
            totalESJur = 0 #só podemos enviar 9999 registros; priorizamos as equipes jurisdicionadas
        elif (totalEJur+totalESJur)>9999:
            totalESJur = 9999 - totalEJur 
        total = totalEJur + totalESJur
        if total==0:
            resposta = "630000"
            enviaResposta(resposta, c)
            conn.close()
            return
        nnnn = str(total).rjust(4, "0")
        i = 0
        registro = ""
        resposta = "63"+nnnn
        consultaEquipe = "Select Nome From Equipes Where Equipe=%s"        
        for linha in linhasEJur:
            i+=1
            orgao = str(linha[0]).rjust(4, "0")
            equipe = linha[1].strip()
            cursor.execute(consultaEquipe, (equipe,))
            rowEquipe = cursor.fetchone()
            if rowEquipe:
                nomeEquipe = rowEquipe[0][:50]
            else:
                nomeEquipe = " "
            equipe = equipe.ljust(25)
            nomeEquipe = nomeEquipe.ljust(50)           
            registro += orgao+equipe+nomeEquipe
            if i%nRegistros==0 or i==total:
                resposta += registro
                if i<total:
                    enviaRespostaSemFechar(resposta, c)
                    registro = ""
                    resposta = "63"
                else:
                    enviaResposta(resposta, c)
                    conn.close()
                    return
                try:
                    mensagemRec = c.recv(256)
                    requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                    if requisicao!="6312345678909":
                        resposta = "99REQUISIÇÃO INVÁLIDA (63C)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 63B - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return  
        orgao = "0000"                
        for linha in linhasESJ:
            i+=1
            equipe = linha[0].strip()
            cursor.execute(consultaEquipe, (equipe,))
            rowEquipe = cursor.fetchone()
            if rowEquipe:
                nomeEquipe = rowEquipe[0][:50]
            else:
                nomeEquipe = " "
            equipe = equipe.ljust(25)
            nomeEquipe = nomeEquipe.ljust(50)  
            registro += orgao+equipe+nomeEquipe
            if i%nRegistros==0 or i==total:
                resposta += registro
                if i<total:
                    enviaRespostaSemFechar(resposta, c)
                    registro = ""
                    resposta = "63"
                else:
                    enviaResposta(resposta, c)
                    conn.close()
                    return
                try:
                    mensagemRec = c.recv(256)
                    requisicao, chaveCriptoAES = descriptografa(mensagemRec, addr, c)
                    if requisicao!="6312345678909":
                        resposta = "99REQUISIÇÃO INVÁLIDA (63D)"
                        enviaResposta(resposta, c) 
                        conn.close()
                        return
                except:
                    c.close()
                    conn.close()
                    logging.info("Erro de time out 63C - provavelmente cliente não respondeu no prazo. Abandonando operação.")
                    return                            

    if codigo==64: #inclui/exclui/altera órgão
        if not len(msgRecebida) in [(50+tamChave), (54+tamChave), (28+tamChave)]: #I, A, E
            resposta = "99REQUISIÇÃO INVÁLIDA (64A)"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        if cadastrador!='S':
            resposta = "64N"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        senha = msgRecebida[13+tamChave:23+tamChave]
        #horaSenha = str(int(int(datetime.now().strftime('%H%M'))/10)).rjust(3,"0")
        if senha[-10:-3]!=SENHADCCS or SENHADCCS==None: 
            resposta = "64N"
            enviaResposta(resposta, c) 
            conn.close()
            return
        operacao = msgRecebida[23+tamChave:24+tamChave]
        if not operacao in ["I", "A", "E"]:
            resposta = "99REQUISIÇÃO INVÁLIDA (64B)"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        if operacao=="A":
            chaveOrgao = msgRecebida[24+tamChave:-26].strip()           
        elif operacao=="E": #exclusão
            chaveOrgao = msgRecebida[-4:].strip()  
        if operacao in ["A", "E"]:
            if not chaveOrgao.isdigit():
                resposta = "64I"
                enviaResposta(resposta, c) 
                conn.close()
                return 
            else:
                chaveOrgao = int(chaveOrgao)                      
        if operacao in ["I", "A"]:
            if operacao=="I":                
                orgao = msgRecebida[24+tamChave:-1].strip()
            else:
                orgao = msgRecebida[-26:-1].strip()
            if len(orgao)<4:
                resposta = "64I"
                enviaResposta(resposta, c) 
                conn.close()
                return
            orgao = orgao.upper()  
            consulta = "Select Codigo from Orgaos Where Orgao=%s"  
            cursor.execute(consulta, (orgao, ))   
            linha = cursor.fetchone()
            if linha: #descrição do órgão a ser incluído não pode existir, nem pode haver alteração para um nome já existente 
                if operacao=="I" or linha[0]!=chaveOrgao: 
                    resposta = "64X"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return                 
            tipo = msgRecebida[-1:].upper()
        if operacao in ["I", "A"]:
            if tipo not in ["L", "R"]: #não é permitida a alteração ou inclusão de órgão nacional (N) (só via MySQL) e não há outros além destes (L, R, N)
                resposta = "64T"
                enviaResposta(resposta, c) 
                conn.close()
                return 
        if operacao in ['A', "E"]:
            consulta = "Select Codigo, Tipo from Orgaos Where Codigo=%s"  
            cursor.execute(consulta, (chaveOrgao, ))   
            linha = cursor.fetchone()
            if not linha:
                resposta = "64I"
                enviaResposta(resposta, c) 
                conn.close()
                return 
            tipoPrevio = linha[1]
            if tipoPrevio=='N':
                resposta = "64X"
                enviaResposta(resposta, c) 
                conn.close()
                return                 
        if operacao=="E":
            cursor.execute("Delete From Orgaos Where Codigo=%s", (chaveOrgao,))
            cursor.execute("Update Usuarios Set Orgao=0 Where Orgao=%s", (chaveOrgao,)) #usuário do órgão excluído ficam locais
            cursor.execute("Delete From Jurisdicao Where Orgao=%s", (chaveOrgao,)) #apagamos todas as jurisdições do órgão excluído
        elif operacao=="A":
            cursor.execute("Update Orgaos Set Tipo=%s, Orgao=%s Where Codigo=%s", (tipo, orgao, chaveOrgao))
            if tipoPrevio=="R" and tipo=='L': #mudou de regional para local - apagamos as jurisdições dele
                cursor.execute("Delete From Jurisdicao Where Orgao=%s", (chaveOrgao,))
        else: #inclusão
            cursor.execute("Insert Into Orgaos (Orgao, Tipo) Values (%s, %s)", (orgao, tipo))
        try:
            conn.commit()
            resposta = "64S"
        except:
            conn.rollback()
            resposta = "64E"
        enviaResposta(resposta, c) 
        conn.close()
        return             
        
    if codigo==65: #Inclui ou exclui equipe na jurisdição de um órgão
        if len(msgRecebida)!=(53+tamChave):
            resposta = "99REQUISIÇÃO INVÁLIDA (65A)"
            enviaResposta(resposta, c) 
            conn.close()
            return          
        if cadastrador!='S':
            resposta = "65N"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        senha = msgRecebida[13+tamChave:23+tamChave]
        #horaSenha = str(int(int(datetime.now().strftime('%H%M'))/10)).rjust(3,"0")
        if senha[-10:-3]!=SENHADCCS or SENHADCCS==None: 
            resposta = "65N"
            enviaResposta(resposta, c) 
            conn.close()
            return 
        operacao = msgRecebida[23+tamChave:24+tamChave]
        if not operacao in ["I", "E"]:
            resposta = "99REQUISIÇÃO INVÁLIDA (65B)"
            enviaResposta(resposta, c) 
            conn.close()
            return    
        chaveOrgao = msgRecebida[-29:-25].strip()
        if not chaveOrgao.isdigit():
            resposta = "65O"
            enviaResposta(resposta, c) 
            conn.close()
            return  
        chaveOrgao = int(chaveOrgao)             
        equipe = msgRecebida[-25:].strip()
        if len(equipe)<13 or not equipe.isdigit():
            resposta = "65Q"
            enviaResposta(resposta, c) 
            conn.close()
            return   
        consulta = "Select Orgao, Tipo From Orgaos Where Codigo=%s"        
        cursor.execute(consulta, (chaveOrgao,))
        linha = cursor.fetchone()
        if not linha:
            resposta = "65O"
            enviaResposta(resposta, c) 
            conn.close()
            return
        tipoOrgao = linha[1] 
        if tipoOrgao in ["L", "N"] and operacao=="I": #orgão local (usuário é vinculado apenas aos seus TDPFS) não possui jurisdição e não há necessidade de incluir uma para órgão nacional
            resposta = "65T"
            enviaResposta(resposta, c) 
            conn.close()
            return                    
        consulta = "Select Codigo From Jurisdicao Where Orgao=%s and Equipe=%s"     
        cursor.execute(consulta, (chaveOrgao, equipe))
        linha = cursor.fetchone()
        if linha:
            if operacao=="I": #equipe a ser incluída já existe na jurisdição do órgão
                resposta = "65X"
                enviaResposta(resposta, c) 
                conn.close()
                return   
            else:
                cursor.execute("Delete From Jurisdicao Where Orgao=%s and Equipe=%s", (chaveOrgao, equipe))
        else:
            if operacao=="E": #equipe a ser excluída NÃO existe na jurisdição do órgão
                resposta = "65X"
                enviaResposta(resposta, c) 
                conn.close()
                return  
            else:
                #temos que ver se a equipe a ser incluída existe em algum TDPF em andamento
                consulta = "Select Numero From TDPFS Where Grupo=%s and Encerramento Is Null"
                cursor.execute(consulta, (equipe,))
                linha = cursor.fetchone()
                if not linha: #equipe não existe em TDPF em andamento - não pode ser incluída na jurisdição de um órgão
                    resposta = "65Q"
                    enviaResposta(resposta, c) 
                    conn.close()
                    return                      
                cursor.execute("Insert Into Jurisdicao (Orgao, Equipe) Values (%s, %s)", (chaveOrgao, equipe))  
        try:
            conn.commit()
            resposta = "65S"
        except:
            conn.rollback()
            resposta = "65E"
        enviaResposta(resposta, c) 
        conn.close()
        return             
                      
    return #não chega aqui, mas ...
      
def espera34(n, c, conn, addr):
    try:
        mensagemRec = c.recv(512)
        requisicao = descriptografa(mensagemRec, addr, c)
        if requisicao!="3412345678909":
            resposta = "99REQUISIÇÃO INVÁLIDA (34B)"
            enviaResposta(resposta, c) 
            conn.close()
            return False
        else:
            return True
    except:
        c.close()
        conn.close()
        logging.info("Erro de time out 34 "+n+" - provavelmente cliente não respondeu no prazo. Abandonando operação.")
        return False    

def meta(cursor, chaveFiscal, ano, trimestreInicial, trimestreFinal): #retorna a meta acumulada entre o trimestre inicial e o final e tb toda a meta entre o trimestre inicial e o restante do ano
    pontosPeriodo = set()
    pontos = None
    consultaPontos = "Select Pontuacao, Ano, Trimestre From Metas Where Fiscal=%s and Ano=%s and Trimestre>=%s and Trimestre<=%s Order By DataMetas DESC, Ano, Trimestre" #computamos as metas mais recentemente cadastradas
    cursor.execute(consultaPontos, (chaveFiscal, ano, trimestreInicial, trimestreFinal))
    rowPontos = cursor.fetchall()
    if rowPontos:
        if len(rowPontos)>0:
            pontos = 0
            for row in rowPontos:
                anoCons = row[1]
                trimestre = row[2]
                if not (anoCons, trimestre) in pontosPeriodo: #para não repetir o cômputo da pontuação na meta do período
                    pontos+=float(row[0])
                    pontosPeriodo.add((anoCons, trimestre))
    metaAnual = None                  
    if trimestreFinal<4 or trimestreInicial>1: #pesquisamos toda a meta do ano
        consultaPontos = "Select Sum(Pontuacao) From Metas Where Fiscal=%s and Ano=%s" #computamos as metas mais recentemente cadastradas
        cursor.execute(consultaPontos, (chaveFiscal, ano))
        rowPontos = cursor.fetchone()
        if rowPontos:
            metaAnual=rowPontos[0]          
        if not metaAnual:
            metaAnual = float(0)
        else:
            metaAnual = float(metaAnual)
    #a meta foi calculada até do 1º ao 4o trimestre
    else:
        metaAnual = pontos
    return pontos, metaAnual #pontos aqui é a meta do período


def consultaMediaPontosEquipe(cursor, equipe, dataInicial, dataFinal): #retorna a média de pontos de uma equipe num certo período
    #somamos os pontos dos TDPFs da equipe encerrados no período e que tenham alguma quantidade de horas alocadas no RHAF
    comando = """Select Sum(Pontos), Sum(PontosMalha)
                 from TDPFS, Alocacoes
                 Where Grupo=%s and Encerramento Is Not Null and (SemExame!='S' or SemExame Is Null) and Encerramento>=%s and Encerramento<=%s and 
                 Alocacoes.TDPF=TDPFS.Codigo and Alocacoes.Horas Is Not Null and Alocacoes.Horas>0 and TDPFS.Pontos Is Not Null and TDPFS.Pontos>0
                 and TDPFS.TDPFPrincipal Is Null"""
    cursor.execute(comando, (equipe, dataInicial, dataFinal))
    tdpfsEquipe = cursor.fetchone()
    if tdpfsEquipe:
        total = tdpfsEquipe[0]+tdpfsEquipe[1]
        if total in [0, None]:
            return 0
    else:
        return 0

    #obtemos a quantidade de fiscais que foram alocados na equipe em algum momento no período
    comando = """Select Count(Distinct Alocacoes.Fiscal) 
                From Alocacoes, TDPFS 
                Where Alocacoes.TDPF=TDPFS.Codigo and TDPFS.Grupo=%s and Alocacoes.Horas Is Not Null and Alocacoes.Horas>0 and
                ((Encerramento Is Null and TDPFS.Emissao<=%s) or (Encerramento Is Not Null and Encerramento>=%s and Encerramento<=%s))"""
    cursor.execute(comando, (equipe, dataFinal, dataInicial, dataFinal))
    reg = cursor.fetchone()
    if reg:
        totalFiscais = reg[0]
        if totalFiscais==None:
            return 0
    else:
        return 0
    #print(total, totalFiscais)
    if totalFiscais!=0:
        mediaEquipe = int(total / totalFiscais)
    else:
        mediaEquipe = 0   
    return mediaEquipe

def calculaPontosFiscal(cursor, chaveFiscal, dataInicial, dataFinal): #calcula os pontos do fiscal no período
    #seleciona os TDPFs em que o usuário esteja alocado
    comando = """Select TDPFS.Codigo, TDPFS.Grupo, TDPFS.Pontos
                    From TDPFS, Alocacoes, Fiscais
                    Where Alocacoes.Fiscal=%s and Alocacoes.TDPF=TDPFS.Codigo and Alocacoes.Fiscal=Fiscais.Codigo and
                    Encerramento Is Not Null and Encerramento>=%s and Encerramento<=%s and (SemExame!='S' or SemExame Is Null) and TDPFS.Pontos Is Not Null 
                    and (TDPFS.TDPFPrincipal Is Null or TDPFS.Tipo!='D')
                    Order by TDPFS.Grupo, Fiscais.CPF"""  
    cursor.execute(comando, (chaveFiscal, dataInicial, dataFinal))
    linhas = cursor.fetchall()
    equipe = ""
    totalFiscal = float(0)     
    equipes = set()       
    for linha in linhas:
        if equipe=="":
            equipe = linha[1]      
        if equipe!=linha[1]:              
            if not equipe in equipes:
                equipes.add(equipe) #temos que informar as equipes que o fiscal faz parte 
            equipe = linha[1]
        chaveTdpf = linha[0]
        pontos = linha[2]
        if pontos in [0, None]:
            continue
        comando = "Select Sum(Alocacoes.Horas) from Alocacoes Where Alocacoes.TDPF=%s"
        cursor.execute(comando, (chaveTdpf,))     #busca as horas totais alocadas ao TDPF
        reg = cursor.fetchone()
        if reg:
            horas = reg[0]   
            if horas in [0, None]: #horas RHAF zeradas - vai para o próximo
                continue 
        else:
            continue #não vai ter como calcular os pontos do fiscal, pois não há horas informadas no RHAF para o TDPF
        comando =  "Select Alocacoes.Horas from Alocacoes Where Alocacoes.TDPF=%s and Alocacoes.Fiscal=%s"   
        cursor.execute(comando, (chaveTdpf, chaveFiscal))     #busca as horas alocadas ao TDPF pelo fiscal
        reg = cursor.fetchone()        
        if reg:
            horasFiscal = reg[0]    
            if horasFiscal in [None, 0]:
                continue
        else:
            continue #fiscal não tem pontos neste TDPF, pois não possui horas alocadas  
        totalFiscal +=  float(pontos) * (float(horasFiscal)/float(horas))   #soma os pontos do TDPF 'devidos' ao fiscal ao total de pontos
    #adicionamos os dados da última equipe do fiscal               
    if not equipe in equipes and equipe!="":
        equipes.add(equipe) #mais uma equipe que o fiscal faz parte
    #retornamos também os pontos de malha que o fiscal fez
    selectMalha = "Select Count(Malha.Codigo), InfoMalha.Valor From Malha, InfoMalha Where Malha.Tipo=InfoMalha.Codigo and Malha.Fiscal=%s and Malha.Data>=%s and Malha.Data<=%s Group By InfoMalha.Valor"
    cursor.execute(selectMalha, (chaveFiscal, dataInicial, dataFinal))
    resultado = cursor.fetchall()
    pontosMalha = 0
    for linha in resultado:
        pontosMalha += float(linha[0])*float(linha[1])
    return totalFiscal, pontosMalha, equipes

def disparaMediaPontos(): #nas datas abaixo (dia 25 do mês posterior ao final do trimestre) manda e-mails para os fiscais informando suas pontuações
    global ambiente
    start = time.time()
    dictMesTrimestre = {1:1, 2:1, 3:1, 4:2, 5:2, 6:2, 7:3, 8:3, 9:3, 10:4, 11:4, 12:4}
    diaMes = datetime.now().strftime("%d/%m")
    #if not diaMes in ["25/04", "25/07", "25/10", "25/01"]: #só enviamos os e-mails nestas datas
    #    return
    conn = conecta() 
    if conn==None:      
        print("Erro de conexão - disparaMediaPontos()")
        logging.info("Erro ao conectar ao BD para enviar e-mails com média de pontos.")  
        return        
    ano = datetime.now().year
    mes = datetime.now().month - 1
    if mes<=0:
        ano -= 1
        mes += 12
    dataFim = datetime(ano, mes, calendar.monthrange(ano, mes)[1]) #último dia do último mês do trimestre anterior
    dataIni = datetime(ano, 1, 1)
    trimestre = dictMesTrimestre[mes]
    cursor = conn.cursor(buffered=True)
    dictMediaEquipes = dict()  
    print("Calculando as médias das equipes e regiões ...")    
    cursor.execute("Select Sum(Pontos), Sum(PontosMalha), Count(Distinct Fiscal), PontosMetas.Equipe, Equipes.Equipe, Equipes.Nome From PontosMetas, Equipes Where Ano=%s and Trimestre<=%s and PontosMetas.Equipe=Equipes.Codigo Group By PontosMetas.Equipe Order By PontosMetas.Equipe", (ano, trimestre))
    linhas = cursor.fetchall()
    for linha in linhas: #média das equipes
        pontos = linha[0]+linha[1]
        fiscais = linha[2]
        chaveEquipe = linha[3]
        equipe = linha[4].strip()
        nomeEquipe = linha[5].strip()
        if fiscais>0:
            dictMediaEquipes[chaveEquipe] = [round(pontos/fiscais, 2), equipe, nomeEquipe]
    dictRegioes = dict() #guarda média de cada equipe e número delas de uma região fiscal
    for chaveEquipe in dictMediaEquipes: #média das regiões
        media = dictMediaEquipes[chaveEquipe][0]
        rf = dictMediaEquipes[chaveEquipe][1][:2]
        if dictRegioes.get(rf, -1)==-1:
            mediaEquipes = media
            nEquipes = 1
        else:
            mediaEquipes = dictRegioes[rf][0]+media
            nEquipes = dictRegioes[rf][1]+1
        dictRegioes[rf] = [mediaEquipes, nEquipes]
    
    nEquipes = 0
    mediaEquipes = 0
    for rf in dictRegioes: #calcula a média nacional
        nEquipes += dictRegioes[rf][1]
        mediaEquipes += dictRegioes[rf][0]
    if nEquipes>0:
        mediaNacional = round(mediaEquipes/nEquipes,2)
    else:
        mediaNacional = 0
    print("Média das regiões e nacional calculadas em "+str(time.time()-start)[:8]+" segundos") 
    print("Selecionando fiscais e enviando e-mails ...")
    comando = """Select Distinctrow email, Fiscais.Codigo From Usuarios, Fiscais, Alocacoes 
                 Where email!='' and email Is Not Null and Adesao Is Not Null and Saida Is Null and idTelegram!=0 and 
                 Usuarios.CPF=Fiscais.CPF and Alocacoes.Fiscal=Fiscais.Codigo and Alocacoes.Desalocacao Is Null""" #somente usuários ativos que estejam alocados em
                                                                                                                   #pelo menos um TDPF
    cursor.execute(comando)    
    linhas = cursor.fetchall()
    cabecalho = "Sr. Usuário,\n\nEstamos encaminhando algumas informações sobre as pontuações de fiscalizações calculadas conforme Portaria Cofis nº 46/2020.\n\n"
    periodo = "Período de Referência: de "+dataIni.strftime("%d/%m/%Y")+" a "+dataFim.strftime("%d/%m/%Y")+"\n\n"
    rodape = "Atenciosamente,\n\nCofis/Disav"
    if ambiente!="PRODUÇÃO":
        rodape += "\n\nAmbiente: "+ambiente
    for linha in linhas:
        email = linha[0] 
        if email==None:
            continue
        email = email.strip()       
        if email.upper()[-11:]!="@RFB.GOV.BR": #por algum acaso, não é email institucional - não enviamos
            continue
        chaveFiscal = linha[1]
        cursor.execute("Select Sum(Pontos), Sum(PontosMalha), Max(MetaFiscal), MetaAnual, Equipe From PontosMetas Where Fiscal=%s and Ano=%s and Trimestre<=%s", (chaveFiscal, ano, trimestre))
        resultFiscal = cursor.fetchone()
        if not resultFiscal: #fiscal não tem informações de pontos nem metas cadastradas
            continue
        pontosFiscal = resultFiscal[0]+resultFiscal[1]
        metaFiscal = resultFiscal[2]
        metaAnual = resultFiscal[3]
        chaveEquipe = resultFiscal[4]
        equipe = dictMediaEquipes[chaveEquipe][1]
        nomeEquipe = dictMediaEquipes[chaveEquipe][2]
        texto = "Pontos do Fiscal: "+str(round(pontosFiscal, 2))+"\n\n"
        texto += "Meta do Fiscal (acumulada até o final do trimestre): "+str(round(metaFiscal, 2))+"\n\n"
        texto += "Meta Anual do Fiscal: "+str(round(metaAnual, 2))+"\n\n"
        mediaEquipe = dictMediaEquipes[chaveEquipe][0] 
        texto += "Média de Pontos dos Fiscais de sua Equipe - "+formataEquipe(equipe)+" - "+nomeEquipe+": "+str(mediaEquipe)+"\n\n"
        texto += "Média de Pontos dos Fiscais da RF "+equipe[:2]+": "+str(round(dictRegioes[equipe[:2]][0]/dictRegioes[equipe[:2]][1], 2))+"\n\n"
        texto += "Média de Pontos dos Fiscais do País: "+str(mediaNacional)+"\n\n"
        texto = cabecalho+periodo+texto+rodape
        if ambiente=="PRODUÇÃO": #enviamos e-mails somente no ambiente de produção
            if enviaEmail(email, texto, "Informações sobre pontuação do trimestre - Fiscalização")!=3:
                logging.info("Erro ao enviar e-mail com dados sobre pontos - "+email)
        else:
            print(email)
            print(texto)
            print("-------------------------------------------------------------------------------")
    print("Finalizado envio de e-mails ...")
    end = time.time() 
    print("Tempo total decorrido: "+str(end-start)[:8]+" segundos")
    return

def consultaCorreios(): #faz a atualização da situação nos correios dos termos postados    
    url = "https://www2.correios.com.br/sistemas/rastreamento/ctrl/ctrlRastreamento.cfm?"
    comando = """Select ControlePostal.Codigo, CodRastreamento, SituacaoAtual, DataSituacao
                 from ControlePostal, TDPFS
                 Where TDPFS.Encerramento Is Null and ControlePostal.TDPF=TDPFS.Codigo and DataEnvio>cast((now() - interval 75 day) as date)
                 and (Upper(SituacaoAtual) Not Like '%ENTREGUE%' or SituacaoAtual Is Null) and DataRecebimento Is Null""" 
                 #consultamos a situação somente dos enviados há até 75 dias e de TDPFs em andamento, sem que o AR ou correspondência tenha sido recebido(a)
    conn = conecta() 
    if conn==None:      
        logging.info("Erro ao conectar ao BD para consultar os correios.")  
        return 
    cursor = conn.cursor(buffered=True)
    cursor.execute(comando)    
    linhas = cursor.fetchall()
    i = 0
    f = open("RastreamentoResultado.txt", "w")
    tentativas = 0
    atualizou = False
    nAtualizados = 0
    print("Iniciando a atualização do controle postal. Total de registros a serem verificados: ", len(linhas))
    while i<len(linhas):
        qtdeRastreamento = 0
        rastreamento = ""
        while qtdeRastreamento<50 and (i+qtdeRastreamento)<len(linhas): #podemos consultar o site dos correios de 50 em 50 códigos, no máximo
            linha = linhas[i+qtdeRastreamento]
            if qtdeRastreamento>=1:
                rastreamento += ";"
            rastreamento += linha[1].strip()
            chave = linha[0]
            qtdeRastreamento+=1
        payload = {"acao": "tracks", "objetos": rastreamento, "btnPesq": "Buscar"} #consulta em lote (no máximo 50 códigos)
        #payload = {"Form Data": "", "acao": "track", "objetos": rastreamento, "btnPesq": "Buscar"} #consulta individual
        resultado = requests.post(url, data=payload)
        if resultado==None:
            tentativas+=1      
            if tentativas>20: #fazemos 20 tentativas
                conn.close()
                logging.info("Site dos correios não está respondendo")
                print("Site dos correios não está respondendo")
                f.close()
                return
            print("Site dos correios não está respondendo - Aguardando 2 s e tentando novamente")
            time.sleep(2)
            continue              
        if "TEMPORARIAMENTE INDISPONÍVEL" in resultado.text.upper() or "SERVICE UNAVAILABLE" in resultado.text.upper():
            tentativas+=1
            if tentativas>20: #fazemos 20 tentativas
                conn.close()
                logging.info("Serviço de rastreamento dos correios está fora do ar")
                print("Serviço de rastreamento dos correios está fora do ar")
                f.close()
                return
            print("Serviço de rastreamento dos correios está fora do ar - Aguardando 10 s e tentado novamente")
            time.sleep(10) #se o serviço estiver temporariamente indisponível, esperamos 10 segundos
            continue
        tentativas = 0
        try:
            df_list = pd.read_html(resultado.text) # this parses all the tables in webpages to a list
        except:
            df_list = None
            print("Não obteve a tabela do site dos correios")
        if df_list:  
            df = df_list[0]
            df.to_excel("TabelaRastreamento.xlsx")
            for j in range(qtdeRastreamento):
                rastreamento = df.iat[j, 0]
                if rastreamento==None:
                    rastreamento = "VAZIO"
                rastreamento = rastreamento.replace(" ","")
                situacao = df.iat[j,1]
                dataSituacao = df.iat[j,2]
                if situacao!=None and not "sistema não possui dados sobre o objeto informado" in situacao:
                    situacao = situacao.strip()[:100].strip()
                    if dataSituacao!=None:
                        dataSituacao = dataSituacao.strip()
                    else:
                        dataSituacao = ""
                else:
                    situacao = ""
                    dataSituacao = ""
                if len(dataSituacao)>=10 and len(situacao)>=5: #deve ser uma data seguido de um local
                    try:
                        f.write(rastreamento+" - "+dataSituacao[:10]+" - "+situacao+"\n")                          
                        dataSituacao = datetime.strptime(dataSituacao[:10], "%d/%m/%Y") 
                        linha = linhas[i+j]
                        if linha[1].strip()==rastreamento: #testamos para ver se os registros estão correspondendo
                            if situacao!=linha[2] or dataSituacao>linha[3]: #só atualizamos se a situação foi alterada ou a data é posterior
                                chave = linha[0]    
                                cursor.execute("Update ControlePostal Set SituacaoAtual=%s, DataSituacao=%s Where Codigo=%s", (situacao, dataSituacao, chave))   
                                atualizou = True        
                                nAtualizados+=1         
                    except:
                        print(rastreamento, "Falhou a conversão da data (2)")
                        logging.info("Falhou a conversão da data (2) - "+rastreamento)
                else:
                    print(rastreamento, "Data ou situação inválida ou inexistente")                    
                    logging.info("Data ou situação inválida ou inexistente - "+rastreamento)                        
        else:
            print("Não há informação de rastreamento para as linhas de "+str(i)+" em diante ...")
            logging.info("Não há informação de rastreamento para as linhas de "+str(i)+" em diante ...")
        i+=50  
    if atualizou:
        try:
            conn.commit()
            print("Registros postais atualizados: "+str(nAtualizados))
        except:
            conn.rollback()
            logging.info("Falhou a atualizão do controle postal")
            print("Falhou a atualização do controle postal")    
    else:
        print("Não houve atualização do controle postal")     
    f.close()
    conn.close()
    return

def alocaFiscaisEquipesTrimestre(ano, trimestre):  #determina a que equipe os fiscais devem ser considerados num determinado trimestre e salva na tabela FiscaisEquipes
    conn = conecta() 
    if conn==None:      
        logging.info("Erro ao conectar ao BD para fazer as alocacões dos fiscais a equipes.")  
        print("Erro ao conectar ao BD para fazer as alocacões dos fiscais a equipes.") 
        return 
    dictMesTrimestre = {1:1, 2:1, 3:1, 4:2, 5:2, 6:2, 7:3, 8:3, 9:3, 10:4, 11:4, 12:4}
    anoTrimestre = str(ano)+str(trimestre)
    cursor = conn.cursor(buffered=True)    
    cursor.execute("Select Max(AnoTrimestre) From FiscaisEquipes")
    linha = cursor.fetchone()
    if linha[0]!=None:
        if linha[0]>=anoTrimestre: #o trimestre já foi atualizado
            conn.close()
            print("Trimestre "+anoTrimestre[:4]+"/"+anoTrimestre[-1:]+" já foi atualizado na tabela FiscaisEquipes")
            print("-------------------------------------------------------------------------------")
            return
    start = time.time()
    dictTrimestreMes = {1: [1, 3], 2:[4, 6], 3:[7, 9], 4:[10, 12]}
    dataInicioTrimestre = datetime(ano, dictTrimestreMes[trimestre][0], 1)
    mesFinal = dictTrimestreMes[trimestre][1]
    diaFinal = calendar.monthrange(ano, mesFinal)[1]
    dataFimTrimestre = datetime(ano, mesFinal, diaFinal)
    total = 0
    cursor.execute("Select Distinctrow Codigo, Nome, CPF From Fiscais") #só selecionamos fiscais com vinculo ativo no período
    linhasFiscais = cursor.fetchall()

    consultaVinculoEfetivo = """Select Vinculos.Equipe, Inicio, Vinculo From Vinculos, Equipes, TipoEquipes                    
                                Where Fiscal=%s and (Vinculo='EFETIVO' or Vinculo='SUPERVISÃO') and Inicio<=%s and (Fim Is Null or Fim>%s) 
                                and Equipes.Codigo=Vinculos.Equipe and Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo and 
                                (TipoEquipes.Descricao Like 'EXECUÇÃO FISC%' or TipoEquipes.Descricao Like 'EXECUÇÃO REVIS%')
                                Order By Vinculo ASC, Inicio ASC""" #para procurar a equipe efetiva no início do trimestre e depois a de supervisão no início
                                
    consultaVinculoEfetivoRecente = """Select Vinculos.Equipe, Inicio From Vinculos, Equipes, TipoEquipes                    
                                       Where Fiscal=%s and Vinculo='EFETIVO' and Fim Is Null  
                                       and Equipes.Codigo=Vinculos.Equipe and Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo and 
                                       (TipoEquipes.Descricao Like 'EXECUÇÃO FISC%' or TipoEquipes.Descricao Like 'EXECUÇÃO REVIS%')
                                       Order By Inicio DESC"""   #para procurar a equipe efetiva mais recente  

    #se a equipe do vínculo não funcionar, verificamos a equipe na data de cadastro das metas
    consultaDataMetas = "Select DataMetas From Metas Where Fiscal=%s and Ano=%s and Trimestre=%s Order By DataMetas DESC" #meta cadastrada mais recentemente

    #SE TIVER METAS, vamos para a regra do RPF mais recentemente distribuído, desde que tenha sido emitido antes do último dia do trimestre de referência e estivesse em andamento nele
    consultaVinculoRPF = """Select Distinctrow Equipes.Codigo, TDPFS.Emissao, Alocacoes.Alocacao From TDPFS, Alocacoes, Equipes, TipoEquipes    
                            Where Alocacoes.Fiscal=%s and Alocacoes.TDPF=TDPFS.Codigo and (Alocacoes.Desalocacao Is Null or Alocacoes.Desalocacao<%s) and
                            (TDPFS.Encerramento>=%s or TDPFS.Encerramento Is Null) and TDPFS.Emissao<=%s and TDPFS.Grupo=Equipes.Equipe and 
                            Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo and 
                            (TipoEquipes.Descricao Like 'EXECUÇÃO FISC%' or TipoEquipes.Descricao Like 'EXECUÇÃO REVIS%')
                            Order By Alocacoes.Alocacao DESC, TDPFS.Emissao DESC
                            """ #Order By Alocacoes.Alocacao DESC (?)

    #SE TIVER METAS, depois equipe do RPF mais recentemente distribuído (em qualquer período)
    consultaVinculoRPFQualquer = """Select Distinctrow Equipes.Codigo, TDPFS.Emissao, Alocacoes.Alocacao From TDPFS, Alocacoes, Equipes, TipoEquipes    
                                    Where Alocacoes.Fiscal=%s and Alocacoes.TDPF=TDPFS.Codigo and TDPFS.Grupo=Equipes.Equipe and 
                                    Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo and 
                                    (TipoEquipes.Descricao Like 'EXECUÇÃO FISC%' or TipoEquipes.Descricao Like 'EXECUÇÃO REVIS%')
                                    Order By Alocacoes.Alocacao DESC, TDPFS.Emissao DESC"""   

    #MESMO SEM METAS E SEM VÍNCULOS, pegamos a equipe do RPF mais recentmente distribuído para o fiscal se ele tiver sido encerrado no trimestre com alguma hora alocada no RHAF
    consultaVinculoRPFEncerrado = """Select Distinctrow Equipes.Codigo, TDPFS.Emissao, Alocacoes.Alocacao From TDPFS, Alocacoes, Equipes, TipoEquipes    
                                     Where Alocacoes.Fiscal=%s and Alocacoes.Horas>0 and Alocacoes.TDPF=TDPFS.Codigo and TDPFS.Grupo=Equipes.Equipe and 
                                     Encerramento Is Not Null and Encerramento>=%s and Encerramento<=%s and
                                     Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo and 
                                     (TipoEquipes.Descricao Like 'EXECUÇÃO FISC%' or TipoEquipes.Descricao Like 'EXECUÇÃO REVIS%')
                                     Order By Alocacoes.Alocacao DESC, TDPFS.Emissao DESC"""                                                           
    totalNE = 0
    for linhaFiscal in linhasFiscais:
        regra = ""
        fiscal = linhaFiscal[0]
        #nomeFiscal = linhaFiscal[1]
        #cpfFiscal = linhaFiscal[2]
        equipe = None
        inicio = None
        cursor.execute(consultaVinculoEfetivo, (fiscal, dataFimTrimestre, dataInicioTrimestre))
        linhasEquipes = cursor.fetchall()
        i = 0
        equipeAnt = None
        tipoAnt = None
        for linhaEquipe in linhasEquipes:
            inicio = linhaEquipe[1]
            tipo = linhaEquipe[2]            
            if inicio>dataInicioTrimestre:
                if i>0:
                    equipe = equipeAnt #pegamos a equipe ativa no primeiro dia do trimestre
                    tipo = tipoAnt
                else:
                    equipe = linhaEquipe[0] #se não houver aquela, pegamos a primeira equipe ativa no trimestre
                break
            elif tipoAnt!=None and tipo!=tipoAnt: #mudou o tipo de vínculo, então pegamos a equipe anterior de vínculo efetivo (prioridade sob supervisão)
                equipe = equipeAnt
                tipo = tipoAnt
                break
            equipeAnt = linhaEquipe[0]
            tipoAnt = tipo                            
            i+=1
        if equipe==None and equipeAnt!=None: #não achou nenhuma equipe, mas achou o fiscal em equipe ativa, mas com início em data anterior ao início do trimestre
            #pegamos a equipe da última linha
            equipe = equipeAnt
            tipo = tipoAnt
        if equipe!=None:
            regra = 'VINCULO '+tipo+' - PRIMEIRO NO TRIMESTRE OU ANTERIOR MAIS RECENTE (1)'
        bAchouMetas = False
        if equipe==None: #se não achou pela regra do efetivo, partimos para a próxima - equipe efetiva na data de registro das metas (mais recente)
            cursor.execute(consultaDataMetas, (fiscal, ano, trimestre))
            linhaData = cursor.fetchone()
            if linhaData:
                bAchouMetas = True
                dataMeta = linhaData[0]              
                cursor.execute(consultaVinculoEfetivo, (fiscal, dataMeta, dataMeta))
                linhasEquipes = cursor.fetchall()
                if linhasEquipes:
                    if len(linhasEquipes)>0:
                        linhaEquipe = linhasEquipes[len(linhasEquipes)-1]
                        equipe = linhaEquipe[0]
                        regra = 'EQUIPE EFETIVA NA DATA DE REGISTRO DA META (2)'
            else:
                bAchouMetas = False
        if equipe==None and bAchouMetas: #não achou ainda a equipe e tem metas - exige meta
            #vamos para a regra do RPF mais recentemente distribuído, desde que tenha sido emitido antes do último dia do trimestre de referência e estivesse em andamento nele
            #podemos procurar o RPF distribuído mais recentemente para o fiscal, conforme regra acima em equipes da fiscalização
            cursor.execute(consultaVinculoRPF, (fiscal, dataFimTrimestre, dataInicioTrimestre, dataFimTrimestre))
            rowMeta = cursor.fetchone()
            if rowMeta:
                if rowMeta[0]!=None:
                    equipe = rowMeta[0] 
                    regra = 'RPF MAIS RECENTEMENTE DISTRIBUÍDO NO TRIMESTRE - METAS (3)'
        if equipe==None and bAchouMetas: #regra RESIDUAL 1 - equipe efetiva na data de início de vínculo mais recente - exige meta
            cursor.execute(consultaVinculoEfetivoRecente, (fiscal, ))
            linhaEquipe = cursor.fetchone()
            if linhaEquipe:
                equipe = linhaEquipe[0]           
                regra = 'EQUIPE EFETIVA MAIS RECENTE - METAS (4)' 
        if equipe==None and bAchouMetas: #regra RESIDUAL 2 - equipe do RPF mais recente distribuído para o fiscal - exige meta
            cursor.execute(consultaVinculoRPFQualquer, (fiscal, ))
            rowMeta = cursor.fetchone()
            if rowMeta:
                if rowMeta[0]!=None:
                    equipe = rowMeta[0] 
                    regra = 'RPF MAIS RECENTEMENTE DISTRIBUÍDO - METAS (5)'
        if equipe==None: #regra RESIDUAL 3 - equipe do RPF mais recentemente distribuído para o fiscal que tenha sido encerrado no trimestre (não exige alocação ativa, mas exige horas no RHAF)                    
            cursor.execute(consultaVinculoRPFEncerrado, (fiscal, dataInicioTrimestre, dataFimTrimestre))
            rowMeta = cursor.fetchone()
            if rowMeta:
                if rowMeta[0]!=None:
                    equipe = rowMeta[0] 
                    regra = 'RPF MAIS RECENTEMENTE DISTRIBUÍDO ENCERRADO NO TRIMESTRE (6)'            
        if equipe!=None: #se achou por alguma das regras, atualizamos a lotação ou a inserimos
            total+=1
            cursor.execute("Select Codigo From FiscaisEquipes Where Fiscal=%s and Ano=%s and Trimestre=%s", (fiscal, ano, trimestre))
            linhaLotacao = cursor.fetchone()
            if not linhaLotacao:
                cursor.execute("Insert Into FiscaisEquipes (Fiscal, Equipe, Ano, Trimestre, AnoTrimestre, Processamento, Regra) Values (%s, %s, %s, %s, %s, %s, %s)", (fiscal, equipe, ano, trimestre, anoTrimestre, datetime.now(), regra)) 
            else:
                cursor.execute("Update Fiscais Equipes Set Equipe=%s, Processamento=%s, Regra=%s Where Codigo=%s", (equipe, datetime.now(), regra, linhaLotacao[0]))
    try:  
        conn.commit()
        print("Equipes dos fiscais para fins de meta foram atualizadas - tabela Fiscais Equipes - "+anoTrimestre[:4]+"/"+anoTrimestre[-1:])
        logging.info("Equipes dos fiscais para fins de meta foram atualizadas - tabela Fiscais Equipes - "+anoTrimestre[:4]+"/"+anoTrimestre[-1:])
        print(total, " registros incluídos/atualizados")
    except:
        conn.rollback()
        logging.info("Houve algum erro na atualização das equipes dos fiscais na tabela FiscaisEquipes - "+anoTrimestre[:4]+"/"+anoTrimestre[-1:])
        print("Houve algum erro na atualização das equipes dos fiscais na tabela FiscaisEquipes - "+anoTrimestre[:4]+"/"+anoTrimestre[-1:])
    print("Total de fiscais NÃO encontrados:", totalNE)
    print("Tempo decorrido: ", time.time()-start)
    print("-------------------------------------------------------------------------------")
    return

def rodaAlocacoesPontos(primeiraCarga=False): #faz a carga das alocacoes dos fiscais nas equipes e dos pontos em cada trimestre
    dictMesTrimestre = {1:1, 2:1, 3:1, 4:2, 5:2, 6:2, 7:3, 8:3, 9:3, 10:4, 11:4, 12:4}  
    if not primeiraCarga:
        mesAtual = datetime.now().month
        mesConsulta = mesAtual - 1
        anoConsulta = datetime.now().year
        if mesConsulta<=0:
            mesConsulta += 12
            anoConsulta -= 1    
        trimestreConsulta = dictMesTrimestre[mesConsulta]         
        alocaFiscaisEquipesTrimestre(anoConsulta, trimestreConsulta)                
        calculaPontosMetasTabela(anoConsulta, trimestreConsulta)
    else:
        for ano in range(2020, datetime.now().year+1, 1):
            for trimestre in range(4):
                if trimestre+1>dictMesTrimestre[datetime.now().month] and ano==datetime.now().year:
                    return
                alocaFiscaisEquipesTrimestre(ano, trimestre+1)
                if ano>=2021:
                    calculaPontosMetasTabela(ano, trimestre+1)
    return

def calculaPontosMetasTabela(anoConsulta, trimestreConsulta): #registra em tabela o cálculo do trimestreConsulta - periodicidade semanal- todo sábado (conforme agendamento)
    conn = conecta() 
    if conn==None:      
        logging.info("Erro ao conectar ao BD para fazer os cálculos dos pontos e obtenção das metas do trimestre anterior.")  
        print("Erro ao conectar ao BD para fazer os cálculos dos pontos e obtenção das metas do trimestre anterior.")  
        return
    start = time.time()
    print("Iniciando atualização da tabela PontosMetas (ano trimestre) ", anoConsulta, trimestreConsulta)
    cursor = conn.cursor(buffered=True)
    dictTrimestreMes = {1: [1, 3], 2:[4, 6], 3:[7, 9], 4:[10, 12]}
    dataInicioTrimestre = datetime(anoConsulta, dictTrimestreMes[trimestreConsulta][0], 1)
    mesFinal = dictTrimestreMes[trimestreConsulta][1]
    diaFinal = calendar.monthrange(anoConsulta, mesFinal)[1]
    dataFimTrimestre = datetime(anoConsulta, mesFinal, diaFinal)    
    selecaoEquipesPais = """Select Distinctrow Equipes.Codigo From Equipes, TipoEquipes 
                            Where Equipes.Sistema=6 and Equipes.Tipo=TipoEquipes.Codigo and TipoEquipes.Tipo In (3, 4)"""
    cursor.execute(selecaoEquipesPais,)   
    fiscais = set()      
    dictEquipes = {}     
    equipes = set() 
    consulta = "Select Fiscal, Regra From FiscaisEquipes Where Equipe=%s and Ano=%s and Trimestre=%s"          
    linhas = cursor.fetchall()
    for linha in linhas:
        chaveEquipe = linha[0]
        if not chaveEquipe in equipes:
            equipes.add(chaveEquipe) #os fiscais desta equipe estão sendo considerados
            #selecionamos todos os fiscais daquela equipe para fins de metas, inclusive os que não as tem (tabela vínculos)
            dados = (chaveEquipe, anoConsulta, trimestreConsulta)
            cursor.execute(consulta, dados)
            linhasFiscais = cursor.fetchall()
            for linha in linhasFiscais:
                fiscal = linha[0]
                regra = linha[1]
                if not fiscal in fiscais:
                    fiscais.add(fiscal)
                    dictEquipes[fiscal] = [chaveEquipe, regra]                                                  
    #depois de termos adicionados todos os fiscais e respectivas equipes do país, pesquisamos seus pontos e suas metas
    totalAtu = 0
    totalIns = 0
    for fiscal in fiscais:           
        pontosFiscal, pontosMalha, _ = calculaPontosFiscal(cursor, fiscal, dataInicioTrimestre, dataFimTrimestre)
        #procuramos as metas do fiscal 
        metaFiscal, metaAnual = meta(cursor, fiscal, anoConsulta, trimestreConsulta, trimestreConsulta)  #sempre do primeiro trimestre até o trimestre atual (meta acumulada) 
        if metaFiscal==None and pontosFiscal==0 and pontosMalha==0: #não há interesse neste fiscal
            continue 
        if metaFiscal==None: #se não há metas para o fiscal, as metas são os pontos
            metaFiscal = pontosFiscal+pontosMalha
            metaAnual = pontosFiscal+pontosMalha
        chaveEquipe = dictEquipes[fiscal][0]
        regra = dictEquipes[fiscal][1]
        cursor.execute("Select Codigo From PontosMetas Where Fiscal=%s and Trimestre=%s and Ano=%s", (fiscal, trimestreConsulta, anoConsulta))
        linhaPontosMetas = cursor.fetchone()
        if linhaPontosMetas:
            cursor.execute("Update PontosMetas Set Equipe=%s, Pontos=%s, PontosMalha=%s MetaFiscal=%s, MetaAnual=%s, Regra=%s, Atualizacao=%s Where Codigo=%s", (chaveEquipe, pontosFiscal, pontosMalha, metaFiscal, metaAnual, regra, datetime.now(), linhaPontosMetas[0]))
            totalAtu+=1
        else:
            cursor.execute("Insert Into PontosMetas (Fiscal, Equipe, Pontos, PontosMalha, MetaFiscal, MetaAnual, Trimestre, Ano, Regra, Atualizacao) Values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", (fiscal, chaveEquipe, pontosFiscal, pontosMalha, metaFiscal, metaAnual, trimestreConsulta, anoConsulta, regra, datetime.now()))
            totalIns+=1
    try:
        conn.commit()
        print(totalIns, "registros incluídos")
        print(totalAtu, "registros atualizados")
        logging.info("Tabela PontosMetas atualizada")
    except:
        conn.rollback()
        print("Erro ao atualizar a tabela PontosMetas")
        logging.info("Erro ao atualizar a tabela PontosMetas")
    print("Tempo decorrido na atualização da tabela PontosMetas: "+str(time.time()-start)[:8]+" segundos")
    print("----------------------------------------------------------------------------")
    return

def disparador(): #para disparar a tarefa agendada (schedule)
    logging.info("Disparador (thread) iniciado ...")
    while True:
        schedule.run_pending() 
        logging.info("Disparador (thread) indo 'dormir' às "+datetime.now().strftime("%d/%m/%Y %H:%M"))
        time.sleep(1*60*60) #dorme por 2 h
    return 

def servidor(): 
    global  threads, s, pubKey

    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)          
    logging.info("Socket successfully created")
     
    port = 1352             
      
    # Next bind to the port 
    # we have not typed any ip in the ip field 
    # instead we have inputted an empty string 
    # this makes the server listen to requests  
    # coming from other computers on the network 
    s.bind(('', port))         
    logging.info("socket binded to %s" %(port))
      
    # put the socket into listening mode 
    if ambiente=="TESTE":
        s.listen(5)   
    else:
        s.listen(10)  
    logging.info("socket is listening")
      
    #só fica escutando a rede, pega a mensagem e encaminha para tratamento 
    while True:  
        try:
            c, addr = s.accept()      
            #print(addr)
            #print(type(addr))
            logging.info('Got connection from ' + str(addr))    
            c.settimeout(10)
            try:
                msgRecebida = c.recv(2048)#.decode('utf-8') #recebe a mensagem  (era 1024)
                #logging.info(binascii.hexlify(msgRecebida))
                #logging.info(len(msgRecebida))
    
                threadTrata = threading.Thread(target=trataMsgRecebida, args=(msgRecebida,c, addr[0]))
                threads.append(threadTrata)
                threadTrata.start()   #inicia a thread que vai tratar a requisição       
                #trataMsgRecebida(msgRecebida, c) #monta a resposta e a envia
               
            except:
                c.close()
                logging.info("Time out " +str(addr))
        except:
            logging.info("Socket que estava 'ouvindo' a rede parou de funcionar. Se fechou o programa, tudo bem. Senão ...")
            return
       
       #c.close() #com as threads, vai ter que fechar lá na função (thread)
       

#encryptor = PKCS1_OAEP.new(pubKey) #<-- lado cliente (abaixo exemplo)
#msgCripto = encryptor.encrypt(str.encode("Testando Cripto 1111111111111 00000000000000 44444444444444444 55555555555555555555555 6666666666666666666666 77777777777777777777777777"))
#decrypted = chavesCripto[0].decrypt(msgCripto).decode("utf-8") #<-- exemplo para o servidor (aqui)
#print(decrypted)
#h = open('mykeyPrivada.pem','rb') <-- antigamente
#privKey = RSA.import_key(h.read()) <-- antigamente
#decryptor = PKCS1_OAEP.new(privKey) <-- antigamente

s = None
interrompe = False
sistema = sys.platform.upper()
if "WIN32" in sistema or "WIN64" in sistema or "WINDOWS" in sistema:
    hostSrv = 'localhost'
    dirLog = 'log\\'
else:
    hostSrv = 'mysqlsrv'
    dirLog = '/Log/' 
logging.basicConfig(filename=dirLog+datetime.now().strftime('%Y-%m-%d %H_%M')+' Serv'+sistema+'.log', format='%(asctime)s - %(message)s', level=logging.INFO)    
MYSQL_ROOT_PASSWORD = os.getenv("MYSQL_ROOT_PASSWORD", "EXAMPLE")
MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234")
CPF1 = os.getenv("CPF1", None)
SENHADCCS = os.getenv("SENHADCCS", None)
CPF2 = os.getenv("CPF2", None)
CPF3 = os.getenv("CPF3", None)
CPF4 = os.getenv("CPF4", None)
CPF5 = os.getenv("CPF5", None)
token = os.getenv("TOKEN", "ERRO")
ambiente = os.getenv("AMBIENTE", "TESTE")
conn = conecta() #testa a conexão com o BD 
if conn!=None:    
    #consultaCorreios()  
    cursor = conn.cursor(buffered=True)
    cursor.execute("Select Data from Extracoes Order By Data DESC")
    row = cursor.fetchone() #data de extração dos dados do Ação Fiscal, via DW ou Receita Data
    if row:
        dataExtracao = row[0]
    else:
        dataExtracao = datetime.strptime("01/01/2021", "%d/%m/%Y")
    #calculaPontosTDPFsEncerrados()   
    diaAtual = datetime.now().date() #será utilizado para criar um arquivo de Log p/ cada dia    
    ultimaVerificacao = datetime.now() #buscamos a última data de extração dos dados agora; só fazemos uma nova pesquisa daqui a uma hora
    cursor.execute("Select Codigo From FiscaisEquipes")
    linha = cursor.fetchone()
    if not linha:
        rodaAlocacoesPontos(primeiraCarga=True)
    else:
        mes = datetime.now().month-1
        ano = datetime.now().year
        if mes<=0:
            mes+=12
            ano-=1
        dictMesTrim = {1:1, 2:1, 3:1, 4:2, 5:2, 6:2, 7:3, 8:3, 9:3, 10:4, 11:4, 12:4}
        cursor.execute("Select Codigo From PontosMetas Where Ano=%s and Trimestre=%s", (ano, dictMesTrim[mes])) #temos que atualizar a tabela de PontosMetas com a parcial do mês
        linha = cursor.fetchone()
        if not linha:
            rodaAlocacoesPontos()
    conn.close()        
    #contém as chaves [0], a função de descriptografia [1] e a data/hora de vencimento [2] para cada segmento IP (resto da divisão da soma das partes do IP por 10 ou 20; 10/20 segmentos)
    chavesCripto = dict()
    print("Gerando chaves criptográficas ...")
    start = time.time()
    inicializaChaves()
    end = time.time()
    print(str(len(chavesCripto))+" chaves geradas em "+str(end - start)[:7]+" segundos.")          
    threads = list() 
    threadServ = threading.Thread(target=servidor, daemon=True) #ativa o servidor
    threadServ.start()
    schedule.every().saturday.at("19:00").do(rodaAlocacoesPontos) #todo sábado roda as funçÕes
    schedule.every().day.at("23:00").do(disparaMediaPontos) #a função verificará se estamos no dia 25 do primeiro mês do trimestre para buscar as informações do trimestre anterior
    schedule.every().day.at("03:00").do(consultaCorreios) #atualiza a situação dos termos enviados por via postal
    #força a execução das tarefas agendadas
    threadDisparador = threading.Thread(target=disparador, daemon=True) #encerra thread quando sair do programa sem esperá-la
    threadDisparador.start()    
    print("Serviço iniciado [", datetime.now(), "]")
    while True:
        sair = input("Digite QUIT quando quiser sair: ")
        if sair:
            if sair.upper().strip()=="QUIT":
                break
    if s!=None: #fecha o socket principal (ouvindo a rede) se estiver aberto
        s.close() #fecha antes de esperar o término das threads para não entrar mais requisições
    i = 0        
    for thread in (threads):
        if thread.is_alive():
            i+=1              #ao interromper o loop, espera as threads que estão tratando requisições
            thread.join()     #terminarem antes de encerrar o programa (fechar conexões)
    logging.info(str(i) + " threads estavam em andamento.")
else:
    print("Não foi possível conectar ao MySQL. Corrija o problema do Banco de Dados e reinicie este serviço (ELSE).")
    logging.info("Erro ao tentar conectar ao BD - Saindo ...")