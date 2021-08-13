# -*- coding: utf-8 -*-
"""
Created on Mon Aug 24 09:26:09 2020

@author: 53363833172
"""

from __future__ import unicode_literals
from re import S
import pandas as pd
import numpy as np
import sys
import os
import logging  
import mysql.connector
from mysql.connector import errorcode
import schedule
import threading
import time
from datetime import datetime, timedelta

#bibliotecas necessárias para envio de e-mail
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.utils import formatdate
from email import encoders
import smtplib

#para gerar a planilha com a relação de TDPFs vincendos (enviada por e-mail para usuários regionais)
from openpyxl import Workbook
#from openpyxl.styles import colors
from openpyxl.styles import Font #, Color
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


#rotina para envio de e-mail - retirado de Bot_Telegram.py
def enviaEmail(email, texto, assunto, arquivo=None): #envia email, conforme parâmetros - se passar o arquivo (caminho e nome), ele vai como anexo 
    try:
        #server = smtplib.SMTP('INETRFOC.RFOC.SRF: 25') #servidor de email Notes
        server = smtplib.SMTP('exchangerfoc.rfoc.srf: 25')
        #pass
    except:
        return 1	
    # create message object instance
    msg = MIMEMultipart()
    # setup the parameters of the message
    msg['From'] = "botespontaneidade@rfb.gov.br"
    msg['To'] = email
    msg['Subject'] = assunto
    # add in the message body
    msg.attach(MIMEText(texto, 'plain'))  
    if arquivo!=None and arquivo!="":  
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(arquivo, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename='+arquivo)
        msg.attach(part)                 
    # send the message via the server.
    try:
        server.sendmail(msg['From'], msg['To'], msg.as_string())
    except:
        server.quit()
        return 2
    server.quit()  
    return 3 #sucesso

def getAlgarismos(texto): #retorna apenas os algarismos de uma string
    algarismos = ""
    for car in texto:
        if car.isdigit():
            algarismos = algarismos + car
    return algarismos

def paraData(data):
    if not data:
        return None  
    tipo = str(type(data)).upper()
    if "DATE" in tipo:
        return data
    if "STR" in tipo or "UNICODE" in tipo:
        try:
            return datetime.strptime(data, "%d/%m/%Y")
        except:
            return None    
    if "TIMESTAMP" in tipo:
        try:
            return pd.Timestamp.to_pydatetime(data)
        except:
            return None  
    return None

def acrescentaZero(numero, n): 
    if numero==None:
        return  ""
    if not "STR" in str(type(numero)):
        numero = str(numero)    
    numero = numero.rjust(n, "0")
    return numero

def acrescentaZeroCPF(cpf):
    return acrescentaZero(cpf, 11)  
      
def montaGrupoFiscal(linha):
    return acrescentaZero(linha['R028_GRSE_SUA_UA_CD'], 7) + acrescentaZero(linha['R028_GRSE_SUA_CD'], 4) + acrescentaZero(linha['R028_GRSE_CD'], 3)

#calcula DV de um CPF - retorna o CPF completo
def calculaDVCPF(cpfPar):
    cpf = getAlgarismos(cpfPar)
    if cpf==None:
        return None
    if len(cpf)>9:
        return None    
    if len(cpf)<9:
        cpf = acrescentaZero(cpf, 9)    
    # Calculado o primeiro DV
    calc = lambda i: int(cpf[i]) * (10-i)
    somaJ = sum(map(calc, range(9)))
    restoJ = somaJ % 11
    if (restoJ == 0 or restoJ == 1):
       j = 0
    else:
       j = 11 - restoJ  
    if j==10:
        j = 0 
    cpf=cpf+str(j)
    # Calculado o segundo DV
    calc2 = lambda i: int(cpf[i]) * (11-i)
    somaK = sum(map(calc2, range(9))) + j*2
    restoK = somaK % 11
    if (restoK == 0 or restoK == 1):
       k = 0
    else:
       k = 11 - restoK  
    if k==10:
        k = 0    
    cpf = cpf + str(k)
    if len(cpf)>11:
        print("Erro CPF:", cpf)
    return cpf

def realizaCargaMalha():
    global dirExcel, termina, hostSrv
    print("Acionada a função que realiza a carga do trabalho da malha - ", datetime.now())   
    if os.path.exists(dirExcel+"REALIZACARGAMALHA.TXT"): #apagamos esse arquivo que serve como indicador de que é para fazer a carga 
        os.remove(dirExcel+"REALIZACARGAMALHA.TXT")    
    try:
        dfMalha = pd.read_excel(dirExcel+"MALHAIDF.xlsx")
    except:
        print("Erro no acesso ao arquivo xlsx com o trabalho da malha ou só há arquivos já processados")
        logging.info("Arquivo Excel MALHAIDF não foi encontrado") 
        return
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 
    try:
        logging.info("Conectando ao servidor de banco de dados ...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql (CargaMalha)!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD (CargaMalha) - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return           
    cursor = conn.cursor(buffered=True) 
    start = time.time() 
    cursor.execute("Select Max(Data) from Malha")
    linha = cursor.fetchone()
    dataAnt = None
    if linha:
        dataAnt = linha[0]
    if dataAnt==None:
        dataAnt = datetime.strptime("01/01/2000", "%d/%m/%Y")
    insertMalha = """Insert Into Malha (Fiscal, Tipo, Recibo, Data, Processamento)
                     Values (%s, %s, %s, %s, %s)"""
    for linha in range(dfMalha.shape[0]): #percorre a planilha com os dados da malha 
        cpf = getAlgarismos(dfMalha.at[linha, 'CPF'] )
        data = paraData(dfMalha.at[linha, 'Data'])
        if data<=dataAnt: #não incluímos períodos já incluídos
            continue
        recibo = dfMalha.at[linha, 'recibo']
        tipoC = dfMalha.at[linha, 'tipo'].strip()
        cursor.execute("Select Codigo From InfoMalha Where Tipo=%s and Inicio<=%s and Fim>=%s", (tipoC, data, data))
        linhaInfo = cursor.fetchone()
        if linhaInfo==None:
            print("Tipo "+tipoC+" não foi encontrado")
            continue
        tipo = linhaInfo[0]
        cursor.execute("Select Codigo From Fiscais Where CPF=%s", (cpf, ))
        linhaFiscal = cursor.fetchone()     
        if not linhaFiscal: #não achamos o fiscal
            nomeFiscal = dfMalha.at[linha, 'nome'].strip()
            cursor.execute("Insert Into Fiscais (CPF, Nome) Values (%s, %s)", (cpf, nomeFiscal))
            cursor.execute("Select Codigo From Fiscais Where CPF=%s", (cpf,))
            linhaFiscal = cursor.fetchone()
            #totalNE+=1
        chaveFiscal = linhaFiscal[0]
        dados = (chaveFiscal, tipo, recibo, data, datetime.now())
        cursor.execute(insertMalha, dados)
    try:
        conn.commit()
        print("Dados da malha trabalhada foram carregados com sucesso!")
    except:
        conn.rollback()
        print("Erro ao atualizar os dados da malha trabalhada")
    print("Tempo decorrido: "+str(time.time()-start)[:8]+" segundos")       
    return        

def realizaCargaVinculos():
    global dirExcel, termina, hostSrv
    print("Acionada a função que realiza a carga dos vínculos - ", datetime.now())   
    if os.path.exists(dirExcel+"REALIZACARGAVINCULOS.TXT"): #apagamos esse arquivo que serve como indicador de que é para fazer a carga dos vínculos imediatamente
        os.remove(dirExcel+"REALIZACARGAVINCULOS.TXT")    
    try:
        dfVinculos = pd.read_excel(dirExcel+"VINCULOS.xlsx")
    except:
        print("Erro no acesso ao arquivo xlsx com os vínculos ou só há arquivos já processados")
        logging.info("Arquivo Excel VINCULOS não foi encontrado") 
        return
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 
    try:
        logging.info("Conectando ao servidor de banco de dados ...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql (CargaVinculos)!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD (CargaVinculos) - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return   
    cursor = conn.cursor(buffered=True) 
    start = time.time() 
    dictEquipes = {}
    cursor.execute("Select Distinctrow Codigo, Equipe From Equipes")
    linhasEquipes = cursor.fetchall()
    for linhaEquipe in linhasEquipes: #adicionamos todas as equipes a um dicionário para não ter que fazer pesquisas repetidas (não fazemos isso com os fiscais pq é uma tabela maior)
        equipe = linhaEquipe[1].strip()
        dictEquipes[equipe] = linhaEquipe[0]
    insertVinculos = """Insert Into Vinculos (Fiscal, Equipe, Vinculo, Inicio, Fim, Registro, InicioSupervisao, FimSupervisao, Processamento)
                        Values (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""
    cursor.execute("Delete From Vinculos Where Codigo>=0") #até definir uma forma de apenas atualizar, apagamos tudo (vai depender do commit com tudo abaixo)
    for linha in range(dfVinculos.shape[0]): #percorre os vínculos na planilha Excel carregada  
        cpf = getAlgarismos(dfVinculos.at[linha, 'CPF'] )
        equipe = getAlgarismos(dfVinculos.at[linha, 'Equipe'])
        inicio = paraData(dfVinculos.at[linha, 'Inicio'])
        fim = paraData(dfVinculos.at[linha, 'Fim'])
        registro = paraData(dfVinculos.at[linha, 'Registro'])
        inicioSupervisao = paraData(dfVinculos.at[linha, 'Inicio_Supervisao'])
        fimSupervisao = paraData(dfVinculos.at[linha, 'Fim_Supervisao'])
        vinculo = dfVinculos.at[linha, 'vinculo']
        cursor.execute("Select Codigo From Fiscais Where CPF=%s", (cpf, ))
        linhaFiscal = cursor.fetchone()
        if not linhaFiscal:
            continue
        chaveFiscal = linhaFiscal[0]
        chaveEquipe = dictEquipes.get(equipe, -1)
        if chaveEquipe==-1: #equipe não existe
            continue
        dados = (chaveFiscal, chaveEquipe, vinculo, inicio, fim, registro, inicioSupervisao, fimSupervisao, datetime.now())
        cursor.execute(insertVinculos, dados)
    try:
        conn.commit()
        print("Vínculos carregados com sucesso!")
    except:
        conn.rollback()
        print("Erro ao atualizar os vínculos")
    print("Tempo decorrido: "+str(time.time()-start)[:8]+" segundos")       
    return

def realizaCargaMetas():
    global dirExcel, termina, hostSrv
    print("Acionada a função que realiza a carga das metas - ", datetime.now())   
    if os.path.exists(dirExcel+"CARGAMETAS.TXT"): #a existência deste arquivo indica que é para fazer a atualização das informações das equipes (info RD)
        os.remove(dirExcel+"CARGAMETAS.TXT")  
    else: #uma vez por mês
        if datetime.now().day>7: #toda primeira quarta feira do mês
            return
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 
    try:
        logging.info("Conectando ao servidor de banco de dados ...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql (CargaMetas)!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD (CargaMetas) - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return   
    cursor = conn.cursor(buffered=True) 
    start = time.time()   
    selectMeta = "Select Codigo, Pontuacao, DataMetas From Metas Where Fiscal=%s and Ano=%s and Trimestre=%s"
    insereMeta = "Insert Into Metas (Fiscal, Ano, Trimestre, Pontuacao, DataMetas, Atualizacao) Values (%s, %s, %s, %s, %s, %s)" 
    atualizaMeta = "Update Metas Set Pontuacao=%s, DataMetas=%s, Atualizacao=%s Where Codigo=%s"
    totalAtualizado = 0
    totalIncluido = 0
    #totalNE = 0
    for ano in range(2021, datetime.now().year+1, 1): #ano de 2021 em diante
        nomeArq = "METAS"+str(ano)
        try:
            dfMetas = pd.read_excel(dirExcel+nomeArq+".xlsx")
            print("Planilha do ano de "+str(ano)+" foi carregada")
        except:
            print("Não foi possível abrir o arquivo "+nomeArq+".xlsx. Prosseguindo ...")
            continue
        #print(dfMetas.dtypes)
        #dfMetas['Sit Fiscal'] = ""
        dfMetas['Sit Equipe'] = ""
        dfMetas['CPF']=dfMetas['cpf_auditor_sem_dv'].astype(str).map(calculaDVCPF)   
        for linha in range(dfMetas.shape[0]): #percorre os fiscais e respectivas metas na planilha Excel carregada
            if (linha+1)%1000==0:
                print("Processando Linha nº "+str(linha+1)+" de "+str(dfMetas.shape[0]))            
            cpfFiscal = dfMetas.at[linha, 'CPF'] 
            cursor.execute("Select Codigo From Fiscais Where CPF=%s", (cpfFiscal,))
            rowFiscal = cursor.fetchone()
            if not rowFiscal:
                nomeFiscal = dfMetas.at[linha, 'nome_auditor'].strip()
                cursor.execute("Insert Into Fiscais (CPF, Nome) Values (%s, %s)", (cpfFiscal, nomeFiscal))
                cursor.execute("Select Codigo From Fiscais Where CPF=%s", (cpfFiscal,))
                rowFiscal = cursor.fetchone()
            chaveFiscal = rowFiscal[0]            
            for col in range(4):
                trimestre = "meta_"+str(col+1)+"_trim"
                pontuacao = float(dfMetas.at[linha, trimestre])
                if pontuacao==np.nan or pontuacao==None or pd.isna(pontuacao):
                    pontuacao = 0
                dataMetas = paraData(dfMetas.at[linha, "Data_Metas"])               
                cursor.execute(selectMeta, (chaveFiscal, ano, col+1)) 
                rowMeta = cursor.fetchone()
                if not rowMeta: #devemos incluir a meta
                    cursor.execute(insereMeta, (chaveFiscal, ano, col+1, pontuacao, dataMetas, datetime.now()))   
                    totalIncluido+=1   
                else:
                    if rowMeta[1]!=pontuacao or rowMeta[2]==None or rowMeta[2]!=dataMetas: #devemos atualizar a meta que mudou algum de seus itens (pontuação ou atualização)
                        cursor.execute(atualizaMeta, (pontuacao, dataMetas, datetime.now(), rowMeta[0]))
                        totalAtualizado+=1
                    else: #pontuação e data são os mesmos - NÃO atualizamos o registro
                        continue
            conn.commit()
        try: #após processar o arquivo (planilha), salvamos o arquivo com informações sobre processamento e apagamos o antigo
            dfMetas.to_excel(dirExcel+nomeArq+"_Processado_"+datetime.now().strftime('%Y-%m-%d')+".xlsx")
            os.remove(dirExcel+nomeArq+".xlsx")  
        except:
            print("Erro ao salvar (processado) ou apagar o arquivo "+nomeArq+".xlsx")     
            logging.info("Erro ao salvar (processado) ou apagar o arquivo "+nomeArq+".xlsx")  
    print("Carga das Metas: ")
    print("Registros Incluídos:", totalIncluido)
    logging.info("Registros Incluídos:"+str(totalIncluido))
    print("Registros Atualizados:", totalAtualizado)
    logging.info("Registros Atualizados:"+str(totalAtualizado))
    #print("CPFs NÃO Encontrados:", totalNE)
    #logging.info("CPFs NÃO Encontrados:"+str(totalNE))
    print("Tempo de processamento: "+str(time.time()-start)[:8])
    conn.close()
    return
      
def realizaCargaPontosSerpro():
    global dirExcel, termina, hostSrv, hora5
    print("Acionada a função que realiza a carga de pontos calculados pela rotina do Serpro - ", datetime.now())
    if os.path.exists(dirExcel+"REALIZACARGAPONTOS.TXT"): #apagamos esse arquivo que serve como indicador de que é para fazer a carga dos pontos imediatamente
        os.remove(dirExcel+"REALIZACARGAPONTOS.TXT")    
    try:
        dfPontos = pd.read_excel(dirExcel+"PONTUACAOSERPRO.xlsx")
        bErro = False
        print(dfPontos.head())        
    except:
        bErro = True
        print("Erro no acesso ao arquivo xlsx com os pontos calculados pelo Serpro ou só há arquivos já processados; outra tentativa será feita às "+hora5+" (PontosSerpro)")
        logging.info("Arquivo Excel PONTUACAOSERPRO não foi encontrado; outra tentativa será feita às "+hora5) 

    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 
    try:
        logging.info("Conectando ao servidor de banco de dados ...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql (PontosSerpro)!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD (PontosSerpro) - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return
    cursor = conn.cursor(buffered=True)        
    if not bErro:
        print("Realizando carga dos PontosSerpro em "+datetime.now().strftime("%d/%m/%Y %H:%M"))
        start = time.time()
        dfPontos['atualizacao'] = pd.to_datetime(dfPontos['atualizacao'])
        dfPontos.sort_values(by=["atualizacao", "RPF", "Sequencia"], ascending=[False, True, True])
        consultaTDPF = "Select Codigo, Pontos, DataPontos From TDPFS Where Numero=%s"
        tdpfAuxAnt = ""
        totalSemPontos = 0
        totalFatores = 0
        totalNE = 0
        totalTdpfsComFatores = 0
        for linha in range(dfPontos.shape[0]): #percorre os TDPFs e respectivos fatores e pontos na planilha Excel
            if (linha+1)%1000==0:
                print("Processando Linha nº "+str(linha+1)+" de "+str(dfPontos.shape[0]))            
            tdpfAux = dfPontos.iat[linha,0]    
            if tdpfAux==tdpfAuxAnt: #não repetimos TDPFs, já que as atualizações mais recentes vem primeiro no dataframe
                continue
            tdpfAuxAnt = tdpfAux
            tdpf = getAlgarismos(tdpfAux)
            cursor.execute(consultaTDPF, (tdpf,))
            linhaTabela = cursor.fetchone()
            if not linhaTabela:
                print("TDPF "+tdpfAux+" não encontrado")
                totalNE+=1
                continue
            chaveTdpf = linhaTabela[0]
            pontosTabela = linhaTabela[1]               
            descricao = dfPontos.iat[linha,2]  
            atualizacao = paraData(dfPontos.iat[linha, 3])        
            if "não se aplica".upper() in descricao.upper() or "nao se aplica".upper() in descricao.upper():
                totalSemPontos+=1
                if pontosTabela!=0 or pontosTabela==None:
                    cursor.execute("Update TDPFS Set Pontos=0, DataPontos=%s Where Codigo=%s", (atualizacao, chaveTdpf)) #este tipo de TDPF não vai ter fatores
                    cursor.execute("Delete From Fatores Where TDPF=%s", (chaveTdpf,))
                    conn.commit()
                continue        
            if linhaTabela[2]!=None: #já houve registro de data dos pontos
                if linhaTabela[2].date()<atualizacao.date(): #a data dos pontos é mais antiga, atualizamos
                    cursor.execute("Delete From Fatores Where TDPF=%s", (chaveTdpf,)) #apagamos os fatores antigos porventura existentes deste TDPF 
                else:
                    continue #TDPF já foi atualizado na mesma data ou mais recentemente - não precisamos atualizar   
            dfFatores = dfPontos.loc[dfPontos['RPF']==tdpfAux] #filtramos somente o TDPF para obter todos os fatores e incluí-los na tabela de fatores
            sequencia = 0
            for linha2 in range(dfFatores.shape[0]):
                if paraData(dfFatores.iat[linha2, 3])<atualizacao or int(dfFatores.iat[linha2, 1])<=sequencia: #passamos para uma atualização mais antiga do mesmo TDPF - desprezamos 
                    break
                sequencia = int(dfFatores.iat[linha2, 1])
                descricao = str(dfFatores.iat[linha2, 2].strip())
                pontos = 0 if pd.isna(dfFatores.iat[linha2, 4]) else float(dfFatores.iat[linha2, 4])
                elementos = 0 if pd.isna(dfFatores.iat[linha2, 5]) else float(dfFatores.iat[linha2, 5])
                percentual = 0 if pd.isna(dfFatores.iat[linha2, 6]) else float(dfFatores.iat[linha2, 6])
                percentual = percentual * 100
                totalFatores+=1
                cursor.execute("Insert Into Fatores (TDPF, Sequencia, Descricao, Elementos, Percentual, Pontos) Values (%s, %s, %s, %s, %s, %s)",\
                            (chaveTdpf, sequencia, descricao, elementos, percentual, pontos))
            
            cursor.execute("Update TDPFS Set Pontos=%s, DataPontos=%s Where Codigo=%s", (pontos, atualizacao, chaveTdpf)) #este tipo de TDPF não vai ter fatores   
            conn.commit()        
            totalTdpfsComFatores+=1         
        print("Pontos do Serpro atualizados com sucesso.")
        print("Total de Fatores Incluídos: ", totalFatores)
        print("Total de TDPFs Com Fatores Incluídos: ", totalTdpfsComFatores)
        print("Total de TDPFs sem Pontos: ", totalSemPontos)
        print("Total de TPPFs NÃO Encontrados: ", totalNE)
        logging.info("Pontos do Serpro atualizados com sucesso.")
        logging.info("Total de Fatores Incluídos: " +str(totalFatores))
        logging.info("Total de TDPFs Com Fatores Incluídos: "+str(totalTdpfsComFatores))
        logging.info("Total de TDPFs sem Pontos: "+str(totalSemPontos))
        logging.info("Total de TPPFs NÃO Encontrados: "+str(totalNE))
        end = time.time()
        print("Carga dos pontos realizada em "+str(end-start)[:10]+" segundos")
        try:
            os.rename(dirExcel+"PONTUACAOSERPRO.xlsx", dirExcel+"PONTUACAOSERPRO_Processado_"+datetime.now().strftime('%Y-%m-%d')+".xlsx")
            logging.info("Arquivo de Pontuação foi renomeado")       
        except:
            print("Erro ao tentar renomear o arquivo de pontuação")            
            logging.error("Erro ao tentar renomear o arquivo de pontuação")    
    print("Iniciando a carga dos pontos de cada fiscal - TDPFs em andamento ...") 
    start = time.time()    
    try:
        dfPontosFiscal = pd.read_excel(dirExcel+"PONTUACAOSERPRORH.xlsx")
        print(dfPontosFiscal.head())         
    except:
        print("Erro no acesso ao arquivo xlsx com os pontos de cada fiscal calculados pelo Serpro ou só há arquivos já processados; outra tentativa será feita às "+hora5+" (PontosSerpro)")
        logging.info("Arquivo Excel PONTUACAOSERPRORH não foi encontrado; outra tentativa será feita às "+hora5) 
        return    
    consultaTDPF = "Select Codigo From TDPFS Where Numero=%s and Encerramento Is Null"  #apenas TDPFs em andamento  
    consultaFiscal = "Select Codigo From Fiscais Where CPF=%s"    
    fiscais = {}
    tdpfsProc = set()
    totalAtu = 0
    totalInc = 0
    for linha in range(dfPontosFiscal.shape[0]): #percorre os TDPFs e respectivos pontos dos fiscais na planilha Excel
        if (linha+1)%1000==0:
            print("Processando Linha nº "+str(linha+1)+" de "+str(dfPontosFiscal.shape[0]))            
        tdpfAux = dfPontosFiscal.at[linha,"RPF"]    
        tdpfAuxAnt = tdpfAux
        tdpf = getAlgarismos(tdpfAux)
        cpfAux = dfPontosFiscal.at[linha, "CPF"]
        cpf = getAlgarismos(cpfAux)
        if (tdpf, cpf) in tdpfsProc: #como a ordem de atualização é descendente, não processamos o mesmo tdpf/cpf outra vez, pois a pontuação é mais antiga
            continue
        cursor.execute(consultaTDPF, (tdpf,))
        linhaTabela = cursor.fetchone()
        if not linhaTabela:
            continue 
        tdpfsProc.add((tdpf, cpf))        
        chaveTdpf = linhaTabela[0]  
        if not cpf in fiscais:
            cursor.execute(consultaFiscal, (cpf, ))
            linhaTabela = cursor.fetchone()
            if not linhaTabela:
                print("Não foi encontrado o fiscal com o CPF "+cpfAux)
                continue
            chaveFiscal = linhaTabela[0]
            fiscais[cpf] = chaveFiscal
        else:
            chaveFiscal = fiscais[cpf]
        pontos = float(dfPontosFiscal.at[linha, "pontos"])
        atualizacao = paraData(dfPontosFiscal.at[linha, "Data_Atu"])
        cursor.execute("Select Codigo, Atualizacao, Pontos From PontosFiscais Where TDPF=%s and Fiscal=%s", (chaveTdpf, chaveFiscal))
        linhaTabela = cursor.fetchone()
        if linhaTabela:
            if linhaTabela[1]>atualizacao: #registro na tabela é mais recente
                continue
            if linhaTabela[2]!=pontos: # só atualizamos se houver mudado a pontuação
                cursor.execute("Update PontosFiscais Set Atualizacao=%s, Pontos=%s Where Codigo=%s", (atualizacao, pontos, linhaTabela[0]))
                totalAtu+=1
            else:
                continue
        else:
            cursor.execute("Insert Into PontosFiscais (TDPF, Fiscal, Pontos, Atualizacao) Values (%s, %s, %s, %s)", (chaveTdpf, chaveFiscal, pontos, atualizacao))
            totalInc+=1
        conn.commit()
    conn.close()      
    print("Total de Registros com Pontos Incluídos para Fiscais: ", totalInc)
    print("Total de Registros com Pontos Atualizados para Fiscais: ", totalAtu)
    logging.info("Total de Registros com Pontos Incluídos para Fiscais: "+str(totalInc))
    logging.info("Total de Registros com Pontos Atualizados para Fiscais: "+str(totalAtu))
    end = time.time()
    print("Carga dos pontos de cada fiscal (TDPFs em andamento) realizada em "+str(end-start)[:10]+" segundos")
    try:
        os.rename(dirExcel+"PONTUACAOSERPRORH.xlsx", dirExcel+"PONTUACAOSERPRORH_Processado_"+datetime.now().strftime('%Y-%m-%d')+".xlsx")
        logging.info("Arquivo de Pontuação RH foi renomeado")
    except:
        print("Erro ao tentar renomear o arquivo de pontuação RH")            
        logging.error("Erro ao tentar renomear o arquivo de pontuação RH")     
    return

def atualizaEquipesDadosAd():
    global dirExcel, termina, hostSrv
    print("Acionada a função que realiza a carga de dados adicionais das equipes - ", datetime.now())
    logging.info("Acionada a função que realiza a carga de dados adicionais das equipes - "+str(datetime.now()))
    if os.path.exists(dirExcel+"CARGADADOSAD.TXT"): #apagamos esse arquivo que serve como indicador de que é para fazer a carga dos dados imediatamente
        os.remove(dirExcel+"CARGADADOSAD.TXT")    
    try:
        dfEquipes = pd.read_excel(dirExcel+"EQUIPESDADOSAD.xlsx", dtype={'Equipe': object, 'grse_dt_criacao': datetime, 'grse_dt_extincao': datetime})
    except:
        print("Erro no acesso aos arquivo EQUIPESDADOSAD.xlsx")
        logging.info("Arquivo Excel EQUIPESDADOSAD.xlsx não foi encontrado") 
        return
    dfEquipes['dt_criacao'] = dfEquipes['grse_dt_criacao'].map(pd.Timestamp)
    dfEquipes['dt_extincao'] = dfEquipes['grse_dt_extincao'].map(pd.Timestamp)
    print(dfEquipes.dtypes)
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 
    try:
        logging.info("Conectando ao servidor de banco de dados ...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD atualizaEquipesDadosAd - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return 
    cursor = conn.cursor(buffered=True)
    totalInc = 0
    totalAtu = 0
    totalDesp = 0
    start = time.time()
    for linha in range(dfEquipes.shape[0]):
        equipe = getAlgarismos(dfEquipes.at[linha, "Equipe"].strip())
        sistema = int(dfEquipes.at[linha, "grse_sgrp_sist_cd"])  
        malha = dfEquipes.at[linha, "grse_in_ativ_malha"]
        criacao = paraData(dfEquipes.at[linha, "dt_criacao"])       
        extincao = paraData(dfEquipes.at[linha, "dt_extincao"])
        if extincao!=None:
            extinta = 'S'
        else:
            extinta = 'N'   
        tipoEquipe = int(dfEquipes.at[linha, "grse_sgrp_tgtr_cd"])
        descTipoEquipe = dfEquipes.at[linha, "Tipo Equipe"].strip()
        cursor.execute("Select Codigo From TipoEquipes Where Tipo=%s", (tipoEquipe, ))
        rowTipoEquipe = cursor.fetchone()
        if not rowTipoEquipe:
            cursor.execute("Insert Into TipoEquipes (Tipo, Descricao) Values (%s, %s)", (tipoEquipe, descTipoEquipe))
            cursor.execute("Select Codigo From TipoEquipes Where Tipo=%s", (tipoEquipe, ))
            rowTipoEquipe = cursor.fetchone()   
        codTipoEquipe = rowTipoEquipe[0]         
        cursor.execute("Select Codigo From Equipes Where Equipe=%s", (equipe,))
        rowEquipe = cursor.fetchone()
        if not rowEquipe:
            if sistema!=6: #se não for da fiscalização, vamos para a próxima equipe
                print("Equipe "+equipe+", sistema "+str(sistema)+" ("+extinta+"), não existe na tabela de Equipes")
                totalDesp+=1
                continue #equipe não existe
            #senão, incluímos a equipe com os dados disponíveis
            nomeEquipe = dfEquipes.at[linha, "grse_nm"]
            cursor.execute("Insert Into Equipes (Equipe, Nome, Malha, Criacao, Extincao, Sistema, Tipo) Values (%s, %s, %s, %s, %s, %s, %s)", (equipe, nomeEquipe, malha, criacao, extincao, sistema, codTipoEquipe))
            totalInc+=1
        else:
            cursor.execute("Update Equipes Set Malha=%s, Criacao=%s, Extincao=%s, Sistema=%s, Tipo=%s Where Codigo=%s", (malha, criacao, extincao, sistema, codTipoEquipe, rowEquipe[0]))
            totalAtu+=1            
        conn.commit()
    print(str(totalAtu)+" equipes atualizadas")
    logging.info(str(totalAtu)+" equipes atualizadas")
    print(str(totalInc)+" equipes incluídas")
    logging.info(str(totalInc)+" equipes incluídas")
    print(str(totalDesp)+" equipes desprezadas")
    logging.info(str(totalDesp)+" equipes desprezadas")
    print("Tempo de processamento: "+str(time.time()-start)[:8]+" segundos")
    conn.close()
    return

def realizaCargaDados():
    global dirExcel, termina, hostSrv, hora1
    print("Acionada a função que realiza a carga de dados do Ação Fiscal/DW - TDPFS ", datetime.now())
    if os.path.exists(dirExcel+"REALIZACARGA.TXT"): #apagamos esse arquivo que serve como indicador de que é para fazer a carga dos dados imediatamente
        os.remove(dirExcel+"REALIZACARGA.TXT")    
    try:
        dfTdpf = pd.read_excel(dirExcel+"TDPFSRPFS.xlsx", dtype={'Porte':object, 'Acompanhamento':object, 'Receita Programada(Tributo) Código': int})
        print("Carregou TDPFSRPFS.xlsx")
        dfAloc = pd.read_excel(dirExcel+"ALOCACOESRPFS.xlsx")
        print("Carregou ALOCACOSRPFS.xlsx")
        dfFiscais = pd.read_excel(dirExcel+"FISCAIS.xlsx")
        print("Carregou FISCAIS.xlsx")
        dfSupervisores = pd.read_csv(dirExcel+"SUPERVISORES.csv", sep=";", encoding = "ISO-8859-1")
        print("Carregou SUPERVISORES.csv")
        dfOperacoes = pd.read_excel(dirExcel+"OPERACOES.xlsx")
        print("Carregou OPERACOES.xlsx")
        dfEquipes = pd.read_excel(dirExcel+"EQUIPES.xlsx", dtype={'Qtde RH Equipe Fiscal': int})
        print("Carregou EQUIPES.xlsx")
    except:
        print("Erro no acesso aos arquivos xlsx e/ou csv ou só há arquivos já processados; outra tentativa será feita às "+hora1)
        logging.info("Arquivos Excel não foram encontrados (um ou mais) - TDPFSRPFS.xlsx, ALOCACOESRPFS.xlsx, Fiscais.xlsx, OPERACOES.xlsx ou Supervisores.CSV; outra tentativa será feita às "+hora1) 
        return
    dfFiscais['CPF']=dfFiscais['CPF'].astype(str).map(acrescentaZeroCPF) 
    dfAloc['Ind. RH Superv. Gr. Fiscal RPF'] = dfAloc['Ind. RH Superv. Gr. Fiscal RPF'].astype(str)
    dfSupervisores['CPF']=dfSupervisores['R028_RH_PF_NR'].astype(str).map(calculaDVCPF)
    dfSupervisores['Grupo Fiscal']=dfSupervisores.apply(montaGrupoFiscal, axis=1) 
    dfEquipes['Equipe Fiscal Código']=dfEquipes['Equipe Fiscal Código'].astype(str)
    #dfTdpf['Porte']=dfTdpf['Porte'].astype(str)
    #dfTdpf['Acompanhamento']=dfTdpf['Acompanhamento'].astype(str)
    #MYSQL_ROOT_PASSWORD = os.getenv("MYSQL_ROOT_PASSWORD", "EXAMPLE")
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 
    #print(dfTdpf.head())
    #print(dfTdpf.dtypes)
    #print(dfSupervisores.head())
    #print(dfSupervisores.dtypes)
    #return
    try:
        logging.info("Conectando ao servidor de banco de dados ...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return
    print("Realizando carga dos dados em "+datetime.now().strftime("%d/%m/%Y %H:%M"))
    cursor = conn.cursor(buffered=True)
    #if hostSrv == 'localhost': #no ambiente local de testes (SO Windows), apagamos a base - comentar estas linhas, se necessário
    #    cursor.execute("Delete from TDPFS")
    #    cursor.execute("Delete from Alocacoes")
    #    cursor.execute("Delete from Supervisores")
    #    cursor.execute("Delete from Atividades")
    #    cursor.execute("Delete from Ciencias")
    #    cursor.execute("Delete from AvisosVencimento")
    #    cursor.execute("Delete from CadastroTDPFs")
    #    conn.commit()
    selectFisc = "Select Codigo, CPF, Nome from Fiscais Where CPF=%s"
    insereFisc = "Insert Into Fiscais (CPF, Nome) Values (%s, %s)"

    selectTDPF = "Select Codigo, Grupo, Encerramento, Vencimento, SemExame from TDPFS Where Numero=%s"
    insereTDPF = "Insert Into TDPFS (Numero, Grupo, Emissao, Nome, NI, Vencimento, Porte, Acompanhamento, Encerramento, CasoEspecial, SemExame, Pontos, DataPontos, TDPFPrincipal, Tipo, FAPE) Values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    atualizaTDPFEnc = "Update TDPFS Set Encerramento=%s, SemExame=%s Where Codigo=%s"
    atualizaTDPFEncSemExame = "Update TDPFS Set Encerramento=%s, SemExame=%s, Pontos=0, DataPontos=%s Where Codigo=%s"
    atualizaTDPFGrupoVencto = "Update TDPFS Set Grupo=%s, Vencimento=%s Where Codigo=%s"

    pesquisaTDPFPrincipal ="Select Codigo from TDPFS Where Numero=%s"

    selectAloc = "Select Codigo, Desalocacao, Supervisor from Alocacoes Where TDPF=%s and Fiscal=%s"
    insereAloc = "Insert Into Alocacoes (TDPF, Fiscal, Alocacao, Supervisor, Horas) Values (%s, %s, %s, %s, %s)"
    insereAlocDesaloc = "Insert Into Alocacoes (TDPF, Fiscal, Alocacao, Desalocacao, Supervisor, Horas) Values (%s, %s, %s, %s, %s, %s)"
    atualizaAloc = "Update Alocacoes Set Desalocacao=%s, Supervisor=%s, Horas=%s Where Codigo=%s"
    atualizaAlocHoras = "Update Alocacoes Set Horas=%s Where Codigo=%s"

    selectCiencias = "Select * from Ciencias Where TDPF=%s" 
    insereCiencia = "Insert Into Ciencias (TDPF, Data, Documento) Values (%s, %s, %s)"

    selectOperacoes = """Select Operacoes.Codigo, OperacoesFiscais.Operacao, PeriodoInicial, PeriodoFinal, Tributos.Tributo, Operacoes.Operacao, Operacoes.Tributo
                         from Operacoes, OperacoesFiscais, Tributos 
                         Where Operacoes.TDPF=%s and Operacoes.Operacao=OperacoesFiscais.Codigo and Operacoes.Tributo=Tributos.Codigo"""
    insereOperacao = "Insert Into Operacoes (TDPF, Operacao, PeriodoInicial, PeriodoFinal, Tributo) Values (%s, %s, %s, %s, %s)"
    apagaOperacao = "Delete from Operacoes Where Codigo=%s"

    selectOpFiscal = "Select Codigo from OperacoesFiscais Where Operacao=%s"
    insertOpFiscal = "Insert Into OperacoesFiscais (Operacao, Descricao, Valor) Values (%s, %s, %s)"

    selectTributo = "Select Codigo from Tributos Where Tributo=%s"
    insertTributo = "Insert Into Tributos (Tributo, Descricao) Values (%s, %s)"

    selectUsuario = "Select Codigo, CPF, email from Usuarios Where CPF=%s"
    insereUsuario = "Insert Into Usuarios (CPF, email) Values (%s, %s)"
    updateUsuario = "Update Usuarios Set email=%s Where Codigo=%s"

    selectCaso = "Select Codigo from CasosEspeciais Where CasoEspecial=%s"
    insereCaso = "Insert Into CasosEspeciais (CasoEspecial, Descricao) Values (%s, %s)"

    #tentei criar um trigger, mas não deu certo (Grasiella está de licença) - 14 %s (utilizado quando o TDPF é cancelado)
    apagaTDPF = """
                Delete from Alocacoes Where TDPF=%s;
                Delete from Atividades Where TDPF=%s;
                Delete from AvisosCiencia Where TDPF=%s;
                Delete from AvisosVencimento Where TDPF=%s;
                Delete from CadastroTDPFs Where TDPF=%s;
                Delete from Ciencias Where TDPF=%s;
                Delete from ControlePostal Where TDPF=%s;
                Delete from DiarioFiscalizacao Where TDPF=%s;
                Delete from Juntadas Where TDPF=%s;
                Delete from Operacoes Where TDPF=%s;
                Delete from AssinaturaFiscal Where Prorrogacao In (Select Prorrogacoes.Codigo From Prorrogacoes Where Prorrogacoes.TDPF=%s);
                Delete From Prorrogacoes Where Prorrogacoes.TDPF=%s;
                Delete from Resultados Where TDPF=%s;
                Delete from TDPFS Where Codigo=%s
                """
                   
    logging.info(f"TDPFs: {dfTdpf.shape[0]} linhas e {dfTdpf.shape[1]} colunas")
    logging.info(f"AFRFBs Execução: {dfAloc.shape[0]} linhas e {dfAloc.shape[1]} colunas")
    tabFiscais=0
    tabTdpfs=0
    tabTdpfsAtu=0
    tabCiencias=0
    tabAloc=0
    tabAlocAtu=0
    tabUsuarios=0
    tabUsuariosAtu=0
    gruposAtu=0
    if termina:
        return
    logging.info("Iniciando loop na carga.")
    atualizou = False   
    if datetime.now().strftime("%d/%m/%Y")!="06/08/2021": #nesta data, já fizemos a carga dos RPFs
        for linha in range(dfTdpf.shape[0]): #percorre os TDPFs das planilhas Excel
            if (linha+1)%500==0:
                print("Processando TDPF nº "+str(linha+1)+" de "+str(dfTdpf.shape[0]))
            if atualizou:
                conn.commit()
                atualizou = False
            tdpfAux = dfTdpf.iat[linha,0]
            if tdpfAux=='9999999.9999.99999': #última linha, referente à data de extração dos dados
                dataExtracao = dfTdpf.iat[linha,9]
                cursor.execute(cursor.execute("Insert Into Extracoes (Data) Values (%s)", (paraData(dataExtracao),)))
                break
            tdpf = getAlgarismos(tdpfAux)
            distribuicao = dfTdpf.iat[linha, 9] #na verdade, aqui é a data de assinatura/emissão do TDPF (antes tinha apenas a distribuição) 
                                                #<- a assinatura revelou-se pior que a distribuição - voltei a usar esta
            inicio = dfTdpf.iat[linha, 11]
            encerramento = dfTdpf.iat[linha, 12]
            situacao = dfTdpf.iat[linha, 15]
            if situacao!=None:
                situacao = situacao.upper()
            if "SEM EXAME" in situacao:
                tipoEnc = "S"
            elif 'COM EXAME' in situacao:
                tipoEnc = "N"        
            else:
                tipoEnc = None
            ni = dfTdpf.iat[linha, 17]
            nome = dfTdpf.iat[linha, 18]
            porte = dfTdpf.iat[linha, 27]
            acompanhamento = dfTdpf.iat[linha, 28]
            if porte==np.nan or pd.isna(porte) or porte=="":
                porte = None
            if acompanhamento==np.nan or pd.isna(acompanhamento) or acompanhamento=="":
                acompanhamento = None    
            tipoProc = dfTdpf.iat[linha, 22][:1] #tipo do procedimento
            if tipoProc in ['D', 'F'] and paraData(dfTdpf.iat[linha, 10])==None: #TDPF não assinado - não o incluímos no BD         
                continue
            tdpfPrincipal = dfTdpf.iat[linha, 26]
            if tdpfPrincipal!=None and not pd.isna(tdpfPrincipal) and tdpfPrincipal!=np.nan:
                tdpfPrincipal = getAlgarismos(tdpfPrincipal)
                cursor.execute(pesquisaTDPFPrincipal, (tdpfPrincipal,))
                rowTdpfPrincipal = cursor.fetchone()
                if rowTdpfPrincipal:
                    chaveTdpfPrincipal = rowTdpfPrincipal[0]
                else:
                    chaveTdpfPrincipal = None
            else:
                chaveTdpfPrincipal = None
            fape = dfTdpf.iat[linha, 29]
            if not fape in ['S', 'N']:
                fape = 'N'            
            #busca o caso especial, se houver, ou o insere     
            casoEspecial = dfTdpf.iat[linha, 19]
            if casoEspecial=="" or pd.isna(casoEspecial) or casoEspecial==None or casoEspecial==np.nan:
                casoEspecialCod = None
            else:
                casoEspecialDesc = dfTdpf.iat[linha, 20].strip()
                casoEspecial = int(casoEspecial)
                cursor.execute(selectCaso, (casoEspecial, ))
                linhaCaso = cursor.fetchone()
                if linhaCaso:
                    casoEspecialCod = linhaCaso[0]
                else:
                    atualizou = True
                    cursor.execute(insereCaso, (casoEspecial, casoEspecialDesc))
                    cursor.execute(selectCaso, (casoEspecial, ))
                    linhaCaso = cursor.fetchone()
                    if linhaCaso:
                        casoEspecialCod = linhaCaso[0]    
                    else:
                        casoEspecialCod = 0            

            #comentei o trecho abaixo pq já vem certinho na planilha do Excel
            #tipo = str(type(porte)).upper()
            #if "STR" in tipo or "UNICODE" in tipo: 
            #    porte = porte[:3]    
            #tipo = str(type(acompanhamento)).upper()
            #if "STR" in tipo or "UNICODE" in tipo: 
            #    acompanhamento = acompanhamento[:1]                       
            cursor.execute(selectTDPF, (tdpf,))
            regTdpf = cursor.fetchone()  
            #precisamos incluir os TDPFs encerrados a partir do ano de entrada em produção (2021) e do ano de 2020
            if not regTdpf and encerramento!="SD" and encerramento!="" and paraData(encerramento)!=None:    #TDPF não existe
                if paraData(encerramento).year<2020: #se foi encerrado antes de 2020, desprezamos                
                    continue
                if "CANCELADA" in situacao or tipoEnc=='S': #não devemos cadastrar TDPFs que não constem da base, mas que já estão cancelados ou foram encerrados sem exame                    
                    continue            
            if regTdpf: #TDPF consta da base
                chaveTdpf = regTdpf[0]  #chave do registro do TDPF  - para poder atualizar o registro, se for necessário   
                if "CANCELADA" in situacao: #TDPF existe, mas foi cancelado - devemos apagar tudo a respeito dele
                    parametrosApaga = [chaveTdpf for i in range(14)]
                    print("Apagando ", tdpfAux, situacao)
                    for result in cursor.execute(apagaTDPF, tuple(parametrosApaga), multi=True):
                        pass                
                    #conn.commit()
                    cursor.close()   #se não fecho, dá erro de "out of sync"
                    cursor = conn.cursor(buffered=True)                   
                    continue #apagamos - vamos ao próximo TDPF
                ##cursor.execute("Update TDPFS Set FAPE=%s Where Codigo=%s", (fape, chaveTdpf)) #  <-- APAGAR depois da primeira execução            <----------
                if regTdpf[2]!=None: #TDPF já estava encerrado
                    if regTdpf[4]==None: #está encerrado, mas sem sinalizar se é sem exame ou com <--manter na produção
                        if tipoEnc=='S':
                            cursor.execute("Update TDPFS Set SemExame=%s, Pontos=0, DataPontos=%s Where Codigo=%s", (tipoEnc, datetime.now(), chaveTdpf))
                        else:
                            cursor.execute("Update TDPFS Set SemExame=%s Where Codigo=%s", (tipoEnc, chaveTdpf))
                        atualizou = True
                    if datetime.now().date()>datetime(2021, 8, 12).date():
                        intervalo = timedelta(days=90)
                    else:
                        intervalo = timedelta(days=240)    #retrocedemos mais se acabar de entrar em produção 
                    if regTdpf[2].date()<(datetime.now()-intervalo).date() and regTdpf[2]==paraData(encerramento): #TDPF já encerrado na base há mais de INTERVALO dias - não há interesse em atualizar as demais tabelas                               
                        continue  
            else: #TDPF não existe na base - pedi para ajustar o relatório do gerencial para incluir os encerrados a partir de 2020 (o problema abaixo não deve ocorrer)
                if porte==None or acompanhamento==None: #porte ou acompanhamento especial não foram obtidos do gerencial Ação Fiscal no DW - significa que não há necessidade de 
                                                        #incluir o TDPF na base pois já está encerrado (por isso não consta do gerencial)
                    logging.info(f"TDPFs: {tdpfAux} não tem monitoramento e/ou porte sem constar na base - TDPF foi desprezado.")                     
                    continue            
            df = dfAloc.loc[dfAloc['Número do RPF Expresso']==tdpfAux] #selecionamos as alocações do TDPF
            if df.shape[0]==0:
                logging.info(f"TDPFs: {tdpfAux} não tem fiscal alocado - TDPF foi desprezado.")
                #mas temos que atualizar o status de encerramento dele, caso tenha sido (relatório de alocações não foi atualizado com TDPFs encerrados sem exame ou cancelados)
                if (regTdpf[2]==None or regTdpf[2]!=paraData(encerramento)) and encerramento!="SD" and encerramento!="" and paraData(encerramento)!=None: #TDPF existia na base em andamento, agora é encerrado
                    tabTdpfsAtu+=1
                    atualizou = True
                    if tipoEnc=="S": #Sem exame - zeramos os pontos
                        cursor.execute(atualizaTDPFEncSemExame, (paraData(encerramento), tipoEnc, datetime.now(), chaveTdpf))            
                    else: #com exame
                        cursor.execute(atualizaTDPFEnc, (paraData(encerramento), tipoEnc, chaveTdpf))                                                 
                continue
            if distribuicao: #calculamos a data de vencimento a partir da data de distribuição - o ideal seria calcular a partir da data de emissão, mas o DW não tem a informação
                distData = paraData(distribuicao)             
                if tipoProc in ['F', 'D']:
                    if tipoProc=='F':
                        diasValidade = 120
                    else:
                        diasValidade = 60
                    vencimento = distData + timedelta(days=(diasValidade-1)) #funciona assim no Ação Fiscal - o primeiro vencimento, ocorre 119/59 dias (conta o dia da emissão); os subsequentes, 120/60 dias
                    while vencimento.date()<datetime.now().date():
                        vencimento = vencimento + timedelta(days=diasValidade)
                        if encerramento!="SD" and encerramento!="" and paraData(encerramento)!=None:
                            if vencimento.date()>paraData(encerramento).date(): #assim que passou do encerramento, paramos de acrescentar 120/60 dias ao vencimento
                                break
                else:
                    vencimento = None
            else: #distribuição nulo? Não deve acontecer, mas ...
                vencimento = None   
                distData = None   
            #precisamos percorrer as alocações para descobrir o grupo atual
            grupoAtu = None 
            for linha2 in range(df.shape[0]):
                grupo = df.iat[linha2, 4]
                tipoGrupo = str(type(grupo)).upper()
                if "STR" in tipoGrupo or "UNICODE" in tipoGrupo:
                    grupo = getAlgarismos(grupo)
                else:
                    grupo = None
                desalocacao = df.iat[linha2, 10]            
                if (desalocacao=="SD" or desalocacao=="" or paraData(desalocacao)==None) and grupo!="" and grupo!=None:
                    grupoAtu = grupo 
                    break 
            bInseriuCiencia = False                
            if not regTdpf: #TDPF não consta da base
                tabTdpfs+=1
                atualizou = True
                if tipoEnc=='S' and encerramento!="SD" and encerramento!="" and paraData(encerramento)!=None:
                    pontos = 0
                    dataPontos = datetime.now()
                else:
                    pontos = None
                    dataPontos = None    
                cursor.execute(insereTDPF, (tdpf, grupoAtu, distData, nome, ni, vencimento, porte, acompanhamento, paraData(encerramento), casoEspecialCod, tipoEnc, pontos, dataPontos, chaveTdpfPrincipal, tipoProc, fape))
                cursor.execute(selectTDPF, (tdpf,))
                regTdpf = cursor.fetchone()   
                chaveTdpf = regTdpf[0]  #chave do registro do TDPF                      
                if inicio!="SD" and inicio!="" and paraData(inicio)!=None:
                    tabCiencias+=1
                    cursor.execute(insereCiencia, (chaveTdpf, paraData(inicio), "ACÃO FISCAL"))
                    bInseriuCiencia = True
            elif (regTdpf[2]==None or regTdpf[2]!=paraData(encerramento)) and encerramento!="SD" and encerramento!="" and paraData(encerramento)!=None: #TDPF existia na base em andamento, agora é encerrado
                tabTdpfsAtu+=1
                atualizou = True
                if tipoEnc=="S": #Sem exame - zeramos os pontos
                    cursor.execute(atualizaTDPFEncSemExame, (paraData(encerramento), tipoEnc, datetime.now(), chaveTdpf))              
                else:
                    cursor.execute(atualizaTDPFEnc, (paraData(encerramento), tipoEnc, chaveTdpf))
            elif regTdpf[1]!=grupoAtu or regTdpf[3]==None: #mudou o grupo e/ou TDPF está sem data de vencimento, mas, em ambos os casos, NÃO encerrado - atualiza grupo e vencimento
                gruposAtu+=1
                atualizou = True
                cursor.execute(atualizaTDPFGrupoVencto, (grupoAtu, vencimento, chaveTdpf))
            elif regTdpf[3]!=None: 
                if regTdpf[3].date()<datetime.now().date(): #TDPF está vencido - atualizamos
                    atualizou = True
                    cursor.execute(atualizaTDPFGrupoVencto, (grupoAtu, vencimento, chaveTdpf))                
            if regTdpf and inicio!="SD" and inicio!="" and paraData(inicio)!=None and not bInseriuCiencia:
                cursor.execute(selectCiencias, (chaveTdpf,))
                regCiencia = cursor.fetchone()
                if not regCiencia:
                    tabCiencias+=1
                    atualizou = True
                    cursor.execute(insereCiencia, (chaveTdpf, paraData(inicio), "ACÃO FISCAL"))                
            for linha2 in range(df.shape[0]): #percorre as alocações na planilhas Excel relativas ao TDPF do loop externo
                cpf = getAlgarismos(df.iat[linha2, 6])
                fiscal = df.iat[linha2, 7] #nome do fiscal
                alocacao = df.iat[linha2, 9]
                desalocacao = df.iat[linha2, 10]  
                supervisor = df.iat[linha2, 12]
                if supervisor==np.nan or pd.isna(supervisor) or supervisor=="" or supervisor==None:
                    supervisor = "N"
                else:
                    supervisor = supervisor[:1]
                horas = df.iat[linha2, 16]
                try:    
                    horas = int(horas)    
                except:
                    horas = 0                  
                dfFiscal = dfFiscais.loc[dfFiscais['CPF']==cpf]
                email = None
                if dfFiscal.shape[0]>0:
                    if dfFiscal.iat[0, 4]!=np.nan and not pd.isna(dfFiscal.iat[0, 4]) and dfFiscal.iat[0, 4]!="": #email está na coluna 4 (coluna 'E' do Excel)
                        email = dfFiscal.iat[0, 4]
                #print(email)        
                cursor.execute(selectUsuario, (cpf,))
                regUser = cursor.fetchone()
                if regUser!=None: #achou o usuário - vemos se tem e-mail cadastrado
                    if (regUser[2]==None or regUser[2]=='') and email!=None: #regUser[2] é o email
                        cursor.execute(updateUsuario, (email, regUser[0]))
                        tabUsuariosAtu+=1
                        atualizou = True
                else: #inserimos o novo usuário
                    cursor.execute(insereUsuario, (cpf, email))       
                    tabUsuarios+=1 
                    atualizou = True
                cursor.execute(selectFisc, (cpf,))
                regFisc = cursor.fetchone()
                if not regFisc:
                    tabFiscais+=1
                    atualizou = True
                    cursor.execute(insereFisc, (cpf, fiscal))
                    cursor.execute(selectFisc, (cpf,))
                    regFisc = cursor.fetchone()
                chaveFiscal = regFisc[0] #chave do registro do Fiscal (inserido ou consultado anteriormente)
                cursor.execute(selectAloc, (chaveTdpf, chaveFiscal))
                regAloc = cursor.fetchone()                  
                if not regAloc:
                    tabAloc+=1
                    if desalocacao=="SD" or desalocacao=="" or paraData(desalocacao)==None:
                        atualizou = True
                        cursor.execute(insereAloc, (chaveTdpf, chaveFiscal, paraData(alocacao), supervisor, horas))
                    else:
                        atualizou = True
                        cursor.execute(insereAlocDesaloc, (chaveTdpf, chaveFiscal, paraData(alocacao), paraData(desalocacao), supervisor, horas))
                elif regAloc[1]==None and desalocacao!="SD" and desalocacao!="" and paraData(desalocacao)!=None:
                    tabAlocAtu+=1
                    atualizou = True
                    cursor.execute(atualizaAloc, (paraData(desalocacao), supervisor, horas, regAloc[0])) 
                elif regAloc[1]!=None and (desalocacao=="SD" or desalocacao=="" or paraData(desalocacao)==None):
                    tabAlocAtu+=1
                    atualizou = True
                    cursor.execute(atualizaAloc, (None, supervisor, horas, regAloc[0]))    
                elif regAloc[2]!=supervisor:
                    tabAlocAtu+=1
                    atualizou = True
                    cursor.execute(atualizaAloc, (regAloc[1], supervisor, horas, regAloc[0]))  
                else:
                    cursor.execute(atualizaAlocHoras, (horas, regAloc[0]))    
                    tabAlocAtu+=1
                    atualizou = True 
            if not tipoProc in ['F', 'R', 'L']:
                continue #não processamos operações, a não ser em caso de fiscalização, revisão e lançamento de multa
            #percorremos as operações do TDPF - excluímos as que não mais existirem e incluímos as que não existirem
            dfOp = dfOperacoes.loc[dfOperacoes['Número do RPF Expresso']==tdpfAux] #selecionamos as operações do TDPF
            #Operacoes.Codigo, OperacoesFiscais.Operacao, PeriodoInicial, PeriodoFinal, Tributos.Tributo, Operacoes.Operacao, Operacoes.Tributo
            cursor.execute(selectOperacoes, (chaveTdpf,))
            regOperacoes = cursor.fetchall()
            opExistentes = []
            if dfOp.shape[0]>0: #desde que tenhamos localizado alguma operação do TDPF (se não, o problema está na planilha de operações)
                for regOperacao in regOperacoes: #atualizamos as operações que mudaram algo no período ou excluímos aquelas que não existem mais
                    operacao = regOperacao[1]
                    codigoOperacao = regOperacao[0]
                    perInicial = regOperacao[2]
                    perFinal = regOperacao[3]
                    tributo = regOperacao[4]
                    operTrib = str(operacao).rjust(6, "0")+str(tributo).rjust(5, "0")
                    dfOpAux = dfOp.loc[(dfOp['Operação Fiscal Atual Código']==operacao) & (dfOp['Receita Programada(Tributo) Código']==tributo)]
                    if dfOpAux.shape[0]>0: #operação/tributo existe no TDPF - temos que ver se há alguma divergência no período (aumentou ou diminuiu)
                        if not operTrib in opExistentes: #não foi atualizada/incluída
                            opExistentes.append(operTrib)
                            menorMes = paraData(dfOpAux.loc[dfOpAux['Mês Início'].idxmin()]["Mês Início"])
                            maiorMes = paraData(dfOpAux.loc[dfOpAux['Mês Fim'].idxmax()]["Mês Fim"])
                            if maiorMes!=perFinal or menorMes!=perInicial:
                                #print("Atualizando período da operação ", tdpfAux, codigoOperacao, operacao, tributo, menorMes.date(), maiorMes.date())
                                comando = "Update Operacoes Set PeriodoInicial=%s, PeriodoFinal=%s Where Codigo=%s"
                                try:
                                    cursor.execute(comando, (menorMes, maiorMes, codigoOperacao))
                                except: #provavelmente já existe um outro registro para este tdpf com esta Operacao/Tributo
                                    apaga = "Delete From Operacoes Where Codigo!=%s, TDPF=%s, Operacao=%s, Tributo=%s"
                                    cursor.execute(apaga, (codigoOperacao, chaveTdpf, regOperacao[5], regOperacao[6])) #apagamos os demais registros
                                    #tentamos novamente atualizar
                                    cursor.execute(comando, (menorMes, maiorMes, codigoOperacao))
                        else: #como já foi incluída, apagamos da tabela, pq senão ficará uma operação repetida
                            comando = "Delete From Operacoes Where Codigo=%s"
                            cursor.execute(comando, (codigoOperacao, ))                            
                    else:
                        cursor.execute(apagaOperacao, (codigoOperacao,)) #operação foi removida do TDPF - removemos ela da base
            #incluímos as operações do TDPF (as que não tiverem sido cadastradas)
            for linha2 in range(dfOp.shape[0]):
                operacao = int(dfOp.iat[linha2, 8])
                valor = dfOp.iat[linha2, 11] #peso/valor da operação
                tributo = int(dfOp.iat[linha2, 1])
                operTrib = str(operacao).rjust(6, "0")+str(tributo).rjust(5, "0")
                if operTrib in opExistentes: #operação/tributo já está na base
                    continue
                opExistentes.append(operTrib)
                #não está na base - temos que incluí-la
                #consultamos o tributo e o incluímos, se não existir           
                cursor.execute(selectTributo, (tributo,))
                rowTributo = cursor.fetchone()
                if not rowTributo:
                    cursor.execute(insertTributo, (tributo, dfOp.iat[linha2, 2].upper()))
                    cursor.execute(selectTributo, (tributo,))
                    rowTributo = cursor.fetchone()
                codTributo = rowTributo[0]
                #consultamos a operação fiscal e a incluímos, se não existir
                cursor.execute(selectOpFiscal, (operacao,))
                rowOperacao = cursor.fetchone()
                if not rowOperacao:
                    cursor.execute(insertOpFiscal, (operacao, dfOp.iat[linha2, 9].upper(), float(valor))) #tirei o tributo
                    cursor.execute(selectOpFiscal, (operacao,))
                    rowOperacao = cursor.fetchone()
                codOperacao = rowOperacao[0]
                #inserimos a operação vinculada ao TDPF
                dfOpAux = dfOp.loc[(dfOp['Operação Fiscal Atual Código']==operacao) & (dfOp['Receita Programada(Tributo) Código']==tributo)]
                #selecionamos o menor e o maior mês do período da operação deste TDPF            
                if dfOpAux.shape[0]>0:
                    perInicial = paraData(dfOpAux.loc[dfOpAux['Mês Início'].idxmin()]["Mês Início"])
                    perFinal = paraData(dfOpAux.loc[dfOpAux['Mês Fim'].idxmax()]["Mês Fim"])            
                cursor.execute(insereOperacao, (chaveTdpf, codOperacao, perInicial, perFinal, codTributo)) #incluí o tributo
        conn.commit()
    if termina:
        return
    #atualizamos a tabela de supervisões de grupos/equipes fiscais (Supervisores)
    comando = "Select Distinctrow Grupo from TDPFS"
    cursor.execute(comando)
    gruposRows = cursor.fetchall()
    tabGrupos = 0
    tabGruposAtu = 0
    atualizouSuperv = False
    print("Processados "+str(dfTdpf.shape[0])+" TDPFS.")
    print("Atualizando supervisores ...")
    superv = 0 #número de supervisores que não fazem parte de nenhum grupo - são incluídos na tabela de usuários pq supervisionam ativamente alguma equipe
    for grupoRow in gruposRows:
        df = dfSupervisores.loc[dfSupervisores['Grupo Fiscal']==grupoRow[0]].sort_values(by=['R028_DT_INI_VINCULO'], ascending=False)
        df = df.reset_index(drop=True)       
        if df.shape[0]>0: #pega só o último registro da supervisão da equipe (mais recente início na supervisão)
            cpf = df.at[0, 'CPF']
            if not "STR" in str(type(cpf)):
                cpf = str(cpf).rjust(11,"0")
            nomeSuperv = df.iat[0, 8]
            dataIni = df.iat[0, 9]
            dataFim = df.iat[0, 10]
            #selectFisc = "Select Codigo, CPF, Nome from Fiscais Where CPF=%s"
            #insereFisc = "Insert Into Fiscais (CPF, Nome) Values (%s, %s)"
            cursor.execute(selectFisc, (cpf,))     
            fiscalRow = cursor.fetchone()
            if not fiscalRow:
                cursor.execute(insereFisc, (cpf, nomeSuperv))
                cursor.execute(selectFisc, (cpf,))     
                fiscalRow = cursor.fetchone()
                atualizouSuperv = True  
            chaveFiscal = fiscalRow[0]   
            if dataFim==None or dataFim=="" or dataFim==np.nan or pd.isna(dataFim) or dataFim==float(0): #é supervisor da equipe
                #print(grupoRow[0]+" - "+cpf)                     
                comando = "Select Codigo, Fiscal, Inicio, Fim from Supervisores Where Equipe=%s and Fiscal=%s and Fim Is Null"
                cursor.execute(comando, (grupoRow[0], chaveFiscal))
                supervisoresRows = cursor.fetchall()
                bAchou = True
                if supervisoresRows==None:
                    bAchou = False
                elif len(supervisoresRows)==0:
                    bAchou = False
                #print(bAchou)    
                if not bAchou: #ainda não consta da tabela de Supervisores
                    #verificamos se este grupo não tem outro supervisor ativo - se tiver, colocamos a data final - fazemos isso para garantir caso haja uma descontinuidade
                    #não obtida pelo else abaixo
                    comando = "Select Codigo from Supervisores Where Equipe=%s and Fim Is Null"
                    cursor.execute(comando, (grupoRow[0], ))                    
                    supervisoresRows = cursor.fetchall()
                    if supervisoresRows!=None:
                        if len(supervisoresRows)>0: #há supervisores (antigos) ativos da equipe
                            comando = "Update Supervisores Set Fim=%s Where Equipe=%s and Fim Is Null" #"matamos" os antigos supervisores
                            cursor.execute(comando, (datetime.now().date(), grupoRow[0]))                    
                    comando = "Insert Into Supervisores (Equipe, Fiscal, Inicio) Values (%s, %s, %s)" #inserimos o novo supervisor
                    tabGrupos+=1
                    cursor.execute(comando, (grupoRow[0], chaveFiscal, paraData(dataIni)))

                #verificamos se este supervisor consta da tabela de usuários
                dfFiscal = dfFiscais.loc[dfFiscais['CPF']==cpf]
                email = None
                if dfFiscal.shape[0]>0: 
                    if dfFiscal.iat[0, 4]!=np.nan and not pd.isna(dfFiscal.iat[0,4]) and dfFiscal.iat[0, 4]!="": #email está na coluna 4 (coluna 'E' do Excel)
                        email = dfFiscal.iat[0, 4]                                   
                    cursor.execute(selectUsuario, (cpf,))
                    rows = cursor.fetchall()
                    bAchou = True
                    if rows==None:
                        bAchou = False
                    elif len(rows)==0:
                        bAchou = False
                    if not bAchou:  #não existe o supervisor na tabela de usuários             
                        cursor.execute(insereUsuario, (cpf, email)) 
                        atualizouSuperv = True                               
            else: #não é mais supervisor da equipe
                comando = "Select Codigo, Fiscal, Inicio, Fim from Supervisores Where Equipe=%s and Fiscal=%s and Inicio=%s and Fim Is Null"
                cursor.execute(comando, (grupoRow[0], chaveFiscal, paraData(dataIni)))
                supervisoresRows = cursor.fetchall()
                bAchou = True
                if supervisoresRows==None:
                    bAchou = False
                elif len(supervisoresRows)==0:
                    bAchou = False
                if bAchou: #colocamos um final na supervisão do titular e do substituto (o substituto sempre termina quando o titular acaba tb)
                    comando = "Update Supervisores Set Fim=%s Where Codigo=%s" #titular
                    cursor.execute(comando,(paraData(dataFim), supervisoresRows[0][0]))
                    comando = "Update Supervisores Set Fim=%s Where Titular=%s" #substituto
                    cursor.execute(comando,(paraData(dataFim), supervisoresRows[0][0]))
                    tabGruposAtu+=1
        else:
            logging.info("Grupo não encontrado: "+grupoRow[0])

    if tabGrupos>0 or tabGruposAtu>0 or atualizouSuperv or atualizou:
        try:
            #cursor.execute("Insert Into Extracoes (Data) Values (%s)", (datetime.now(),))
            conn.commit()  
            logging.info("Registros Incluídos:")  
            logging.info(f"TDPFs: {tabTdpfs}")
            logging.info(f"Fiscais: {tabFiscais}")
            logging.info(f"Ciencias: {tabCiencias}")
            logging.info(f"Alocacoes: {tabAloc}")
            logging.info(f"Usuarios: {tabUsuarios}")  
            logging.info(f"Equipes: {tabGrupos}")   

            logging.info("Registros Atualizados:")
            logging.info(f"TDPFs: {tabTdpfsAtu}")
            logging.info(f"Grupos(TDPFs)/Vencimentos: {gruposAtu}")
            logging.info(f"Alocacoes: {tabAlocAtu}")
            logging.info(f"Usuarios: {tabUsuariosAtu}")
            logging.info(f"Equipes: {tabGruposAtu}")   
            textoUNacionais = "Carga de dados (TDPFs, Alocações e Operações) efetuada na base de dados - "+datetime.now().strftime("%d/%m/%Y %H:%M") +"\nRegistros Incluídos:\n" 
            textoUNacionais += f"TDPFs: {tabTdpfs}\n" + f"Fiscais: {tabFiscais}\n" + f"Ciencias: {tabCiencias}\n" + f"Alocacoes: {tabAloc}\n"
            textoUNacionais += f"Usuarios: {tabUsuarios}\n"
            textoUNacionais += f"Equipes: {tabGrupos}\nRegistros Atualizados:\n"
            textoUNacionais += f"TDPFs: {tabTdpfsAtu}\n" + f"Grupos(TDPFs)/Vencimentos: {gruposAtu}\n" + f"Usuarios: {tabUsuariosAtu}\n" + f"Equipes: {tabGruposAtu}\n"
            print("TDPFs, Alocacoes, Ciências e Supervisores/Equipes foram atualizados")          
        except:
            textoUNacionais = "Sr. Usuário,\n\nA tentativa de realizar a carga de dados (TDPFs, Alocações e Operações) FALHOU - "+datetime.now().strftime("%d/%m/%Y %H:%M") + "\n\nAtenciosamente,\n\nCofis/Disav"
            print("Erro ao tentar efetivar as atualizações no Banco de Dados - É necessário verificar o erro e tentar fazer novamente a carga.")
            logging.info("Erro ao tentar efetivar as atualizações no Banco de Dados - Nenhum dado foi atualizado")
            avisaUsuariosNacionais(textoUNacionais, cursor)
            conn.rollback()
            conn.close()
            return 
    #atualizando as equipes
    print("Iniciando a atualização das equipes ...")
    for linha in range(dfEquipes.shape[0]):
        equipe = dfEquipes.iat[linha, 0]
        if equipe==None:
            continue
        equipe = getAlgarismos(equipe.strip())
        if equipe=="":
            continue
        equipe = equipe.rjust(14,"0")
        nomeEquipe = dfEquipes.iat[linha, 1].strip().upper()
        UL = dfEquipes.iat[linha, 2].strip().upper()
        qtdeRH = dfEquipes.iat[linha, 3]
        if qtdeRH==np.nan or pd.isna(qtdeRH) or qtdeRH==None:
            qtdeRH = 0
        qtdeRH = int(qtdeRH)
        cursor.execute("Select Codigo From Equipes Where Equipe=%s", (equipe,))
        rowEquipe = cursor.fetchone()
        if rowEquipe: #equipe existe - atualizamos informações
            codigoEquipe = rowEquipe[0]
            cursor.execute("Update Equipes Set Nome=%s, UL=%s, QtdeRH=%s Where Codigo=%s", (nomeEquipe, UL, qtdeRH, codigoEquipe))
        else: #não existe - incluímos a equipe
            cursor.execute ("Insert Into Equipes (Equipe, Nome, UL, QtdeRH) Values(%s, %s, %s, %s)", (equipe, nomeEquipe, UL, qtdeRH)) 
            
    print("Iniciando atualização do indicador de supervisor nos TDPFs")
    select = """Select Alocacoes.Codigo From Alocacoes, TDPFS, Supervisores 
                Where Alocacoes.TDPF=TDPFS.Codigo and TDPFS.Encerramento Is Null and 
                Supervisores.Equipe=TDPFS.Grupo and Supervisores.Fim Is Null and Supervisores.Fiscal=Alocacoes.Fiscal and Alocacoes.Desalocacao Is Null"""
    cursor.execute(select)
    rows = cursor.fetchall()
    lista = ""
    for row in rows:
        if lista!="":
            lista = lista +", "
        lista = lista + str(row[0])
    if lista!="":
        lista = "(" + lista +")"
        comando = "Update Alocacoes Set Supervisor='N' Where Alocacoes.Supervisor='S' and Alocacoes.Codigo Not In " + lista
        cursor.execute(comando) #indicador de supervisor 'N' nos TDPFS
        comando = "Update Alocacoes Set Supervisor='S' Where Alocacoes.Supervisor='N' and Alocacoes.Codigo In " + lista
        cursor.execute(comando) #indicador de supervisor 'S' nos TDPFS
        #comando = "Update Alocacoes Set Supervisor='N' Where Alocacoes.Desalocacao Is Not Null and Alocacoes.Supervisor='S'" 
        #fiscal desalocado não é supervisor
        #cursor.execute(comando)
        try:
            conn.commit()           
            print("Supervisores não alocados a TDPFs e incluídos: "+str(superv))                      
        except:
            print("Erro ao tentar efetivar as atualização do indicador de supervisor dos TDPFs - É necessário verificar o erro e tentar fazer novamente a carga.")
            logging.info("Erro ao tentar efetivar as atualização do indicador de supervisor dos TDPFs")
            conn.rollback()        
            logging.error(mysql.connector.Error)
            print(mysql.connector.Error)      
    try:
        os.rename(dirExcel+"TDPFSRPFS.xlsx", dirExcel+"TDPFSRPFS_Processado_"+datetime.now().strftime('%Y-%m-%d')+".xlsx")
        os.rename(dirExcel+"ALOCACOESRPFS.xlsx", dirExcel+"ALOCACOESRPFS_Processado_"+datetime.now().strftime('%Y-%m-%d')+".xlsx")
        logging.info("Arquivos renomeados")
        if tabGrupos>0 or tabGruposAtu>0 or atualizouSuperv or atualizou:
            textoUNacionais += "Arquivos foram renomeados."
    except:
        print("Erro ao tentar renomear os arquivos")            
        logging.error("Erro ao tentar renomear os arquivos")   
        if tabGrupos>0 or tabGruposAtu>0 or atualizouSuperv or atualizou:
            textoUNacionais += "Houve um erro ao tentar renomear os arquivos."                       
    print("Carga finalizada ", datetime.now())
    if tabGrupos>0 or tabGruposAtu>0 or atualizouSuperv or atualizou:
        textoUNacionais = "Sr. Usuário,\n\n" + textoUNacionais + "\n\nAtenciosamente,\n\nCofis/Disav"
        avisaUsuariosNacionais(textoUNacionais, cursor) #avisamos os usuários nacionais da realização da carga
    avisaUsuariosRegionais(conn) #avisamos usuários regionais dos TDPFs vincendos em curto prazo (pedido da Débora Difis07)
    cursor.close()
    conn.close() 
    return


def avisaUsuariosNacionais(texto, cursor): #avisamos os usuários nacionais do resultado da carga
    AMBIENTE = os.getenv("AMBIENTE", "TESTE") 
    texto = texto + "\n\nAmbiente: "+AMBIENTE
    consultaUNacionais = "Select email from Usuarios Inner Join Orgaos On Usuarios.Orgao=Orgaos.Codigo Where Usuarios.Orgao Is Not Null and Orgaos.Tipo='N'" 
    cursor.execute(consultaUNacionais)   
    rows = cursor.fetchall()
    #avisamos todos os usuários nacionais que houve uma carga de dados
    for row in rows:
        if AMBIENTE=="PRODUÇÃO":
            if enviaEmail(row[0], texto,"Carga de Dados Alertas Fiscalização")!=3: #se der erro
                logging.info("Erro no envio do e-mail para usuário NACIONAL "+row[0])   
                print("Erro no envio do e-mail para usuário NACIONAL "+row[0]) 
        else:
            print(row[0], texto)
    return

def avisaUsuariosRegionais(conn): #avisamos usuários regionais dos TDPFs vincendos em curto prazo (4 a 10 dias - vai apenas um aviso, portanto, pois a carga é semanal)
    AMBIENTE = os.getenv("AMBIENTE", "TESTE") 
    if not AMBIENTE in ["PRODUÇÃO", "TESTE"]:
        return
    cursor = conn.cursor(buffered=True)
    consultaULocais = """Select email, Usuarios.Orgao, Orgaos.Orgao from Usuarios, Orgaos 
                         Where Usuarios.Orgao Is Not Null and Usuarios.Orgao<>0 and Usuarios.Orgao=Orgaos.Codigo and email Is Not Null and Orgaos.Tipo='R' Order by email"""
    cursor.execute(consultaULocais)
    rows = cursor.fetchall()
    consultaTdpfs = """
                  Select Distinctrow TDPFS.Codigo, TDPFS.Numero, TDPFS.Grupo, Fiscais.Nome, TDPFS.Vencimento, TDPFS.Emissao from Orgaos, TDPFS, Jurisdicao, Supervisores, Fiscais
                  Where TDPFS.Encerramento Is Null and Jurisdicao.Orgao=%s and Jurisdicao.Equipe=TDPFS.Grupo and Supervisores.Equipe=Jurisdicao.Equipe 
                  and Supervisores.Fim Is Null and Fiscais.Codigo=Supervisores.Fiscal and Supervisores.Titular Is Null and
                  TDPFS.Emissao<cast((now() - interval 180 day) as date) and TDPFS.Tipo='F' and
                  (TDPFS.Vencimento>=cast((now() + interval 4 day) as date) and TDPFS.Vencimento<=cast((now() + interval 10 day) as date)) 
                  Order by TDPFS.Grupo, TDPFS.Vencimento,TDPFS.Numero"""  
                  #TDPFs Fiscalização que vencem de 4 a 10 dias (avisamos apenas uma vez, pois na próxima carga estará vencido, faltará menos de 4 dias ou terá sido renovado)
                  #mas somente TDPFs emitidos há mais de 180 dias (60 dias de margem), pois só nos interessa da 2a prorrogação em diante
    consultaAviso = "Select Codigo, Data from AvisosVencimentoDifis Where TDPF=%s"    
    insere = set()
    atualiza = set()              
    for row in rows:
        total = 0
        email = row[0]
        cursor.execute(consultaTdpfs, (row[1], ))
        tdpfs = cursor.fetchall()
        equipeAnt = ""
        texto = ""
        i = 0
        #planilha a ser encaminhada anexa ao e-mail
        book = Workbook()
        sheet = book.active  
        sheet.cell(row=1, column=1).value = "Nº Ordem"
        sheet.cell(row=1, column=2).value = "Nº Ord Equipe"
        sheet.cell(row=1, column=3).value = "Equipe"
        sheet.cell(row=1, column=4).value = "Supervisor"
        sheet.cell(row=1, column=5).value = "TDPF"
        sheet.cell(row=1, column=6).value = "Emissão"
        sheet.cell(row=1, column=7).value = "Vencimento"
        #sheet.row_dimensions[1].height = 42    
        larguras = [13, 16, 20, 35, 20, 15, 15]
        for col in range(len(larguras)):
            sheet.column_dimensions[get_column_letter(col+1)].width = larguras[col]  
            currentCell = sheet.cell(row=1, column=col+1)
            currentCell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
            currentCell.font = Font(bold=True)       
        for tdpf in tdpfs:
            chaveTdpf = tdpf[0]
            numero = tdpf[1]            
            cursor.execute(consultaAviso, (chaveTdpf,))
            aviso = cursor.fetchone()
            if aviso:
                if (aviso[1].date()+timedelta(days=6))>=datetime.now().date(): #não se passaram 7 dias desde o último aviso
                    continue #desprezamos este TDPF, pois já foi avisado há menos de uma semana
                else:
                    atualiza.add(aviso[0])
            else:
                insere.add(chaveTdpf)
            equipe = tdpf[2].strip()
            nome = tdpf[3]
            vencimento = tdpf[4]
            emissao = tdpf[5]
            if equipeAnt!=equipe:
                equipeAnt = equipe
                if texto!="":
                    texto += "\n"                
                texto += "Equipe: "+equipe[:7]+"."+equipe[7:11]+"."+equipe[11:]+" - Supervisor: "+nome+"\n"
                i = 0
            total+=1
            i+=1
            sheet.cell(row=total+1, column=1).value = total
            sheet.cell(row=total+1, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            sheet.cell(row=total+1, column=2).value = i
            sheet.cell(row=total+1, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)             
            sheet.cell(row=total+1, column=3).value = equipe[:7]+"."+equipe[7:11]+"."+equipe[11:]  
            sheet.cell(row=total+1, column=3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)           
            sheet.cell(row=total+1, column=4).value = nome
            sheet.cell(row=total+1, column=5).value = numero[:7]+"."+numero[7:11]+"."+numero[11:]
            sheet.cell(row=total+1, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)           
            sheet.cell(row=total+1, column=6).value = emissao.strftime("%d/%m/%Y")
            sheet.cell(row=total+1, column=6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)           
            sheet.cell(row=total+1, column=7).value = vencimento.strftime("%d/%m/%Y")
            sheet.cell(row=total+1, column=7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)           
            texto = texto +"  "+str(i)+") "+numero[:7]+"."+numero[7:11]+"."+numero[11:]+" - "+vencimento.strftime("%d/%m/%Y")+" - "+emissao.strftime("%d/%m/%Y")+"\n"
        if texto!="":
            texto = "Sr. Usuário Regional,\n\nEstamos enviando abaixo a relação de TDPFs de sua região com vencimento entre 4 e 10 dias (Nº - Vencimento - Emissão):\n\n"+texto
            texto += "\nTotal de TDPFs: "+str(total)+"\n"
            texto += "\nSegue, em anexo, planilha Excel contendo relação destes TDPFs.\n"
            texto += "\nDevido a restrições do DW, o vencimento no Ação Fiscal pode estar um pouco mais distante. Ressaltamos também que a carga de dados neste serviço ocorre semanalmente.\n"
            texto += "\nAtenciosamente,\n\nCofis/Disav"
            nomeArq = "Regiao_"+row[2].strip()+"_"+datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")[:-3]+".xlsx"
            book.save(nomeArq)     
            if AMBIENTE=="PRODUÇÃO":         
                if enviaEmail(email, texto, "TDPFs Vincendos Entre 4 e 10 Dias - Região", nomeArq)!=3:
                    print("Erro no envio do e-mail para usuário REGIONAL "+email)
                    logging.info("Erro no envio do e-mail para usuário REGIONAL "+email) 
                os.remove(nomeArq) 
            else:
                print(texto)                
    insereTuplas = []
    atualizaTuplas = []
    for chaveTdpf in insere:
        insereTuplas.append((chaveTdpf, datetime.now().date()))
    for codigo in atualiza:
        atualizaTuplas.append((datetime.now().date(), codigo))
    try:        
        cursor.executemany("Update AvisosVencimentoDifis Set Data=%s Where Codigo=%s", atualizaTuplas)
        cursor.executemany("Insert Into AvisosVencimentoDifis (TDPF, Data) Values (%s, %s)", insereTuplas)
        conn.commit()
        logging.info("Tabela AvisosVencimentoDifis foi atualizada com sucesso.")
    except:
        conn.rollback()
        logging.info("Atualização da tabela AvisosVencimentoDifis FALHOU!")
    return

def realizaCargaDCCs():
    global dirExcel, termina, hostSrv, hora2
    try:
        dfDCCs= pd.read_excel(dirExcel+"DCCS.xlsx", dtype={'DCC':object, 'Data':object})
    except:
        print("Erro no acesso ao arquivo de DCCS.xlsx ou só há arquivos já processados; outra tentativa será feita às "+hora2)
        logging.info("Arquivo Excel de DCCs não foi encontrado; outra tentativa será feita às "+hora2) 
        return
    dfDCCs['DCC']=dfDCCs['DCC'].astype(str)
    dfDCCs['Data']=dfDCCs['Data'].astype(str)
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 

    try:
        logging.info("Conectando ao servidor de banco de dados (2)...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql (2)!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD (2) - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s) (2).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe (2).")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados (2)")
        return
    print("Realizando carga dos DCCs em "+datetime.now().strftime("%d/%m/%Y %H:%M"))
    cursor = conn.cursor(buffered=True)   
    atualizou = False 
    for linha in range(dfDCCs.shape[0]):
        dcc = dfDCCs.iat[linha,0]        
        dataJuntada = dfDCCs.iat[linha,1]
        if dataJuntada=="HOJE":
            data = datetime.now()
        elif dataJuntada=="NULO":
            data = None
        else:
            data = paraData(dataJuntada)        
        comando = "Select Codigo From TDPFS Where DCC=%s"
        cursor.execute(comando, (dcc,))
        row = cursor.fetchone()
        if not row:
            continue
        if row[0]==None:
            continue
        tdpf = row[0]
        comando = "Select Codigo, Solicitacao from Juntadas Where TDPF=%s"
        cursor.execute(comando, (tdpf,))
        row = cursor.fetchone()
        if row==None:
            comando = "Insert Into Juntadas (TDPF, Solicitacao) Values (%s, %s)"
            cursor.execute(comando, (tdpf, data))
            atualizou = True
        else:
            codigo = row[0]
            ultJuntada = row[1]
            comando = "Update Juntadas Set Solicitacao=%s Where Codigo=%s"
            if ultJuntada==None and data!=None:
                cursor.execute(comando, (data, codigo))
                atualizou = True
            elif ultJuntada!=None and dataJuntada=="HOJE":
                pass
            elif ultJuntada==None and data==None:
                pass
            elif ultJuntada!=None and data==None:
                cursor.execute(comando, (data, codigo))
                atualizou = True                
            elif ultJuntada.date()<data.date():
                cursor.execute(comando, (data, codigo))
                atualizou = True
    bErro = False
    if atualizou:
        try:
            conn.commit()
            logging.info("Tabela de Juntadas atualizada")             
        except:
            conn.rollback()
            logging.error("Erro ao tentar atualizar a tabela de Juntadas")
            logging.error(mysql.connector.Error)
            print(mysql.connector.Error)              
            bErro = True
    if not bErro:
        try:
            os.rename(dirExcel+"DCCS.xlsx", dirExcel+"DCCS_Processado_"+datetime.now().strftime('%Y-%m-%d')+".xlsx")
            logging.info("Arquivo de DCCs renomeado")
        except:
            print("Erro ao tentar renomear o arquivo de DCCs")         
            logging.error("Erro ao tentar renomear o arquivo de DCCs")         
    print("Carga das Juntadas dos DCCs finalizada - ", datetime.now())
    conn.close()
    return

def realizaCargaCienciasPendentes():
    global dirExcel, termina, hostSrv, hora3
    try:
        dfCiencias= pd.read_excel(dirExcel+"CIENCIASPENDENTES.xlsx", 
                                  dtype={'RPF Gerencial - Num.':str, 'Proc. Doc. Lanç. - Numero Inf.': str, 
                                         'Dia Registro Doc. Lançamento':datetime, 'Data Última Extração': datetime})
    except:
        print("Erro no acesso ao arquivo de CIENCIASPENDENTES.xlsx ou só há arquivos já processados; outra tentativa será feita às "+hora3)
        logging.info("Arquivo Excel de CIENCIASPENDENTES não foi encontrado. Outra ciência será tentada às "+hora3) 
        return
    #dfCiencias['Proc. Doc. Lanç. - Numero Inf.']=dfCiencias['Proc. Doc. Lanç. - Numero Inf.'].astype(str)
    print(dfCiencias.dtypes)
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 

    try:
        logging.info("Conectando ao servidor de banco de dados (3)...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql (3)!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD (2) - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s) (3).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe (3).")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados (3)")
        return
    print("Realizando carga das Ciências Pendentes em "+datetime.now().strftime("%d/%m/%Y %H:%M"))   
    cursor = conn.cursor(buffered=True)   
    for linha in range(dfCiencias.shape[0]):   
        tdpf = getAlgarismos(dfCiencias.iat[linha, 0])
        processo = dfCiencias.iat[linha, 2]
        registro = dfCiencias.iat[linha, 6]
        extracao = dfCiencias.iat[linha, 10]  
        #print(extracao, " - ", print(paraData(extracao)))
        meio = dfCiencias.iat[linha, 9] #se o meio estiver preenchido, foi feita a ciência no SIEF Processos
        consultaTdpf = "Select Codigo from TDPFS Where Numero=%s"
        cursor.execute(consultaTdpf, (tdpf,))
        row = cursor.fetchone()
        if not row: #TDPF não encontrado - não interessa
            continue
        chaveTdpf = row[0]
        consultaProcesso = "Select Codigo, Finalizado from AvisosCiencia Where TDPF=%s and Processo=%s"
        cursor.execute(consultaProcesso, (chaveTdpf, processo))
        row = cursor.fetchone()
        if not row and meio in [None, np.nan, ""]: #registro não existe - devemos incluí-lo se meio não estiver preenchido (se meio estiver preenchido, houve ciência)
            #print(chaveTdpf, " - ", processo)
            insere = "Insert Into AvisosCiencia (TDPF, Processo, Integracao, Extracao) Values (%s, %s, %s, %s)"
            try:
                cursor.execute(insere, (chaveTdpf, processo, paraData(registro), paraData(extracao)))
            except:
                print("Processo: " + processo + " - TDPF: "+tdpf)
                logging.error(mysql.connector.Error)
                logging.info("Processo: "+ processo + " - TDPF: "+tdpf)
                print(mysql.connector.Error)
        elif row and not meio in  [None, np.nan, ""]: #meio está preenchido - temos que finalizar o registro existente
            finalizaSQL = "Update AvisosCiencia Set Finalizado=%s, Extracao=%s Where Codigo=%s"
            cursor.execute(finalizaSQL, (datetime.now(), paraData(extracao), row[0]))
        elif row:
            extracaoSQL = "Update AvisosCiencia Set Extracao=%s Where Codigo=%s" #atualizamos a data de extração (a ciência ainda não foi registrada)
            cursor.execute(extracaoSQL, (paraData(extracao), row[0]))
    cursor.execute("Select MAX(Extracao) from AvisosCiencia")
    row = cursor.fetchone() #buscamos a maior data de extração registrada na tabela
    if row:
        extracao = row[0]
        #print("Extração (Max): ", extracao)
        #finalizamos todos os registros cuja extração se deu em data anterior à última extração (não foram atualizados acima)
        cursor.execute("Update AvisosCiencia Set Finalizado=%s Where Extracao<%s", (datetime.now(), extracao.date()))
    try:
        conn.commit()
        logging.error("Tabela de AvisosCiencia foi atualizada.")
        print("Tabela de AvisosCiencia foi atualizada - ", datetime.now())
        try:
            os.rename(dirExcel+"CIENCIASPENDENTES.xlsx", dirExcel+"CIENCIASPENDENTES_Processado_"+datetime.now().strftime('%Y-%m-%d')+".xlsx")
            logging.info("Arquivo de CIENCIASPENDENTES renomeado")
        except:
            print("Erro ao renomear arquivo de Ciências Pendentes")
            logging.error("Erro ao tentar renomear o arquivo de CIENCIASPENDENTES")                     
    except:
        conn.rollback()
        logging.error("Erro ao tentar atualizar a tabela de AvisosCiencia")
        print("Erro ao tentar atualizar a tabela de AvisosCiencia - ", datetime.now())
        logging.error(mysql.connector.Error)
        print(mysql.connector.Error)         
    return

def realizaCargaIndicadores(): #carga dos indicadores obtidos do Ação Fiscal na tabela Resultados
    global dirExcel, termina, hostSrv, hora4
    try:
        dfIndicadores= pd.read_excel(dirExcel+"INDICADORES.xlsx", dtype={'TDPF':object})
    except:
        print("Erro no acesso ao arquivo de INDICADORES.xlsx ou só há arquivos já processados; outra tentativa será feita às "+hora4)
        logging.info("Arquivo Excel de Indicadores não foi encontrado; outra tentativa será feita às "+hora4) 
        return
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 
    print(dfIndicadores.dtypes)
    try:
        logging.info("Conectando ao servidor de banco de dados (2)...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv, database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql (4)!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD (4) - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s) (4).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe (4).")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados (4)")
        return
    print("Realizando carga dos Indicadores em "+datetime.now().strftime("%d/%m/%Y %H:%M"))
    cursor = conn.cursor(buffered=True)   
    atualizou = False 
    consultaTdpf = "Select Codigo from TDPFS Where Numero=%s"
    consultaResultado = """Select Arrolamentos, MedCautelar, RepPenais, Inaptidoes, Baixas, ExcSimples, DigVincs from Resultados 
                           Where Resultados.TDPF=%s"""
    insere = "Insert Into Resultados (TDPF, Arrolamentos, MedCautelar, RepPenais, Inaptidoes, Baixas, ExcSimples, DigVincs, Data, SujPassivos, Situacao11, Interposicao, Situacao15, EstabPrev1, EstabPrev2, Segurados, Prestadores, Tomadores, QtdePER, LancMuldi, Compensacao, CreditoExt) Values (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 'N', 'N', 'N', 0, 0, 0, 0, 0, 0, 'N', 'N', 'N')"
    atualiza = "Update Resultados Set Arrolamentos=%s, MedCautelar=%s, RepPenais=%s, Inaptidoes=%s, Baixas=%s, ExcSimples=%s, DigVincs=%s Where TDPF=%s"
    incluidos = 0
    atualizados = 0
    for linha in range(dfIndicadores.shape[0]):
        if (linha+1)%300==0:
            print(linha+1, " processadas ...")
        tdpf = getAlgarismos(dfIndicadores.iat[linha, 0])
        cursor.execute(consultaTdpf, (tdpf,))
        row = cursor.fetchone()
        if row==None:
            #TDPF não encontrado - não há o que fazer
            continue
        chaveTdpf = row[0]
        parametros = []
        parametros.append(chaveTdpf)
        for i in range(1, 8, 1): #carrega os indicadores nesta lista
            parametros.append(int(dfIndicadores.iat[linha, i]))
            if i==2: #medida cautelar é sim (S) ou não (N)
                parametros[2] = "S" if dfIndicadores.iat[linha, 2]>0 else "N"
                
        #print(tdpf, parametros)
        cursor.execute(consultaResultado, (chaveTdpf,))
        row = cursor.fetchone()
        if row==None:
            parametros.append(datetime.now())
            cursor.execute(insere, tuple(parametros)) #não há a linha na tabela para o TDPF - fazemos sua inserção
            incluidos+=1
        else:
            parametros.pop(0)  #excluo a chaveTdpf da primeira posição para inclui-la na última, por conta da estrutura do SQL Update   
            #iria respeitar a informação outrora prestada, mas resolvi seguir a informação do Ação Fiscal (oficial), por isso comentei o for abaixo <--------
            #for i in range(7):
            #    if i==1:
            #        if parametros[1]=="N" and row[1]=="S": #medida cautelar é S ou N
            #            parametros[1] = "S" #se estava 'S', mantemos
            #    else:
            #        parametros[i] = max(parametros[i], row[i] if row[i]!=None else 0) #não reduzimos o valor já informado
            parametros.append(chaveTdpf)
            cursor.execute(atualiza, tuple(parametros))
            atualizados+=1
        atualizou = True
    print(linha+1, " processadas no total.")   
    print(incluidos, " registro incluídos.")
    print(atualizados, " registros atualizados.")         
    if atualizou:
        try:
            conn.commit()
            logging.error("Tabela de Resultados foi atualizada.")
            print("Tabela de Resultados foi atualizada - ", datetime.now())
            try:
                os.rename(dirExcel+"INDICADORES.xlsx", dirExcel+"INDICADORES_Processado_"+datetime.now().strftime('%Y-%m-%d')+".xlsx")
                logging.info("Arquivo de INDICADORES renomeado")
            except:
                print("Erro ao renomear o arquivo de Indicadores")
                logging.error("Erro ao tentar renomear o arquivo de INDICADORES")             
        except:
            conn.rollback()
            logging.error("Erro ao tentar atualizar a tabela de Resultados")
            print("Erro ao tentar atualizar a tabela de Resultados - ", datetime.now())
            logging.error(mysql.connector.Error)
            print(mysql.connector.Error)  
    else:
        print("Nenhuma atualização de indicadores ocorreu - ", datetime.now())
        logging.info("Nenhuma atualização de indicadores ocorreu.")
    return          

def realizaCargaCasosEspeciais(): #atualização da tabela TDPFs com casos especiais - é para rodar somente uma vez, por conta da inclusão do campo e da respectiva tabela
    global dirExcel, termina, hostSrv
    if datetime.now().date()!=datetime.strptime("20/05/2021", "%d/%m/%Y").date():
        print("Carga dos casos especiais deve ocorrer apenas em 20/05/2021")
        return
    try:
        dfTdpf = pd.read_excel(dirExcel+"TDPFS.xlsx", dtype={'Porte':object, 'Acompanhamento':object, 'Receita Programada(Tributo) Código': int})
    except:
        print("Erro no acesso ao arquivo TDPFS.xlsx - realizaCargaCasosEspeciais")
        logging.info("Erro no acesso ao arquivo TDPFS.xlsx - realizaCargaCasosEspeciais") 
        return
    MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "databasenormal")
    MYSQL_USER = os.getenv("MYSQL_USER", "my_user")
    MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "mypass1234") 
    try:
        logging.info("Conectando ao servidor de banco de dados ...")
        logging.info(MYSQL_DATABASE)
        logging.info(MYSQL_USER)

        conn = mysql.connector.connect(user=MYSQL_USER, password=MYSQL_PASSWORD,
                                    host=hostSrv,
                                    database=MYSQL_DATABASE)
        logging.info("Conexão efetuada com sucesso ao MySql!")                               
    except mysql.connector.Error as err:
        print("Erro na conexão com o BD - veja Log: "+datetime.now().strftime('%d/%m/%Y %H:%M'))
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            logging.info("Usuário ou senha inválido(s).")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            logging.error("Banco de dados não existe.")
        else:
            logging.error(err)
            logging.error("Erro na conexão com o Banco de Dados")
        return
    print("Realizando carga dos Casos Especiais em "+datetime.now().strftime("%d/%m/%Y %H:%M"))
    cursor = conn.cursor(buffered=True)

    selectCaso = "Select Codigo from CasosEspeciais Where CasoEspecial=%s"
    insereCaso = "Insert Into CasosEspeciais (CasoEspecial, Descricao) Values (%s, %s)"
                   
    if termina:
        return
    logging.info("Iniciando loop na carga.")
    atualizou = 0
    incluidos = 0
    for linha in range(dfTdpf.shape[0]): #percorre os TDPFs das planilhas Excel
        if (linha+1)%500==0:
            print("Processando TDPF nº "+str(linha+1)+" de "+str(dfTdpf.shape[0]))
        tdpfAux = dfTdpf.iat[linha,0]
        if tdpfAux=='9999999.9999.99999': #última linha, referente à data de extração dos dados
            break
        tdpf = getAlgarismos(tdpfAux)  
        #busca o caso especial, se houver, ou o insere     
        casoEspecial = dfTdpf.iat[linha, 19]
        if casoEspecial=="" or pd.isna(casoEspecial) or casoEspecial==None or casoEspecial==np.nan:
            continue #não há o que atualizar (não há caso especial)
        else:
            casoEspecialDesc = dfTdpf.iat[linha, 20].strip()
            casoEspecial = int(casoEspecial)
            cursor.execute(selectCaso, (casoEspecial, ))
            linhaCaso = cursor.fetchone()
            if linhaCaso:
                casoEspecialCod = linhaCaso[0]
            else:
                cursor.execute(insereCaso, (casoEspecial, casoEspecialDesc))
                cursor.execute(selectCaso, (casoEspecial, ))
                incluidos+=1
                linhaCaso = cursor.fetchone()
                if linhaCaso:
                    casoEspecialCod = linhaCaso[0]    
                else:
                    casoEspecialCod = 0            
        cursor.execute("Select Codigo from TDPFS Where Numero=%s", (tdpf,))
        regTdpf = cursor.fetchone()  
        if regTdpf: #TDPF consta da base
            chaveTdpf = regTdpf[0]  #chave do registro do TDPF  - para poder atualizar o registro, se for necessário 
        else:
            continue #nào há o que atualizar - TDPF não existe na base
        #atualiza a tabela de TDPFs com o caso especial                    
        cursor.execute("Update TDPFS Set CasoEspecial=%s Where Codigo=%s", (casoEspecialCod, chaveTdpf))
        atualizou+=1
    try:
        conn.commit()
        print("Foram atualizados "+str(atualizou)+" TDPFs com o respectivo caso especial.")
        print("Foram incluídos "+str(incluidos)+" na tabela CasosEspeciais")
    except:
        conn.rollback()
        print("Ocorreu algum erro ao realizar o commit da carga dos casos especiais.")
    return



def disparador():
    global termina, dirExcel
    while not termina:
        schedule.run_pending() 
        time.sleep(5) #aguardo um pouco para começar a rodar as tarefas, inclusive realizaCargaDados que apagará o arquivo abaixo
        if os.path.exists(dirExcel+"REALIZACARGA.TXT"): #a existência deste arquivo indica que é para fazer a carga dos dados imediatamente
            realizaCargaDados()   
        if os.path.exists(dirExcel+"REALIZACARGAPONTOS.TXT"): #a existência deste arquivo indica que é para fazer a carga dos dados imediatamente
            realizaCargaPontosSerpro()  
        if os.path.exists(dirExcel+"CARGADADOSAD.TXT"): #a existência deste arquivo indica que é para fazer a atualização das informações das equipes (info RD)
            atualizaEquipesDadosAd()
        if os.path.exists(dirExcel+"CARGAMETAS.TXT"): #a existência deste arquivo indica que é para fazer a atualização das informações das equipes (info RD)
            realizaCargaMetas()    
        if os.path.exists(dirExcel+"CARGAVINCULOS.TXT"): #a existência deste arquivo indica que é para fazer a atualização das informações das equipes (info RD)
            realizaCargaVinculos()            
        time.sleep(60*60) #a cada hora, vê o que tem de tarefa pendente
    return 
   
sistema = sys.platform.upper()

if "WIN32" in sistema or "WIN64" in sistema or "WINDOWS" in sistema:
    hostSrv = 'localhost'
    dirLog = 'log\\'
    dirExcel = 'Excel\\'
else:
    hostSrv = 'mysqlsrv'
    dirLog = '/Log/' 
    dirExcel = '/Excel/'

print("Bot Carga ativo.")
logging.basicConfig(filename=dirLog+datetime.now().strftime('%Y-%m-%d %H_%M')+' Carga'+sistema+'.log', format='%(asctime)s - %(message)s', level=logging.INFO)
hora1 = "09:30"
schedule.every().day.at(hora1).do(realizaCargaDados) #a cada 24 horas, verifica se há arquivos para fazer a carga
hora2 = "14:30"
schedule.every().day.at(hora2).do(realizaCargaDCCs)
hora3 = "12:00"
schedule.every().day.at(hora3).do(realizaCargaCienciasPendentes)
hora4 = "10:15"
#schedule.every().day.at(hora4).do(realizaCargaIndicadores)
hora5 = "11:00"
schedule.every().day.at(hora5).do(realizaCargaPontosSerpro)
schedule.every().wednesday.at("12:00").do(realizaCargaMetas)
termina = False
threadDisparador = threading.Thread(target=disparador, daemon=True) #encerra thread quando sair do programa sem esperá-la
threadDisparador.start()
#realizaCargaDados() #faz a primeira tentativa de carga das planilhas logo no acionamento do programa
if os.path.exists(dirExcel+"REALIZACARGA.TXT"): #a existência deste arquivo indica que é para fazer a carga dos dados imediatamente
    realizaCargaDados()  
#realizaCargaDCCs() #idem para os DCCs
#realizaCargaCienciasPendentes() #ciencias pendentes de AI 
###realizaCargaIndicadores() #carga de indicadores (parâmetros) dos TDPFs encerrados <--NÃO descomentar
#realizaCargaCasosEspeciais() #só vai executar no dia 20/05/2021
#atualizaEquipesDadosAd()    
if os.path.exists(dirExcel+"CARGADADOSAD.TXT"): #a existência deste arquivo indica que é para fazer a atualização das informações das equipes (info RD)
    atualizaEquipesDadosAd()  
realizaCargaMalha()
if os.path.exists(dirExcel+"CARGAMALHA.TXT"): #a existência deste arquivo indica que é para fazer a atualização dos pontos da malha (info RD)
    realizaCargaMalha()          
realizaCargaVinculos()  
if os.path.exists(dirExcel+"CARGAVINCULOS.TXT"): #a existência deste arquivo indica que é para fazer a atualização dos vínculos de fiscais com equipes (info RD)
    realizaCargaVinculos()  
realizaCargaPontosSerpro()
if os.path.exists(dirExcel+"REALIZACARGAPONTOS.TXT"): #a existência deste arquivo indica que é para fazer a carga dos dados imediatamente
    realizaCargaPontosSerpro()   
#realizaCargaMetas()
if os.path.exists(dirExcel+"CARGAMETAS.TXT"): #a existência deste arquivo indica que é para fazer a atualização das metas (info RD)
    realizaCargaMetas()      
while not termina:
    entrada = input("Digite QUIT para terminar o serviço Carga BOT: ")
    if entrada:
        if entrada.strip().upper()=="QUIT":
            termina = True
schedule.clear()        